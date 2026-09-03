"""EMPLOYEE PPE REGISTER — shoes, blankets, FRCs and coveralls issued to staff.

Purpose
=======
A practical staff-issue register for personal items that are often posted on
normal Delivery Notes but later need to be tracked by employee code:

  · Safety shoes / safety boots
  · Blankets
  · FRCs / flame-resistant clothing
  · Coveralls

The register is stored in its own SQLite file under the storage root, just like
Company Issuance and Cable Records. It does not move stock on its own. Instead,
it can IMPORT matching lines from the inventory Delivery Notes and keep a clean
employee-wise record for HR / camp / store follow-up.
"""
from __future__ import annotations

import csv
import datetime as _dt
import io
import re
import sqlite3
from pathlib import Path
from typing import Any, Sequence

from . import config
from .database import Database

FOLDER = "Employee PPE Register"
DB_NAME = "employee_ppe.db"
SCHEMA_VERSION = 2

GROUP_SHOES = "Safety Shoes"
GROUP_BLANKET = "Blanket"
GROUP_FRC = "FRC"
GROUP_COVERALL = "Coverall"
GROUP_OTHER = "Other PPE"
GROUPS = [GROUP_SHOES, GROUP_BLANKET, GROUP_FRC, GROUP_COVERALL, GROUP_OTHER]

ST_ISSUED = "Issued"
ST_RETURNED = "Returned"
ST_NEEDS_INFO = "Needs Employee Info"
STATUSES = [ST_ISSUED, ST_RETURNED, ST_NEEDS_INFO]

REPORT_LIST = [
    "Full PPE Register",
    "Safety Shoes Register",
    "Blanket Register",
    "FRC Register",
    "Coverall Register",
    "By Employee",
    "By Delivery Note",
    "Missing Employee Codes",
    "Synced Delivery Note PPE",
]

FIELDS = [
    ("issue_date", "Issue Date"),
    ("employee_code", "Employee Code"),
    ("employee_name", "Employee Name"),
    ("department", "Department"),
    ("project", "Project / Site"),
    ("item_group", "Item Group"),
    ("item_code", "Item Code"),
    ("item_desc", "Description"),
    ("size_text", "Size"),
    ("qty", "Qty"),
    ("uom", "UOM"),
    ("dn_no", "Delivery Note No."),
    ("return_date", "Return Date"),
    ("remarks", "Remarks"),
    ("status", "Status"),
]
LABELS = dict(FIELDS)
ALL_FIELDS = FIELDS
NUMERIC_FIELDS = {"qty"}
DATE_FIELDS = {"issue_date", "return_date"}

HEADER_MAP = {
    "date": "issue_date",
    "issuedate": "issue_date",
    "dateofissue": "issue_date",
    "dateofissuance": "issue_date",
    "issuedon": "issue_date",
    "employeecode": "employee_code",
    "employeeno": "employee_code",
    "empcode": "employee_code",
    "empid": "employee_code",
    "employeeid": "employee_code",
    "code": "employee_code",
    "employeename": "employee_name",
    "employee": "employee_name",
    "name": "employee_name",
    "staffname": "employee_name",
    "staff": "employee_name",
    "department": "department",
    "dept": "department",
    "project": "project",
    "projectsite": "project",
    "site": "project",
    "location": "project",
    "itemgroup": "item_group",
    "group": "item_group",
    "category": "item_group",
    "itemcode": "item_code",
    "codeitem": "item_code",
    "productcode": "item_code",
    "stockcode": "item_code",
    "description": "item_desc",
    "itemdescription": "item_desc",
    "item": "item_desc",
    "itemname": "item_desc",
    "size": "size_text",
    "sizetext": "size_text",
    "qty": "qty",
    "quantity": "qty",
    "uom": "uom",
    "unit": "uom",
    "dn": "dn_no",
    "dnno": "dn_no",
    "dnnumber": "dn_no",
    "deliverynote": "dn_no",
    "deliverynoteno": "dn_no",
    "deliverynotenumber": "dn_no",
    "returndate": "return_date",
    "dateofreturn": "return_date",
    "remarks": "remarks",
    "remark": "remarks",
    "notes": "remarks",
    "status": "status",
}

DDL = """
PRAGMA journal_mode=WAL;
PRAGMA foreign_keys=ON;

CREATE TABLE IF NOT EXISTS settings (
    key   TEXT PRIMARY KEY,
    value TEXT
);

CREATE TABLE IF NOT EXISTS records (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    issue_no        TEXT NOT NULL UNIQUE,
    issue_date      TEXT NOT NULL DEFAULT '',
    employee_code   TEXT DEFAULT '',
    employee_name   TEXT DEFAULT '',
    department      TEXT DEFAULT '',
    project         TEXT DEFAULT '',
    item_group      TEXT DEFAULT '',
    item_code       TEXT DEFAULT '',
    item_desc       TEXT DEFAULT '',
    size_text       TEXT DEFAULT '',
    qty             REAL NOT NULL DEFAULT 0,
    uom             TEXT DEFAULT '',
    dn_no           TEXT DEFAULT '',
    doc_date        TEXT DEFAULT '',
    pdf_path        TEXT DEFAULT '',
    source_type     TEXT NOT NULL DEFAULT 'MANUAL',
    source_doc_id   INTEGER,
    source_line_id  INTEGER,
    status          TEXT NOT NULL DEFAULT 'Issued',
    return_date     TEXT DEFAULT '',
    issued_by       TEXT DEFAULT '',
    remarks         TEXT DEFAULT '',
    batch_id        INTEGER,
    created_by      TEXT DEFAULT '',
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    updated_at      TEXT DEFAULT (datetime('now','localtime')),
    UNIQUE(source_type, source_doc_id, source_line_id)
);
CREATE INDEX IF NOT EXISTS ix_ppe_emp   ON records(employee_code, employee_name);
CREATE INDEX IF NOT EXISTS ix_ppe_group ON records(item_group);
CREATE INDEX IF NOT EXISTS ix_ppe_dn    ON records(dn_no);
CREATE INDEX IF NOT EXISTS ix_ppe_date  ON records(issue_date);
CREATE INDEX IF NOT EXISTS ix_ppe_stat  ON records(status);
CREATE INDEX IF NOT EXISTS ix_ppe_batch ON records(batch_id);

CREATE TABLE IF NOT EXISTS batches (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    ts         TEXT DEFAULT (datetime('now','localtime')),
    source     TEXT DEFAULT '',
    rows       INTEGER DEFAULT 0,
    skipped    INTEGER DEFAULT 0,
    username   TEXT DEFAULT ''
);

CREATE TABLE IF NOT EXISTS audit (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    ts         TEXT DEFAULT (datetime('now','localtime')),
    username   TEXT DEFAULT '',
    action     TEXT NOT NULL,
    entity     TEXT DEFAULT '',
    entity_id  TEXT DEFAULT '',
    details    TEXT DEFAULT ''
);
CREATE INDEX IF NOT EXISTS ix_ppe_audit ON audit(ts);
"""

_DATE_PATTERNS = (
    "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y",
    "%Y/%m/%d", "%d-%b-%Y", "%d %b %Y", "%b %d, %Y", "%d-%b-%y",
    "%d-%m-%y", "%d/%m/%y", "%Y-%m-%d %H:%M:%S",
)


def _now() -> str:
    return _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def today() -> str:
    return _dt.date.today().isoformat()


def db_path() -> Path:
    return config.folder(FOLDER) / DB_NAME


def norm(text: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(text or "").lower()).strip()


def norm_key(text: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(text or "").lower())


def to_float(v: Any) -> float:
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    t = str(v).strip().lower()
    if t in ("-", "nil", "none", "n/a", "na"):
        return 0.0
    try:
        return float(re.sub(r"[^\d.\-]", "", t) or 0)
    except ValueError:
        return 0.0


def to_date(v: Any) -> str:
    if v in (None, ""):
        return ""
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    t = str(v).strip()
    if not t or t in ("-", "N/A", "n/a"):
        return ""
    for f in _DATE_PATTERNS:
        try:
            d = _dt.datetime.strptime(t, f).date()
            if d.year < 100:
                d = d.replace(year=d.year + 2000)
            return d.isoformat()
        except ValueError:
            continue
    try:
        n = float(t)
        if 20000 < n < 60000:
            return (_dt.date(1899, 12, 30) + _dt.timedelta(days=int(n))).isoformat()
    except ValueError:
        pass
    return t


class PPEIssueDB:
    def __init__(self, path: str | Path | None = None, current_user: str = "admin"):
        self.path = Path(path or db_path())
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(str(self.path), check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self.current_user = current_user or "admin"
        self.conn.executescript(DDL)
        self.conn.commit()
        self._apply_schema_fixes()
        self.set_setting("schema_version", str(SCHEMA_VERSION))

    def _apply_schema_fixes(self) -> None:
        cols = {str(r[1]) for r in self.query("PRAGMA table_info(records)")}
        if "batch_id" not in cols:
            self.execute("ALTER TABLE records ADD COLUMN batch_id INTEGER")
            self.conn.commit()

    def close(self) -> None:
        self.conn.close()

    def execute(self, sql: str, params: Any = ()):
        return self.conn.execute(sql, params)

    def query(self, sql: str, params: Any = ()) -> list[sqlite3.Row]:
        return list(self.conn.execute(sql, params))

    def one(self, sql: str, params: Any = ()):
        return self.conn.execute(sql, params).fetchone()

    def scalar(self, sql: str, params: Any = (), default: Any = 0) -> Any:
        row = self.one(sql, params)
        if row is None:
            return default
        if isinstance(row, sqlite3.Row):
            return row[0]
        return row[0] if row else default

    def commit(self) -> None:
        self.conn.commit()

    def set_setting(self, key: str, value: Any) -> None:
        self.execute(
            "INSERT INTO settings(key,value) VALUES(?,?)"
            " ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, str(value)),
        )
        self.conn.commit()

    def get_setting(self, key: str, default: str = "") -> str:
        r = self.one("SELECT value FROM settings WHERE key=?", (key,))
        return str(r[0]) if r else default

    def audit(self, action: str, entity: str = "", entity_id: str = "", details: str = "") -> None:
        self.execute(
            "INSERT INTO audit(username,action,entity,entity_id,details) VALUES(?,?,?,?,?)",
            (self.current_user, action, entity, entity_id, details),
        )
        self.conn.commit()

    def next_issue_no(self) -> str:
        year = _dt.date.today().year
        pref = f"PPE-{year}-"
        n = int(self.scalar("SELECT COUNT(*) FROM records WHERE issue_no LIKE ?", (pref + "%",), 0)) + 1
        while self.one("SELECT 1 FROM records WHERE issue_no=?", (f"{pref}{n:05d}",)):
            n += 1
        return f"{pref}{n:05d}"


def get_db(current_user: str = "admin") -> PPEIssueDB:
    return PPEIssueDB(current_user=current_user)


def item_group_for(text: str, category: str = "", code: str = "") -> str:
    low = norm(" ".join(x for x in (text, category, code) if x))
    if any(k in low for k in ("safety shoe", "safety shoes", "shoe", "shoes", "boot", "boots")):
        return GROUP_SHOES
    if "blanket" in low:
        return GROUP_BLANKET
    if any(k in low for k in ("frc", "flame resistant", "fire resistant", "fire retardant")):
        return GROUP_FRC
    if any(k in low for k in ("coverall", "coveralls", "overall", "overalls")):
        return GROUP_COVERALL
    if any(k in low for k in ("ppe", "uniform", "safety wear", "protective")):
        return GROUP_OTHER
    return ""


def detect_size(text: str) -> str:
    src = str(text or "")
    for pat in (
            r"\b(?:size|sz)\s*[:#-]?\s*(\d{2})\b",
            r"\b(XXXL|XXL|XL|XS|S|M|L)\b",
            r"\b(3[6-9]|4\d|50|52)\b"):
        m = re.search(pat, src, flags=re.I)
        if m:
            return m.group(1).upper()
    return ""


def distinct(db: PPEIssueDB, field: str) -> list[str]:
    rows = db.query(f"SELECT DISTINCT {field} FROM records WHERE COALESCE({field},'')<>'' ORDER BY {field}")
    return [str(r[0]) for r in rows]


def compute_status(rec: dict | sqlite3.Row) -> str:
    r = dict(rec)
    status = str(r.get("status") or "").strip()
    if status in STATUSES:
        return status
    if r.get("return_date"):
        return ST_RETURNED
    if not str(r.get("employee_code") or r.get("employee_name") or "").strip():
        return ST_NEEDS_INFO
    return ST_ISSUED


def save_record(db: PPEIssueDB, data: dict[str, Any], record_id: int | None = None) -> int:
    rec = dict(data)
    rec["issue_date"] = str(rec.get("issue_date") or today())
    rec["qty"] = float(rec.get("qty") or 0)
    if rec["qty"] <= 0:
        raise ValueError("Quantity must be greater than zero.")
    rec["item_group"] = (
        rec.get("item_group")
        or item_group_for(rec.get("item_desc", ""), code=rec.get("item_code", ""))
        or GROUP_OTHER
    )
    rec["size_text"] = rec.get("size_text") or detect_size(rec.get("item_desc", ""))
    rec["status"] = compute_status(rec)
    if record_id:
        cols = [k for k in rec.keys() if k not in {"id", "issue_no", "created_by", "created_at"}]
        sets = ", ".join(f"{c}=?" for c in cols)
        db.execute(
            f"UPDATE records SET {sets}, updated_at=datetime('now','localtime') WHERE id=?",
            [rec[c] for c in cols] + [record_id],
        )
        db.commit()
        db.audit("EDITED", "record", str(record_id), rec.get("employee_code", ""))
        return record_id
    rec.setdefault("issue_no", db.next_issue_no())
    rec.setdefault("source_type", "MANUAL")
    rec.setdefault("created_by", db.current_user)
    rec.setdefault("issued_by", db.current_user)
    cols = [
        "issue_no", "issue_date", "employee_code", "employee_name", "department", "project",
        "item_group", "item_code", "item_desc", "size_text", "qty", "uom", "dn_no", "doc_date",
        "pdf_path", "source_type", "source_doc_id", "source_line_id", "status", "return_date",
        "issued_by", "remarks", "batch_id", "created_by",
    ]
    nullables = {"source_doc_id", "source_line_id", "batch_id"}
    values = [rec.get(c) if (c in nullables and rec.get(c) not in ("", None)) else (None if c in nullables else rec.get(c, ""))
              for c in cols]
    cur = db.execute(f"INSERT INTO records({','.join(cols)}) VALUES({','.join('?' * len(cols))})", values)
    db.commit()
    new_id = int(cur.lastrowid)
    db.audit("CREATED", "record", str(new_id), rec.get("issue_no", ""))
    return new_id


def delete_record(db: PPEIssueDB, record_id: int) -> None:
    row = db.one("SELECT issue_no FROM records WHERE id=?", (record_id,))
    if not row:
        return
    db.execute("DELETE FROM records WHERE id=?", (record_id,))
    db.commit()
    db.audit("DELETED", "record", str(record_id), row[0])


def get_record(db: PPEIssueDB, record_id: int) -> dict | None:
    r = db.one("SELECT * FROM records WHERE id=?", (record_id,))
    return dict(r) if r else None


def mark_returned(db: PPEIssueDB, record_id: int, return_date: str = "", remarks: str = "") -> None:
    row = get_record(db, record_id)
    if not row:
        return
    db.execute(
        "UPDATE records SET status=?, return_date=?, remarks=?, updated_at=datetime('now','localtime') WHERE id=?",
        (ST_RETURNED, return_date or today(), remarks or row.get("remarks", ""), record_id),
    )
    db.commit()
    db.audit("EDITED", "record", str(record_id), f"returned {row['issue_no']}")


def list_records(db: PPEIssueDB, text: str = "", item_group: str = "", status: str = "",
                 source_type: str = "", date_from: str = "", date_to: str = "") -> list[dict]:
    sql = "SELECT * FROM records WHERE 1=1"
    p: list[Any] = []
    if date_from:
        sql += " AND issue_date>=?"
        p.append(date_from)
    if date_to:
        sql += " AND issue_date<=?"
        p.append(date_to)
    if item_group:
        sql += " AND item_group=?"
        p.append(item_group)
    if status:
        sql += " AND status=?"
        p.append(status)
    if source_type:
        sql += " AND source_type=?"
        p.append(source_type)
    if text.strip():
        like = f"%{text.strip()}%"
        sql += (
            " AND (issue_no LIKE ? OR employee_code LIKE ? OR employee_name LIKE ? OR project LIKE ?"
            " OR item_code LIKE ? OR item_desc LIKE ? OR dn_no LIKE ? OR remarks LIKE ?)"
        )
        p += [like] * 8
    sql += " ORDER BY issue_date DESC, id DESC LIMIT 5000"
    out = [dict(r) for r in db.query(sql, p)]
    for r in out:
        r["status"] = compute_status(r)
    return out


def dashboard_data(db: PPEIssueDB) -> dict[str, Any]:
    rows = list_records(db)
    groups = {g: 0 for g in GROUPS}
    employees = set()
    synced = 0
    missing = 0
    for r in rows:
        groups[r.get("item_group") or GROUP_OTHER] = groups.get(r.get("item_group") or GROUP_OTHER, 0) + 1
        if r.get("employee_code") or r.get("employee_name"):
            employees.add((r.get("employee_code") or "", r.get("employee_name") or ""))
        if r.get("source_type") == "DN":
            synced += 1
        if compute_status(r) == ST_NEEDS_INFO:
            missing += 1
    return {
        "total_records": len(rows),
        "employees": len(employees),
        "synced": synced,
        "missing_info": missing,
        "by_group": groups,
        "recent": rows[:20],
    }


def _existing_dn_lines(db: PPEIssueDB) -> set[tuple[int, int]]:
    return {
        (int(r[0]), int(r[1]))
        for r in db.query(
            "SELECT COALESCE(source_doc_id,0), COALESCE(source_line_id,0) FROM records WHERE source_type='DN'"
        )
        if int(r[0] or 0) and int(r[1] or 0)
    }


def sync_candidates(main_db: Database, ppe_db: PPEIssueDB, date_from: str = "", date_to: str = "",
                    text: str = "") -> list[dict]:
    sql = """SELECT d.id AS doc_id, d.doc_no, d.doc_date, d.project, d.department,
                    d.issued_to, d.handover_to, d.handover_id, d.pdf_path, d.created_by,
                    l.id AS line_id, l.item_code, l.description AS line_desc, l.qty, l.uom,
                    l.remarks, COALESCE(i.description,'') AS item_desc, COALESCE(i.category,'') AS category
               FROM documents d
               JOIN document_lines l ON l.doc_id=d.id
          LEFT JOIN items i ON i.id=l.item_id
              WHERE d.doc_type='DN' AND d.status='FINAL'"""
    p: list[Any] = []
    if date_from:
        sql += " AND d.doc_date>=?"
        p.append(date_from)
    if date_to:
        sql += " AND d.doc_date<=?"
        p.append(date_to)
    if text.strip():
        like = f"%{text.strip()}%"
        sql += (
            " AND (d.doc_no LIKE ? OR d.project LIKE ? OR d.issued_to LIKE ? OR d.handover_to LIKE ?"
            " OR d.handover_id LIKE ? OR l.item_code LIKE ? OR l.description LIKE ? OR i.description LIKE ?)"
        )
        p += [like] * 8
    sql += " ORDER BY d.doc_date DESC, d.id DESC, l.id"
    existing = _existing_dn_lines(ppe_db)
    out = []
    for r in main_db.query(sql, p):
        desc = (r["item_desc"] or r["line_desc"] or "").strip()
        group = item_group_for(desc, r["category"], r["item_code"])
        if not group:
            continue
        emp_code = (r["handover_id"] or "").strip()
        emp_name = (r["handover_to"] or r["issued_to"] or "").strip()
        rec = {
            "doc_id": int(r["doc_id"]),
            "line_id": int(r["line_id"]),
            "doc_no": r["doc_no"],
            "doc_date": r["doc_date"],
            "project": r["project"] or "",
            "department": r["department"] or "",
            "employee_code": emp_code,
            "employee_name": emp_name,
            "item_group": group,
            "item_code": r["item_code"] or "",
            "item_desc": desc,
            "size_text": detect_size(desc),
            "qty": float(r["qty"] or 0),
            "uom": r["uom"] or "",
            "pdf_path": r["pdf_path"] or "",
            "issued_by": r["created_by"] or main_db.current_user,
            "remarks": r["remarks"] or "",
            "imported": (int(r["doc_id"]), int(r["line_id"])) in existing,
            "source_type": "DN",
        }
        rec["status"] = compute_status(rec)
        out.append(rec)
    return out


def import_from_delivery_notes(main_db: Database, ppe_db: PPEIssueDB, date_from: str = "",
                               date_to: str = "", text: str = "") -> tuple[int, int]:
    rows = sync_candidates(main_db, ppe_db, date_from, date_to, text)
    ins = 0
    skipped = 0
    for r in rows:
        if r.get("imported"):
            skipped += 1
            continue
        try:
            save_record(ppe_db, {
                "issue_date": r["doc_date"],
                "employee_code": r["employee_code"],
                "employee_name": r["employee_name"],
                "department": r["department"],
                "project": r["project"],
                "item_group": r["item_group"],
                "item_code": r["item_code"],
                "item_desc": r["item_desc"],
                "size_text": r["size_text"],
                "qty": r["qty"],
                "uom": r["uom"],
                "dn_no": r["doc_no"],
                "doc_date": r["doc_date"],
                "pdf_path": r["pdf_path"],
                "source_type": "DN",
                "source_doc_id": r["doc_id"],
                "source_line_id": r["line_id"],
                "status": r["status"],
                "issued_by": r["issued_by"],
                "remarks": r["remarks"],
            })
            ins += 1
        except sqlite3.IntegrityError:
            skipped += 1
    ppe_db.audit("IMPORTED", "delivery_notes", "", f"{ins} imported, {skipped} skipped")
    return ins, skipped


def sniff(text: str) -> tuple[list[str], list[list[str]]]:
    raw = [ln for ln in text.replace("\r\n", "\n").replace("\r", "\n").split("\n") if ln.strip()]
    if not raw:
        return [], []
    if "\t" in raw[0]:
        rows = [ln.split("\t") for ln in raw]
    elif raw[0].count(",") >= 2:
        rows = list(csv.reader(io.StringIO("\n".join(raw))))
    elif raw[0].count("|") >= 2:
        rows = [[c.strip() for c in ln.strip().strip("|").split("|")] for ln in raw]
    else:
        rows = [re.split(r"\s{2,}", ln.strip()) for ln in raw]
    rows = [[str(c).strip() for c in r] for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return [], []
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    head = rows[0]
    if sum(1 for c in head if norm_key(c) in HEADER_MAP) >= max(2, min(4, width)):
        return head, rows[1:]
    return [f"Column {i + 1}" for i in range(width)], rows


def read_file(path: str | Path) -> tuple[list[str], list[list[Any]]]:
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(p, data_only=True, read_only=True)
        ws = wb.active
        data = [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
        data = [r for r in data if any(str(c).strip() for c in r)]
        if not data:
            return [], []
        head = [str(c).strip() for c in data[0]]
        if sum(1 for c in head if norm_key(c) in HEADER_MAP) >= 2:
            return head, data[1:]
        return [f"Column {i + 1}" for i in range(len(head))], data
    return sniff(p.read_text(encoding="utf-8", errors="ignore"))


def auto_map(headers: Sequence[str]) -> dict[int, str]:
    out: dict[int, str] = {}
    used: set[str] = set()
    for i, h in enumerate(headers):
        f = HEADER_MAP.get(norm_key(h))
        if f and f not in used:
            out[i] = f
            used.add(f)
    return out


def preview(headers: Sequence[str], rows: Sequence[Sequence[Any]],
            mapping: dict[int, str], defaults: dict | None = None) -> list[dict]:
    defaults = defaults or {}
    out = []
    for row in rows:
        rec = {f: "" for f, _ in ALL_FIELDS}
        rec.update({k: v for k, v in defaults.items() if v not in (None, "")})
        for i, field in mapping.items():
            if i < len(row):
                rec[field] = row[i]
        for f in NUMERIC_FIELDS:
            rec[f] = to_float(rec.get(f))
        for f in DATE_FIELDS:
            rec[f] = to_date(rec.get(f))
        rec["issue_date"] = rec.get("issue_date") or today()
        rec["employee_code"] = str(rec.get("employee_code") or "").strip()
        rec["employee_name"] = str(rec.get("employee_name") or "").strip()
        rec["department"] = str(rec.get("department") or "").strip()
        rec["project"] = str(rec.get("project") or "").strip()
        rec["item_code"] = str(rec.get("item_code") or "").strip()
        rec["item_desc"] = str(rec.get("item_desc") or "").strip()
        rec["uom"] = str(rec.get("uom") or "").strip() or "PCS"
        rec["dn_no"] = str(rec.get("dn_no") or "").strip()
        rec["remarks"] = str(rec.get("remarks") or "").strip()
        group_hint = str(rec.get("item_group") or "").strip()
        rec["item_group"] = item_group_for(rec["item_desc"], group_hint, rec["item_code"]) or group_hint or GROUP_OTHER
        if rec["item_group"] not in GROUPS:
            rec["item_group"] = item_group_for(rec["item_group"], rec["item_group"], rec["item_code"]) or GROUP_OTHER
        rec["size_text"] = str(rec.get("size_text") or "").strip() or detect_size(rec["item_desc"])
        status_hint = norm(" ".join(str(rec.get(k) or "") for k in ("status", "remarks")))
        if rec.get("return_date") or ("returned" in status_hint and "not returned" not in status_hint):
            rec["status"] = ST_RETURNED
        else:
            rec["status"] = compute_status(rec)
        rec["qty"] = float(rec.get("qty") or 0)
        if rec["qty"] <= 0 and rec["item_desc"]:
            rec["qty"] = 1.0
        if not rec["item_desc"] and not rec["item_code"]:
            continue
        if rec["qty"] <= 0:
            continue
        rec["source_type"] = "MANUAL"
        out.append(rec)
    return out


def _import_key(rec: dict | sqlite3.Row) -> tuple:
    r = dict(rec)
    return (
        r.get("issue_date") or "",
        norm(r.get("employee_code")),
        norm(r.get("employee_name")),
        norm(r.get("item_group")),
        norm(r.get("item_code")),
        norm(r.get("item_desc")),
        round(to_float(r.get("qty")), 3),
        norm(r.get("dn_no")),
    )


def import_records(db: PPEIssueDB, records: Sequence[dict], source: str = "",
                   skip_duplicates: bool = True) -> tuple[int, int]:
    cur = db.execute(
        "INSERT INTO batches(ts,source,rows,skipped,username) VALUES(?,?,?,?,?)",
        (_now(), str(source), 0, 0, db.current_user),
    )
    batch_id = int(cur.lastrowid)

    existing: set[tuple] = set()
    if skip_duplicates:
        for r in db.query(
                "SELECT issue_date, employee_code, employee_name, item_group, item_code, item_desc, qty, dn_no FROM records"):
            existing.add(_import_key(r))

    inserted = skipped = 0
    for rec in records:
        key = _import_key(rec)
        if skip_duplicates and key in existing:
            skipped += 1
            continue
        existing.add(key)
        save_record(db, {**dict(rec), "batch_id": batch_id, "source_type": "MANUAL"})
        inserted += 1
    db.execute("UPDATE batches SET rows=?, skipped=? WHERE id=?", (inserted, skipped, batch_id))
    db.commit()
    db.audit("IMPORTED", "sheet", str(batch_id),
             f"{inserted} inserted, {skipped} duplicate(s) skipped from {source}")
    return inserted, skipped


def undo_batch(db: PPEIssueDB, batch_id: int) -> int:
    ids = [int(r[0]) for r in db.query("SELECT id FROM records WHERE batch_id=?", (batch_id,))]
    for record_id in ids:
        db.execute("DELETE FROM records WHERE id=?", (record_id,))
    db.commit()
    if ids:
        db.audit("DELETED", "batch", str(batch_id), f"{len(ids)} record(s) removed (undo import)")
    return len(ids)


def batches(db: PPEIssueDB) -> list[dict]:
    return [dict(r) for r in db.query(
        "SELECT b.*, (SELECT COUNT(*) FROM records r WHERE r.batch_id=b.id) live"
        " FROM batches b ORDER BY b.id DESC LIMIT 100")]


def template_rows() -> tuple[list[str], list[list[Any]]]:
    cols = [lbl for _, lbl in FIELDS]
    td = _dt.date.today().isoformat()
    return cols, [
        [td, "EMP-1001", "Ahmed Salem", "Camp", "Camp A", GROUP_SHOES,
         "PPE-SHOE-42", "Safety Shoes Size 42", "42", 1, "PAIR", "DN-2026-00123", "", "New issue", ST_ISSUED],
        [td, "EMP-1002", "Bilal Khan", "Laundry", "Camp B", GROUP_COVERALL,
         "PPE-COV-L", "Coverall L", "L", 2, "PCS", "DN-2026-00124", "", "Two sets issued", ST_ISSUED],
        [td, "EMP-1003", "Rashid", "Welfare", "Camp C", GROUP_BLANKET,
         "PPE-BLK-01", "Blanket Single Bed", "", 1, "PCS", "DN-2026-00125", td, "Returned", ST_RETURNED],
    ]


def build_report(db: PPEIssueDB, name: str, f: dict | None = None) -> tuple[str, list[str], list[list[Any]]]:
    f = f or {}
    rows = list_records(db, text=str(f.get("text", "")), item_group=str(f.get("item_group", "")),
                        status=str(f.get("status", "")), source_type=str(f.get("source_type", "")),
                        date_from=str(f.get("date_from", "")), date_to=str(f.get("date_to", "")))

    def base(src: list[dict]) -> list[list[Any]]:
        return [[r["issue_no"], r["issue_date"], r["employee_code"], r["employee_name"],
                 r["project"], r["item_group"], r["item_code"], r["item_desc"], r["size_text"],
                 round(float(r["qty"] or 0), 2), r["uom"], r["dn_no"], r["source_type"],
                 compute_status(r), r["remarks"]] for r in src]

    if name == "Full PPE Register":
        cols = ["Issue No", "Date", "Employee Code", "Employee", "Project / Dept", "Group",
                "Item Code", "Description", "Size", "Qty", "UOM", "DN No", "Source",
                "Status", "Remarks"]
        return name, cols, base(rows)

    mapping = {
        "Safety Shoes Register": GROUP_SHOES,
        "Blanket Register": GROUP_BLANKET,
        "FRC Register": GROUP_FRC,
        "Coverall Register": GROUP_COVERALL,
    }
    if name in mapping:
        sel = [r for r in rows if r["item_group"] == mapping[name]]
        cols = ["Date", "Employee Code", "Employee", "Item Code", "Description", "Size",
                "Qty", "DN No", "Project", "Status"]
        return name, cols, [[r["issue_date"], r["employee_code"], r["employee_name"], r["item_code"],
                             r["item_desc"], r["size_text"], round(float(r["qty"] or 0), 2),
                             r["dn_no"], r["project"], compute_status(r)] for r in sel]

    if name == "By Employee":
        agg: dict[tuple[str, str], dict[str, Any]] = {}
        for r in rows:
            k = (r["employee_code"], r["employee_name"])
            box = agg.setdefault(k, {"project": r["project"], "groups": set(), "qty": 0.0, "rows": 0})
            box["groups"].add(r["item_group"])
            box["qty"] += float(r["qty"] or 0)
            box["rows"] += 1
        cols = ["Employee Code", "Employee", "Project / Dept", "Records", "Total Qty", "Groups"]
        data = [[k[0], k[1], v["project"], v["rows"], round(v["qty"], 2), ", ".join(sorted(v["groups"]))]
                for k, v in sorted(agg.items(), key=lambda kv: ((kv[0][1] or "").lower(), (kv[0][0] or "").lower()))]
        return name, cols, data

    if name == "By Delivery Note":
        agg: dict[str, dict[str, Any]] = {}
        for r in rows:
            k = r["dn_no"] or "Manual"
            box = agg.setdefault(k, {"date": r["doc_date"] or r["issue_date"], "groups": set(), "qty": 0.0, "rows": 0})
            box["groups"].add(r["item_group"])
            box["qty"] += float(r["qty"] or 0)
            box["rows"] += 1
        cols = ["DN No", "Date", "Records", "Total Qty", "Groups"]
        data = [[k, v["date"], v["rows"], round(v["qty"], 2), ", ".join(sorted(v["groups"]))]
                for k, v in sorted(agg.items(), key=lambda kv: (kv[1]["date"], kv[0]), reverse=True)]
        return name, cols, data

    if name == "Missing Employee Codes":
        sel = [r for r in rows if compute_status(r) == ST_NEEDS_INFO or not str(r["employee_code"] or "").strip()]
        cols = ["Date", "DN No", "Employee Code", "Employee", "Group", "Item Code", "Description", "Qty"]
        return name, cols, [[r["issue_date"], r["dn_no"], r["employee_code"], r["employee_name"],
                             r["item_group"], r["item_code"], r["item_desc"], round(float(r["qty"] or 0), 2)]
                            for r in sel]

    if name == "Synced Delivery Note PPE":
        sel = [r for r in rows if r["source_type"] == "DN"]
        cols = ["Date", "DN No", "Employee Code", "Employee", "Group", "Description", "Size", "Qty", "Project"]
        return name, cols, [[r["issue_date"], r["dn_no"], r["employee_code"], r["employee_name"],
                             r["item_group"], r["item_desc"], r["size_text"], round(float(r["qty"] or 0), 2),
                             r["project"]] for r in sel]

    return name, ["No Data"], []
