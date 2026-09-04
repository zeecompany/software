"""Searchable text index for document files, attachments and library files."""
from __future__ import annotations

import datetime as _dt
import html
import io
import re
import time
import zipfile
from pathlib import Path
from typing import Any

from .database import Database

DDL = """
CREATE TABLE IF NOT EXISTS file_search_index (
    path           TEXT PRIMARY KEY,
    source_type    TEXT DEFAULT '',
    doc_no         TEXT DEFAULT '',
    doc_type       TEXT DEFAULT '',
    file_name      TEXT DEFAULT '',
    title          TEXT DEFAULT '',
    content_text   TEXT DEFAULT '',
    content_norm   TEXT DEFAULT '',
    modified_ts    REAL DEFAULT 0,
    size_bytes     INTEGER DEFAULT 0,
    updated_at     TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS ix_file_search_doc  ON file_search_index(doc_no);
CREATE INDEX IF NOT EXISTS ix_file_search_type ON file_search_index(source_type);
"""

TEXT_SUFFIXES = {
    ".txt", ".csv", ".md", ".log", ".json", ".xml", ".html", ".htm",
    ".docx", ".xlsx", ".xlsm", ".pdf",
}

_REFRESHED_AT: dict[str, float] = {}


def ensure_schema(db: Database) -> None:
    db.conn.executescript(DDL)
    db.conn.commit()


def _norm(text: str) -> str:
    return " ".join(str(text or "").lower().split())


def _now() -> str:
    return _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _table_exists(db: Database, name: str) -> bool:
    return bool(db.one("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (name,)))


def _iter_known_files(db: Database):
    seen: set[str] = set()

    def add(path: str, source_type: str, doc_no: str = "", doc_type: str = "",
            title: str = ""):
        p = str(Path(path))
        if not p or p in seen:
            return
        seen.add(p)
        yield {
            "path": p,
            "source_type": source_type,
            "doc_no": doc_no,
            "doc_type": doc_type,
            "title": title,
        }

    for r in db.query("SELECT doc_no, doc_type, status, pdf_path FROM documents WHERE COALESCE(pdf_path,'')<>''"):
        title = f"{r['doc_type']} {r['doc_no']} ({r['status']})"
        yield from add(r["pdf_path"], "Document PDF", r["doc_no"], r["doc_type"], title)
    for r in db.query(
            "SELECT doc_no, doc_type, COALESCE(source,'file') AS source, file_path"
            " FROM attachments WHERE COALESCE(file_path,'')<>''"):
        src = "Pasted Attachment" if str(r["source"]).lower() == "clipboard" else "Attachment"
        title = f"{r['doc_type']} {r['doc_no']} — {src}"
        yield from add(r["file_path"], src, r["doc_no"] or "", r["doc_type"] or "", title)
    if _table_exists(db, "gdn_documents"):
        for r in db.query("SELECT doc_no, status, title, pdf_path FROM gdn_documents WHERE COALESCE(pdf_path,'')<>''"):
            title = f"GDN {r['doc_no']} — {r['title'] or 'DELIVERY NOTE'} ({r['status']})"
            yield from add(r["pdf_path"], "General DN PDF", r["doc_no"], "GDN", title)
    if _table_exists(db, "library_files"):
        for r in db.query(
                "SELECT path, name, doc_no, doc_type, kind FROM library_files"
                " WHERE COALESCE(path,'')<>'' AND status<>'MISSING'"):
            title = f"Library {r['kind']} — {r['name']}"
            yield from add(r["path"], "Library File", r["doc_no"] or "", r["doc_type"] or "", title)


def _read_text_file(path: Path, max_chars: int) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="ignore")[:max_chars]
    except OSError:
        return ""


def _read_docx(path: Path, max_chars: int) -> str:
    try:
        with zipfile.ZipFile(path) as z:
            parts = []
            for name in ("word/document.xml", "word/header1.xml", "word/footer1.xml"):
                try:
                    parts.append(z.read(name).decode("utf-8", errors="ignore"))
                except KeyError:
                    continue
    except OSError:
        return ""
    text = re.sub(r"<[^>]+>", " ", " ".join(parts))
    return html.unescape(" ".join(text.split()))[:max_chars]


def _read_xlsx(path: Path, max_chars: int) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return ""
    out: list[str] = []
    size = 0
    try:
        for ws in wb.worksheets[:8]:
            out.append(ws.title)
            for row in ws.iter_rows(values_only=True):
                bits = [str(v).strip() for v in row if v not in (None, "")]
                if not bits:
                    continue
                line = " | ".join(bits)
                out.append(line)
                size += len(line)
                if size >= max_chars:
                    raise StopIteration
    except StopIteration:
        pass
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return " ".join(out)[:max_chars]


def _read_pdf(path: Path, max_chars: int) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
    except Exception:
        return ""
    out: list[str] = []
    size = 0
    try:
        for page in reader.pages[:30]:
            txt = page.extract_text() or ""
            if txt:
                out.append(txt)
                size += len(txt)
            if size >= max_chars:
                break
    except Exception:
        return ""
    return "\n".join(out)[:max_chars]


def extract_text(path: str | Path, max_chars: int = 40000) -> str:
    p = Path(path)
    if not p.exists() or not p.is_file():
        return ""
    suf = p.suffix.lower()
    if suf not in TEXT_SUFFIXES:
        return ""
    if suf == ".pdf":
        return _read_pdf(p, max_chars)
    if suf in {".xlsx", ".xlsm"}:
        return _read_xlsx(p, max_chars)
    if suf == ".docx":
        return _read_docx(p, max_chars)
    return _read_text_file(p, max_chars)


def refresh_index(db: Database, force: bool = False, min_interval: float = 45.0) -> int:
    ensure_schema(db)
    key = str(db.path)
    now = time.monotonic()
    if not force and (now - _REFRESHED_AT.get(key, 0.0)) < min_interval:
        return 0
    changed = 0
    live_paths: set[str] = set()
    for rec in _iter_known_files(db):
        p = Path(rec["path"])
        if not p.exists() or not p.is_file():
            continue
        try:
            st = p.stat()
        except OSError:
            continue
        live_paths.add(str(p))
        row = db.one("SELECT modified_ts, size_bytes, doc_no, doc_type, source_type, title"
                     " FROM file_search_index WHERE path=?", (str(p),))
        same_meta = bool(row and float(row["modified_ts"] or 0) == float(st.st_mtime)
                         and int(row["size_bytes"] or 0) == int(st.st_size)
                         and str(row["doc_no"] or "") == str(rec["doc_no"] or "")
                         and str(row["doc_type"] or "") == str(rec["doc_type"] or "")
                         and str(row["source_type"] or "") == str(rec["source_type"] or "")
                         and str(row["title"] or "") == str(rec["title"] or ""))
        if same_meta and not force:
            continue
        content_text = " ".join(extract_text(p).split())
        content_norm = _norm(content_text)
        db.execute(
            """INSERT INTO file_search_index(path,source_type,doc_no,doc_type,file_name,title,
                    content_text,content_norm,modified_ts,size_bytes,updated_at)
               VALUES(?,?,?,?,?,?,?,?,?,?,?)
               ON CONFLICT(path) DO UPDATE SET
                    source_type=excluded.source_type,
                    doc_no=excluded.doc_no,
                    doc_type=excluded.doc_type,
                    file_name=excluded.file_name,
                    title=excluded.title,
                    content_text=excluded.content_text,
                    content_norm=excluded.content_norm,
                    modified_ts=excluded.modified_ts,
                    size_bytes=excluded.size_bytes,
                    updated_at=excluded.updated_at""",
            (str(p), rec["source_type"], rec["doc_no"], rec["doc_type"], p.name,
             rec["title"], content_text, content_norm, float(st.st_mtime), int(st.st_size), _now()),
        )
        changed += 1
    stale = [r[0] for r in db.query("SELECT path FROM file_search_index") if r[0] not in live_paths]
    if stale:
        db.execute(f"DELETE FROM file_search_index WHERE path IN ({','.join('?' * len(stale))})", stale)
        changed += len(stale)
    db.commit()
    _REFRESHED_AT[key] = now
    return changed


def _snippet(content_text: str, query: str, width: int = 260) -> str:
    text = " ".join(str(content_text or "").split())
    if not text:
        return ""
    q = str(query or "").strip().lower()
    low = text.lower()
    pos = low.find(q)
    if pos < 0:
        return text[:width].rstrip()
    start = max(0, pos - width // 3)
    end = min(len(text), pos + len(q) + width // 2)
    out = text[start:end].strip()
    if start > 0:
        out = "… " + out
    if end < len(text):
        out += " …"
    return out


def search(db: Database, text: str, limit: int = 80) -> list[dict]:
    ensure_schema(db)
    query = _norm(text)
    if not query:
        return []
    refresh_index(db)
    like = f"%{query}%"
    rows = db.query(
        """SELECT * FROM file_search_index
           WHERE content_norm LIKE ?
              OR LOWER(file_name) LIKE ?
              OR LOWER(title) LIKE ?
              OR LOWER(doc_no) LIKE ?
           ORDER BY CASE WHEN content_norm LIKE ? THEN 0 ELSE 1 END,
                    doc_no<>'', doc_no, file_name
           LIMIT ?""",
        (like, like, like, like, like, int(limit)),
    )
    out = []
    for r in rows:
        d = dict(r)
        d["snippet"] = _snippet(d.get("content_text", ""), text)
        d["exists"] = Path(d["path"]).exists()
        out.append(d)
    return out
