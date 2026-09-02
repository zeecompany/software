"""PDF, Excel, CSV generation + Email / WhatsApp / file-manager sharing."""
from __future__ import annotations

import csv
import datetime as _dt
import io
import os
import platform
import subprocess
import urllib.parse
import webbrowser
from pathlib import Path
from typing import Any, Sequence

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import (Flowable, Image, KeepTogether, PageBreak, Paragraph,
                                SimpleDocTemplate, Spacer, Table, TableStyle)

from . import arabic as AR
from . import config
from .database import Database

BRAND_BLUE = colors.HexColor("#0b3d6b")
BRAND_ACCENT = colors.HexColor("#f5a300")
LIGHT = colors.HexColor("#eef3f8")

class BottomAnchored(Flowable):
    """Draw the wrapped flowable flush with the BOTTOM of the current frame.

    Used for the authorised signature strip so it always sits just above the
    footer line, no matter how many item rows the document has.
    """

    def __init__(self, inner):
        Flowable.__init__(self)
        self.inner = inner

    def wrap(self, avail_w, avail_h):
        self.width = avail_w
        _, ih = self.inner.wrap(avail_w, avail_h)
        self._ih = ih
        if ih > avail_h:
            # genuinely no room left -> move to the next page and anchor there
            return (avail_w, avail_h + 1)
        # Consume the rest of the frame so the block lands on the page bottom.
        # When only a sliver of the page is left we simply sit inline, which
        # avoids pushing the signatures onto a page of their own.
        self._anchor = avail_h >= ih * 1.35
        return (avail_w, avail_h if self._anchor else ih)

    def drawOn(self, canvas, x, y, _sW=0):
        # y is the bottom of the space we were given -> draw the content there
        self.inner.drawOn(canvas, x, y, _sW)


_styles = getSampleStyleSheet()
P_SM = ParagraphStyle("sm", parent=_styles["Normal"], fontSize=8, leading=10)
P_MD = ParagraphStyle("md", parent=_styles["Normal"], fontSize=9.5, leading=12)
P_TITLE = ParagraphStyle("t", parent=_styles["Title"], fontSize=17, textColor=BRAND_BLUE,
                         spaceAfter=2)
P_SUB = ParagraphStyle("s", parent=_styles["Normal"], fontSize=9, textColor=colors.grey,
                       alignment=TA_CENTER)


def safe_name(text: str) -> str:
    return "".join(c for c in str(text) if c.isalnum() or c in " -_.").strip() or "document"


# Characters Windows forbids in a file name. Everything else (including
# parentheses and dashes, which the AURCO naming pattern needs) is kept.
_BAD_FILE_CHARS = '<>:"/\\|?*'


def safe_file_part(text: str, fallback: str = "") -> str:
    """Sanitise one piece of a file name while keeping ( ) - and spaces."""
    out = "".join(("-" if c in _BAD_FILE_CHARS else c)
                  for c in str(text or "") if ord(c) >= 32)
    # collapse runs of whitespace so the name stays tidy
    out = " ".join(out.split()).strip(" .-_")
    return out or fallback


# Windows path limit is 260 chars; keep the base name comfortably below it.
MAX_NAME_LEN = 150


def unique_prs(lines) -> list[str]:
    """Every distinct PR / MR number on the document, in first-seen order."""
    seen, out = set(), []
    for l in lines:
        try:
            pr = (l["pr_no"] if "pr_no" in l.keys() else "") or ""
        except (AttributeError, TypeError):
            pr = (l.get("pr_no") if isinstance(l, dict) else "") or ""
        pr = str(pr).strip()
        if pr and pr not in seen:
            seen.add(pr)
            out.append(pr)
    return out


# Default pattern requested by AURCO:
#   DN-0821 Material Delivered (Main WH - Project Name) PR Numbers
DEFAULT_DN_TEMPLATE = "{docno} Material Delivered ({warehouse} - {project}) {prs}"

FILENAME_TOKENS = [
    ("{docno}", "Full document number, e.g. DN-2026-00821"),
    ("{docno_short}", "Prefix + sequence only, e.g. DN-00821"),
    ("{seq}", "Just the running number, e.g. 00821"),
    ("{prefix}", "Document prefix, e.g. DN"),
    ("{date}", "Document date, yyyy-mm-dd"),
    ("{ddmm}", "Day and month, e.g. 0821"),
    ("{year}", "Four-digit year"),
    ("{warehouse}", "Issuing store / From location"),
    ("{project}", "Project or site"),
    ("{party}", "Issued to / supplier / receiver"),
    ("{vehicle}", "Vehicle number"),
    ("{driver}", "Handover / driver name"),
    ("{reference}", "Reference or linked document"),
    ("{prs}", "Every PR / MR number on the document"),
    ("{type}", "Document type code, e.g. DN"),
]


def _doc_field(doc_row, key: str) -> str:
    try:
        return str(doc_row[key] or "")
    except (IndexError, KeyError, TypeError):
        return ""


def filename_context(db: Database, doc_row, lines) -> dict:
    """Values available to a file-name template."""
    doc_no = _doc_field(doc_row, "doc_no")
    parts = doc_no.split("-")
    prefix = parts[0] if parts else doc_no
    seq = parts[-1] if len(parts) > 1 else ""
    date = _doc_field(doc_row, "doc_date")
    ddmm = ""
    if len(date) >= 10:
        ddmm = f"{date[8:10]}{date[5:7]}"
    prs = unique_prs(lines)
    return {
        "docno": doc_no,
        "docno_short": f"{prefix}-{seq}" if seq else doc_no,
        "seq": seq,
        "prefix": prefix,
        "date": date,
        "ddmm": ddmm,
        "year": date[:4] if len(date) >= 4 else "",
        "warehouse": (_doc_field(doc_row, "from_location")
                      or _doc_field(doc_row, "warehouse")),
        "project": (_doc_field(doc_row, "project")
                    or _doc_field(doc_row, "department")),
        "party": (_doc_field(doc_row, "issued_to") or _doc_field(doc_row, "supplier")
                  or _doc_field(doc_row, "received_by")
                  or _doc_field(doc_row, "returned_by")),
        "vehicle": _doc_field(doc_row, "vehicle"),
        "driver": (_doc_field(doc_row, "handover_to")
                   or _doc_field(doc_row, "driver")),
        "reference": (_doc_field(doc_row, "reference")
                      or _doc_field(doc_row, "linked_doc")),
        "type": _doc_field(doc_row, "doc_type"),
        "_prs": prs,
    }


def render_filename(template: str, ctx: dict, sep: str = " ",
                    max_len: int = MAX_NAME_LEN) -> str:
    """Fill a file-name template, tidy the result and keep it Windows-safe.

    Empty tokens collapse gracefully: an unused "( - )" group is removed rather
    than left as empty brackets, and PR numbers are dropped whole (never cut in
    half) if the name would exceed the length limit.
    """
    import re

    prs = list(ctx.get("_prs") or [])

    def build(keep: int) -> str:
        vals = dict(ctx)
        shown = prs[:keep]
        if keep < len(prs):
            shown = shown + [f"+{len(prs) - keep}-more"]
        vals["prs"] = sep.join(shown)
        out = template
        for k, v in vals.items():
            if k.startswith("_"):
                continue
            out = out.replace("{" + k + "}", safe_file_part(v))
        # drop leftover unknown tokens
        out = re.sub(r"\{[a-zA-Z_]+\}", "", out)
        # tidy empty bracket groups and stray separators
        out = re.sub(r"\(\s*[-–—]?\s*\)", "", out)
        out = re.sub(r"\(\s*([^()]*?)\s*\)", lambda m: f"({m.group(1).strip(' -–—')})"
                     if m.group(1).strip(" -–—") else "", out)
        out = re.sub(r"\s{2,}", " ", out)
        out = out.replace(" - )", ")").replace("( - ", "(")
        return safe_file_part(out, "document")

    keep = len(prs)
    name = build(keep)
    while keep > 0 and len(name) > max_len:
        keep -= 1
        name = build(keep)
    return name[:max_len].strip(" .-_") or "document"


def document_basename(db: Database, doc_row, lines) -> str:
    """File name for a document, driven by a configurable template.

    Delivery Notes default to the AURCO pattern:
        DN-2026-00821 Material Delivered (Main WH - Project) 001582 001601
    Other document types keep the plain document number unless a template is
    configured for them.
    """
    dtype = _doc_field(doc_row, "doc_type")
    ctx = filename_context(db, doc_row, lines)
    sep = db.get_setting("dn_filename_separator", " ") or " "
    tpl = (db.get_setting(f"filename_template_{dtype}", "") or "").strip()
    if not tpl and dtype == "DN":
        if db.get_bool("dn_filename_use_pattern", True):
            tpl = db.get_setting("dn_filename_template",
                                 DEFAULT_DN_TEMPLATE) or DEFAULT_DN_TEMPLATE
        elif db.get_bool("dn_filename_include_pr", True):
            tpl = "{docno}" + (sep + "{prs}" if ctx["_prs"] else "")
        else:
            tpl = "{docno}"
    if not tpl:
        tpl = "{docno}"
    return render_filename(tpl, ctx, sep)


# --------------------------------------------------------------- page frame
def _shift(c, factor: float):
    """Lighten (>1) or darken (<1) a reportlab colour."""
    try:
        return colors.Color(min(1.0, c.red * factor), min(1.0, c.green * factor),
                            min(1.0, c.blue * factor))
    except Exception:
        return c


def _mix(c1, c2, t: float):
    return colors.Color(c1.red + (c2.red - c1.red) * t,
                        c1.green + (c2.green - c1.green) * t,
                        c1.blue + (c2.blue - c1.blue) * t)


def _logo_panel(canvas, db: Database, x, y, w_, h_):
    """Optional rounded plate behind the logo.

    Logos drawn in dark ink disappear on a dark header band; a light plate keeps
    them crisp. Controlled by Settings -> 'Logo backing plate'.
    """
    mode = db.get_setting("logo_backing", "Auto")
    if mode == "None":
        return
    if mode == "Auto":
        # only add a plate when the band itself is dark
        primary, _ = _brand_colors(db)
        lum = 0.299 * primary.red + 0.587 * primary.green + 0.114 * primary.blue
        if lum > 0.62:
            return
    fill = colors.white if mode != "Dark" else colors.HexColor("#101418")
    pad = 1.6 * mm
    canvas.saveState()
    canvas.setFillColor(fill)
    canvas.roundRect(x - pad, y - pad, w_ + pad * 2, h_ + pad * 2, 1.8 * mm,
                     stroke=0, fill=1)
    canvas.restoreState()


def _brand_colors(db: Database):
    """PDF brand colours follow the application theme."""
    try:
        primary = colors.HexColor(db.get_setting("ui_primary", "#0b3d6b"))
        accent = colors.HexColor(db.get_setting("ui_accent", "#f5a300"))
        return primary, accent
    except Exception:
        return BRAND_BLUE, BRAND_ACCENT


def _paint_band(canvas, design: dict, x0, y0, w, h) -> None:
    """Solid / gradient / transparent band background."""
    style = design.get("bg_style", "Gradient")
    if style == "None":
        return
    try:
        c1 = colors.HexColor(design.get("bg_color1", "#12161c"))
    except Exception:
        c1 = colors.HexColor("#12161c")
    try:
        c2 = colors.HexColor(design.get("bg_color2", design.get("bg_color1", "#12161c")))
    except Exception:
        c2 = c1
    if style == "Solid" or c1.rgb() == c2.rgb():
        canvas.setFillColor(c1)
        canvas.rect(x0, y0, w, h, stroke=0, fill=1)
        return
    steps = 60
    if design.get("bg_angle", "Horizontal") == "Vertical":
        for i in range(steps):
            canvas.setFillColor(_mix(c1, c2, i / (steps - 1)))
            canvas.rect(x0, y0 + h * i / steps, w, h / steps + 0.6, stroke=0, fill=1)
    else:
        for i in range(steps):
            canvas.setFillColor(_mix(c1, c2, i / (steps - 1)))
            canvas.rect(x0 + w * i / steps, y0, w / steps + 0.6, h, stroke=0, fill=1)


def _draw_logo(canvas, db: Database, design: dict, logo: str, x, y, lw, lh) -> None:
    mode = design.get("logo_backing", "Auto")
    if mode != "None":
        show_plate = True
        if mode == "Auto":
            if design.get("bg_style", "Gradient") == "None":
                show_plate = False
            else:
                try:
                    c = colors.HexColor(design.get("bg_color1", "#12161c"))
                    lum = 0.299 * c.red + 0.587 * c.green + 0.114 * c.blue
                    show_plate = lum <= 0.62
                except Exception:
                    show_plate = True
        if show_plate:
            pad = 1.6 * mm
            canvas.saveState()
            canvas.setFillColor(colors.white if mode != "Dark"
                                else colors.HexColor("#101418"))
            canvas.roundRect(x - pad, y - pad, lw + pad * 2, lh + pad * 2,
                             1.8 * mm, stroke=0, fill=1)
            canvas.restoreState()
    canvas.drawImage(logo, x, y, width=lw, height=lh,
                     preserveAspectRatio=True, mask="auto")


def _render_band(canvas, doc, db: Database, design: dict, ctx: dict, *,
                 is_footer: bool, logo: str = "") -> None:
    """Draw one designed band (header or footer) on the page."""
    from . import header_design as HD
    AR.configure(db)      # load the Arabic typeface chosen in Settings

    w, page_h = doc.pagesize
    band = float(design.get("height", 24)) * mm
    pad = float(design.get("padding", 12)) * mm
    y0 = 0 if is_footer else page_h - band

    _paint_band(canvas, design, 0, y0, w, band)

    # accent bar / rule
    if design.get("accent_bar", True):
        try:
            ac = colors.HexColor(design.get("accent_color", "#f5a300"))
        except Exception:
            ac = colors.HexColor("#f5a300")
        ah = float(design.get("accent_height", 1.5)) * mm
        canvas.setFillColor(ac)
        canvas.rect(0, (y0 + band) if is_footer else (y0 - ah), w, ah, stroke=0, fill=1)
    if design.get("bottom_border"):
        try:
            bc = colors.HexColor(design.get("border_color", "#c1121f"))
        except Exception:
            bc = colors.HexColor("#c1121f")
        canvas.setStrokeColor(bc)
        canvas.setLineWidth(0.7)
        yline = (y0 + band) if is_footer else y0
        canvas.line(pad, yline, w - pad, yline)

    # ---- logo
    left_x, right_x = pad, w - pad
    center_shift = 0.0
    if design.get("logo_show", True) and logo and Path(logo).exists():
        lw = float(design.get("logo_width", 34)) * mm
        lh = float(design.get("logo_height", 9)) * mm
        ly = y0 + (band - lh) / 2
        slot = design.get("logo_slot", "Left")
        try:
            if slot == "Left":
                _draw_logo(canvas, db, design, logo, left_x, ly, lw, lh)
                left_x += lw + 5 * mm
            elif slot == "Right":
                _draw_logo(canvas, db, design, logo, right_x - lw, ly, lw, lh)
                right_x -= lw + 5 * mm
            else:
                _draw_logo(canvas, db, design, logo, w / 2 - lw / 2, ly, lw, lh)
                center_shift = lh
        except Exception:
            pass

    # ---- text elements, grouped into stacked rows
    els = [e for e in design.get("elements", []) if e.get("visible", True)]
    if not els:
        return
    n_rows = max(int(e.get("row", 0)) for e in els) + 1
    gap = float(design.get("row_gap", 1.0)) * mm
    row_h = [max([float(e.get("size", 10)) for e in els
                  if int(e.get("row", 0)) == r] or [10]) * 1.22 for r in range(n_rows)]
    total = sum(row_h) + gap * (n_rows - 1)
    top = y0 + (band + total) / 2
    if center_shift:
        top = y0 + (band + total) / 2

    cursor = top
    for r in range(n_rows):
        cursor -= row_h[r]
        baseline = cursor + row_h[r] * 0.22
        for el in [e for e in els if int(e.get("row", 0)) == r]:
            txt = HD.resolve_text(el, ctx).strip()
            if not txt:
                continue
            try:
                canvas.setFillColor(colors.HexColor(el.get("color", "#ffffff")))
            except Exception:
                canvas.setFillColor(colors.white)
            base_font = HD.font_name(el)
            if AR.is_rtl(txt):
                txt, base_font = AR.prepare(
                    txt, bool(el.get("bold")), base_font,
                    eastern_digits=db.get_bool("arabic_eastern_digits", True))
            canvas.setFont(base_font, float(el.get("size", 10)))
            slot = el.get("slot", "Left")
            if slot == "Right":
                canvas.drawRightString(right_x, baseline, txt)
            elif slot == "Center":
                canvas.drawCentredString(w / 2, baseline, txt)
            else:
                canvas.drawString(left_x, baseline, txt)
        cursor -= gap


def _build_with_totals(out: Path, story, land: bool, db: Database, doc_type: str,
                       make_hf) -> None:
    """Build twice when the design needs a page total, once otherwise."""
    from . import header_design as HD
    needs_total = any(
        e.get("source") == "pageof" or "{pages}" in (e.get("text") or "")
        for kind in ("header", "footer")
        for e in HD.get_design(db, kind, doc_type).get("elements", [])
        if e.get("visible", True))
    if not needs_total:
        _doc(out, land, db, doc_type).build(story, onFirstPage=make_hf(None),
                                            onLaterPages=make_hf(None))
        return
    import copy
    probe = copy.deepcopy(story)
    counter = {"n": 0}

    class _Counter(SimpleDocTemplate):
        def afterPage(self):
            counter["n"] = self.page

    tmp = out.with_suffix(".count.pdf")
    c = _Counter(str(tmp), pagesize=_doc(out, land, db, doc_type).pagesize,
                 topMargin=_doc(out, land, db, doc_type).topMargin,
                 bottomMargin=_doc(out, land, db, doc_type).bottomMargin,
                 leftMargin=12 * mm, rightMargin=12 * mm)
    try:
        c.build(probe, onFirstPage=make_hf(None), onLaterPages=make_hf(None))
        total = counter["n"] or 1
    except Exception:
        total = 1
    finally:
        try:
            tmp.unlink()
        except OSError:
            pass
    _doc(out, land, db, doc_type).build(story, onFirstPage=make_hf(total),
                                        onLaterPages=make_hf(total))


def _header_footer(db: Database, doc_title: str = "", is_report: bool = False,
                   layout: dict | None = None, doc_row=None, doc_type: str = "__default__",
                   total_pages=None):
    """Return a page-decorator that paints the designed header and footer."""
    from . import header_design as HD

    layout = layout or {}
    header = HD.get_design(db, "header", doc_type)
    footer = HD.get_design(db, "footer", doc_type)

    # per-document-type overrides coming from the Document Designer tab
    if (layout.get("header_band_color") or "").strip():
        header = dict(header)
        header["bg_color1"] = layout["header_band_color"]
    if (layout.get("header_band_color2") or "").strip():
        header = dict(header)
        header["bg_color2"] = layout["header_band_color2"]
    if (layout.get("footer_note") or "").strip():
        pass  # picked up through the context

    logo = db.get_setting("logo_path", "")
    show_key = "logo_show_on_reports" if is_report else "logo_show_on_docs"
    if not db.get_bool(show_key, True) or \
            str(layout.get("show_logo", "1")) not in ("1", "True", "true"):
        header = dict(header)
        header["logo_show"] = False
    watermark = db.get_bool("logo_watermark", False)
    base_ctx = HD.context(db, doc_title, doc_row,
                          {"footer_note": (layout.get("footer_note") or "").strip()
                           or db.get_setting("doc_footer", "")})

    def draw(canvas, doc):
        canvas.saveState()
        w, h = doc.pagesize
        if watermark and logo and Path(logo).exists():
            try:
                canvas.saveState()
                canvas.setFillAlpha(0.06)
                canvas.drawImage(logo, w / 2 - 55 * mm, h / 2 - 55 * mm, width=110 * mm,
                                 height=110 * mm, preserveAspectRatio=True, mask="auto")
                canvas.restoreState()
            except Exception:
                pass
        ctx = dict(base_ctx)
        ctx["page"] = str(doc.page)
        ctx["pages"] = str(total_pages or getattr(doc, "_aurco_total_pages", "") or "")
        _render_band(canvas, doc, db, header, ctx, is_footer=False, logo=logo)
        _render_band(canvas, doc, db, footer, ctx, is_footer=True, logo=logo)
        canvas.restoreState()

    return draw


def _doc(path: Path, landscape_mode: bool = False, db: Database | None = None,
         doc_type: str = "__default__") -> SimpleDocTemplate:
    size = landscape(A4) if landscape_mode else A4
    top, bot = 28 * mm, 17 * mm
    if db is not None:
        try:
            from . import header_design as HD
            hd = HD.get_design(db, "header", doc_type or "__default__")
            fd = HD.get_design(db, "footer", doc_type or "__default__")
            top = (float(hd.get("height", 24)) + 6) * mm
            bot = (float(fd.get("height", 14)) + 3) * mm
        except Exception:
            pass
    return SimpleDocTemplate(str(path), pagesize=size, topMargin=top,
                             bottomMargin=bot, leftMargin=12 * mm, rightMargin=12 * mm,
                             title=path.stem, author=f"AURCO / {config.CREATED_BY}")


def _rule(primary, accent, width: float) -> Table:
    """Thin two-tone rule used under a report title."""
    t = Table([[""]], colWidths=[width], rowHeights=[1.4 * mm])
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (0, 0), primary),
                           ("LINEBELOW", (0, 0), (0, 0), 1.1, accent)]))
    return t


def _kv_block(pairs: Sequence[tuple[str, Any]], cols: int = 2,
              total_width: float = 186 * mm) -> Table:
    """Label / value grid that always fits the page.

    The column widths used to be hardcoded at (28+63)mm per pair, so cols=3 was
    273mm on a 186mm text width and the first label was clipped off the left
    edge. Widths are now derived from the page width.
    """
    cols = max(1, int(cols))
    pair_w = total_width / cols
    label_w = max(20 * mm, min(30 * mm, pair_w * 0.34))
    value_w = pair_w - label_w
    def _safe(x: Any) -> str:
        """Escape text for ReportLab's mini-HTML.

        A value like "Jafura L&T" is not valid markup: it was silently rendered
        as "Jafura L&T;". Values that already carry intentional markup (<b>,
        <font ...>) are passed through untouched.
        """
        t = str(x)
        if "<" in t and ">" in t:
            return t
        return t.replace("&", "&amp;")

    data, row = [], []
    for k, v in pairs:
        row += [Paragraph(f"<b>{_safe(k)}</b>", P_SM),
                Paragraph(_safe(v if str(v or "").strip() else "-"), P_SM)]
        if len(row) >= cols * 2:
            data.append(row)
            row = []
    if row:
        row += [""] * (cols * 2 - len(row))
        data.append(row)
    if not data:
        data = [[""] * (cols * 2)]
    t = Table(data, colWidths=[label_w, value_w] * cols)
    style = [
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c9d6e2")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]
    for c in range(cols):                      # tint every label column
        style.append(("BACKGROUND", (c * 2, 0), (c * 2, -1), LIGHT))
    t.setStyle(TableStyle(style))
    return t


def _is_number(v: Any) -> bool:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return True
    t = str(v or "").strip().replace(",", "").replace("%", "").replace("+", "")
    if not t or t in ("-", "."):
        return False
    try:
        float(t)
        return True
    except ValueError:
        return False


def _num_value(v: Any) -> float:
    try:
        return float(str(v).replace(",", "").replace("%", "").replace("+", "").strip())
    except (TypeError, ValueError):
        return 0.0


_STATUS_TINT = {
    "normal": "#1a9c52", "available": "#1a9c52", "full available": "#1a9c52",
    "delivered": "#1a9c52", "done": "#1a9c52", "final": "#1a9c52", "match": "#1a9c52",
    "warning": "#e0a300", "partial": "#e0a300", "partial available": "#e0a300",
    "partially delivered": "#e0a300", "draft": "#e0a300", "preparing": "#1098ad",
    "ready": "#0b7285", "in progress": "#1098ad",
    "critical": "#e8590c", "shortage": "#e8590c", "excess": "#1a9c52",
    "out of stock": "#c92a2a", "not available": "#c92a2a", "no stock": "#c92a2a",
    "not found": "#7048e8", "item not found": "#7048e8", "reversed": "#c92a2a",
    "cancelled": "#868e96", "blocked": "#c92a2a", "urgent": "#c92a2a",
    "damaged": "#c92a2a", "usable": "#1a9c52",
}


def _grid(cols: list[str], rows: list[list[Any]], widths=None, font=7.6,
          header_color=None, stripe: bool = True, *, db: Database | None = None,
          zebra: str = "#f4f8fb", totals_row: bool = False,
          highlight_status: bool = True, compact: bool = False) -> Table:
    """A polished data table.

    Numbers are right-aligned and thousands-separated, status words are tinted,
    negative values turn red, an optional TOTAL row is emphasised, and long text
    wraps instead of overflowing.
    """
    header_color = header_color or BRAND_BLUE
    try:
        head_dark = (0.299 * header_color.red + 0.587 * header_color.green
                     + 0.114 * header_color.blue) < 0.62
    except Exception:
        head_dark = True
    head_fg = colors.white if head_dark else colors.HexColor("#101418")

    pad = 1.8 if compact else 2.6
    h_style = ParagraphStyle("gh", parent=P_SM, fontSize=font, leading=font * 1.25,
                             textColor=head_fg, alignment=TA_CENTER)
    head = [Paragraph(f"<b>{c}</b>", h_style) for c in cols]

    # decide column alignment from the data itself
    ncol = len(cols)
    numeric_col = [True] * ncol
    for r in rows:
        for c in range(min(ncol, len(r))):
            v = r[c]
            if str(v or "").strip() and not _is_number(v):
                numeric_col[c] = False
    for c, name in enumerate(cols):
        key = str(name).strip().lower()
        if key in ("sr.", "sr", "#", "line", "no"):
            numeric_col[c] = True
        elif any(tok in key for tok in ("no.", "number", "code", "ref", "batch",
                                        "serial", "phone", "iqama", "project")):
            # identifiers stay verbatim: "001582" must not become "1,582"
            numeric_col[c] = False

    body = []
    for r in rows:
        line = []
        for c in range(ncol):
            v = r[c] if c < len(r) else ""
            txt = "" if v is None else str(v)
            colour = None
            if highlight_status and txt:
                colour = _STATUS_TINT.get(txt.strip().lower())
            if numeric_col[c] and _is_number(v):
                n = _num_value(v)
                if isinstance(v, float) or "." in str(v):
                    txt = f"{n:,.2f}"
                else:
                    txt = f"{n:,.0f}"
                if str(v).strip().startswith("+"):
                    txt = "+" + txt
                if n < 0:
                    colour = colour or "#c92a2a"
            align = TA_RIGHT if numeric_col[c] else TA_LEFT
            st = ParagraphStyle(f"c{c}", parent=P_SM, fontSize=font,
                                leading=font * 1.28, alignment=align)
            if colour:
                st = ParagraphStyle(f"c{c}s", parent=st, textColor=colors.HexColor(colour))
                txt = f"<b>{txt}</b>"
            line.append(Paragraph(txt, st))
        body.append(line)

    t = Table([head] + body, colWidths=widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), header_color),
        ("LINEBELOW", (0, 0), (-1, 0), 0.9, header_color),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#c9d6e2")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), pad),
        ("BOTTOMPADDING", (0, 0), (-1, -1), pad),
        ("LEFTPADDING", (0, 0), (-1, -1), 3.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3.5),
    ]
    if stripe:
        try:
            zcol = colors.HexColor(zebra)
        except Exception:
            zcol = colors.HexColor("#f4f8fb")
        for i in range(1, len(body) + 1):
            if i % 2 == 0:
                style.append(("BACKGROUND", (0, i), (-1, i), zcol))
    if totals_row and body:
        last = len(body)
        style += [("BACKGROUND", (0, last), (-1, last), colors.HexColor("#eef3f8")),
                  ("LINEABOVE", (0, last), (-1, last), 0.9, header_color),
                  ("TOPPADDING", (0, last), (-1, last), pad + 1.5),
                  ("BOTTOMPADDING", (0, last), (-1, last), pad + 1.5)]
    t.setStyle(TableStyle(style))
    return t


def _signatures(db: Database, blocks: list[dict], layout: dict | None = None) -> Table:
    """Signature strip: role, optional signature image, printed name + designation.

    `blocks` = [{role, name, designation, signature_path}] from signatories.py
    """
    layout = layout or {}
    n = max(1, len(blocks))
    try:
        sig_h = float(layout.get("signature_height", 18) or 18)
    except (TypeError, ValueError):
        sig_h = 18.0
    show_dt = db.get_bool("show_signature_datetime", True)

    img_row, role_row, name_row, desg_row, date_row = [], [], [], [], []
    any_image = False
    for b in blocks:
        path = b.get("signature_path") or ""
        cell: Any = ""
        if path and Path(path).exists():
            try:
                cell = Image(path, width=(170 / n) * mm * 0.72, height=sig_h * mm * 0.7,
                             kind="proportional")
                any_image = True
            except Exception:
                cell = ""
        img_row.append(cell)
        role_row.append(Paragraph(f"<b>{b.get('role','')}</b>", P_SM))
        name_row.append(Paragraph(b.get("name") or "&nbsp;", P_SM))
        sub = b.get("designation") or "Name / Signature"
        if db.get_bool("show_handover_id", True):
            ident = " · ".join(x for x in (b.get("id_number") or "",
                                           b.get("phone") or "") if x)
            if ident:
                sub = f"{sub}<br/>ID: {ident}" if b.get("designation") else f"ID: {ident}"
        desg_row.append(Paragraph(f"<font size=6.5 color='#6b7c8f'>{sub}</font>", P_SM))
        date_row.append(Paragraph("<font size=6.5 color='#6b7c8f'>Date: "
                                  "________________</font>", P_SM) if show_dt else "")

    # when nobody has a stored signature image, keep a smaller blank strip for
    # wet signing instead of a large empty band
    strip = sig_h * mm if any_image else min(sig_h, 12) * mm
    has_ident = any((b.get("id_number") or b.get("phone")) for b in blocks) and \
        db.get_bool("show_handover_id", True)
    data = [img_row, role_row, name_row, desg_row]
    heights = [strip, 6 * mm, 6 * mm, (8.5 if has_ident else 5) * mm]
    if show_dt:
        data.append(date_row)
        heights.append(5 * mm)

    col_w = (186 / n) * mm
    t = Table(data, colWidths=[col_w] * n, rowHeights=heights)
    style_name = (layout.get("signature_line_style")
                  or db.get_setting("signature_line_style", "Line"))
    primary, accent = _brand_colors(db)
    cmds = [
        ("VALIGN", (0, 0), (-1, 0), "BOTTOM"),
        ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]
    if style_name == "Box":
        cmds += [("BOX", (0, 0), (-1, -1), 0.6, primary),
                 ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#c9d6e2"))]
    elif style_name != "None":
        # a single signing rule under each column -- no vertical dividers, which
        # otherwise print as stray lines through the signature area
        cmds += [("LINEBELOW", (0, 0), (-1, 0), 0.8, primary),
                 ("LEFTPADDING", (0, 0), (-1, -1), 4),
                 ("RIGHTPADDING", (0, 0), (-1, -1), 4)]
    t.setStyle(TableStyle(cmds))
    return t


def _attachment_rows(db: Database, doc_type: str, doc_no: str):
    return db.query(
        "SELECT file_path, added_at, COALESCE(source,'file') AS source,"
        " COALESCE(page_order,1) AS page_order"
        " FROM attachments WHERE doc_type=? AND doc_no=?"
        " ORDER BY COALESCE(page_order,1), id",
        (doc_type, doc_no))


def _attachment_block(db: Database, doc_type: str, doc_no: str) -> list[Any]:
    """List the supporting documents attached to this document."""
    rows = _attachment_rows(db, doc_type, doc_no)
    if not rows:
        return []
    data = []
    for i, r in enumerate(rows, 1):
        label = Path(r["file_path"]).name
        if (r["source"] or "").strip().lower() == "clipboard":
            label += " (pasted)"
        data.append([i, label, r["added_at"] or ""])
    return [Spacer(1, 3 * mm),
            Paragraph("<b>Supporting documents attached</b>", P_MD),
            Spacer(1, 1.5 * mm),
            _grid(["#", "File", "Attached on"], data, [10 * mm, 120 * mm, 40 * mm], font=7.5)]


# ------------------------------------------------------------ document PDFs
DOC_TITLES = {"DN": "Delivery Note", "GRN": "Goods Receipt Note", "RET": "Return Note",
              "ADJ": "Stock Adjustment Note", "TRF": "Stock Transfer Note",
              "CNT": "Physical Stock Count Sheet"}
DOC_FOLDERS = {"DN": "Delivery Notes", "GRN": "Inventory", "RET": "Returns",
               "ADJ": "Stock Adjustments", "TRF": "Stock Transfers", "CNT": "Stock Counts"}
REVERSED_DOC_FOLDERS = {
    "DN": "Reversed Delivery Notes",
    "GRN": "Reversed Inventory",
    "RET": "Reversed Returns",
    "ADJ": "Reversed Stock Adjustments",
    "TRF": "Reversed Stock Transfers",
    "CNT": "Reversed Stock Counts",
}


def document_output_folder(doc_type: str, status: str = "FINAL") -> Path:
    status = str(status or "").upper()
    if status == "REVERSED":
        return config.folder(REVERSED_DOC_FOLDERS.get(doc_type, "Reports"))
    return config.folder(DOC_FOLDERS.get(doc_type, "Reports"))


def reversed_document_basename(db: Database, doc_row, lines) -> str:
    base = document_basename(db, doc_row, lines)
    suffix = " - REVERSED"
    if len(base) > MAX_NAME_LEN - len(suffix):
        base = base[:MAX_NAME_LEN - len(suffix)].rstrip(" .-_")
    return f"{base}{suffix}"


def _append_attachments(db: Database, doc_type: str, doc_no: str, pdf_path: Path) -> int:
    """Append the document's attachments as extra pages after the document.

    PDFs are merged page-for-page; images are placed on their own A4 page. Files
    that cannot be embedded (Word, Excel, e-mail...) get a reference page so the
    printed pack still records that they exist.
    Returns the number of pages added.
    """
    rows = _attachment_rows(db, doc_type, doc_no)
    if not rows:
        return 0
    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError:
        return 0

    writer = PdfWriter()
    try:
        for page in PdfReader(str(pdf_path)).pages:
            writer.add_page(page)
    except Exception:
        return 0

    added = 0
    for r in rows:
        f = Path(r["file_path"])
        if not f.exists():
            continue
        suffix = f.suffix.lower()
        try:
            if suffix == ".pdf":
                for page in PdfReader(str(f)).pages:
                    writer.add_page(page)
                    added += 1
            elif suffix in (".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tif", ".tiff", ".webp"):
                buf = io.BytesIO()
                c = rl_canvas.Canvas(buf, pagesize=A4)
                _attachment_caption(c, db, doc_no, f.name)
                try:
                    img = ImageReader(str(f))
                    iw, ih = img.getSize()
                    max_w, max_h = A4[0] - 24 * mm, A4[1] - 46 * mm
                    scale = min(max_w / iw, max_h / ih)
                    w_, h_ = iw * scale, ih * scale
                    c.drawImage(img, (A4[0] - w_) / 2, (A4[1] - h_) / 2 - 6 * mm,
                                width=w_, height=h_, preserveAspectRatio=True, mask="auto")
                except Exception:
                    c.setFont("Helvetica", 10)
                    c.drawCentredString(A4[0] / 2, A4[1] / 2, f"[ {f.name} ]")
                c.showPage()
                c.save()
                buf.seek(0)
                for page in PdfReader(buf).pages:
                    writer.add_page(page)
                    added += 1
            else:
                buf = io.BytesIO()
                c = rl_canvas.Canvas(buf, pagesize=A4)
                _attachment_caption(c, db, doc_no, f.name)
                c.setFont("Helvetica", 11)
                c.drawCentredString(A4[0] / 2, A4[1] / 2 + 6 * mm,
                                    "This attachment cannot be embedded in the PDF.")
                c.setFont("Helvetica-Bold", 12)
                c.drawCentredString(A4[0] / 2, A4[1] / 2 - 2 * mm, f.name)
                c.setFont("Helvetica", 8)
                c.drawCentredString(A4[0] / 2, A4[1] / 2 - 12 * mm, str(f))
                c.showPage()
                c.save()
                buf.seek(0)
                for page in PdfReader(buf).pages:
                    writer.add_page(page)
                    added += 1
        except Exception:
            continue

    if added:
        tmp = pdf_path.with_suffix(".merged.pdf")
        with open(tmp, "wb") as fh:
            writer.write(fh)
        tmp.replace(pdf_path)
    return added


def _attachment_caption(c, db: Database, doc_no: str, name: str) -> None:
    """Small banner at the top of an appended attachment page."""
    primary, accent = _brand_colors(db)
    c.setFillColor(primary)
    c.rect(0, A4[1] - 14 * mm, A4[0], 14 * mm, stroke=0, fill=1)
    c.setFillColor(accent)
    c.rect(0, A4[1] - 15 * mm, A4[0], 1 * mm, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(12 * mm, A4[1] - 9 * mm, f"ATTACHMENT — {doc_no}")
    c.setFont("Helvetica", 9)
    c.drawRightString(A4[0] - 12 * mm, A4[1] - 9 * mm, name[:70])


def document_pdf(db: Database, doc_id: int, out_path: str | Path | None = None,
                 include_attachments: bool | None = None) -> Path:
    d = db.one("SELECT * FROM documents WHERE id=?", (doc_id,))
    if d is None:
        raise ValueError("Document not found")
    lines = db.query("SELECT * FROM document_lines WHERE doc_id=? ORDER BY id", (doc_id,))
    dtype = d["doc_type"]
    title = DOC_TITLES.get(dtype, dtype)
    from . import signatories as SG

    def _g(k):
        """Safe column read (older databases may lack newer columns)."""
        try:
            return d[k] or ""
        except (IndexError, KeyError):
            return ""

    layout = SG.get_layout(db, dtype)
    status = str(d["status"] or "").upper()
    base_name = (reversed_document_basename(db, d, lines) if status == "REVERSED"
                 else document_basename(db, d, lines))
    out = Path(out_path) if out_path else (document_output_folder(dtype, status) /
                                           f"{base_name}.pdf")
    out.parent.mkdir(parents=True, exist_ok=True)
    story: list[Any] = [Paragraph(title, P_TITLE),
                        Paragraph(f"Document No: <b>{d['doc_no']}</b> &nbsp;|&nbsp; Status: "
                                  f"<b>{d['status']}</b>", P_SUB), Spacer(1, 5 * mm)]

    common = [("Document No", d["doc_no"]), ("Date", d["doc_date"])]
    if dtype == "DN":
        def _g(k):
            try:
                return d[k] or ""
            except (IndexError, KeyError):
                return ""
        # Gate-pass style header: only the fields the storekeeper needs at the
        # gate. Everything else (requester, purpose, signatories) sits in the
        # body / footer of the document.
        pairs = [("DN Number", d["doc_no"]), ("Date", d["doc_date"]),
                 ("From", _g("from_location") or d["warehouse"]),
                 ("Project", d["project"]),
                 ("Vehicle", d["vehicle"]),
                 ("In Time", _g("in_time")),
                 ("Out Time", _g("out_time")),
                 ("Reference / MR", d["reference"])]
        _hid, _hph = _g("handover_id"), _g("handover_phone")
        if _g("handover_to") or _hid or _hph:
            pairs += [("Handover To (Driver)", _g("handover_to")),
                      ("ID / Iqama · Phone",
                       " · ".join(x for x in (_hid, _hph) if x))]
        if SG.layout_bool(layout, "show_extra_header", False):
            pairs += [("Department", d["department"]),
                      ("Requested By", d["requested_by"]),
                      ("Issued To", d["issued_to"]), ("Purpose", d["purpose"]),
                      ("Driver", d["driver"]), ("Prepared By", d["created_by"])]
        cols = ["Sr.", "Item Code", "Description", "UOM", "Quantity", "PR / MR No.", "Remarks"]
        widths = [10 * mm, 27 * mm, 60 * mm, 15 * mm, 19 * mm, 30 * mm, 25 * mm]
        rows = [[i, l["item_code"], l["description"], l["uom"], f"{l['qty']:g}",
                 (l["pr_no"] if "pr_no" in l.keys() else "") or "", l["remarks"]]
                for i, l in enumerate(lines, 1)]
        sigs = None
    elif dtype == "GRN":
        pairs = common + [("Supplier", d["supplier"]), ("PO / MR Reference", d["reference"]),
                          ("Warehouse", d["warehouse"]), ("Location", d["location"]),
                          ("Received By", d["received_by"]), ("Prepared By", d["created_by"])]
        cols = ["Sr.", "Item Code", "Description", "UOM", "Qty", "Unit Cost", "Total", "Batch"]
        widths = [10 * mm, 27 * mm, 60 * mm, 15 * mm, 17 * mm, 20 * mm, 22 * mm, 15 * mm]
        rows = [[i, l["item_code"], l["description"], l["uom"], f"{l['qty']:g}",
                 f"{l['unit_cost']:,.2f}", f"{l['total_cost']:,.2f}", l["batch"]]
                for i, l in enumerate(lines, 1)]
        sigs = None
    elif dtype == "RET":
        pairs = common + [("Original DN", d["linked_doc"]), ("Returned By", d["returned_by"]),
                          ("Received By", d["received_by"]), ("Project / Site", d["project"]),
                          ("Department", d["department"]), ("Warehouse", d["warehouse"])]
        cols = ["Sr.", "Item Code", "Description", "UOM", "Issued Qty", "Returned Qty",
                "Condition", "Remarks"]
        widths = [10 * mm, 26 * mm, 55 * mm, 14 * mm, 20 * mm, 22 * mm, 20 * mm, 19 * mm]
        rows = [[i, l["item_code"], l["description"], l["uom"], f"{l['issued_qty']:g}",
                 f"{l['qty']:g}", l["condition"], l["remarks"]] for i, l in enumerate(lines, 1)]
        sigs = None
    elif dtype == "TRF":
        pairs = common + [("From Warehouse", d["warehouse"]), ("To Warehouse", d["to_warehouse"]),
                          ("From Location", d["location"]), ("To Location", d["to_location"]),
                          ("Issued By", d["issued_to"]), ("Responsible / Received By", d["received_by"])]
        cols = ["Sr.", "Item Code", "Description", "UOM", "Quantity", "Remarks"]
        widths = [12 * mm, 30 * mm, 78 * mm, 18 * mm, 22 * mm, 26 * mm]
        rows = [[i, l["item_code"], l["description"], l["uom"], f"{l['qty']:g}", l["remarks"]]
                for i, l in enumerate(lines, 1)]
        sigs = None
    elif dtype == "ADJ":
        pairs = common + [("Reason", d["reason"]), ("Warehouse", d["warehouse"]),
                          ("Reference", d["linked_doc"]), ("Prepared By", d["created_by"])]
        cols = ["Sr.", "Item Code", "Description", "UOM", "Adjustment (+/-)", "Remarks"]
        widths = [12 * mm, 30 * mm, 70 * mm, 18 * mm, 28 * mm, 28 * mm]
        rows = [[i, l["item_code"], l["description"], l["uom"], f"{l['qty']:+g}", l["remarks"]]
                for i, l in enumerate(lines, 1)]
        sigs = None
    else:  # CNT
        pairs = common + [("Warehouse", d["warehouse"]), ("Location", d["location"]),
                          ("Counted By", d["received_by"]), ("Status", d["status"])]
        cols = ["Sr.", "Item Code", "Description", "UOM", "System Qty", "Counted Qty", "Variance"]
        widths = [12 * mm, 30 * mm, 66 * mm, 18 * mm, 22 * mm, 22 * mm, 22 * mm]
        rows = [[i, l["item_code"], l["description"], l["uom"], f"{l['system_qty']:g}",
                 f"{l['counted_qty']:g}" if d["status"] != "DRAFT" or l["counted_qty"] else "",
                 f"{l['variance']:+g}" if l["variance"] else ""]
                for i, l in enumerate(lines, 1)]
        sigs = None

    try:
        _fs = float(layout.get("font_size", 7.6) or 7.6)
    except (TypeError, ValueError):
        _fs = 7.6
    _hdr = layout.get("header_color") or ""
    story += [_kv_block(pairs), Spacer(1, 4 * mm),
              _grid(cols, rows, widths, font=_fs,
                    header_color=colors.HexColor(_hdr) if _hdr else _brand_colors(db)[0],
                    stripe=SG.layout_bool(layout, "row_stripe", True))]

    # Optional per-PR recap (off by default -- the PR number is already on every
    # line of the items table). Enable in Settings -> Document Designer.
    if dtype == "DN" and SG.layout_bool(layout, "show_pr_recap", False):
        groups: dict[str, dict] = {}
        for l in lines:
            pr = ((l["pr_no"] if "pr_no" in l.keys() else "") or "").strip() or "(no PR)"
            g = groups.setdefault(pr, {"lines": 0, "qty": 0.0})
            g["lines"] += 1
            g["qty"] += float(l["qty"] or 0)
        if groups:
            recap = [[pr, g["lines"], f"{g['qty']:g}"] for pr, g in sorted(groups.items())]
            story += [Spacer(1, 4 * mm),
                      Paragraph("<b>Purchase / Material Requests covered by this Delivery Note</b>", P_MD),
                      Spacer(1, 1.5 * mm),
                      _grid(["PR / MR Number", "Lines", "Total Qty"], recap,
                            [60 * mm, 25 * mm, 30 * mm], font=8)]

    tot_qty = sum(float(l["qty"] or 0) for l in lines)
    tot_val = sum(float(l["total_cost"] or 0) for l in lines)
    summary = f"Total Lines: <b>{len(lines)}</b> &nbsp;&nbsp; Total Quantity: <b>{tot_qty:g}</b>"
    if tot_val:
        summary += f" &nbsp;&nbsp; Total Value: <b>{db.get_setting('currency','')} {tot_val:,.2f}</b>"
    story += [Spacer(1, 3 * mm), Paragraph(summary, P_MD)]
    if d["remarks"]:
        story += [Spacer(1, 2 * mm), Paragraph(f"<b>Remarks:</b> {d['remarks']}", P_SM)]

    if SG.layout_bool(layout, "show_attachments", True):
        story += _attachment_block(db, dtype, d["doc_no"])

    if SG.layout_bool(layout, "show_terms", False) and (layout.get("terms_text") or "").strip():
        story += [Spacer(1, 3 * mm),
                  Paragraph("<b>Terms &amp; conditions</b>", P_MD),
                  Paragraph(layout["terms_text"].replace("\n", "<br/>"), P_SM)]

    overrides = {r["role"]: dict(r) for r in db.query(
        "SELECT * FROM document_signatures WHERE doc_type=? AND doc_no=?", (dtype, d["doc_no"]))}
    # Fall back to the names stored on the document header itself, so a document
    # created outside the form (import, material request, API) still prints the
    # right people against each signature block.
    _hdr_map = {
        SG.ROLE_ISSUED_BY: _g("issued_by"),
        SG.ROLE_DELIVERED_BY: _g("delivered_by"),
        SG.ROLE_HANDOVER_TO: _g("handover_to"),
        SG.ROLE_RECEIVED_BY: d["received_by"] or "",
        SG.ROLE_RETURNED_BY: d["returned_by"] or "",
    } if dtype in ("DN", "RET", "TRF") else {}
    for _role, _name in _hdr_map.items():
        if _name and not overrides.get(_role, {}).get("name"):
            known = SG.find_signatory(db, _name) or {}
            entry = {"role": _role, "name": _name,
                     "designation": known.get("designation", ""),
                     "signature_path": (known.get("signature_path", "")
                                        if db.get_bool("print_signature_images", True) else ""),
                     "id_number": known.get("id_number", ""),
                     "phone": known.get("phone", "")}
            if _role == SG.ROLE_HANDOVER_TO:
                entry["id_number"] = _g("handover_id") or entry["id_number"]
                entry["phone"] = _g("handover_phone") or entry["phone"]
            overrides[_role] = entry
    blocks = SG.resolve_blocks(db, dtype, overrides)

    # ---- footer area: the authorised signature sits at the TOP of it, so it is
    # the first thing seen under the item table rather than buried at the page
    # bottom. `_signatures` is pushed to the bottom of the frame only when the
    # user asks for it in the Document Designer.
    _cap = layout.get("signature_caption") or "Authorised Signatures"
    sig_flow = Table(
        [[Paragraph(f"<b>{_cap}</b>", P_MD)],
         [_signatures(db, blocks, layout)]],
        colWidths=[186 * mm])
    sig_flow.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (0, 0), 0),
        ("BOTTOMPADDING", (0, 0), (0, 0), 3),
        ("TOPPADDING", (0, 1), (0, 1), 0),
        ("BOTTOMPADDING", (0, 1), (0, 1), 0),
    ]))
    if SG.layout_bool(layout, "signature_inline", False):
        # legacy behaviour: signatures immediately under the item table
        story += [Spacer(1, 6 * mm), sig_flow]
    else:
        # default: pin the signatures to the bottom of the page, directly above
        # the footer line
        story += [Spacer(1, 6 * mm), BottomAnchored(sig_flow)]

    land = (layout.get("orientation", "Portrait") == "Landscape")
    _build_with_totals(out, story, land, db, dtype,
                       lambda total: _header_footer(db, title, False, layout, d, dtype,
                                                    total_pages=total))
    if include_attachments is None:
        include_attachments = SG.layout_bool(layout, "merge_attachments", True)
    n_att = 0
    if include_attachments:
        n_att = _append_attachments(db, dtype, d["doc_no"], out)

    db.execute("UPDATE documents SET pdf_path=? WHERE id=?", (str(out), doc_id))
    db.commit()
    db.audit("PRINTED", dtype, d["doc_no"],
             f"PDF -> {out.name}" + (f" (+{n_att} attachment page(s))" if n_att else ""))
    return out


def _summary_cards(db: Database, stats: list[tuple[str, str, str]], width: float) -> Table:
    """Row of KPI tiles printed under a report title."""
    n = max(1, len(stats))
    cells = []
    for label, value, colour in stats:
        inner = Table([[Paragraph(f"<font size=7.5 color='#6b7c8f'>{label.upper()}</font>",
                                  P_SM)],
                       [Paragraph(f"<b><font size=13 color='{colour}'>{value}</font></b>",
                                  P_SM)]],
                      colWidths=[width / n - 3 * mm])
        inner.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#d8e1ea")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (0, 0), 4),
            ("BOTTOMPADDING", (0, 1), (0, 1), 5),
        ]))
        cells.append(inner)
    t = Table([cells], colWidths=[width / n] * n)
    t.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0),
                           ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))
    return t


def _auto_stats(db: Database, cols: list[str], rows: list[list[Any]]
                ) -> list[tuple[str, str, str]]:
    """Pick a few meaningful figures out of any report table."""
    out: list[tuple[str, str, str]] = []
    cur = db.get_setting("currency", "")
    body = rows
    if rows:
        first = str(rows[-1][0] if rows[-1] else "").strip().lower()
        joined = " ".join(str(x).strip().lower() for x in rows[-1])
        if first in ("", "total") or joined.startswith("total") or " total " in f" {joined} ":
            body = rows[:-1]          # ignore an existing TOTAL line
    wanted = {"balance": "Total Balance", "quantity": "Total Qty", "qty": "Total Qty",
              "total qty": "Total Qty", "value": "Total Value",
              "stock value": "Total Value", "requested": "Requested",
              "short by": "Shortage", "issued qty": "Issued", "variance": "Variance"}
    for i, c in enumerate(cols):
        key = str(c).strip().lower()
        for k, label in wanted.items():
            if key == k or key.startswith(k + " ("):
                total = sum(_num_value(r[i]) for r in body
                            if i < len(r) and _is_number(r[i]))
                if not total:
                    break
                money = "value" in key
                colour = "#c92a2a" if ("short" in key and total) else "#0b3d6b"
                out.append((label, (f"{cur} {total:,.0f}" if money else f"{total:,.2f}"),
                            colour))
                break
        if len(out) >= 5:
            break
    out.insert(0, ("Records", f"{len(body):,}", "#0b3d6b"))
    if "Status" in cols:
        i = cols.index("Status")
        bad = sum(1 for r in body if i < len(r)
                  and str(r[i]).strip().lower() in ("out of stock", "critical",
                                                    "not available", "no stock"))
        if bad:
            out.append(("Needs Attention", f"{bad:,}", "#c92a2a"))
    return out[:6]


def report_pdf(db: Database, title: str, cols: list[str], rows: list[list[Any]],
               out_path: str | Path | None = None, subtitle: str = "",
               stats: list[tuple[str, str, str]] | None = None,
               totals_row: bool = False) -> Path:
    out = Path(out_path) if out_path else (config.folder("Reports") /
                                           f"{safe_name(title)}_{_dt.datetime.now():%Y%m%d_%H%M%S}.pdf")
    out.parent.mkdir(parents=True, exist_ok=True)
    land = len(cols) > 7
    page_w = (landscape(A4)[0] if land else A4[0]) - 24 * mm
    primary, accent = _brand_colors(db)

    # column widths proportional to the longest content, within sane limits
    sample = rows[:200]
    weights = []
    for i, c in enumerate(cols):
        longest = len(str(c))
        for r in sample:
            if i < len(r):
                longest = max(longest, min(46, len(str(r[i] if r[i] is not None else ""))))
        weights.append(max(4.0, float(longest)))
    total_w = sum(weights) or 1
    widths = [max(14 * mm, page_w * w / total_w) for w in weights]
    over = sum(widths) - page_w
    if over > 0:                      # shrink the widest columns back to the page
        order = sorted(range(len(widths)), key=lambda i: -widths[i])
        for i in order:
            room = widths[i] - 14 * mm
            take = min(room, over)
            widths[i] -= take
            over -= take
            if over <= 0.1:
                break

    fs = 7.4 if land else 7.9
    if len(cols) > 12:
        fs = 6.6

    story: list[Any] = [
        Paragraph(title, P_TITLE),
        _rule(primary, accent, page_w),
    ]
    if subtitle:
        story.append(Paragraph(subtitle, P_SUB))
    story.append(Spacer(1, 3 * mm))
    cards = stats if stats is not None else _auto_stats(db, cols, rows)
    if cards:
        story += [_summary_cards(db, cards, page_w), Spacer(1, 4 * mm)]
    if rows and not totals_row:
        joined = " ".join(str(x).strip().lower() for x in rows[-1])
        totals_row = joined.startswith("total") or " total " in f" {joined} "
    if rows:
        story.append(_grid(cols, rows, widths, font=fs, header_color=primary,
                           totals_row=totals_row, compact=len(rows) > 30))
    else:
        story.append(Paragraph("<i>No records match the selected filters.</i>", P_MD))

    _build_with_totals(out, story, land, db, "__default__",
                       lambda total: _header_footer(db, "Report", True,
                                                    total_pages=total))
    db.audit("EXPORTED", "report", title, f"PDF -> {out.name}")
    return out


def item_history_pdf(db: Database, item_id: int, out_path: str | Path | None = None) -> Path:
    from . import services as S
    it = db.one("SELECT * FROM items WHERE id=?", (item_id,))
    hist = S.item_history(db, item_id)
    s = S.item_movement_summary(db, item_id)
    out = Path(out_path) if out_path else (config.folder("Reports") /
                                           f"History_{safe_name(it['code'])}.pdf")
    pairs = [("Item Code", it["code"]), ("Description", it["description"]), ("UOM", it["uom"]),
             ("Category", it["category"]), ("Warehouse", it["warehouse"]),
             ("Location / Rack", f"{it['location']} / {it['rack']}"),
             ("Current Balance", f"{it['balance']:g}"), ("Status", S.stock_status(db, it))]
    calc = [["Opening Balance", f"{s['opening']:g}"],
            ["(+) Stock Received", f"{s['received']:g}"],
            ["(+) Returns", f"{s['returned']:g}"],
            ["(+) Transfers In", f"{s['transfer_in']:g}"],
            ["(-) Stock Issued", f"{s['issued']:g}"],
            ["(-) Transfers Out", f"{s['transfer_out']:g}"],
            ["(±) Adjustments", f"{s['adj_in'] - s['adj_out']:+g}"],
            ["= Current Balance", f"{s['closing']:g}"]]
    cols = ["Date", "Type", "Document", "In", "Out", "Balance", "Party", "Reason", "User"]
    rows = [[h["txn_date"], h["txn_type"], h["doc_no"], f"{h['qty_in']:g}" if h["qty_in"] else "",
             f"{h['qty_out']:g}" if h["qty_out"] else "", f"{h['balance_after']:g}", h["party"],
             h["reason"], h["username"]] for h in hist]
    story = [Paragraph("Item Stock Movement History", P_TITLE), Spacer(1, 4 * mm),
             _kv_block(pairs), Spacer(1, 4 * mm),
             _grid(["Calculation", "Quantity"], calc, [60 * mm, 40 * mm], font=8.5),
             Spacer(1, 5 * mm), _grid(cols, rows, [22 * mm] + [None] * 8, font=7)]
    _doc(out, True, db).build(story, onFirstPage=_header_footer(db, "Transaction History"),
                              onLaterPages=_header_footer(db, "Transaction History"))
    db.audit("EXPORTED", "item", it["code"], f"history PDF -> {out.name}")
    return out


# -------------------------------------------------------------------- Excel
def export_excel(db: Database, title: str, cols: list[str], rows: list[list[Any]],
                 out_path: str | Path | None = None, freeze: bool = True,
                 totals: bool = True) -> Path:
    """Branded Excel export: styled header, banded rows, auto-filter, frozen
    panes, number formats, conditional colours and a totals row."""
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule, DataBarRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo

    out = Path(out_path) if out_path else (config.folder("Exports") /
                                           f"{safe_name(title)}_{_dt.datetime.now():%Y%m%d_%H%M%S}.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    primary = (db.get_setting("ui_primary", "#0b3d6b") or "#0b3d6b").lstrip("#")
    accent = (db.get_setting("ui_accent", "#f5a300") or "#f5a300").lstrip("#")

    wb = Workbook()
    ws = wb.active
    ws.title = safe_name(title)[:31] or "Report"
    company = db.get_setting("company_name", "AURCO")

    ws.append([f"{company}"])
    ws["A1"].font = Font(size=15, bold=True, color=primary)
    ws.append([title])
    ws["A2"].font = Font(size=12, bold=True, color="333333")
    ws.append([f"Generated {_dt.datetime.now():%d-%m-%Y %H:%M}  ·  {len(rows)} record(s)"
               f"  ·  AURCO Inventory Manager"])
    ws["A3"].font = Font(size=9, italic=True, color="808080")
    ws.append([])
    head_row = 5

    ws.append(list(cols))
    thin = Side(style="thin", color="BFCEDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[head_row]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=primary)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[head_row].height = 26

    numeric = [True] * len(cols)
    for r in rows:
        for i in range(min(len(cols), len(r))):
            if str(r[i] or "").strip() and not _is_number(r[i]):
                numeric[i] = False
    for r in rows:
        ws.append([(_num_value(v) if (i < len(numeric) and numeric[i] and _is_number(v))
                    else v) for i, v in enumerate(r)])

    band = PatternFill("solid", fgColor="F4F8FB")
    last = ws.max_row
    for ri, row in enumerate(ws.iter_rows(min_row=head_row + 1, max_row=last,
                                          max_col=len(cols)), start=1):
        for ci, c in enumerate(row):
            c.border = border
            c.font = Font(size=10)
            if ri % 2 == 0:
                c.fill = band
            if isinstance(c.value, (int, float)):
                c.number_format = "#,##0.00" if isinstance(c.value, float) else "#,##0"
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(vertical="center")
                txt = str(c.value or "").strip().lower()
                tint = _STATUS_TINT.get(txt)
                if tint:
                    c.font = Font(size=10, bold=True, color=tint.lstrip("#"))

    # totals row for numeric columns
    if totals and rows and len(rows) > 1:
        trow = last + 1
        ws.cell(row=trow, column=1, value="TOTAL")
        for i, isnum in enumerate(numeric, start=1):
            cell = ws.cell(row=trow, column=i)
            cell.border = border
            cell.fill = PatternFill("solid", fgColor="EEF3F8")
            cell.font = Font(bold=True, size=10, color=primary)
            if isnum and i > 1:
                col = get_column_letter(i)
                cell.value = f"=SUBTOTAL(109,{col}{head_row + 1}:{col}{last})"
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    # widths, filter, freeze
    for i, col in enumerate(cols, 1):
        longest = len(str(col)) + 4
        for r in rows[:500]:
            if i - 1 < len(r):
                longest = max(longest, len(str(r[i - 1] if r[i - 1] is not None else "")) + 2)
        ws.column_dimensions[get_column_letter(i)].width = min(52, max(10, longest))
    ref = f"A{head_row}:{get_column_letter(len(cols))}{last}"
    ws.auto_filter.ref = ref
    if freeze:
        ws.freeze_panes = f"A{head_row + 1}"

    # data bars on the biggest numeric column, red for negatives
    if rows:
        best, width = None, 0
        for i, isnum in enumerate(numeric, start=1):
            if not isnum or i == 1:
                continue
            tot = sum(abs(_num_value(r[i - 1])) for r in rows if i - 1 < len(r))
            if tot > width:
                best, width = i, tot
        if best:
            col = get_column_letter(best)
            rng = f"{col}{head_row + 1}:{col}{last}"
            ws.conditional_formatting.add(rng, DataBarRule(
                start_type="min", end_type="max", color=accent, showValue=True))
        for i, isnum in enumerate(numeric, start=1):
            if isnum:
                col = get_column_letter(i)
                ws.conditional_formatting.add(
                    f"{col}{head_row + 1}:{col}{last}",
                    CellIsRule(operator="lessThan", formula=["0"],
                               font=Font(color="C92A2A", bold=True)))
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = f"{head_row}:{head_row}"
    wb.save(out)
    db.audit("EXPORTED", "report", title, f"Excel -> {out.name}")
    return out


def export_csv(db: Database, title: str, cols: list[str], rows: list[list[Any]],
               out_path: str | Path | None = None) -> Path:
    out = Path(out_path) if out_path else (config.folder("Exports") /
                                           f"{safe_name(title)}_{_dt.datetime.now():%Y%m%d_%H%M%S}.csv")
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(cols)
        w.writerows(rows)
    db.audit("EXPORTED", "report", title, f"CSV -> {out.name}")
    return out


# ------------------------------------------------------------------ sharing
def open_path(path: str | Path) -> None:
    p = Path(path)
    if p.suffix.lower() == ".pdf":
        try:
            from PySide6.QtWidgets import QApplication
            if QApplication.instance() is not None:
                from ..ui import pdf_viewer as _PV
                if _PV.show_pdf(p):
                    return
        except Exception:
            pass
    sp = str(p)
    if platform.system() == "Windows":
        os.startfile(sp)  # type: ignore[attr-defined]
    elif platform.system() == "Darwin":
        subprocess.Popen(["open", sp])
    else:
        subprocess.Popen(["xdg-open", sp])


def open_file_location(path: str | Path) -> None:
    p = Path(path)
    if platform.system() == "Windows":
        subprocess.Popen(["explorer", "/select,", str(p)])
    else:
        open_path(p.parent)


def print_file(db: Database, path: str | Path) -> None:
    """Send a PDF to the default/configured Windows printer."""
    p = Path(path)
    if platform.system() == "Windows":
        printer = db.get_setting("printer_name", "")
        try:
            if printer:
                os.startfile(str(p), "printto", f'"{printer}"')  # type: ignore
            else:
                os.startfile(str(p), "print")  # type: ignore
        except Exception:
            os.startfile(str(p))  # type: ignore
    else:
        open_path(p)
    db.audit("PRINTED", "file", p.name)


def whatsapp_url(number: str = "", message: str = "") -> str:
    num = "".join(ch for ch in str(number or "") if ch.isdigit())
    q = urllib.parse.quote(str(message or ""))
    return f"https://wa.me/{num}?text={q}" if num else f"https://wa.me/?text={q}"


def whatsapp_share(db: Database, path: str | Path, number: str = "", message: str = "") -> str:
    """Practical, API-free workflow: copy the file, open WhatsApp with the message
    pre-filled, user attaches the file (already on the clipboard/explorer)."""
    p = Path(path)
    msg = message or db.get_setting("wa_message", "")
    msg = f"{msg}\n\nDocument: {p.name}"
    num = "".join(ch for ch in (number or db.get_setting("wa_default_number", "")) if ch.isdigit())
    url = whatsapp_url(num, msg)
    try:
        open_file_location(p)
    except Exception:
        pass
    webbrowser.open(url)
    db.audit("EXPORTED", "whatsapp", p.name, f"to {num or 'chooser'}")
    return url


def email_pdf(db: Database, path: str | Path, to_addr: str, subject: str = "",
              body: str = "") -> str:
    """Send via configured SMTP; fall back to the default mail client."""
    import smtplib
    from email.message import EmailMessage

    p = Path(path)
    subject = subject or f"{db.get_setting('company_name','AURCO')} - {p.stem}"
    body = body or f"Please find attached: {p.name}\n\nAURCO Inventory Manager"
    host = db.get_setting("smtp_host", "")
    if not host:
        webbrowser.open(f"mailto:{to_addr}?subject={urllib.parse.quote(subject)}"
                        f"&body={urllib.parse.quote(body + chr(10) + chr(10) + str(p))}")
        try:
            open_file_location(p)
        except Exception:
            pass
        return "Opened your default email client (SMTP is not configured in Settings)."
    msg = EmailMessage()
    msg["From"] = db.get_setting("smtp_from") or db.get_setting("smtp_user")
    msg["To"] = to_addr
    msg["Subject"] = subject
    msg.set_content(body)
    msg.add_attachment(p.read_bytes(), maintype="application",
                       subtype="pdf" if p.suffix.lower() == ".pdf" else "octet-stream",
                       filename=p.name)
    port = int(db.get_setting("smtp_port", 587) or 587)
    if port == 465:
        srv = smtplib.SMTP_SSL(host, port, timeout=30)
    else:
        srv = smtplib.SMTP(host, port, timeout=30)
        if db.get_bool("smtp_tls", True):
            srv.starttls()
    with srv:
        if db.get_setting("smtp_user"):
            srv.login(db.get_setting("smtp_user"), db.get_setting("smtp_pass", ""))
        srv.send_message(msg)
    db.audit("EXPORTED", "email", p.name, f"to {to_addr}")
    return f"Email sent to {to_addr}."


# ------------------------------------------------------------------ barcode
def barcode_pdf(db: Database, items: list[dict], out_path: str | Path | None = None) -> Path:
    """Printable Code128 label sheet (3 columns)."""
    from reportlab.graphics.barcode import code128

    out = Path(out_path) if out_path else (config.folder("Reports") /
                                           f"Barcodes_{_dt.datetime.now():%Y%m%d_%H%M%S}.pdf")
    story: list[Any] = [Paragraph("Item Barcode Labels", P_TITLE), Spacer(1, 4 * mm)]
    cells, row = [], []
    for it in items:
        value = it.get("barcode") or it.get("code")
        bc = code128.Code128(str(value), barHeight=13 * mm, barWidth=0.38 * mm, humanReadable=True)
        inner = Table([[Paragraph(f"<b>{it.get('code','')}</b>", P_SM)],
                       [Paragraph(str(it.get("description", ""))[:38], P_SM)], [bc]],
                      colWidths=[60 * mm])
        inner.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                                   ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                   ("TOPPADDING", (0, 0), (-1, -1), 2)]))
        row.append(inner)
        if len(row) == 3:
            cells.append(row)
            row = []
    if row:
        row += [""] * (3 - len(row))
        cells.append(row)
    if cells:
        t = Table(cells, colWidths=[62 * mm] * 3)
        t.setStyle(TableStyle([("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
        story.append(t)
    _doc(out, False, db).build(story, onFirstPage=_header_footer(db, "Barcode Labels"),
                               onLaterPages=_header_footer(db, "Barcode Labels"))
    db.audit("PRINTED", "barcodes", "", f"{len(items)} label(s)")
    return out


# ------------------------------------------------------ general delivery note
def general_dn_pdf(db: Database, doc_id: int, out_path: str | Path | None = None) -> Path:
    """Delivery Note for the *General DN Maker* — no inventory link at all.

    Uses the same letterhead, grid engine and signature strip as a real DN so the
    printed result is indistinguishable in quality, but nothing here reads or
    writes stock.
    """
    from . import gdn as G
    from . import signatories as SG

    d, lines = G.get(db, doc_id)
    if d is None:
        raise ValueError("General delivery note not found")
    layout = SG.get_layout(db, "DN")
    title = (d.get("title") or "DELIVERY NOTE").strip()
    out = Path(out_path) if out_path else (
        config.folder(G.FOLDER) / f"{safe_name(d['doc_no'])}.pdf")
    out.parent.mkdir(parents=True, exist_ok=True)

    show_val = bool(d.get("show_values"))
    cur = d.get("currency") or db.get_setting("currency", "")

    story: list[Any] = [
        Paragraph(title.title() if title.isupper() else title, P_TITLE),
        Paragraph(f"Document No: <b>{d['doc_no']}</b> &nbsp;|&nbsp; Status: "
                  f"<b>{d.get('status','FINAL')}</b> &nbsp;|&nbsp; "
                  "<i>Standalone document — not linked to inventory stock</i>", P_SUB),
        Spacer(1, 5 * mm)]

    pairs = [("DN Number", d["doc_no"]), ("Date", d["doc_date"]),
             ("From", d.get("from_location", "")),
             ("Deliver To", d.get("to_party", "")),
             ("Project", d.get("project", "")),
             ("Vehicle", d.get("vehicle", "")),
             ("In Time", d.get("in_time", "")),
             ("Out Time", d.get("out_time", "")),
             ("Reference", d.get("reference", ""))]
    if d.get("to_address"):
        pairs.append(("Address", d["to_address"]))
    if d.get("purpose"):
        pairs.append(("Purpose", d["purpose"]))
    if d.get("handover_to") or d.get("handover_id") or d.get("handover_phone"):
        pairs += [("Handover To (Driver)", d.get("handover_to", "")),
                  ("ID / Iqama · Phone",
                   " · ".join(x for x in (d.get("handover_id", ""),
                                          d.get("handover_phone", "")) if x))]

    if show_val:
        cols = ["Sr.", "Item Code", "Description", "UOM", "Quantity",
                f"Unit Price ({cur})", f"Amount ({cur})", "Remarks"]
        widths = [10 * mm, 24 * mm, 55 * mm, 14 * mm, 18 * mm, 22 * mm, 24 * mm, 19 * mm]
        rows = [[i, l.get("item_code", ""), l.get("description", ""), l.get("uom", ""),
                 f"{float(l.get('qty') or 0):g}", f"{float(l.get('unit_cost') or 0):,.2f}",
                 f"{float(l.get('total_cost') or 0):,.2f}", l.get("remarks", "")]
                for i, l in enumerate(lines, 1)]
    else:
        cols = ["Sr.", "Item Code", "Description", "UOM", "Quantity", "Ref / PR",
                "Remarks"]
        widths = [10 * mm, 27 * mm, 62 * mm, 15 * mm, 20 * mm, 27 * mm, 25 * mm]
        rows = [[i, l.get("item_code", ""), l.get("description", ""), l.get("uom", ""),
                 f"{float(l.get('qty') or 0):g}", l.get("pr_no", ""), l.get("remarks", "")]
                for i, l in enumerate(lines, 1)]

    try:
        _fs = float(layout.get("font_size", 7.6) or 7.6)
    except (TypeError, ValueError):
        _fs = 7.6
    story += [_kv_block(pairs), Spacer(1, 4 * mm),
              _grid(cols, rows, widths, font=_fs, header_color=_brand_colors(db)[0],
                    stripe=SG.layout_bool(layout, "row_stripe", True))]

    tot_qty = sum(float(l.get("qty") or 0) for l in lines)
    summary = (f"Total Lines: <b>{len(lines)}</b> &nbsp;&nbsp; "
               f"Total Quantity: <b>{tot_qty:g}</b>")
    if show_val:
        tot = sum(float(l.get("total_cost") or 0) for l in lines)
        summary += f" &nbsp;&nbsp; Total Value: <b>{cur} {tot:,.2f}</b>"
    story += [Spacer(1, 3 * mm), Paragraph(summary, P_MD)]
    if d.get("remarks"):
        story += [Spacer(1, 2 * mm),
                  Paragraph(f"<b>Remarks:</b> {d['remarks']}", P_SM)]
    story += _attachment_block(db, "GDN", d["doc_no"])
    if (d.get("terms") or "").strip():
        story += [Spacer(1, 3 * mm), Paragraph("<b>Terms &amp; conditions</b>", P_MD),
                  Paragraph(d["terms"].replace("\n", "<br/>"), P_SM)]

    overrides = {}
    for role, name, ident, phone in (
            (SG.ROLE_ISSUED_BY, d.get("issued_by", ""), "", ""),
            (SG.ROLE_DELIVERED_BY, d.get("delivered_by", ""), "", ""),
            (SG.ROLE_HANDOVER_TO, d.get("handover_to", ""), d.get("handover_id", ""),
             d.get("handover_phone", "")),
            (SG.ROLE_RECEIVED_BY, d.get("received_by", ""), "", "")):
        if not name:
            continue
        known = SG.find_signatory(db, name) or {}
        overrides[role] = {
            "role": role, "name": name,
            "designation": known.get("designation", ""),
            "signature_path": (known.get("signature_path", "")
                               if db.get_bool("print_signature_images", True) else ""),
            "id_number": ident or known.get("id_number", ""),
            "phone": phone or known.get("phone", "")}
    blocks = SG.resolve_blocks(db, "DN", overrides)

    _cap = layout.get("signature_caption") or "Authorised Signatures"
    sig_flow = Table([[Paragraph(f"<b>{_cap}</b>", P_MD)],
                      [_signatures(db, blocks, layout)]], colWidths=[186 * mm])
    sig_flow.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0),
                                  ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                                  ("TOPPADDING", (0, 0), (0, 0), 0),
                                  ("BOTTOMPADDING", (0, 0), (0, 0), 3),
                                  ("TOPPADDING", (0, 1), (0, 1), 0),
                                  ("BOTTOMPADDING", (0, 1), (0, 1), 0)]))
    if SG.layout_bool(layout, "signature_inline", False):
        story += [Spacer(1, 6 * mm), sig_flow]
    else:
        story += [Spacer(1, 6 * mm), BottomAnchored(sig_flow)]

    land = (layout.get("orientation", "Portrait") == "Landscape")

    class _Row(dict):
        """Give header_design.context() the doc_no / date / project it expects."""

        def __getitem__(self, k):
            return dict.get(self, k, "")

    row = _Row(d)
    _build_with_totals(out, story, land, db, "DN",
                       lambda total: _header_footer(db, title, False, layout, row, "DN",
                                                    total_pages=total))
    n_att = _append_attachments(db, "GDN", d["doc_no"], out)
    db.execute("UPDATE gdn_documents SET pdf_path=? WHERE id=?", (str(out), doc_id))
    db.commit()
    db.audit("PRINTED", "general-dn", d["doc_no"],
             f"PDF -> {out.name}" + (f" (+{n_att} attachment page(s))" if n_att else ""))
    return out


# ----------------------------------------------------------- admin station PDF
def admin_report_pdf(db: Database, title: str, cols: list[str], rows: list[list[Any]],
                     out_path: str | Path | None = None, subtitle: str = "",
                     stats: list[tuple[str, str, str]] | None = None) -> Path:
    """Report PDF for the Admin Station, written into its own folder.

    `db` is the *inventory* database, used read-only for the letterhead and theme
    only — no admin data ever reaches it.
    """
    out = Path(out_path) if out_path else (
        config.folder("Admin Station") /
        f"{safe_name(title)}_{_dt.datetime.now():%Y%m%d_%H%M%S}.pdf")
    out.parent.mkdir(parents=True, exist_ok=True)
    return report_pdf(db, title, cols, rows, out, subtitle=subtitle, stats=stats)


# ------------------------------------------- tools, instruments & devices
def tool_report_pdf(db: Database, title: str, cols: list[str],
                    rows: list[list[Any]], out_path: str | Path | None = None,
                    subtitle: str = "",
                    stats: list[tuple[str, str, str]] | None = None) -> Path:
    """Report PDF for the Tools, Instruments & Devices module, written into its own folder.

    `db` is the *inventory* database, used read-only for the letterhead and
    theme only — no tool-custody data ever reaches it.
    """
    from . import toolstation as _T
    out = Path(out_path) if out_path else (
        _T.module_folder() /
        f"{safe_name(title)}_{_dt.datetime.now():%Y%m%d_%H%M%S}.pdf")
    out.parent.mkdir(parents=True, exist_ok=True)
    return report_pdf(db, title, cols, rows, out, subtitle=subtitle, stats=stats)


def cable_report_pdf(db: Database, title: str, cols: list[str],
                     rows: list[list[Any]], out_path: str | Path | None = None,
                     subtitle: str = "",
                     stats: list[tuple[str, str, str]] | None = None) -> Path:
    """Report PDF for the Cable Records module, written into its own folder.

    `db` is the *inventory* database, used read-only for the letterhead and
    theme only — no cable data ever reaches it.
    """
    from . import cables as _C
    out = Path(out_path) if out_path else (
        _C.module_folder() /
        f"{safe_name(title)}_{_dt.datetime.now():%Y%m%d_%H%M%S}.pdf")
    out.parent.mkdir(parents=True, exist_ok=True)
    return report_pdf(db, title, cols, rows, out, subtitle=subtitle, stats=stats)


def handover_pdf(db: Database, tdb, handover_id: int,
                 out_path: str | Path | None = None) -> Path:
    """Reprint one handover as the controlled form WH-FRM-001.

    Follows the paper layout the user supplied: A handover details ·
    B recipient · C item grid · D acknowledgement, with the signature block
    sitting directly above the footer rule as required.
    """
    from . import toolstation as _T

    h = _T.get_handover(tdb, handover_id)
    if h is None:
        raise ValueError("Handover not found.")
    out = Path(out_path) if out_path else (
        _T.module_folder() / "Forms" /
        f"{safe_file_part(h['ref_no'])}.pdf")
    out.parent.mkdir(parents=True, exist_ok=True)

    primary, accent = _brand_colors(db)
    page_w = landscape(A4)[0] - 24 * mm

    def _section(text: str) -> Table:
        t = Table([[Paragraph(
            f"<font color='white' size=8.4><b>{text}</b></font>", P_SM)]],
            colWidths=[page_w])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), accent),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
        return t

    # Helvetica has no ballot-box glyph (U+2610/2611) -- it prints as a black
    # box, exactly like the old U+26A0 bug. Use markup instead so the ticked
    # option is unmistakable in print and in photocopies.
    def _tick(on: bool) -> str:
        return "[X]" if on else "[  ]"

    types = " &nbsp;".join(
        (f"<b>[X] {t}</b>" if h["txn_type"] == t
         else f"<font color='#8a97a5'>[&nbsp;&nbsp;] {t}</font>")
        for t in _T.TXN_TYPES)

    story: list[Any] = [
        Paragraph("TOOLS, DEVICES &amp; INSTRUMENTS HANDOVER FORM", P_TITLE),
        Paragraph(f"{h['warehouse'] or 'MAIN'} WAREHOUSE  ·  Form No.: "
                  f"WH-FRM-001  Rev: 00", P_SUB),
        _rule(primary, accent, page_w), Spacer(1, 3 * mm),
        _section("A — HANDOVER DETAILS"),
        _kv_block([("Form No.", h["form_no"] or "-"),
                   ("Handover Reference No.", h["ref_no"]),
                   ("Date", _T.fmt_date(h["doc_date"])),
                   ("Time", h["doc_time"] or "-"),
                   ("Transaction Type", types),
                   ("Expected Return Date", _T.fmt_date(h["expected_return"]) or "-"),
                   ("Project ID", h["project_id"] or "-"),
                   ("Project Name", h["project_name"] or "-"),
                   ("Project / Site Location", h["location"] or "-"),
                   ("Custody Status", h["status"])],
                  cols=3, total_width=page_w),
        Spacer(1, 3 * mm),
        _section("B — RECIPIENT / CUSTODIAN DETAILS"),
        _kv_block([("Handed To", h["handed_to"] or "-"),
                   ("Employee / Iqama ID", h["iqama_id"] or "-"),
                   ("Job Title", h["job_title"] or "-"),
                   ("Mobile No.", h["mobile"] or "-"),
                   ("Company / Department", h["company"] or "-"),
                   ("Email", h["email"] or "-"),
                   ("Supervisor / Manager", h["supervisor"] or "-"),
                   ("Cost Code / WBS", h["cost_code"] or "-")],
                  cols=3, total_width=page_w),
        Spacer(1, 3 * mm),
        _section("C — ITEM DETAILS"), Spacer(1, 1.5 * mm)]

    cols = ["No.", "Asset / Tool ID", "Category", "Description", "Make / Model",
            "Serial No.", "Qty", "Returned", "Accessories / Components",
            "Cond.", "Calib. Due", "Remarks / Defects"]
    rows = [[l["line_no"], l["asset_id"], l["category"], l["description"],
             l["make_model"], l["serial_no"] or "-",
             round(float(l["qty"] or 0), 2),
             round(float(l["qty_returned"] or 0), 2),
             l["accessories"] or "-", l["condition"],
             _T.fmt_date(l["calib_due"]) or "-", l["remarks"]]
            for l in h["lines"]]
    weights = [3, 9, 8, 16, 8, 9, 4, 5, 14, 4, 8, 12]
    widths = [page_w * w / sum(weights) for w in weights]
    story.append(_grid(cols, rows, widths, 7.2, db=db, compact=True))
    story.append(Spacer(1, 1.5 * mm))
    story.append(Paragraph(
        "<font size=6.6 color='#6b7c8f'>*Condition grade:&nbsp;&nbsp; "
        "A – New / Excellent&nbsp;&nbsp;&nbsp; B – Good&nbsp;&nbsp;&nbsp; "
        "C – Fair / Usable&nbsp;&nbsp;&nbsp; D – Damaged / Not Usable</font>",
        P_SM))
    story.append(Spacer(1, 2 * mm))
    story.append(_kv_block(
        [("Serial / Asset ID checked", _tick(h["v_serial"])),
         ("Accessories checked", _tick(h["v_accessories"])),
         ("Calibration valid", _tick(h["v_calibration"])),
         ("Photos attached", _tick(h["v_photos"]))],
        cols=4, total_width=page_w))

    story.append(Spacer(1, 4 * mm))
    story.append(_section("D — ACKNOWLEDGEMENT & AUTHORIZATION"))
    story.append(Paragraph(
        "<font size=7>I acknowledge receipt of the items listed above in the "
        "stated condition and with the stated accessories. I accept "
        "responsibility for their safe custody, proper use, and return (where "
        "applicable). I will immediately report any loss, damage, malfunction "
        "or change in condition to the Main Warehouse and my supervisor. Items "
        "shall not be transferred to another person or project without "
        "authorization.</font>", P_SM))
    story.append(Spacer(1, 2.5 * mm))

    sig = Table([[Paragraph("<font size=7.4><b>ISSUED BY — WAREHOUSE</b><br/><br/>"
                            f"{h['issued_by'] or ''}<br/>"
                            "____________________________<br/>"
                            f"<font size=6.6 color='#6b7c8f'>Signature &amp; Date: "
                            f"{h['issued_at'] or ''}</font></font>", P_SM),
                 Paragraph("<font size=7.4><b>RECEIVED BY — CUSTODIAN</b><br/><br/>"
                           f"{h['received_by'] or ''}<br/>"
                           "____________________________<br/>"
                           f"<font size=6.6 color='#6b7c8f'>Signature &amp; Date: "
                           f"{h['received_at'] or ''}</font></font>", P_SM)]],
                colWidths=[page_w / 2] * 2)
    sig.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                             ("LEFTPADDING", (0, 0), (-1, -1), 6),
                             ("TOPPADDING", (0, 0), (-1, -1), 4)]))
    # the signature block must sit at the bottom, immediately above the footer
    # line -- the same rule the delivery note follows
    story += [Spacer(1, 6 * mm), BottomAnchored(sig)]

    _build_with_totals(out, story, True, db, "__default__",
                       lambda total: _header_footer(
                           db, f"Tools Handover — {h['ref_no']}", True,
                           total_pages=total))
    db.audit("EXPORTED", "tool-handover", h["ref_no"], f"PDF -> {out.name}")
    return out


# ------------------------------------------------- material request check PDF
_MR_TINT = {
    "Full Available": ("#0f7b3d", "#e6f6ec"),
    "Partial Available": ("#9a6700", "#fff6e0"),
    "Not Available": ("#b3261e", "#fdecea"),
    "Item Not Found": ("#5b2bc9", "#f0eaff"),
    "Pending": ("#4a5a6a", "#eef1f4"),
    "Preparing": ("#0b6e83", "#e4f6fa"),
    "Ready": ("#075e70", "#e2f4f8"),
    "Partially Delivered": ("#9a6700", "#fff6e0"),
    "Delivered": ("#0f7b3d", "#e6f6ec"),
    "Cancelled": ("#6b7280", "#f0f1f3"),
}


def _mr_summary_strip(db: Database, stats: Sequence[tuple[str, str, str]],
                      width: float) -> Table:
    """The KPI strip from the on-screen check view, rendered for print.

    Mirrors the app: small grey caption over a large coloured figure, each tile
    boxed and evenly spaced across the page.
    """
    n = max(1, len(stats))
    cells = []
    for label, value, colour in stats:
        col = colour or "#12283f"
        inner = Table(
            [[Paragraph(f"<font size=6.6 color='#6b7c8f'>{str(label).upper()}</font>", P_SM)],
             [Paragraph(f"<b><font size=12.5 color='{col}'>{value}</font></b>", P_SM)]],
            colWidths=[width / n - 2.6 * mm])
        inner.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#d8e1ea")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fbfcfe")),
            ("LINEABOVE", (0, 0), (-1, 0), 1.5, colors.HexColor(col)),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (0, 0), 3.5),
            ("BOTTOMPADDING", (0, 0), (0, 0), 0),
            ("TOPPADDING", (0, 1), (0, 1), 0),
            ("BOTTOMPADDING", (0, 1), (0, 1), 4.5),
        ]))
        cells.append(inner)
    t = Table([cells], colWidths=[width / n] * n)
    t.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0),
                           ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                           ("VALIGN", (0, 0), (-1, -1), "TOP")]))
    return t


def _mr_grid(db: Database, cols: list[str], rows: list[list[Any]], widths,
             font: float, status_cols: dict[int, bool]) -> Table:
    """Item grid for MR documents: pill-tinted status cells, banded rows.

    `status_cols` maps a column index -> True when that column holds an
    availability / fulfilment word that should print as a coloured pill.
    """
    primary, accent = _brand_colors(db)
    head_style = ParagraphStyle("mrh", parent=P_SM, fontSize=font, leading=font * 1.28,
                                textColor=colors.white, alignment=TA_CENTER)
    head = [Paragraph(f"<b>{c}</b>", head_style) for c in cols]

    ncol = len(cols)
    numeric = [True] * ncol
    for r in rows:
        for c in range(min(ncol, len(r))):
            v = r[c]
            if str(v or "").strip() and not _is_number(v):
                numeric[c] = False
    # Identifier columns are digit strings but must never be reformatted --
    # "001582" would otherwise print as "1,582" and lose its leading zeros.
    for c, name in enumerate(cols):
        key = str(name).strip().lower()
        if str(name).strip().lower() in ("line", "sr.", "sr", "#"):
            numeric[c] = True
        elif any(tok in key for tok in ("no.", "number", "code", "ref", "pr /",
                                        "project", "batch", "serial", "phone",
                                        "iqama", "id")):
            numeric[c] = False
        if c in status_cols:
            numeric[c] = False

    body, tint_cmds = [], []
    for ri, r in enumerate(rows, start=1):
        line = []
        for c in range(ncol):
            v = r[c] if c < len(r) else ""
            txt = "" if v is None else str(v)
            if c in status_cols and txt.strip():
                fg, bg = _MR_TINT.get(txt.strip(), ("#12283f", "#eef1f4"))
                tint_cmds.append(("BACKGROUND", (c, ri), (c, ri), colors.HexColor(bg)))
                st = ParagraphStyle(f"m{c}", parent=P_SM, fontSize=font,
                                    leading=font * 1.28, alignment=TA_CENTER,
                                    textColor=colors.HexColor(fg))
                line.append(Paragraph(f"<b>{txt}</b>", st))
                continue
            colour = None
            if numeric[c] and _is_number(v):
                num = _num_value(v)
                txt = f"{num:,.2f}" if (isinstance(v, float) or "." in str(v)) \
                    else f"{num:,.0f}"
                if num < 0:
                    colour = "#b3261e"
                # a non-zero shortage should be impossible to miss
                if "short" in str(cols[c]).lower() and num > 0:
                    colour = "#b3261e"
            if "\u26a0" in txt:
                # Helvetica has no warning sign glyph -- it prints as a black box
                txt = txt.replace("\u26a0", "").replace("  ", " ").strip()
                txt = f"{txt}"
                colour = colour or "#8a5a00"
            st = ParagraphStyle(
                f"m{c}", parent=P_SM, fontSize=font, leading=font * 1.28,
                alignment=TA_RIGHT if numeric[c] else TA_LEFT,
                textColor=colors.HexColor(colour) if colour else colors.HexColor("#16202b"))
            if colour:
                txt = f"<b>{txt}</b>"
            line.append(Paragraph(txt, st))
        body.append(line)

    t = Table([head] + body, colWidths=widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), primary),
        ("LINEBELOW", (0, 0), (-1, 0), 1.1, accent),
        ("GRID", (0, 0), (-1, -1), 0.28, colors.HexColor("#ccd8e4")),
        ("BOX", (0, 0), (-1, -1), 0.7, primary),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3.0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.0),
        ("LEFTPADDING", (0, 0), (-1, -1), 3.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3.5),
    ]
    for i in range(1, len(body) + 1):
        if i % 2 == 0:
            style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#f5f8fb")))
    style += tint_cmds          # pills paint over the banding
    t.setStyle(TableStyle(style))
    return t


def mr_title(mr_no: str, prs: Sequence[str] | None = None, project: str = "") -> str:
    """Heading for a Material Request print.

    Sites quote their own PR/MR number and the project — not AURCO's internal
    MR-2026-xxxxx — so both are put in the heading. Long PR lists are trimmed so
    the title never wraps off the page.
    """
    prs = [str(p).strip() for p in (prs or []) if str(p).strip()]
    bits = f"Material Request {mr_no}".strip()
    if prs:
        shown = prs[:3]
        tail = f" +{len(prs) - 3} more" if len(prs) > 3 else ""
        label = "PR / MR No." if len(prs) == 1 else "PR / MR Nos."
        bits += f"  ·  {label} {', '.join(shown)}{tail}"
    project = str(project or "").strip()
    if project:
        bits += f"  ·  {project}"
    return bits


def mr_request_path(db: Database, mr_no: str, prs: Sequence[str] | None = None,
                    project: str = "") -> Path:
    """File name for a Material Request print: the PR numbers and project are
    in the name too, so the PDF can be found without opening it."""
    prs = [str(p).strip() for p in (prs or []) if str(p).strip()]
    parts = [safe_file_part(mr_no) or "Material Request"]
    if project:
        parts.append(f"({safe_file_part(project)})")
    if prs:
        shown = prs[:4]
        if len(prs) > 4:
            shown = shown + [f"+{len(prs) - 4}-more"]
        parts.append("PR " + " ".join(safe_file_part(p) for p in shown))
    name = " ".join(x for x in parts if x)
    if len(name) > MAX_NAME_LEN:
        name = name[:MAX_NAME_LEN].rstrip(" -_")
    return config.folder("Reports") / f"{name}.pdf"


def material_check_pdf(db: Database, title: str, cols: list[str], rows: list[list[Any]],
                       stats: Sequence[tuple[str, str, str]] | None = None,
                       header_pairs: Sequence[tuple[str, Any]] | None = None,
                       out_path: str | Path | None = None,
                       subtitle: str = "", legend: bool = True) -> Path:
    """Print view for Material Request checking — styled like the app screen.

    Layout: letterhead · title · brand rule · request header block · KPI strip ·
    colour-coded item grid · legend · signature-ready footer note.
    """
    out = Path(out_path) if out_path else (
        config.folder("Reports") /
        f"{safe_name(title)}_{_dt.datetime.now():%Y%m%d_%H%M%S}.pdf")
    out.parent.mkdir(parents=True, exist_ok=True)

    land = len(cols) > 7
    page_w = (landscape(A4)[0] if land else A4[0]) - 24 * mm
    primary, accent = _brand_colors(db)

    # proportional widths, with tight columns kept tight
    sample = rows[:250]
    weights = []
    for i, c in enumerate(cols):
        longest = len(str(c))
        for r in sample:
            if i < len(r):
                longest = max(longest, min(44, len(str(r[i] if r[i] is not None else ""))))
        weights.append(max(4.0, float(longest)))
    total = sum(weights) or 1
    widths = [max(13 * mm, page_w * w / total) for w in weights]
    over = sum(widths) - page_w
    if over > 0:
        for i in sorted(range(len(widths)), key=lambda i: -widths[i]):
            take = min(widths[i] - 13 * mm, over)
            widths[i] -= take
            over -= take
            if over <= 0.1:
                break

    fs = 7.1 if land else 7.7
    if len(cols) > 13:
        fs = 6.5

    status_cols = {i for i, c in enumerate(cols)
                   if str(c).strip().lower() in ("availability", "fulfilment", "status")}

    story: list[Any] = [Paragraph(title, P_TITLE), _rule(primary, accent, page_w)]
    if subtitle:
        story.append(Paragraph(subtitle, P_SUB))
    story.append(Spacer(1, 3.5 * mm))
    if header_pairs:
        story += [_kv_block([(k, v) for k, v in header_pairs
                             if str(v or "").strip() or v == 0],
                            cols=3, total_width=page_w), Spacer(1, 3.5 * mm)]
    if stats:
        story += [_mr_summary_strip(db, stats, page_w), Spacer(1, 4 * mm)]
    if rows:
        story.append(_mr_grid(db, cols, rows, widths, fs, status_cols))
    else:
        story.append(Paragraph("<i>No request lines to display.</i>", P_MD))

    if legend and rows:
        chips = []
        for name in ("Full Available", "Partial Available", "Not Available",
                     "Item Not Found"):
            fg, bg = _MR_TINT[name]
            chip = Table([[Paragraph(
                f"<font size=6.8 color='{fg}'><b>{name}</b></font>", P_SM)]],
                colWidths=[33 * mm])
            chip.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(bg)),
                ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(fg)),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2)]))
            chips.append(chip)
        leg = Table([[Paragraph("<font size=7 color='#6b7c8f'><b>Legend</b></font>",
                                P_SM)] + chips],
                    colWidths=[18 * mm] + [35 * mm] * len(chips))
        leg.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                 ("LEFTPADDING", (0, 0), (-1, -1), 1),
                                 ("RIGHTPADDING", (0, 0), (-1, -1), 1)]))
        story += [Spacer(1, 3.5 * mm), leg]

    _build_with_totals(out, story, land, db, "__default__",
                       lambda total: _header_footer(db, "Material Request", True,
                                                    total_pages=total))
    db.audit("EXPORTED", "MR-check", title, f"PDF -> {out.name}")
    return out


# ------------------------------------------------- company issuance register
def issuance_report_pdf(db: Database, title: str, cols: list[str],
                        rows: list[list[Any]],
                        stats: Sequence[tuple[str, str, str]] | None = None,
                        header_pairs: Sequence[tuple[str, Any]] | None = None,
                        out_path: str | Path | None = None,
                        subtitle: str = "") -> Path:
    """Report for the Company Issuance Register, on the company letterhead."""
    from . import issuance as ISS
    out = Path(out_path) if out_path else (
        config.folder(ISS.FOLDER) /
        f"{safe_name(title)}_{_dt.datetime.now():%Y%m%d_%H%M%S}.pdf")
    out.parent.mkdir(parents=True, exist_ok=True)
    return material_check_pdf(db, title, cols, rows, stats=stats,
                              header_pairs=header_pairs, out_path=out,
                              subtitle=subtitle, legend=False)


def issuance_receipt_pdf(idb, db: Database, issue_id: int,
                         out_path: str | Path | None = None,
                         include_evidence: bool = True) -> Path:
    """Signed hand-over receipt for one issuance, with the photo proof attached.

    `idb` is the standalone IssuanceDB; `db` is the inventory database, read
    only for the letterhead, theme and signature settings.
    """
    from . import issuance as ISS
    from . import signatories as SG

    rec = ISS.get_issue(idb, issue_id)
    if rec is None:
        raise ValueError("Issuance record not found")
    ev = ISS.evidence_for(idb, issue_id)
    status = ISS.compute_status(rec)
    out = Path(out_path) if out_path else (
        config.folder(ISS.FOLDER) /
        f"{safe_name(rec['issue_no'])} {safe_file_part(rec['company'])}.pdf")
    out.parent.mkdir(parents=True, exist_ok=True)

    layout = SG.get_layout(db, "DN")
    is_perm = rec["issue_type"] == ISS.PERMANENT
    title = ("Material Issue Receipt — Permanent" if is_perm
             else "Material Issue & Return Receipt")

    story: list[Any] = [
        Paragraph(title, P_TITLE),
        Paragraph(f"Issue No: <b>{rec['issue_no']}</b> &nbsp;|&nbsp; Status: "
                  f"<b>{status}</b> &nbsp;|&nbsp; "
                  "<i>Company issuance register — not linked to inventory stock</i>",
                  P_SUB),
        Spacer(1, 5 * mm)]

    pairs = [("Issue No", rec["issue_no"]), ("Date", rec["record_date"]),
             ("Company", rec["company"]), ("MR / Reference", rec["mr_no"] or "-"),
             ("Recipient", rec["recipient"]), ("Iqama / ID", rec["iqama"] or "-"),
             ("Phone", rec["phone"] or "-"),
             ("Issue Type", rec["issue_type"]),
             ("Date of Issuance", rec["issue_date"]),
             ("Expected Return", rec["expected_return"] or ("N/A" if is_perm else "-")),
             ("Date of Return", rec["return_date"] or "-"),
             ("DN / Gate Pass", rec["dn_no"] or "-"),
             ("Project / Site", rec["project"] or "-"),
             ("Issued By", rec["issued_by"] or "-")]
    story += [_kv_block([(k, v) for k, v in pairs], cols=3), Spacer(1, 4 * mm)]

    cur = db.get_setting("currency", "")
    qty = ISS.to_float(rec["qty"])
    back = ISS.to_float(rec["qty_returned"])
    icols = ["Item Issued", "Item Code", "UOM", "Qty Issued", "Qty Returned",
             "Still Out", "Condition Out", "Condition In"]
    irows = [[rec["item"], rec["item_code"] or "-", rec["uom"] or "-", f"{qty:g}",
              f"{back:g}", f"{ISS.outstanding_qty(rec):g}",
              rec["condition_out"] or "-", rec["condition_in"] or "-"]]
    story.append(_grid(icols, irows,
                       [52 * mm, 22 * mm, 14 * mm, 20 * mm, 22 * mm, 18 * mm,
                        20 * mm, 18 * mm],
                       font=8, header_color=_brand_colors(db)[0]))

    d_out = ISS.days_out(rec)
    bits = [f"Total Quantity: <b>{qty:g}</b>"]
    if not is_perm:
        bits.append(f"Outstanding: <b>{ISS.outstanding_qty(rec):g}</b>")
        if d_out is not None:
            bits.append(f"Days out: <b>{d_out}</b>")
        if status == ISS.ST_OVERDUE:
            bits.append(f"<font color='#c92a2a'>Overdue by "
                        f"<b>{ISS.days_overdue(rec)}</b> day(s)</font>")
    if ISS.to_float(rec["unit_value"]):
        bits.append(f"Value: <b>{cur} "
                    f"{qty * ISS.to_float(rec['unit_value']):,.2f}</b>")
    story += [Spacer(1, 3 * mm), Paragraph(" &nbsp;&nbsp;·&nbsp;&nbsp; ".join(bits),
                                           P_MD)]
    if rec["remarks"]:
        story += [Spacer(1, 2 * mm),
                  Paragraph(f"<b>Remarks:</b> {rec['remarks']}", P_SM)]

    # ---- proof thumbnails inline, so the receipt itself carries the evidence
    imgs = [e for e in ev
            if Path(e["file_path"]).suffix.lower() in ISS.IMAGE_SUFFIXES
            and Path(e["file_path"]).exists()]
    if imgs:
        story += [Spacer(1, 4 * mm),
                  Paragraph("<b>Photographic evidence</b>", P_MD), Spacer(1, 2 * mm)]
        cells, row = [], []
        for e in imgs[:6]:
            try:
                thumb = Image(e["file_path"], width=56 * mm, height=42 * mm,
                              kind="proportional")
            except Exception:
                continue
            cap = f"{e['kind'].title()} · {Path(e['file_path']).name[:26]}"
            inner = Table([[thumb], [Paragraph(f"<font size=6.5 color='#6b7c8f'>"
                                               f"{cap}</font>", P_SM)]],
                          colWidths=[58 * mm])
            inner.setStyle(TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#c9d6e2")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
            row.append(inner)
            if len(row) == 3:
                cells.append(row)
                row = []
        if row:
            row += [""] * (3 - len(row))
            cells.append(row)
        if cells:
            t = Table(cells, colWidths=[60 * mm] * 3)
            t.setStyle(TableStyle([("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
            story.append(t)
    elif not (rec["dn_no"] or "").strip():
        story += [Spacer(1, 4 * mm),
                  Paragraph("<font color='#c92a2a'><b>No photographic evidence is "
                            "attached to this issuance.</b></font>", P_MD)]

    # ---- signatures
    blocks = [
        {"role": "Issued By", "name": rec["issued_by"] or "",
         "designation": "Store", "signature_path": "", "id_number": "", "phone": ""},
        {"role": "Received By", "name": rec["recipient"] or "",
         "designation": rec["company"] or "", "signature_path": "",
         "id_number": rec["iqama"] or "", "phone": rec["phone"] or ""},
    ]
    if not is_perm:
        blocks.append({"role": "Returned / Received Back By",
                       "name": rec["received_back_by"] or "", "designation": "Store",
                       "signature_path": "", "id_number": "", "phone": ""})
    sig = Table([[Paragraph("<b>Authorised Signatures</b>", P_MD)],
                 [_signatures(db, blocks, layout)]], colWidths=[186 * mm])
    sig.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0),
                             ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                             ("TOPPADDING", (0, 1), (0, 1), 0)]))
    story += [Spacer(1, 6 * mm), BottomAnchored(sig)]

    class _Row(dict):
        def __getitem__(self, k):
            return dict.get(self, k, "")

    ctx = _Row({"doc_no": rec["issue_no"], "doc_date": rec["issue_date"],
                "project": rec["project"], "warehouse": rec["location"],
                "created_by": rec["created_by"]})
    _build_with_totals(out, story, False, db, "DN",
                       lambda total: _header_footer(db, title, False, layout, ctx,
                                                    "DN", total_pages=total))
    if include_evidence:
        _append_issuance_evidence(idb, db, issue_id, out)
    idb.audit("PRINTED", "issue", rec["issue_no"], f"receipt -> {out.name}")
    return out


def _append_issuance_evidence(idb, db: Database, issue_id: int,
                              pdf_path: Path) -> int:
    """Append every proof file full-page after the receipt."""
    from . import issuance as ISS
    ev = [e for e in ISS.evidence_for(idb, issue_id)
          if Path(e["file_path"]).exists()]
    if not ev:
        return 0
    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError:
        return 0
    writer = PdfWriter()
    try:
        for page in PdfReader(str(pdf_path)).pages:
            writer.add_page(page)
    except Exception:
        return 0
    rec = ISS.get_issue(idb, issue_id) or {}
    added = 0
    for e in ev:
        f = Path(e["file_path"])
        suffix = f.suffix.lower()
        try:
            if suffix == ".pdf":
                for page in PdfReader(str(f)).pages:
                    writer.add_page(page)
                    added += 1
            elif suffix in ISS.IMAGE_SUFFIXES:
                buf = io.BytesIO()
                c = rl_canvas.Canvas(buf, pagesize=A4)
                _attachment_caption(
                    c, db, str(rec.get("issue_no", "")),
                    f"{e['kind'].title()} evidence — {f.name}")
                try:
                    img = ImageReader(str(f))
                    iw, ih = img.getSize()
                    max_w, max_h = A4[0] - 30 * mm, A4[1] - 55 * mm
                    sc = min(max_w / iw, max_h / ih)
                    c.drawImage(img, (A4[0] - iw * sc) / 2, 22 * mm,
                                width=iw * sc, height=ih * sc,
                                preserveAspectRatio=True, mask="auto")
                except Exception:
                    c.setFont("Helvetica", 10)
                    c.drawCentredString(A4[0] / 2, A4[1] / 2,
                                        f"Could not render {f.name}")
                c.showPage()
                c.save()
                buf.seek(0)
                for page in PdfReader(buf).pages:
                    writer.add_page(page)
                    added += 1
        except Exception:
            continue
    if added:
        with open(pdf_path, "wb") as fh:
            writer.write(fh)
    return added
