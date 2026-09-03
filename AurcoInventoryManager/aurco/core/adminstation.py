"""ADMIN STATION — a completely separate camp / office record register.

Design rule requested by the user: *this module must remain completely separate
from the inventory system*. That is enforced physically, not by convention:

    ·  its own SQLite file            <storage>/Admin Station/admin_station.db
    ·  its own schema, numbering, audit trail and backups
    ·  no foreign key, join or import from items / stock_ledger / documents
    ·  nothing here ever posts a stock movement

The only thing shared with the inventory application is the *look* (company
letterhead, theme colours) when a PDF is printed, which comes from the main
settings table read-only.

Record shape follows the sheet the user supplied:

    SR# · Camp/Office Name · Date of Record · Item Category · Item Description
    UOM · Quantity · Return · Destination Location · Remarks
"""
from __future__ import annotations

import csv
import datetime as _dt
import io
import json
import os
import re
import shutil
import sqlite3
from pathlib import Path
from typing import Any, Iterable, Sequence

from . import config

FOLDER = "Admin Station"
DB_NAME = "admin_station.db"

SCHEMA_VERSION = 2

DDL = """
PRAGMA journal_mode=WAL;

CREATE TABLE IF NOT EXISTS settings (
    key   TEXT PRIMARY KEY,
    value TEXT
);

CREATE TABLE IF NOT EXISTS records (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    sr_no        TEXT DEFAULT '',
    camp         TEXT DEFAULT '',
    record_date  TEXT DEFAULT '',
    category     TEXT DEFAULT '',
    description  TEXT DEFAULT '',
    uom          TEXT DEFAULT '',
    qty          REAL NOT NULL DEFAULT 0,
    qty_return   REAL NOT NULL DEFAULT 0,
    destination  TEXT DEFAULT '',
    remarks      TEXT DEFAULT '',
    ref_no       TEXT DEFAULT '',
    custodian    TEXT DEFAULT '',
    condition    TEXT DEFAULT '',
    unit_cost    REAL NOT NULL DEFAULT 0,
    status       TEXT NOT NULL DEFAULT 'Active',
    batch_id     INTEGER,
    created_by   TEXT DEFAULT '',
    created_at   TEXT DEFAULT (datetime('now','localtime')),
    updated_at   TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS ix_ar_camp ON records(camp);
CREATE INDEX IF NOT EXISTS ix_ar_cat  ON records(category);
CREATE INDEX IF NOT EXISTS ix_ar_date ON records(record_date);
CREATE INDEX IF NOT EXISTS ix_ar_dest ON records(destination);

CREATE TABLE IF NOT EXISTS batches (
    id        INTEGER PRIMARY KEY AUTOINCREMENT,
    ts        TEXT DEFAULT (datetime('now','localtime')),
    source    TEXT DEFAULT '',
    rows      INTEGER DEFAULT 0,
    skipped   INTEGER DEFAULT 0,
    mapping   TEXT DEFAULT '',
    username  TEXT DEFAULT ''
);

CREATE TABLE IF NOT EXISTS audit (
    id        INTEGER PRIMARY KEY AUTOINCREMENT,
    ts        TEXT DEFAULT (datetime('now','localtime')),
    username  TEXT DEFAULT '',
    action    TEXT NOT NULL,
    entity    TEXT DEFAULT '',
    entity_id TEXT DEFAULT '',
    details   TEXT DEFAULT ''
);
CREATE INDEX IF NOT EXISTS ix_ar_audit ON audit(ts);
"""

# canonical field -> printable label (the user's sheet, in her order)
FIELDS: list[tuple[str, str]] = [
    ("sr_no", "SR#"),
    ("camp", "Camp / Office Name"),
    ("record_date", "Date of Record"),
    ("category", "Item Category"),
    ("description", "Item Description"),
    ("uom", "UOM"),
    ("qty", "Quantity"),
    ("qty_return", "Return"),
    ("destination", "Destination Location"),
    ("remarks", "Remarks"),
]
EXTRA_FIELDS: list[tuple[str, str]] = [
    ("ref_no", "Reference"),
    ("custodian", "Custodian"),
    ("condition", "Condition"),
    ("unit_cost", "Unit Cost"),
    ("status", "Status"),
]
ALL_FIELDS = FIELDS + EXTRA_FIELDS
LABELS = dict(ALL_FIELDS)

NUMERIC = {"qty", "qty_return", "unit_cost"}

STATUSES = ["Active", "Returned", "Partially Returned", "Transferred", "Damaged",
            "Disposed", "On Hold"]

CONDITIONS = ["", "New", "Good", "Used", "Needs Repair", "Damaged", "Scrap"]

# Header aliases: matched after stripping everything except a-z0-9, so
# "Camp/Office Name", "camp office name" and "CAMP_OFFICE" all land on `camp`.
HEADER_MAP = {
    "sr": "sr_no", "srno": "sr_no", "sr#": "sr_no", "serial": "sr_no", "sl": "sr_no",
    "slno": "sr_no", "no": "sr_no", "s": "sr_no", "sno": "sr_no", "line": "sr_no",
    "campofficename": "camp", "camp": "camp", "campname": "camp", "office": "camp",
    "officename": "camp", "camposite": "camp", "location": "camp", "site": "camp",
    "campoffice": "camp",
    "dateofrecord": "record_date", "date": "record_date", "recorddate": "record_date",
    "entrydate": "record_date", "docdate": "record_date", "day": "record_date",
    "itemcategory": "category", "category": "category", "type": "category",
    "itemtype": "category", "group": "category",
    "itemdescription": "description", "description": "description", "item": "description",
    "itemname": "description", "productname": "description", "particulars": "description",
    "material": "description", "materialdescription": "description", "desc": "description",
    "uom": "uom", "unit": "uom", "unitofmeasure": "uom", "units": "uom",
    "quantity": "qty", "qty": "qty", "quntity": "qty", "quanity": "qty", "nos": "qty",
    "return": "qty_return", "returns": "qty_return", "returned": "qty_return",
    "returnqty": "qty_return", "returnedqty": "qty_return", "qtyreturn": "qty_return",
    "returnquantity": "qty_return",
    "destinationlocation": "destination", "destination": "destination",
    "deliveredto": "destination", "issuedto": "destination", "sentto": "destination",
    "toposition": "destination", "dest": "destination",
    "remarks": "remarks", "reamrks": "remarks", "remark": "remarks", "note": "remarks",
    "notes": "remarks", "comment": "remarks", "comments": "remarks",
    "reference": "ref_no", "ref": "ref_no", "refno": "ref_no", "referenceno": "ref_no",
    "dn": "ref_no", "dnno": "ref_no", "document": "ref_no",
    "custodian": "custodian", "responsible": "custodian", "incharge": "custodian",
    "receivedby": "custodian", "holder": "custodian",
    "condition": "condition", "itemcondition": "condition",
    "unitcost": "unit_cost", "cost": "unit_cost", "price": "unit_cost",
    "rate": "unit_cost", "unitprice": "unit_cost",
    "status": "status",
}


def _now() -> str:
    return _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def norm(s: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def db_path() -> Path:
    return config.folder(FOLDER) / DB_NAME


# --------------------------------------------------------------------- value parsing
def to_float(v: Any) -> float:
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    t = str(v).strip().lower()
    if t in ("yes", "y", "true", "returned", "full", "all"):
        return -1.0                        # sentinel: "everything was returned"
    if t in ("no", "n", "false", "-", "nil", "none"):
        return 0.0
    try:
        return float(re.sub(r"[^\d.\-]", "", t) or 0)
    except ValueError:
        return 0.0


_DATE_PATTERNS = ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y",
                  "%Y/%m/%d", "%d-%b-%Y", "%d %b %Y", "%b %d, %Y", "%d-%m-%y",
                  "%d/%m/%y", "%Y-%m-%d %H:%M:%S")


def to_date(v: Any) -> str:
    """Best-effort conversion of anything date-like to ISO yyyy-mm-dd."""
    if v in (None, ""):
        return ""
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    t = str(v).strip()
    if not t:
        return ""
    for f in _DATE_PATTERNS:
        try:
            return _dt.datetime.strptime(t, f).date().isoformat()
        except ValueError:
            continue
    # Excel serial number
    try:
        n = float(t)
        if 20000 < n < 60000:
            return (_dt.date(1899, 12, 30) + _dt.timedelta(days=int(n))).isoformat()
    except ValueError:
        pass
    return t          # keep whatever the user typed rather than losing it


# --------------------------------------------------------------------- database
class AdminDB:
    """Standalone database for the Admin Station."""

    def __init__(self, path: str | Path | None = None):
        self.path = Path(path or db_path())
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(str(self.path), check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA busy_timeout=15000")
        self.current_user = "admin"
        self._init()

    # ------------------------------------------------------------- schema
    def _init(self) -> None:
        self.conn.executescript(DDL)
        self.conn.commit()
        self._migrate()

    def _migrate(self) -> None:
        cols = {r["name"] for r in self.conn.execute("PRAGMA table_info(records)")}
        added = []
        for col, ddl in (("ref_no", "TEXT DEFAULT ''"),
                         ("custodian", "TEXT DEFAULT ''"),
                         ("condition", "TEXT DEFAULT ''"),
                         ("unit_cost", "REAL NOT NULL DEFAULT 0"),
                         ("status", "TEXT NOT NULL DEFAULT 'Active'"),
                         ("batch_id", "INTEGER")):
            if col not in cols:
                self.conn.execute(f"ALTER TABLE records ADD COLUMN {col} {ddl}")
                added.append(col)
        if added:
            self.conn.commit()
            self.audit("MIGRATED", "admin-station", "", "added " + ", ".join(added))
        self.set_setting("schema_version", str(SCHEMA_VERSION))

    # --------------------------------------------------------- primitives
    def execute(self, sql: str, params: Sequence = ()) -> sqlite3.Cursor:
        return self.conn.execute(sql, params)

    def query(self, sql: str, params: Sequence = ()) -> list[sqlite3.Row]:
        return list(self.conn.execute(sql, params).fetchall())

    def one(self, sql: str, params: Sequence = ()):
        return self.conn.execute(sql, params).fetchone()

    def scalar(self, sql: str, params: Sequence = (), default=0):
        r = self.conn.execute(sql, params).fetchone()
        return default if r is None or r[0] is None else r[0]

    def commit(self) -> None:
        self.conn.commit()

    def close(self) -> None:
        try:
            self.conn.commit()
            self.conn.close()
        except Exception:
            pass

    # ----------------------------------------------------------- settings
    def get_setting(self, key: str, default: Any = None) -> Any:
        r = self.one("SELECT value FROM settings WHERE key=?", (key,))
        return r["value"] if r else default

    def set_setting(self, key: str, value: Any) -> None:
        self.execute("INSERT INTO settings(key,value) VALUES(?,?) "
                     "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                     (key, str(value)))
        self.conn.commit()

    def audit(self, action: str, entity: str = "", entity_id: str = "",
              details: str = "") -> None:
        self.execute("INSERT INTO audit(ts,username,action,entity,entity_id,details)"
                     " VALUES(?,?,?,?,?,?)",
                     (_now(), self.current_user, action, entity, str(entity_id), details))
        self.conn.commit()

    # ------------------------------------------------------------ backups
    def backup(self, dest_folder: str | Path | None = None, note: str = "") -> Path:
        dest = Path(dest_folder or config.folder(FOLDER) / "Backups")
        dest.mkdir(parents=True, exist_ok=True)
        out = dest / f"admin_station_{_dt.datetime.now():%Y%m%d_%H%M%S}.db"
        # two backups inside the same second must not collide -- the safety
        # copy taken by restore() would otherwise overwrite the very file it
        # is about to restore from, silently wiping it.
        n = 2
        while out.exists():
            out = dest / (f"admin_station_{_dt.datetime.now():%Y%m%d_%H%M%S}"
                          f"_{n}.db")
            n += 1
        self.conn.commit()
        tgt = sqlite3.connect(str(out))
        with tgt:
            self.conn.backup(tgt)
        tgt.close()
        self.audit("BACKUP", "database", out.name, note)
        return out

    def restore(self, src: str | Path) -> None:
        src = Path(src)
        if not src.exists():
            raise FileNotFoundError(src)
        safety = self.backup(note=f"safety copy before restoring {src.name}")
        if safety.resolve() == src.resolve():
            raise ValueError("Refusing to restore a file over itself.")
        self.conn.close()
        shutil.copy2(src, self.path)
        self.conn = sqlite3.connect(str(self.path), check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._init()
        self.audit("RESTORE", "database", src.name)


_admin_db: AdminDB | None = None


def get_admin_db() -> AdminDB:
    global _admin_db
    if _admin_db is None:
        _admin_db = AdminDB()
    return _admin_db


def set_admin_db(db: AdminDB | None) -> None:
    global _admin_db
    _admin_db = db


def reset_admin_db() -> None:
    global _admin_db
    if _admin_db is not None:
        _admin_db.close()
    _admin_db = None


# --------------------------------------------------------------------- CRUD
def next_sr(db: AdminDB, camp: str = "") -> str:
    """Next SR# — continues the highest numeric SR already stored."""
    rows = db.query("SELECT sr_no FROM records")
    mx = 0
    for r in rows:
        t = re.sub(r"[^\d]", "", str(r["sr_no"] or ""))
        if t.isdigit():
            mx = max(mx, int(t))
    return str(mx + 1)


def save_record(db: AdminDB, data: dict, rec_id: int | None = None) -> int:
    fields = [f for f, _ in ALL_FIELDS]
    payload = {f: data.get(f, "") for f in fields}
    for f in NUMERIC:
        payload[f] = to_float(payload.get(f))
        if payload[f] < 0:                       # "Yes" in the Return column
            payload[f] = payload["qty"] if f == "qty_return" else 0.0
    payload["record_date"] = to_date(payload.get("record_date"))
    payload["status"] = payload.get("status") or _auto_status(payload)
    if rec_id:
        sets = ", ".join(f"{f}=?" for f in fields)
        db.execute(f"UPDATE records SET {sets}, updated_at=? WHERE id=?",
                   [payload[f] for f in fields] + [_now(), rec_id])
        db.commit()
        db.audit("EDITED", "record", rec_id, payload.get("description", ""))
        return rec_id
    if not payload.get("sr_no"):
        payload["sr_no"] = next_sr(db)
    cols = ", ".join(fields + ["created_by", "created_at", "updated_at", "batch_id"])
    marks = ", ".join("?" * (len(fields) + 4))
    cur = db.execute(f"INSERT INTO records({cols}) VALUES({marks})",
                     [payload[f] for f in fields] +
                     [db.current_user, _now(), _now(), data.get("batch_id")])
    db.commit()
    db.audit("CREATED", "record", cur.lastrowid, payload.get("description", ""))
    return int(cur.lastrowid)


def _auto_status(p: dict) -> str:
    q, r = to_float(p.get("qty")), to_float(p.get("qty_return"))
    if r <= 0:
        return "Active"
    if r >= q > 0:
        return "Returned"
    return "Partially Returned"


def delete_records(db: AdminDB, ids: Iterable[int]) -> int:
    ids = list(ids)
    if not ids:
        return 0
    db.execute(f"DELETE FROM records WHERE id IN ({','.join('?' * len(ids))})", ids)
    db.commit()
    db.audit("DELETED", "record", ",".join(str(i) for i in ids), f"{len(ids)} record(s)")
    return len(ids)


def get_record(db: AdminDB, rec_id: int) -> dict | None:
    r = db.one("SELECT * FROM records WHERE id=?", (rec_id,))
    return dict(r) if r else None


def search(db: AdminDB, text: str = "", camp: str = "", category: str = "",
           destination: str = "", status: str = "", date_from: str = "",
           date_to: str = "", only_open: bool = False) -> list[dict]:
    sql = "SELECT * FROM records WHERE 1=1"
    p: list[Any] = []
    if text.strip():
        like = f"%{text.strip()}%"
        sql += (" AND (description LIKE ? OR camp LIKE ? OR category LIKE ?"
                " OR destination LIKE ? OR remarks LIKE ? OR sr_no LIKE ?"
                " OR ref_no LIKE ? OR custodian LIKE ?)")
        p += [like] * 8
    for col, val in (("camp", camp), ("category", category),
                     ("destination", destination), ("status", status)):
        if val:
            sql += f" AND {col}=?"
            p.append(val)
    if date_from:
        sql += " AND record_date>=?"
        p.append(date_from)
    if date_to:
        sql += " AND record_date<=?"
        p.append(date_to)
    if only_open:
        sql += " AND qty_return < qty"
    sql += " ORDER BY record_date DESC, id DESC"
    return [dict(r) for r in db.query(sql, p)]


def distinct(db: AdminDB, column: str) -> list[str]:
    if column not in {f for f, _ in ALL_FIELDS}:
        return []
    return [r[0] for r in db.query(
        f"SELECT DISTINCT {column} FROM records WHERE COALESCE({column},'')<>''"
        f" ORDER BY {column}")]


# ------------------------------------------------------------------ dashboard
def _where(f: dict | None) -> tuple[str, list]:
    """Build a shared WHERE clause so every tile and chart obeys the filters."""
    f = f or {}
    sql, p = "", []
    if f.get("text"):
        like = f"%{f['text'].strip()}%"
        sql += (" AND (description LIKE ? OR camp LIKE ? OR category LIKE ?"
                " OR destination LIKE ? OR remarks LIKE ? OR sr_no LIKE ?"
                " OR ref_no LIKE ? OR custodian LIKE ?)")
        p += [like] * 8
    for col in ("camp", "category", "destination", "status"):
        if f.get(col):
            sql += f" AND {col}=?"
            p.append(f[col])
    if f.get("date_from"):
        sql += " AND record_date>=?"
        p.append(f["date_from"])
    if f.get("date_to"):
        sql += " AND record_date<=?"
        p.append(f["date_to"])
    if f.get("only_open"):
        sql += " AND qty_return < qty"
    return sql, p


def dashboard(db: AdminDB, f: dict | None = None) -> dict:
    w, p = _where(f)

    def sc(expr: str, extra: str = "", params: list | None = None):
        return db.scalar(f"SELECT {expr} FROM records WHERE 1=1{w}{extra}",
                         p + (params or []), default=0)

    total = sc("COUNT(*)")
    qty = float(sc("COALESCE(SUM(qty),0)"))
    ret = float(sc("COALESCE(SUM(qty_return),0)"))
    value = float(sc("COALESCE(SUM(qty*unit_cost),0)"))
    month = _dt.date.today().strftime("%Y-%m")
    open_lines = int(sc("COUNT(*)", " AND qty_return < qty"))
    return {
        "records": int(total),
        "camps": int(sc("COUNT(DISTINCT camp)", " AND camp<>''")),
        "categories": int(sc("COUNT(DISTINCT category)", " AND category<>''")),
        "destinations": int(sc("COUNT(DISTINCT destination)", " AND destination<>''")),
        "custodians": int(sc("COUNT(DISTINCT custodian)", " AND custodian<>''")),
        "qty": qty,
        "returned": ret,
        "on_site": qty - ret,
        "value": value,
        "avg_qty": (qty / total) if total else 0.0,
        "open_lines": open_lines,
        "returned_lines": int(sc("COUNT(*)", " AND qty_return >= qty AND qty>0")),
        "return_rate": (ret / qty * 100.0) if qty else 0.0,
        "this_month": int(sc("COUNT(*)", " AND substr(record_date,1,7)=?", [month])),
        "this_month_qty": float(sc("COALESCE(SUM(qty),0)",
                                   " AND substr(record_date,1,7)=?", [month])),
        "damaged": int(sc("COUNT(*)",
                          " AND lower(condition) IN ('damaged','scrap')")),
        "no_date": int(sc("COUNT(*)", " AND COALESCE(record_date,'')=''")),
        "last_import": (db.scalar("SELECT ts FROM batches ORDER BY id DESC LIMIT 1",
                                  default="") or ""),
    }


def by_column(db: AdminDB, column: str, limit: int = 10,
              measure: str = "qty", f: dict | None = None) -> list[tuple[str, float]]:
    if column not in {x for x, _ in ALL_FIELDS}:
        return []
    agg = {"qty": "SUM(qty)", "count": "COUNT(*)",
           "value": "SUM(qty*unit_cost)",
           "outstanding": "SUM(qty - qty_return)"}.get(measure, "SUM(qty)")
    w, p = _where(f)
    rows = db.query(
        f"SELECT COALESCE(NULLIF({column},''),'(blank)') k, {agg} v FROM records"
        f" WHERE 1=1{w} GROUP BY k ORDER BY v DESC LIMIT ?", p + [limit])
    return [(r["k"], float(r["v"] or 0)) for r in rows]


def monthly(db: AdminDB, months: int = 12,
            f: dict | None = None) -> list[tuple[str, float]]:
    w, p = _where(f)
    rows = db.query(
        f"SELECT substr(record_date,1,7) m, COALESCE(SUM(qty),0) q FROM records"
        f" WHERE length(record_date)>=7{w} GROUP BY m ORDER BY m DESC LIMIT ?",
        p + [months])
    return [(r["m"], float(r["q"])) for r in reversed(rows)]


def monthly_in_out(db: AdminDB, months: int = 12,
                   f: dict | None = None) -> list[tuple[str, float, float]]:
    """Issued vs returned per month — feeds the grouped bar chart."""
    w, p = _where(f)
    rows = db.query(
        f"SELECT substr(record_date,1,7) m, COALESCE(SUM(qty),0) q,"
        f" COALESCE(SUM(qty_return),0) r FROM records"
        f" WHERE length(record_date)>=7{w} GROUP BY m ORDER BY m DESC LIMIT ?",
        p + [months])
    return [(r["m"][-2:], float(r["q"]), float(r["r"])) for r in reversed(rows)]


def top_items(db: AdminDB, limit: int = 10,
              f: dict | None = None) -> list[tuple[str, float]]:
    w, p = _where(f)
    rows = db.query(
        f"SELECT COALESCE(NULLIF(description,''),'(blank)') k, SUM(qty) v"
        f" FROM records WHERE 1=1{w} GROUP BY lower(k) ORDER BY v DESC LIMIT ?",
        p + [limit])
    return [(r["k"], float(r["v"] or 0)) for r in rows]


def ageing(db: AdminDB, f: dict | None = None) -> list[tuple[str, float]]:
    """Outstanding lines bucketed by how long they have been out."""
    w, p = _where(f)
    rows = db.query(
        f"SELECT record_date d, (qty - qty_return) o FROM records"
        f" WHERE qty_return < qty AND length(record_date)=10{w}", p)
    today = _dt.date.today()
    buckets = {"0-30 d": 0.0, "31-60 d": 0.0, "61-90 d": 0.0,
               "91-180 d": 0.0, "180+ d": 0.0}
    for r in rows:
        try:
            age = (today - _dt.date.fromisoformat(r["d"])).days
        except ValueError:
            continue
        o = float(r["o"] or 0)
        if age <= 30:
            buckets["0-30 d"] += o
        elif age <= 60:
            buckets["31-60 d"] += o
        elif age <= 90:
            buckets["61-90 d"] += o
        elif age <= 180:
            buckets["91-180 d"] += o
        else:
            buckets["180+ d"] += o
    return [(k, v) for k, v in buckets.items()]


def condition_split(db: AdminDB, f: dict | None = None) -> list[tuple[str, float]]:
    w, p = _where(f)
    rows = db.query(
        f"SELECT COALESCE(NULLIF(condition,''),'(not set)') k, COUNT(*) v"
        f" FROM records WHERE 1=1{w} GROUP BY k ORDER BY v DESC", p)
    return [(r["k"], float(r["v"])) for r in rows]


def camp_matrix(db: AdminDB) -> tuple[list[str], list[list[Any]]]:
    """Camp x category pivot of quantities — an 'advanced' cross-tab view."""
    cats = distinct(db, "category") or ["(blank)"]
    camps = distinct(db, "camp") or ["(blank)"]
    grid = {(c, k): 0.0 for c in camps for k in cats}
    for r in db.query("SELECT COALESCE(NULLIF(camp,''),'(blank)') camp,"
                      " COALESCE(NULLIF(category,''),'(blank)') cat, SUM(qty) q"
                      " FROM records GROUP BY camp, cat"):
        grid[(r["camp"], r["cat"])] = float(r["q"] or 0)
    cols = ["Camp / Office"] + cats + ["Total"]
    rows: list[list[Any]] = []
    for c in camps:
        line: list[Any] = [c]
        tot = 0.0
        for k in cats:
            v = grid.get((c, k), 0.0)
            tot += v
            line.append(round(v, 2))
        line.append(round(tot, 2))
        rows.append(line)
    if rows:
        totals: list[Any] = ["TOTAL"]
        for i in range(1, len(cols)):
            totals.append(round(sum(float(r[i] or 0) for r in rows), 2))
        rows.append(totals)
    return cols, rows


# -------------------------------------------------------------------- reports
REPORT_LIST = [
    "Full Record Register",
    "Camp / Office Summary",
    "Category Summary",
    "Destination Summary",
    "Pending Returns",
    "Completed Returns",
    "Monthly Movement",
    "Camp × Category Matrix",
    "Duplicate Suspects",
    "Recently Added",
    "Damaged / Scrap Register",
    "Value by Camp",
]


def build_report(db: AdminDB, name: str, f: dict | None = None
                 ) -> tuple[str, list[str], list[list[Any]]]:
    f = f or {}
    recs = search(db, f.get("text", ""), f.get("camp", ""), f.get("category", ""),
                  f.get("destination", ""), f.get("status", ""),
                  f.get("date_from", ""), f.get("date_to", ""))

    def base_cols() -> list[str]:
        return [lbl for _, lbl in FIELDS] + ["Status"]

    def base_rows(src) -> list[list[Any]]:
        return [[r["sr_no"], r["camp"], r["record_date"], r["category"], r["description"],
                 r["uom"], round(r["qty"] or 0, 2), round(r["qty_return"] or 0, 2),
                 r["destination"], r["remarks"], r["status"]] for r in src]

    if name == "Full Record Register":
        return name, base_cols(), base_rows(recs)

    if name == "Recently Added":
        rows = sorted(recs, key=lambda r: r.get("created_at") or "", reverse=True)[:200]
        return name, base_cols(), base_rows(rows)

    if name == "Pending Returns":
        rows = [r for r in recs if (r["qty_return"] or 0) < (r["qty"] or 0)]
        cols = ["SR#", "Camp / Office", "Date", "Category", "Description", "UOM",
                "Quantity", "Returned", "Outstanding", "Destination", "Remarks"]
        data = [[r["sr_no"], r["camp"], r["record_date"], r["category"], r["description"],
                 r["uom"], round(r["qty"] or 0, 2), round(r["qty_return"] or 0, 2),
                 round((r["qty"] or 0) - (r["qty_return"] or 0), 2), r["destination"],
                 r["remarks"]] for r in rows]
        if data:
            data.append(["", "TOTAL", "", "", f"{len(data)} line(s)", "",
                         round(sum(x[6] for x in data), 2),
                         round(sum(x[7] for x in data), 2),
                         round(sum(x[8] for x in data), 2), "", ""])
        return name, cols, data

    if name == "Completed Returns":
        rows = [r for r in recs if (r["qty_return"] or 0) >= (r["qty"] or 0) and (r["qty"] or 0) > 0]
        return name, base_cols(), base_rows(rows)

    if name == "Damaged / Scrap Register":
        rows = [r for r in recs
                if str(r.get("condition", "")).lower() in ("damaged", "scrap")
                or "damag" in str(r.get("remarks", "")).lower()]
        cols = base_cols() + ["Condition"]
        return name, cols, [row + [r.get("condition", "")]
                            for row, r in zip(base_rows(rows), rows)]

    if name in ("Camp / Office Summary", "Category Summary", "Destination Summary",
                "Value by Camp"):
        key = {"Camp / Office Summary": "camp", "Category Summary": "category",
               "Destination Summary": "destination", "Value by Camp": "camp"}[name]
        label = {"camp": "Camp / Office", "category": "Item Category",
                 "destination": "Destination Location"}[key]
        agg: dict[str, dict] = {}
        for r in recs:
            k = r.get(key) or "(blank)"
            a = agg.setdefault(k, {"lines": 0, "qty": 0.0, "ret": 0.0, "val": 0.0,
                                   "cats": set()})
            a["lines"] += 1
            a["qty"] += float(r["qty"] or 0)
            a["ret"] += float(r["qty_return"] or 0)
            a["val"] += float(r["qty"] or 0) * float(r["unit_cost"] or 0)
            a["cats"].add(r.get("category") or "")
        if name == "Value by Camp":
            cols = [label, "Lines", "Quantity", "Estimated Value"]
            data = [[k, a["lines"], round(a["qty"], 2), round(a["val"], 2)]
                    for k, a in sorted(agg.items(), key=lambda kv: -kv[1]["val"])]
            if data:
                data.append(["TOTAL", sum(x[1] for x in data),
                             round(sum(x[2] for x in data), 2),
                             round(sum(x[3] for x in data), 2)])
            return name, cols, data
        cols = [label, "Lines", "Categories", "Quantity", "Returned", "On Site"]
        data = [[k, a["lines"], len([c for c in a["cats"] if c]), round(a["qty"], 2),
                 round(a["ret"], 2), round(a["qty"] - a["ret"], 2)]
                for k, a in sorted(agg.items(), key=lambda kv: -kv[1]["qty"])]
        if data:
            data.append(["TOTAL", sum(x[1] for x in data), "",
                         round(sum(x[3] for x in data), 2),
                         round(sum(x[4] for x in data), 2),
                         round(sum(x[5] for x in data), 2)])
        return name, cols, data

    if name == "Monthly Movement":
        agg: dict[str, dict] = {}
        for r in recs:
            k = (r.get("record_date") or "")[:7] or "(no date)"
            a = agg.setdefault(k, {"lines": 0, "qty": 0.0, "ret": 0.0})
            a["lines"] += 1
            a["qty"] += float(r["qty"] or 0)
            a["ret"] += float(r["qty_return"] or 0)
        cols = ["Month", "Lines", "Quantity", "Returned", "Net"]
        data = [[k, a["lines"], round(a["qty"], 2), round(a["ret"], 2),
                 round(a["qty"] - a["ret"], 2)] for k, a in sorted(agg.items())]
        if data:
            data.append(["TOTAL", sum(x[1] for x in data),
                         round(sum(x[2] for x in data), 2),
                         round(sum(x[3] for x in data), 2),
                         round(sum(x[4] for x in data), 2)])
        return name, cols, data

    if name == "Camp × Category Matrix":
        cols, rows = camp_matrix(db)
        return name, cols, rows

    if name == "Duplicate Suspects":
        seen: dict[tuple, list[dict]] = {}
        for r in recs:
            k = (norm(r.get("camp")), norm(r.get("description")),
                 r.get("record_date") or "", round(float(r["qty"] or 0), 2))
            seen.setdefault(k, []).append(r)
        cols = ["Camp / Office", "Date", "Description", "Quantity", "Times", "SR numbers"]
        data = [[g[0]["camp"], g[0]["record_date"], g[0]["description"],
                 round(g[0]["qty"] or 0, 2), len(g),
                 ", ".join(str(x["sr_no"]) for x in g)]
                for g in seen.values() if len(g) > 1]
        return name, cols, data

    return name, base_cols(), base_rows(recs)


# --------------------------------------------------------------------- import
def sniff(text: str) -> tuple[list[str], list[list[str]]]:
    """Split pasted clipboard text into (headers, rows) — tab, csv, pipe or 2+ spaces."""
    raw = [ln for ln in text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
           if ln.strip()]
    if not raw:
        return [], []
    if "\t" in raw[0]:
        rows = [ln.split("\t") for ln in raw]
    elif raw[0].count(",") >= 2:
        rows = list(csv.reader(io.StringIO("\n".join(raw))))
    elif raw[0].count("|") >= 2:
        rows = [[c.strip() for c in ln.strip().strip("|").split("|")] for ln in raw]
    else:
        rows = [re.split(r"\s{2,}", ln.strip()) for ln in raw]
    rows = [[str(c).strip() for c in r] for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return [], []
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    head = rows[0]
    if sum(1 for c in head if norm(c) in HEADER_MAP) >= max(2, width // 3):
        return head, rows[1:]
    return [f"Column {i + 1}" for i in range(width)], rows


def read_file(path: str | Path) -> tuple[list[str], list[list[Any]]]:
    """Read xlsx / xls(m) / csv / txt into (headers, rows)."""
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(p, data_only=True, read_only=True)
        ws = wb.active
        data = [[("" if c is None else c) for c in row]
                for row in ws.iter_rows(values_only=True)]
        wb.close()
        data = [r for r in data if any(str(c).strip() for c in r)]
        if not data:
            return [], []
        head = [str(c).strip() for c in data[0]]
        if sum(1 for c in head if norm(c) in HEADER_MAP) >= 2:
            return head, data[1:]
        return [f"Column {i + 1}" for i in range(len(head))], data
    return sniff(p.read_text(encoding="utf-8", errors="ignore"))


def auto_map(headers: Sequence[str]) -> dict[int, str]:
    """Column index -> canonical field."""
    out: dict[int, str] = {}
    used: set[str] = set()
    for i, h in enumerate(headers):
        f = HEADER_MAP.get(norm(h))
        if f and f not in used:
            out[i] = f
            used.add(f)
    return out


def preview(headers: Sequence[str], rows: Sequence[Sequence[Any]],
            mapping: dict[int, str], defaults: dict | None = None) -> list[dict]:
    """Turn raw rows + mapping into canonical record dicts (not yet saved)."""
    defaults = defaults or {}
    out = []
    for r in rows:
        rec = {f: "" for f, _ in ALL_FIELDS}
        rec.update({k: v for k, v in defaults.items() if v})
        for i, field in mapping.items():
            if i < len(r):
                rec[field] = r[i]
        for f in NUMERIC:
            rec[f] = to_float(rec.get(f))
        if rec["qty_return"] < 0:
            rec["qty_return"] = rec["qty"]
        rec["record_date"] = to_date(rec.get("record_date")) or defaults.get("record_date", "")
        rec["status"] = rec.get("status") or _auto_status(rec)
        if not any(str(rec.get(k, "")).strip() for k in
                   ("camp", "description", "category", "destination")) and not rec["qty"]:
            continue
        out.append(rec)
    return out


def import_records(db: AdminDB, records: Sequence[dict], source: str = "",
                   mapping: dict | None = None, skip_duplicates: bool = True
                   ) -> tuple[int, int]:
    """Insert prepared records. Returns (inserted, skipped)."""
    cur = db.execute(
        "INSERT INTO batches(ts,source,rows,skipped,mapping,username) VALUES(?,?,?,?,?,?)",
        (_now(), str(source), 0, 0, json.dumps(mapping or {}), db.current_user))
    batch_id = int(cur.lastrowid)

    existing: set[tuple] = set()
    if skip_duplicates:
        for r in db.query("SELECT camp, description, record_date, qty, destination"
                          " FROM records"):
            existing.add((norm(r["camp"]), norm(r["description"]), r["record_date"] or "",
                          round(float(r["qty"] or 0), 3), norm(r["destination"])))

    ins = skipped = 0
    auto_sr = next_sr(db)
    n = int(re.sub(r"[^\d]", "", auto_sr) or 0)
    for rec in records:
        key = (norm(rec.get("camp")), norm(rec.get("description")),
               rec.get("record_date") or "", round(to_float(rec.get("qty")), 3),
               norm(rec.get("destination")))
        if skip_duplicates and key in existing:
            skipped += 1
            continue
        existing.add(key)
        data = dict(rec)
        data["batch_id"] = batch_id
        if not str(data.get("sr_no") or "").strip():
            n += 1
            data["sr_no"] = str(n)
        save_record(db, data)
        ins += 1
    db.execute("UPDATE batches SET rows=?, skipped=? WHERE id=?", (ins, skipped, batch_id))
    db.commit()
    db.audit("IMPORTED", "records", batch_id,
             f"{ins} inserted, {skipped} duplicate(s) skipped from {source}")
    return ins, skipped


def undo_batch(db: AdminDB, batch_id: int) -> int:
    n = db.execute("DELETE FROM records WHERE batch_id=?", (batch_id,)).rowcount
    db.commit()
    db.audit("DELETED", "batch", batch_id, f"{n} record(s) removed (undo import)")
    return int(n)


def batches(db: AdminDB) -> list[dict]:
    return [dict(r) for r in db.query(
        "SELECT b.*, (SELECT COUNT(*) FROM records r WHERE r.batch_id=b.id) live"
        " FROM batches b ORDER BY b.id DESC LIMIT 100")]


def template_rows() -> tuple[list[str], list[list[Any]]]:
    """The recommended upload format (also used for the downloadable template)."""
    cols = [lbl for _, lbl in FIELDS]
    today = _dt.date.today().isoformat()
    sample = [
        ["1", "Camp 1 - Jubail", today, "Furniture & Fittings", "Single Steel Bed",
         "No", 8, 0, "Room B-12", "New arrival"],
        ["2", "Camp 1 - Jubail", today, "Accommodation", "Window AC", "No", 2, 1,
         "Room B-12", "One unit returned for service"],
        ["3", "Head Office - Dammam", today, "IT Equipment", "Laptop", "No", 1, 0,
         "Admin Dept", ""],
    ]
    return cols, sample


# ------------------------------------------------------- shared drop folder
SUPPORTED_SUFFIXES = (".xlsx", ".xlsm", ".csv", ".txt")


def get_watch_folder(db: AdminDB) -> str:
    return db.get_setting("watch_folder", "") or ""


def set_watch_folder(db: AdminDB, path: str) -> None:
    db.set_setting("watch_folder", str(path or ""))
    db.audit("EDITED", "watch-folder", "", str(path))


def folder_status(path: str | Path) -> tuple[bool, str]:
    """Check a shared folder is usable before we promise anything.

    Returns (ok, message). A network share that is offline, or one the user can
    only read, is reported plainly instead of failing later mid-import.
    """
    if not str(path or "").strip():
        return False, "No folder selected."
    p = Path(path)
    if not p.exists():
        return False, f"The folder does not exist or is offline:\n{p}"
    if not p.is_dir():
        return False, f"That path is a file, not a folder:\n{p}"
    if not os.access(p, os.R_OK):
        return False, f"No permission to read:\n{p}"
    writable = os.access(p, os.W_OK)
    note = "Read and write access." if writable else \
        "Read-only — files can be imported but not archived."
    return True, note


def scan_folder(db: AdminDB, path: str | Path,
                include_done: bool = False) -> list[dict]:
    """List uploadable files in the shared folder, newest first.

    Each entry records whether that exact file has already been imported, so a
    site admin dropping the same sheet twice cannot double-post it.
    """
    ok, _ = folder_status(path)
    if not ok:
        return []
    done = {r["source"]: r for r in db.query(
        "SELECT source, MAX(ts) ts, SUM(rows) rows FROM batches GROUP BY source")}
    out = []
    for f in sorted(Path(path).iterdir(), key=lambda x: x.name.lower()):
        if not f.is_file() or f.suffix.lower() not in SUPPORTED_SUFFIXES:
            continue
        if f.name.startswith("~$"):          # Excel lock file
            continue
        try:
            st = f.stat()
        except OSError:
            continue
        key = str(f)
        prev = done.get(key)
        rec = {
            "path": key, "name": f.name,
            "size_kb": round(st.st_size / 1024.0, 1),
            "modified": _dt.datetime.fromtimestamp(st.st_mtime
                                                   ).strftime("%Y-%m-%d %H:%M"),
            "imported_at": (prev["ts"] if prev else ""),
            "imported_rows": int(prev["rows"] or 0) if prev else 0,
            "status": "Imported" if prev else "New",
            "site": _site_from_name(f),
        }
        if prev and not include_done:
            continue
        out.append(rec)
    out.sort(key=lambda r: r["modified"], reverse=True)
    return out


def _site_from_name(f: Path) -> str:
    """Guess which site sent a file from its name or its sub-folder."""
    stem = f.stem.replace("_", " ").replace("-", " ")
    for token in stem.split():
        if token.lower() in ("camp", "site", "office"):
            idx = stem.split().index(token)
            rest = stem.split()[idx:idx + 2]
            return " ".join(rest).title()
    parent = f.parent.name
    return parent if parent and len(parent) < 40 else ""


def import_from_folder(db: AdminDB, files: Sequence[str], defaults: dict | None = None,
                       skip_duplicates: bool = True,
                       archive_to: str | Path | None = None) -> dict:
    """Import several dropped files in one pass.

    Returns a per-file summary so the operator can see exactly what happened,
    including files that could not be read at all.
    """
    summary = {"files": [], "inserted": 0, "skipped": 0, "failed": 0}
    for path in files:
        entry = {"file": Path(path).name, "inserted": 0, "skipped": 0, "error": ""}
        try:
            headers, rows = read_file(path)
            if not rows:
                raise ValueError("no data rows found")
            mapping = auto_map(headers)
            if not mapping:
                raise ValueError("no recognisable columns")
            base = dict(defaults or {})
            if not base.get("camp"):
                guess = _site_from_name(Path(path))
                if guess:
                    base["camp"] = guess
            recs = preview(headers, rows, mapping, base)
            ins, sk = import_records(db, recs, str(path), mapping, skip_duplicates)
            entry["inserted"], entry["skipped"] = ins, sk
            summary["inserted"] += ins
            summary["skipped"] += sk
            if archive_to and ins:
                try:
                    dest = Path(archive_to)
                    dest.mkdir(parents=True, exist_ok=True)
                    target = dest / Path(path).name
                    if target.exists():
                        target = dest / (f"{Path(path).stem}_"
                                         f"{_dt.datetime.now():%Y%m%d_%H%M%S}"
                                         f"{Path(path).suffix}")
                    shutil.move(str(path), str(target))
                    entry["archived"] = str(target)
                except OSError as exc:
                    entry["error"] = f"imported, but could not archive: {exc}"
        except Exception as exc:  # noqa: BLE001
            entry["error"] = str(exc)
            summary["failed"] += 1
        summary["files"].append(entry)
    db.audit("IMPORTED", "watch-folder", "",
             f"{summary['inserted']} row(s) from {len(files)} file(s), "
             f"{summary['failed']} failed")
    return summary
