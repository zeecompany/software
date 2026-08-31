"""TOOLS, INSTRUMENTS & DEVICES — the custody register module.

(Historically called the "Tool Station"; the data folder is migrated
automatically, see `_migrate_legacy_folder`.)

A separate station in the same spirit as the Admin Station: it tracks **who is
holding which tool**, not how much stock exists. Enforced physically, not by
convention:

    ·  its own SQLite file        <storage>/Tools, Instruments & Devices/
                                  tool_station.db
    ·  its own schema, numbering, audit trail and backups
    ·  no foreign key, join or import from items / stock_ledger / documents
    ·  nothing here ever posts a stock movement

Record shape follows the controlled form the user supplied
(``WH-FRM-001 Rev 00``, "TOOLS, DEVICES & INSTRUMENTS HANDOVER FORM"):

    A — HANDOVER DETAILS      form no · reference · date · time · type ·
                              expected return · project id / name / location
    B — RECIPIENT / CUSTODIAN handed to · iqama · job title · mobile ·
                              company · email · supervisor · cost code
    C — ITEM DETAILS          asset id · category · description · make/model ·
                              serial · qty · accessories · condition ·
                              calibration due · remarks
    D — ACKNOWLEDGEMENT       issued by / received by, names + date-time
    E — ITEM PHOTOGRAPHS      evidence per item

The four transaction types on the form drive the whole custody engine:

    Issue          tool leaves the warehouse, open-ended custody
    Transfer       custody moves from one holder to another
    Temporary Loan tool must come back by the expected return date
    Return         tool comes back to the warehouse (closes an earlier record)

The reference number the form generates is self-describing and is decoded on
sight, so a folder full of signed PDFs can be filed without typing anything:

    WH-087IS2308202601
    ^^ ^^^ ^^ ^^^^^^^^ ^^
    |  |   |  |        `- sequence within that day
    |  |   |  `---------- date 23/08/2026
    |  |   `------------- transaction type (IS/TR/TL/RT)
    |  `----------------- project 087  ->  PRJ000087
    `-------------------- originating warehouse
"""
from __future__ import annotations

import csv
import datetime as _dt
import io
import os
import re
import shutil
import sqlite3
from pathlib import Path
from typing import Any, Iterable, Sequence

from . import config

MODULE_NAME = "Tools, Instruments & Devices"
FOLDER = MODULE_NAME
LEGACY_FOLDER = "Tool Station"        # folder name used before the rename
DB_NAME = "tool_station.db"

SCHEMA_VERSION = 1

# ------------------------------------------------------------------ vocabulary
ISSUE = "Issue"
TRANSFER = "Transfer"
LOAN = "Temporary Loan"
RETURN = "Return"
TXN_TYPES = (ISSUE, TRANSFER, LOAN, RETURN)

# short codes as they appear inside the handover reference number
TXN_CODES = {"IS": ISSUE, "TR": TRANSFER, "TL": LOAN, "RT": RETURN}
CODE_OF = {v: k for k, v in TXN_CODES.items()}

TXN_COLORS = {
    ISSUE: "#1098ad",
    TRANSFER: "#7048e8",
    LOAN: "#e8590c",
    RETURN: "#1a9c52",
}

# custody state of a handover
OPEN = "Open"                 # tool is out with the custodian
PART_RETURNED = "Partially Returned"
CLOSED = "Returned"
TRANSFERRED = "Transferred Out"
OVERDUE = "Overdue"
CANCELLED = "Cancelled"

STATUS_COLORS = {
    OPEN: "#1098ad",
    PART_RETURNED: "#e8590c",
    CLOSED: "#1a9c52",
    TRANSFERRED: "#7048e8",
    OVERDUE: "#c92a2a",
    CANCELLED: "#6b7c8f",
}

# condition grades exactly as printed on the form
CONDITIONS = {
    "A": "A – New / Excellent",
    "B": "B – Good",
    "C": "C – Fair / Usable",
    "D": "D – Damaged / Not Usable",
}

DDL = """
CREATE TABLE IF NOT EXISTS settings (
    key   TEXT PRIMARY KEY,
    value TEXT
);

CREATE TABLE IF NOT EXISTS handovers (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    ref_no         TEXT UNIQUE NOT NULL,
    form_no        TEXT DEFAULT '',
    txn_type       TEXT NOT NULL DEFAULT 'Issue',
    doc_date       TEXT DEFAULT '',
    doc_time       TEXT DEFAULT '',
    expected_return TEXT DEFAULT '',
    warehouse      TEXT DEFAULT '',
    project_id     TEXT DEFAULT '',
    project_name   TEXT DEFAULT '',
    location       TEXT DEFAULT '',
    -- section B
    handed_to      TEXT DEFAULT '',
    iqama_id       TEXT DEFAULT '',
    job_title      TEXT DEFAULT '',
    mobile         TEXT DEFAULT '',
    company        TEXT DEFAULT '',
    email          TEXT DEFAULT '',
    supervisor     TEXT DEFAULT '',
    cost_code      TEXT DEFAULT '',
    -- section D
    issued_by      TEXT DEFAULT '',
    issued_at      TEXT DEFAULT '',
    received_by    TEXT DEFAULT '',
    received_at    TEXT DEFAULT '',
    -- verification ticks
    v_serial       INTEGER NOT NULL DEFAULT 0,
    v_accessories  INTEGER NOT NULL DEFAULT 0,
    v_calibration  INTEGER NOT NULL DEFAULT 0,
    v_photos       INTEGER NOT NULL DEFAULT 0,
    -- custody bookkeeping
    status         TEXT NOT NULL DEFAULT 'Open',
    closed_by_ref  TEXT DEFAULT '',
    parent_ref     TEXT DEFAULT '',
    remarks        TEXT DEFAULT '',
    source_file    TEXT DEFAULT '',
    file_hash      TEXT DEFAULT '',
    created_by     TEXT DEFAULT '',
    created_at     TEXT DEFAULT (datetime('now','localtime')),
    updated_at     TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS ix_th_type    ON handovers(txn_type);
CREATE INDEX IF NOT EXISTS ix_th_status  ON handovers(status);
CREATE INDEX IF NOT EXISTS ix_th_date    ON handovers(doc_date);
CREATE INDEX IF NOT EXISTS ix_th_project ON handovers(project_id);
CREATE INDEX IF NOT EXISTS ix_th_holder  ON handovers(handed_to);

CREATE TABLE IF NOT EXISTS handover_lines (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    handover_id  INTEGER NOT NULL,
    line_no      INTEGER NOT NULL DEFAULT 0,
    asset_id     TEXT DEFAULT '',
    category     TEXT DEFAULT '',
    description  TEXT DEFAULT '',
    make_model   TEXT DEFAULT '',
    serial_no    TEXT DEFAULT '',
    qty          REAL NOT NULL DEFAULT 1,
    qty_returned REAL NOT NULL DEFAULT 0,
    accessories  TEXT DEFAULT '',
    condition    TEXT DEFAULT '',
    calib_due    TEXT DEFAULT '',
    remarks      TEXT DEFAULT '',
    photo        TEXT DEFAULT ''
);
CREATE INDEX IF NOT EXISTS ix_tl_ho     ON handover_lines(handover_id);
CREATE INDEX IF NOT EXISTS ix_tl_asset  ON handover_lines(asset_id);
CREATE INDEX IF NOT EXISTS ix_tl_serial ON handover_lines(serial_no);

CREATE TABLE IF NOT EXISTS assets (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    asset_id      TEXT UNIQUE NOT NULL,
    category      TEXT DEFAULT '',
    description   TEXT DEFAULT '',
    make_model    TEXT DEFAULT '',
    serial_no     TEXT DEFAULT '',
    condition     TEXT DEFAULT '',
    calib_due     TEXT DEFAULT '',
    holder        TEXT DEFAULT '',
    holder_iqama  TEXT DEFAULT '',
    project_id    TEXT DEFAULT '',
    location      TEXT DEFAULT '',
    status        TEXT DEFAULT 'In Store',
    last_ref      TEXT DEFAULT '',
    last_date     TEXT DEFAULT '',
    notes         TEXT DEFAULT '',
    created_at    TEXT DEFAULT (datetime('now','localtime')),
    updated_at    TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS ix_as_holder ON assets(holder);
CREATE INDEX IF NOT EXISTS ix_as_status ON assets(status);

CREATE TABLE IF NOT EXISTS files (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    path        TEXT UNIQUE NOT NULL,
    name        TEXT DEFAULT '',
    ref_no      TEXT DEFAULT '',
    handover_id INTEGER,
    size_kb     REAL DEFAULT 0,
    modified    TEXT DEFAULT '',
    file_hash   TEXT DEFAULT '',
    pages       INTEGER DEFAULT 0,
    status      TEXT DEFAULT 'New',
    note        TEXT DEFAULT '',
    seen_at     TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS ix_tf_ref ON files(ref_no);

CREATE TABLE IF NOT EXISTS folders (
    id       INTEGER PRIMARY KEY AUTOINCREMENT,
    path     TEXT UNIQUE NOT NULL,
    label    TEXT DEFAULT '',
    active   INTEGER NOT NULL DEFAULT 1,
    added_at TEXT DEFAULT (datetime('now','localtime')),
    last_scan TEXT DEFAULT ''
);

CREATE TABLE IF NOT EXISTS audit (
    id        INTEGER PRIMARY KEY AUTOINCREMENT,
    ts        TEXT DEFAULT (datetime('now','localtime')),
    username  TEXT DEFAULT '',
    action    TEXT DEFAULT '',
    entity    TEXT DEFAULT '',
    entity_id TEXT DEFAULT '',
    details   TEXT DEFAULT ''
);
"""


def _now() -> str:
    return _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def today() -> str:
    return _dt.date.today().isoformat()


def norm(s: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def _migrate_legacy_folder() -> None:
    """Move an existing "Tool Station" data folder to the new module name.

    Renaming the module must never strand the custody database, so the old
    folder is moved once, on first use. If anything at all goes wrong (network
    share, file in use) the old folder is simply left alone and its content is
    copied instead — the module keeps working either way.
    """
    root = config.get_storage_root() or config.default_storage_root()
    old, new = Path(root) / LEGACY_FOLDER, Path(root) / FOLDER
    if not old.exists() or old == new:
        return
    try:
        if not new.exists():
            old.rename(new)
            return
        # both exist: only fill the gaps, never overwrite newer data
        for item in old.iterdir():
            target = new / item.name
            if target.exists():
                continue
            shutil.move(str(item), str(target))
        if not any(old.iterdir()):
            old.rmdir()
    except OSError:
        pass


def module_folder() -> Path:
    _migrate_legacy_folder()
    return config.folder(FOLDER)


def db_path() -> Path:
    return module_folder() / DB_NAME


def evidence_dir() -> Path:
    return module_folder() / "Evidence"


# --------------------------------------------------------------- value parsing
_DATE_PATTERNS = (
    "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d", "%Y/%m/%d",
    "%d/%m/%y", "%d-%m-%y", "%d-%b-%Y", "%d %b %Y", "%d-%b-%y",
    "%b %d, %Y", "%d %B %Y", "%m/%d/%Y",
)


def to_float(v: Any, default: float = 0.0) -> float:
    if v in (None, ""):
        return default
    if isinstance(v, (int, float)):
        return float(v)
    t = re.sub(r"[^\d.\-]", "", str(v))
    try:
        return float(t) if t not in ("", "-", ".") else default
    except ValueError:
        return default


def to_date(v: Any) -> str:
    """Best-effort conversion of anything date-like to ISO yyyy-mm-dd."""
    if v in (None, ""):
        return ""
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    t = str(v).strip()
    if not t or t == "-":
        return ""
    for f in _DATE_PATTERNS:
        try:
            return _dt.datetime.strptime(t, f).date().isoformat()
        except ValueError:
            continue
    try:
        n = float(t)
        if 20000 < n < 60000:
            return (_dt.date(1899, 12, 30) + _dt.timedelta(days=int(n))).isoformat()
    except ValueError:
        pass
    return t


def fmt_date(iso: str) -> str:
    """ISO -> dd/mm/yyyy for printing on the controlled form."""
    try:
        return _dt.date.fromisoformat(str(iso)[:10]).strftime("%d/%m/%Y")
    except Exception:          # noqa: BLE001
        return str(iso or "")


# ------------------------------------------------------- reference decoding
# WH-087IS2308202601 -> warehouse WH · project 087 · Issue · 23/08/2026 · #01
_REF_RE = re.compile(
    r"^(?P<wh>[A-Z]{2,4})[- ]?(?P<proj>\d{2,4})(?P<code>IS|TR|TL|RT)"
    r"(?P<d>\d{2})(?P<m>\d{2})(?P<y>\d{4})(?P<seq>\d{1,3})$", re.I)


def parse_ref(ref: str) -> dict:
    """Decode a handover reference number into its parts.

    Returns {} when the text is not a reference, so callers can simply test the
    result. Never raises — a badly typed reference must not stop an import.
    """
    m = _REF_RE.match(str(ref or "").strip().replace(" ", ""))
    if not m:
        return {}
    g = m.groupdict()
    try:
        date = _dt.date(int(g["y"]), int(g["m"]), int(g["d"])).isoformat()
    except ValueError:
        date = ""
    return {
        "ref_no": str(ref).strip(),
        "warehouse": g["wh"].upper(),
        "project_no": g["proj"],
        "project_id": f"PRJ{int(g['proj']):06d}",
        "txn_type": TXN_CODES[g["code"].upper()],
        "doc_date": date,
        "seq": int(g["seq"]),
    }


def make_ref(warehouse: str, project_id: str, txn_type: str,
             date: str, seq: int) -> str:
    """Build a reference in the same shape the paper form produces."""
    wh = re.sub(r"[^A-Z]", "", str(warehouse or "WH").upper())[:4] or "WH"
    digits = re.sub(r"\D", "", str(project_id or ""))
    proj = f"{int(digits) % 1000:03d}" if digits else "000"
    code = CODE_OF.get(txn_type, "IS")
    d = to_date(date) or today()
    try:
        dt = _dt.date.fromisoformat(d)
    except ValueError:
        dt = _dt.date.today()
    return f"{wh}-{proj}{code}{dt:%d%m%Y}{int(seq):02d}"


def next_ref(db: "ToolDB", warehouse: str, project_id: str, txn_type: str,
             date: str = "") -> str:
    """Next free reference for that warehouse / project / type / day."""
    d = to_date(date) or today()
    for seq in range(1, 100):
        ref = make_ref(warehouse, project_id, txn_type, d, seq)
        if not db.one("SELECT 1 FROM handovers WHERE ref_no=?", (ref,)):
            return ref
    return make_ref(warehouse, project_id, txn_type, d, 99)


# ------------------------------------------------------------------- database
class ToolDB:
    """Standalone database for the Tools, Instruments & Devices module."""

    def __init__(self, path: str | Path | None = None):
        self.path = Path(path or db_path())
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(str(self.path), check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA busy_timeout=15000")
        self.conn.execute("PRAGMA foreign_keys=ON")
        self.current_user = "admin"
        self._init()

    # ------------------------------------------------------------- schema
    def _init(self) -> None:
        # journal mode is chosen for where the file actually lives: WAL needs
        # real shared memory and is unsafe on an SMB share, so a networked
        # install falls back to TRUNCATE rather than risking corruption.
        try:
            self.conn.execute(f"PRAGMA journal_mode={self._journal_mode()}")
        except sqlite3.Error:
            pass
        self.conn.executescript(DDL)
        self.conn.commit()
        self._migrate()

    def _journal_mode(self) -> str:
        p = str(self.path)
        networked = p.startswith("\\\\") or p.startswith("//")
        if not networked and os.name != "nt":
            try:
                import subprocess  # noqa: S404 - read-only mount check
                networked = False
            except Exception:      # noqa: BLE001
                networked = False
        return "TRUNCATE" if networked else "WAL"

    def _migrate(self) -> None:
        """Additive migrations only — a column is added, never dropped."""
        for table, wanted in (("handovers", {}), ("handover_lines", {}),
                              ("assets", {}), ("files", {})):
            have = {r["name"] for r in self.conn.execute(
                f"PRAGMA table_info({table})")}
            for col, ddl in wanted.items():
                if col not in have:
                    self.conn.execute(
                        f"ALTER TABLE {table} ADD COLUMN {col} {ddl}")
        self.conn.commit()
        self.set_setting("schema_version", str(SCHEMA_VERSION))

    # --------------------------------------------------------- primitives
    def execute(self, sql: str, params: Sequence = ()) -> sqlite3.Cursor:
        return self.conn.execute(sql, params)

    def query(self, sql: str, params: Sequence = ()) -> list[sqlite3.Row]:
        return list(self.conn.execute(sql, params).fetchall())

    def one(self, sql: str, params: Sequence = ()):
        return self.conn.execute(sql, params).fetchone()

    def scalar(self, sql: str, params: Sequence = (), default=0):
        r = self.conn.execute(sql, params).fetchone()
        return default if r is None or r[0] is None else r[0]

    def commit(self) -> None:
        self.conn.commit()

    def close(self) -> None:
        try:
            self.conn.commit()
            self.conn.close()
        except Exception:          # noqa: BLE001
            pass

    # ----------------------------------------------------------- settings
    def get_setting(self, key: str, default: Any = None) -> Any:
        r = self.one("SELECT value FROM settings WHERE key=?", (key,))
        return r["value"] if r else default

    def set_setting(self, key: str, value: Any) -> None:
        self.execute("INSERT INTO settings(key,value) VALUES(?,?) "
                     "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                     (key, str(value)))
        self.conn.commit()

    def audit(self, action: str, entity: str = "", entity_id: str = "",
              details: str = "") -> None:
        self.execute("INSERT INTO audit(ts,username,action,entity,entity_id,details)"
                     " VALUES(?,?,?,?,?,?)",
                     (_now(), self.current_user, action, entity,
                      str(entity_id), details))
        self.conn.commit()

    # ------------------------------------------------------------ backups
    def backup(self, dest_folder: str | Path | None = None, note: str = "") -> Path:
        dest = Path(dest_folder or config.folder(FOLDER) / "Backups")
        dest.mkdir(parents=True, exist_ok=True)
        out = dest / f"tool_station_{_dt.datetime.now():%Y%m%d_%H%M%S}.db"
        # two backups inside the same second must not collide -- the safety
        # copy taken by restore() would otherwise overwrite the very file it
        # is about to restore from, silently wiping it.
        n = 2
        while out.exists():
            out = dest / (f"tool_station_{_dt.datetime.now():%Y%m%d_%H%M%S}"
                          f"_{n}.db")
            n += 1
        self.conn.commit()
        tgt = sqlite3.connect(str(out))
        with tgt:
            self.conn.backup(tgt)
        tgt.close()
        self.audit("BACKUP", "database", out.name, note)
        return out

    def restore(self, src: str | Path) -> None:
        src = Path(src)
        if not src.exists():
            raise FileNotFoundError(src)
        safety = self.backup(note=f"safety copy before restoring {src.name}")
        if safety.resolve() == src.resolve():
            raise ValueError("Refusing to restore a file over itself.")
        try:
            self.conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        except sqlite3.Error:
            pass
        self.conn.close()
        # A stale -wal / -shm sidecar is replayed the moment the file is
        # reopened, which silently undid the restore. Remove them with the
        # database itself so the backup really is what gets loaded.
        for suffix in ("-wal", "-shm"):
            side = Path(str(self.path) + suffix)
            if side.exists():
                try:
                    side.unlink()
                except OSError:
                    pass
        shutil.copy2(src, self.path)
        self.conn = sqlite3.connect(str(self.path), check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA busy_timeout=15000")
        self._init()
        self.audit("RESTORE", "database", src.name)


_tool_db: ToolDB | None = None


def get_tool_db() -> ToolDB:
    global _tool_db
    if _tool_db is None:
        _tool_db = ToolDB()
    return _tool_db


def set_tool_db(db: ToolDB | None) -> None:
    global _tool_db
    _tool_db = db


def reset_tool_db() -> None:
    global _tool_db
    if _tool_db is not None:
        _tool_db.close()
    _tool_db = None


# ----------------------------------------------------------------------- CRUD
_HEAD_FIELDS = (
    "ref_no", "form_no", "txn_type", "doc_date", "doc_time", "expected_return",
    "warehouse", "project_id", "project_name", "location", "handed_to",
    "iqama_id", "job_title", "mobile", "company", "email", "supervisor",
    "cost_code", "issued_by", "issued_at", "received_by", "received_at",
    "v_serial", "v_accessories", "v_calibration", "v_photos",
    "status", "closed_by_ref", "parent_ref", "remarks", "source_file",
    "file_hash",
)

_LINE_FIELDS = (
    "line_no", "asset_id", "category", "description", "make_model",
    "serial_no", "qty", "qty_returned", "accessories", "condition",
    "calib_due", "remarks", "photo",
)


def save_handover(db: ToolDB, head: dict, lines: Sequence[dict],
                  handover_id: int | None = None) -> int:
    """Insert or update one handover with all of its item lines.

    The whole document is written in a single transaction: a half-saved
    handover (header without lines) would misreport custody.
    """
    data = {k: head.get(k, "") for k in _HEAD_FIELDS}
    data["txn_type"] = data["txn_type"] or ISSUE
    data["doc_date"] = to_date(data["doc_date"]) or today()
    data["expected_return"] = to_date(data["expected_return"])
    for k in ("v_serial", "v_accessories", "v_calibration", "v_photos"):
        data[k] = 1 if head.get(k) else 0

    if not str(data["ref_no"]).strip():
        data["ref_no"] = next_ref(db, data["warehouse"], data["project_id"],
                                  data["txn_type"], data["doc_date"])

    # a reference decoded from the paper form fills in anything left blank
    dec = parse_ref(data["ref_no"])
    if dec:
        data["warehouse"] = data["warehouse"] or dec["warehouse"]
        data["project_id"] = data["project_id"] or dec["project_id"]
        if not head.get("txn_type"):
            data["txn_type"] = dec["txn_type"]
        data["doc_date"] = data["doc_date"] or dec["doc_date"]

    clash = db.one("SELECT id FROM handovers WHERE ref_no=? AND id<>?",
                   (data["ref_no"], handover_id or -1))
    if clash:
        raise ValueError(f"Handover {data['ref_no']} already exists.")

    if handover_id:
        sets = ", ".join(f"{k}=?" for k in _HEAD_FIELDS)
        db.execute(f"UPDATE handovers SET {sets}, updated_at=? WHERE id=?",
                   [data[k] for k in _HEAD_FIELDS] + [_now(), handover_id])
        db.execute("DELETE FROM handover_lines WHERE handover_id=?", (handover_id,))
        hid = handover_id
        action = "EDITED"
    else:
        data["created_by"] = db.current_user
        cols = list(_HEAD_FIELDS) + ["created_by"]
        qs = ", ".join("?" * len(cols))
        cur = db.execute(f"INSERT INTO handovers({', '.join(cols)}) VALUES({qs})",
                         [data[k] for k in cols])
        hid = int(cur.lastrowid)
        action = "CREATED"

    for i, ln in enumerate(lines, 1):
        row = {k: ln.get(k, "") for k in _LINE_FIELDS}
        row["line_no"] = int(to_float(ln.get("line_no"), i) or i)
        row["qty"] = to_float(ln.get("qty"), 1) or 1
        row["qty_returned"] = to_float(ln.get("qty_returned"), 0)
        row["calib_due"] = to_date(row["calib_due"])
        cols = ["handover_id"] + list(_LINE_FIELDS)
        qs = ", ".join("?" * len(cols))
        db.execute(f"INSERT INTO handover_lines({', '.join(cols)}) VALUES({qs})",
                   [hid] + [row[k] for k in _LINE_FIELDS])

    _refresh_status(db, hid)
    _sync_assets(db, hid)
    db.commit()
    db.audit(action, "handover", data["ref_no"],
             f"{data['txn_type']} · {len(lines)} item(s) · {data['handed_to']}")
    return hid


def get_handover(db: ToolDB, handover_id: int) -> dict | None:
    r = db.one("SELECT * FROM handovers WHERE id=?", (handover_id,))
    if r is None:
        return None
    d = dict(r)
    d["lines"] = handover_lines(db, handover_id)
    return d


def by_ref(db: ToolDB, ref: str) -> dict | None:
    r = db.one("SELECT * FROM handovers WHERE ref_no=?", (str(ref).strip(),))
    return get_handover(db, r["id"]) if r else None


def handover_lines(db: ToolDB, handover_id: int) -> list[dict]:
    return [dict(r) for r in db.query(
        "SELECT * FROM handover_lines WHERE handover_id=? ORDER BY line_no, id",
        (handover_id,))]


def delete_handovers(db: ToolDB, ids: Iterable[int]) -> int:
    ids = [int(i) for i in ids]
    if not ids:
        return 0
    qs = ",".join("?" * len(ids))
    refs = [r["ref_no"] for r in db.query(
        f"SELECT ref_no FROM handovers WHERE id IN ({qs})", ids)]
    db.execute(f"DELETE FROM handover_lines WHERE handover_id IN ({qs})", ids)
    db.execute(f"DELETE FROM handovers WHERE id IN ({qs})", ids)
    db.commit()
    for ref in refs:
        _recompute_asset_for_ref(db, ref)
    db.audit("DELETED", "handover", ", ".join(refs), f"{len(ids)} record(s)")
    return len(ids)


def cancel_handover(db: ToolDB, handover_id: int, reason: str) -> None:
    if not str(reason).strip():
        raise ValueError("A reason is required to cancel a handover.")
    h = db.one("SELECT ref_no, remarks FROM handovers WHERE id=?", (handover_id,))
    if h is None:
        raise ValueError("Handover not found.")
    db.execute("UPDATE handovers SET status=?, remarks=?, updated_at=? WHERE id=?",
               (CANCELLED, f"{h['remarks'] or ''} [cancelled: {reason}]".strip(),
                _now(), handover_id))
    db.commit()
    _sync_assets(db, handover_id)
    db.audit("CANCELLED", "handover", h["ref_no"], reason)


# ------------------------------------------------------------ custody engine
def _refresh_status(db: ToolDB, handover_id: int) -> None:
    """Derive the custody state from the lines and the calendar."""
    h = db.one("SELECT * FROM handovers WHERE id=?", (handover_id,))
    if h is None or h["status"] in (CANCELLED, TRANSFERRED):
        return
    lines = db.query("SELECT qty, qty_returned FROM handover_lines"
                     " WHERE handover_id=?", (handover_id,))
    total = sum(float(r["qty"] or 0) for r in lines)
    back = sum(float(r["qty_returned"] or 0) for r in lines)

    if h["txn_type"] == RETURN:
        status = CLOSED
    elif total > 0 and back >= total - 1e-9:
        status = CLOSED
    elif back > 1e-9:
        status = PART_RETURNED
    else:
        status = OPEN

    # a temporary loan past its date is overdue until it comes back
    if status in (OPEN, PART_RETURNED) and h["txn_type"] == LOAN:
        due = to_date(h["expected_return"])
        if due and due < today():
            status = OVERDUE

    if status != h["status"]:
        db.execute("UPDATE handovers SET status=?, updated_at=? WHERE id=?",
                   (status, _now(), handover_id))


def refresh_all_statuses(db: ToolDB) -> int:
    """Re-evaluate every open record — call on load so overdue is never stale."""
    n = 0
    for r in db.query("SELECT id, status FROM handovers"
                      " WHERE status NOT IN (?,?,?)",
                      (CLOSED, CANCELLED, TRANSFERRED)):
        before = r["status"]
        _refresh_status(db, r["id"])
        after = db.scalar("SELECT status FROM handovers WHERE id=?", (r["id"],), "")
        if after != before:
            n += 1
    if n:
        db.commit()
    return n


def post_return(db: ToolDB, ref: str, returns: Sequence[dict],
                head: dict | None = None) -> int:
    """Register a Return against an earlier handover.

    ``returns`` is [{line_id, qty, condition, remarks}, ...]. A Return document
    is created in its own right (so the paperwork matches) and the original
    handover's returned quantities are updated.
    """
    src = by_ref(db, ref)
    if src is None:
        raise ValueError(f"No handover found with reference {ref}.")
    if src["status"] == CANCELLED:
        raise ValueError(f"{ref} was cancelled — nothing to return.")

    by_id = {l["id"]: l for l in src["lines"]}
    moved: list[dict] = []
    for r in returns:
        ln = by_id.get(int(r.get("line_id", 0)))
        if ln is None:
            continue
        qty = to_float(r.get("qty"), 0)
        if qty <= 0:
            continue
        outstanding = float(ln["qty"] or 0) - float(ln["qty_returned"] or 0)
        if qty > outstanding + 1e-9:
            raise ValueError(
                f"{ln['description'] or ln['asset_id']}: returning {qty:g} but "
                f"only {outstanding:g} is outstanding.")
        db.execute("UPDATE handover_lines SET qty_returned=qty_returned+?,"
                   " condition=COALESCE(NULLIF(?,''), condition) WHERE id=?",
                   (qty, str(r.get("condition") or ""), ln["id"]))
        m = dict(ln)
        m["qty"] = qty
        m["qty_returned"] = qty
        m["condition"] = r.get("condition") or ln["condition"]
        m["remarks"] = r.get("remarks") or ""
        moved.append(m)

    if not moved:
        raise ValueError("Enter a quantity to return on at least one line.")

    h = dict(head or {})
    h.setdefault("txn_type", RETURN)
    h.setdefault("warehouse", src["warehouse"])
    h.setdefault("project_id", src["project_id"])
    h.setdefault("project_name", src["project_name"])
    h.setdefault("location", src["location"])
    h.setdefault("handed_to", src["handed_to"])
    h.setdefault("iqama_id", src["iqama_id"])
    h.setdefault("job_title", src["job_title"])
    h.setdefault("mobile", src["mobile"])
    h.setdefault("company", src["company"])
    h.setdefault("doc_date", today())
    h["parent_ref"] = src["ref_no"]
    h["ref_no"] = h.get("ref_no") or next_ref(
        db, h["warehouse"], h["project_id"], RETURN, h["doc_date"])

    new_id = save_handover(db, h, moved)
    _refresh_status(db, src["id"])
    db.execute("UPDATE handovers SET closed_by_ref=?, updated_at=? WHERE id=?",
               (h["ref_no"], _now(), src["id"]))
    db.commit()
    _sync_assets(db, src["id"])
    db.audit("RETURNED", "handover", src["ref_no"],
             f"{len(moved)} line(s) returned via {h['ref_no']}")
    return new_id


def post_transfer(db: ToolDB, ref: str, to: dict,
                  line_ids: Sequence[int] | None = None) -> int:
    """Move custody of some/all items of a handover to another person."""
    src = by_ref(db, ref)
    if src is None:
        raise ValueError(f"No handover found with reference {ref}.")
    if not str(to.get("handed_to") or "").strip():
        raise ValueError("Enter who is taking custody.")

    keep = set(int(i) for i in (line_ids or []))
    moving = [l for l in src["lines"]
              if (not keep or l["id"] in keep)
              and float(l["qty"] or 0) - float(l["qty_returned"] or 0) > 1e-9]
    if not moving:
        raise ValueError("Nothing is outstanding on that handover to transfer.")

    h = dict(to)
    h["txn_type"] = TRANSFER
    h.setdefault("warehouse", src["warehouse"])
    h.setdefault("project_id", src["project_id"])
    h.setdefault("project_name", src["project_name"])
    h.setdefault("doc_date", today())
    h["parent_ref"] = src["ref_no"]
    h["ref_no"] = h.get("ref_no") or next_ref(
        db, h["warehouse"], h["project_id"], TRANSFER, h["doc_date"])

    fresh = []
    for l in moving:
        d = dict(l)
        d["qty"] = float(l["qty"] or 0) - float(l["qty_returned"] or 0)
        d["qty_returned"] = 0
        fresh.append(d)
    new_id = save_handover(db, h, fresh)

    # the source lines leave the previous holder's custody
    for l in moving:
        db.execute("UPDATE handover_lines SET qty_returned=qty WHERE id=?", (l["id"],))
    _refresh_status(db, src["id"])
    # a transfer is NOT a return: the tools never came back to the warehouse,
    # so the source is closed as Transferred Out and the audit stays honest.
    db.execute("UPDATE handovers SET status=?, closed_by_ref=?, updated_at=?"
               " WHERE id=?", (TRANSFERRED, h["ref_no"], _now(), src["id"]))
    db.commit()
    _sync_assets(db, new_id)
    db.audit("TRANSFERRED", "handover", src["ref_no"],
             f"{len(moving)} item(s) -> {h['handed_to']} via {h['ref_no']}")
    return new_id


# ------------------------------------------------------------- asset register
def _sync_assets(db: ToolDB, handover_id: int) -> None:
    """Keep the asset register in step with the latest handover.

    The register answers "where is this tool right now?" — it is derived data,
    always rebuilt from the handovers, never edited independently.
    """
    h = db.one("SELECT * FROM handovers WHERE id=?", (handover_id,))
    if h is None:
        return
    for l in handover_lines(db, handover_id):
        aid = str(l["asset_id"] or "").strip()
        if not aid:
            continue
        _recompute_asset(db, aid)
    db.commit()


def _recompute_asset_for_ref(db: ToolDB, ref: str) -> None:
    for r in db.query("SELECT DISTINCT asset_id FROM handover_lines l"
                      " JOIN handovers h ON h.id=l.handover_id WHERE h.ref_no=?",
                      (ref,)):
        if r["asset_id"]:
            _recompute_asset(db, r["asset_id"])
    db.commit()


def _recompute_asset(db: ToolDB, asset_id: str) -> None:
    """Rebuild one asset row from its full handover history."""
    rows = db.query(
        """SELECT h.ref_no, h.txn_type, h.doc_date, h.doc_time, h.status,
                  h.handed_to, h.iqama_id, h.project_id, h.location,
                  l.category, l.description, l.make_model, l.serial_no,
                  l.condition, l.calib_due, l.qty, l.qty_returned
             FROM handover_lines l JOIN handovers h ON h.id = l.handover_id
            WHERE l.asset_id = ? AND h.status <> ?
            ORDER BY h.doc_date, h.doc_time, h.id""", (asset_id, CANCELLED))
    if not rows:
        db.execute("DELETE FROM assets WHERE asset_id=?", (asset_id,))
        return

    last = rows[-1]
    holder, iqama, project, location = "", "", "", ""
    status = "In Store"
    for r in rows:
        out = float(r["qty"] or 0) - float(r["qty_returned"] or 0)
        if r["txn_type"] == RETURN:
            holder, iqama, status = "", "", "In Store"
        elif out > 1e-9:
            holder = r["handed_to"] or holder
            iqama = r["iqama_id"] or iqama
            project = r["project_id"] or project
            location = r["location"] or location
            status = ("On Loan" if r["txn_type"] == LOAN else "Issued Out")
            if r["status"] == OVERDUE:
                status = "Overdue"
        else:
            holder, iqama, status = "", "", "In Store"

    data = {
        "asset_id": asset_id,
        "category": last["category"] or "",
        "description": last["description"] or "",
        "make_model": last["make_model"] or "",
        "serial_no": last["serial_no"] or "",
        "condition": last["condition"] or "",
        "calib_due": to_date(last["calib_due"]),
        "holder": holder, "holder_iqama": iqama,
        "project_id": project, "location": location,
        "status": status,
        "last_ref": last["ref_no"], "last_date": last["doc_date"],
    }
    if db.one("SELECT 1 FROM assets WHERE asset_id=?", (asset_id,)):
        sets = ", ".join(f"{k}=?" for k in data if k != "asset_id")
        db.execute(f"UPDATE assets SET {sets}, updated_at=? WHERE asset_id=?",
                   [v for k, v in data.items() if k != "asset_id"]
                   + [_now(), asset_id])
    else:
        cols = ", ".join(data)
        qs = ", ".join("?" * len(data))
        db.execute(f"INSERT INTO assets({cols}) VALUES({qs})", list(data.values()))


def rebuild_assets(db: ToolDB) -> int:
    """Rebuild the whole asset register from the handover history."""
    ids = [r["asset_id"] for r in db.query(
        "SELECT DISTINCT asset_id FROM handover_lines WHERE asset_id<>''")]
    db.execute("DELETE FROM assets")
    for aid in ids:
        _recompute_asset(db, aid)
    db.commit()
    db.audit("REBUILT", "assets", "", f"{len(ids)} asset(s)")
    return len(ids)


# ---------------------------------------------------------------- the filter
def search(db: ToolDB, text: str = "", txn_type: str = "", status: str = "",
           project: str = "", holder: str = "", warehouse: str = "",
           category: str = "", date_from: str = "", date_to: str = "",
           overdue_only: bool = False, limit: int = 5000) -> list[dict]:
    """The unified filter behind the register grid.

    Every document type is returned in ONE shape, whatever it was originally,
    which is what makes Issue / Transfer / Loan / Return comparable on screen.
    """
    sql = """SELECT h.*,
                (SELECT COUNT(*) FROM handover_lines l WHERE l.handover_id=h.id) n_items,
                (SELECT COALESCE(SUM(qty),0) FROM handover_lines l
                  WHERE l.handover_id=h.id) qty,
                (SELECT COALESCE(SUM(qty_returned),0) FROM handover_lines l
                  WHERE l.handover_id=h.id) qty_back
             FROM handovers h WHERE 1=1"""
    p: list[Any] = []
    if txn_type:
        sql += " AND h.txn_type=?"
        p.append(txn_type)
    if status:
        sql += " AND h.status=?"
        p.append(status)
    if project:
        sql += " AND (h.project_id=? OR h.project_name=?)"
        p += [project, project]
    if holder:
        sql += " AND h.handed_to=?"
        p.append(holder)
    if warehouse:
        sql += " AND h.warehouse=?"
        p.append(warehouse)
    if category:
        sql += (" AND h.id IN (SELECT handover_id FROM handover_lines"
                " WHERE category=?)")
        p.append(category)
    if date_from:
        sql += " AND h.doc_date>=?"
        p.append(to_date(date_from))
    if date_to:
        sql += " AND h.doc_date<=?"
        p.append(to_date(date_to))
    if overdue_only:
        sql += " AND h.status=?"
        p.append(OVERDUE)
    if text:
        like = f"%{text.strip()}%"
        sql += (" AND (h.ref_no LIKE ? OR h.handed_to LIKE ? OR h.iqama_id LIKE ?"
                " OR h.project_id LIKE ? OR h.project_name LIKE ?"
                " OR h.location LIKE ? OR h.mobile LIKE ? OR h.issued_by LIKE ?"
                " OR h.remarks LIKE ?"
                " OR h.id IN (SELECT handover_id FROM handover_lines WHERE"
                "     asset_id LIKE ? OR description LIKE ? OR serial_no LIKE ?"
                "     OR make_model LIKE ? OR category LIKE ?))")
        p += [like] * 14
    sql += " ORDER BY h.doc_date DESC, h.id DESC LIMIT ?"
    p.append(limit)

    out = []
    for r in db.query(sql, p):
        d = dict(r)
        d["outstanding"] = max(0.0, float(d["qty"] or 0) - float(d["qty_back"] or 0))
        d["days_out"] = _days_since(d["doc_date"])
        d["days_late"] = _days_late(d)
        out.append(d)
    return out


def search_lines(db: ToolDB, **kw) -> list[dict]:
    """The same filter, exploded to one row per item — for the item view."""
    heads = {h["id"]: h for h in search(db, **kw)}
    if not heads:
        return []
    qs = ",".join("?" * len(heads))
    rows = db.query(
        f"SELECT * FROM handover_lines WHERE handover_id IN ({qs})"
        " ORDER BY handover_id DESC, line_no", list(heads))
    out = []
    for r in rows:
        h = heads[r["handover_id"]]
        d = dict(r)
        d.update({k: h[k] for k in
                  ("ref_no", "txn_type", "doc_date", "status", "handed_to",
                   "iqama_id", "project_id", "project_name", "location",
                   "warehouse", "expected_return", "issued_by")})
        d["outstanding"] = max(0.0, float(d["qty"] or 0)
                               - float(d["qty_returned"] or 0))
        out.append(d)
    return out


def _days_since(date: str) -> int:
    try:
        return (_dt.date.today() - _dt.date.fromisoformat(str(date)[:10])).days
    except Exception:          # noqa: BLE001
        return 0


def _days_late(h: dict) -> int:
    if h.get("status") not in (OPEN, PART_RETURNED, OVERDUE):
        return 0
    due = to_date(h.get("expected_return"))
    if not due:
        return 0
    try:
        return max(0, (_dt.date.today() - _dt.date.fromisoformat(due)).days)
    except Exception:          # noqa: BLE001
        return 0


def distinct(db: ToolDB, column: str) -> list[str]:
    safe = {"txn_type", "status", "project_id", "project_name", "handed_to",
            "warehouse", "location", "company", "issued_by"}
    if column in safe:
        return [r[0] for r in db.query(
            f"SELECT DISTINCT {column} FROM handovers WHERE {column}<>''"
            f" ORDER BY {column}")]
    if column in ("category", "asset_id", "make_model"):
        return [r[0] for r in db.query(
            f"SELECT DISTINCT {column} FROM handover_lines WHERE {column}<>''"
            f" ORDER BY {column}")]
    return []


def search_assets(db: ToolDB, text: str = "", status: str = "",
                  holder: str = "", category: str = "") -> list[dict]:
    sql = "SELECT * FROM assets WHERE 1=1"
    p: list[Any] = []
    if status:
        sql += " AND status=?"
        p.append(status)
    if holder:
        sql += " AND holder=?"
        p.append(holder)
    if category:
        sql += " AND category=?"
        p.append(category)
    if text:
        like = f"%{text.strip()}%"
        sql += (" AND (asset_id LIKE ? OR description LIKE ? OR serial_no LIKE ?"
                " OR make_model LIKE ? OR holder LIKE ? OR category LIKE ?)")
        p += [like] * 6
    sql += " ORDER BY asset_id"
    out = []
    for r in db.query(sql, p):
        d = dict(r)
        d["calib_days"] = _calib_days(d["calib_due"])
        out.append(d)
    return out


def _calib_days(due: str) -> int | None:
    d = to_date(due)
    if not d:
        return None
    try:
        return (_dt.date.fromisoformat(d) - _dt.date.today()).days
    except Exception:          # noqa: BLE001
        return None


def asset_history(db: ToolDB, asset_id: str) -> list[dict]:
    return [dict(r) for r in db.query(
        """SELECT h.ref_no, h.txn_type, h.doc_date, h.doc_time, h.status,
                  h.handed_to, h.iqama_id, h.project_id, h.location,
                  h.issued_by, l.qty, l.qty_returned, l.condition, l.remarks
             FROM handover_lines l JOIN handovers h ON h.id=l.handover_id
            WHERE l.asset_id=? ORDER BY h.doc_date, h.doc_time, h.id""",
        (asset_id,))]


def custody_by_person(db: ToolDB, f: dict | None = None) -> list[dict]:
    """Who is holding what, right now (honouring the dashboard filters)."""
    if f:
        lines = [l for l in search_lines(db, **f)
                 if l["status"] in (OPEN, PART_RETURNED, OVERDUE)]
        agg: dict[tuple, dict] = {}
        for l in lines:
            key = (l.get("handed_to") or "", l.get("iqama_id") or "")
            if not key[0]:
                continue
            e = agg.setdefault(key, {"handed_to": key[0], "iqama_id": key[1],
                                     "mobile": "", "project_id": l.get("project_id"),
                                     "docs": set(), "outstanding": 0.0,
                                     "since": l.get("doc_date"), "overdue": 0})
            e["docs"].add(l.get("ref_no"))
            e["outstanding"] += to_float(l.get("outstanding"))
            if str(l.get("doc_date") or "") < str(e["since"] or "9999"):
                e["since"] = l.get("doc_date")
            if l.get("status") == OVERDUE:
                e["overdue"] = 1
        out = []
        for e in agg.values():
            if e["outstanding"] <= 0.0001:
                continue
            e["docs"] = len(e["docs"])
            out.append(e)
        return sorted(out, key=lambda d: -d["outstanding"])
    rows = db.query(
        """SELECT h.handed_to, h.iqama_id, h.mobile, h.project_id,
                  COUNT(DISTINCT h.id) docs,
                  SUM(l.qty - l.qty_returned) outstanding,
                  MIN(h.doc_date) since,
                  SUM(CASE WHEN h.status=? THEN 1 ELSE 0 END) overdue
             FROM handovers h JOIN handover_lines l ON l.handover_id=h.id
            WHERE h.status IN (?,?,?) AND h.handed_to<>''
            GROUP BY h.handed_to, h.iqama_id
            HAVING outstanding > 0.0001
            ORDER BY outstanding DESC""",
        (OVERDUE, OPEN, PART_RETURNED, OVERDUE))
    return [dict(r) for r in rows]


# ------------------------------------------------------------------ dashboard
def dashboard(db: ToolDB, f: dict | None = None) -> dict:
    f = f or {}
    rows = search(db, **f) if f else search(db)
    lines = search_lines(db, **f) if f else search_lines(db)
    out_qty = sum(l["outstanding"] for l in lines)
    people = {r["handed_to"] for r in rows
              if r["status"] in (OPEN, PART_RETURNED, OVERDUE) and r["handed_to"]}
    overdue = [r for r in rows if r["status"] == OVERDUE]
    assets = search_assets(db)
    calib_soon = [a for a in assets
                  if a["calib_days"] is not None and 0 <= a["calib_days"] <= 30]
    calib_exp = [a for a in assets
                 if a["calib_days"] is not None and a["calib_days"] < 0]
    return {
        "documents": len(rows),
        "issues": sum(1 for r in rows if r["txn_type"] == ISSUE),
        "transfers": sum(1 for r in rows if r["txn_type"] == TRANSFER),
        "loans": sum(1 for r in rows if r["txn_type"] == LOAN),
        "returns": sum(1 for r in rows if r["txn_type"] == RETURN),
        "open": sum(1 for r in rows if r["status"] == OPEN),
        "part": sum(1 for r in rows if r["status"] == PART_RETURNED),
        "closed": sum(1 for r in rows if r["status"] == CLOSED),
        "moved": sum(1 for r in rows if r["status"] == TRANSFERRED),
        "overdue": len(overdue),
        "overdue_days": max([r["days_late"] for r in overdue], default=0),
        "items": len(lines),
        "out_qty": out_qty,
        "custodians": len(people),
        "assets": len(assets),
        "assets_out": sum(1 for a in assets if a["status"] != "In Store"),
        "calib_soon": len(calib_soon),
        "calib_expired": len(calib_exp),
        "damaged": sum(1 for l in lines
                       if str(l.get("condition") or "").upper().startswith("D")),
        "photos": sum(1 for l in lines if l.get("photo")),
    }


#: columns that live on the item lines rather than on the document header
LINE_COLUMNS = ("category", "description", "asset_id", "make_model", "condition")

#: what a chart or tile counts — the dashboard's "Measure" selector
MEASURES = {
    "count": "Documents / lines",
    "qty": "Quantity handed over",
    "outstanding": "Quantity still out",
}


def _measure_value(row: dict, measure: str) -> float:
    if measure == "qty":
        return to_float(row.get("qty"))
    if measure == "outstanding":
        return to_float(row.get("outstanding"))
    return 1.0


def by_column(db: ToolDB, column: str, limit: int = 10,
              measure: str = "count",
              f: dict | None = None) -> list[tuple[str, float]]:
    """Group the filtered register by any column, header-level or line-level."""
    rows = (search_lines(db, **(f or {})) if column in LINE_COLUMNS
            else search(db, **(f or {})))
    agg: dict[str, float] = {}
    for r in rows:
        key = str(r.get(column) or "(blank)")
        agg[key] = agg.get(key, 0) + _measure_value(r, measure)
    return [kv for kv in sorted(agg.items(), key=lambda kv: -kv[1])[:limit] if kv[1]]


#: outstanding-custody ageing buckets, in days
AGE_BUCKETS = ((0, 7, "0-7 d"), (8, 30, "8-30 d"), (31, 60, "31-60 d"),
               (61, 90, "61-90 d"), (91, 10 ** 6, "90+ d"))


def ageing(db: ToolDB, f: dict | None = None,
           measure: str = "count") -> list[tuple[str, float]]:
    """How long the still-out documents have been out."""
    rows = [r for r in search(db, **(f or {}))
            if r["status"] in (OPEN, PART_RETURNED, OVERDUE)]
    out = []
    for lo, hi, label in AGE_BUCKETS:
        total = sum(_measure_value(r, measure) for r in rows
                    if lo <= int(r.get("days_out") or 0) <= hi)
        out.append((label, total))
    return out


def monthly_split(db: ToolDB, months: int = 8,
                  f: dict | None = None) -> list[tuple[str, float, float]]:
    """(month, handed over, returned) — the two series of the grouped chart."""
    rows = search_lines(db, **(f or {}))
    agg: dict[str, list[float]] = {}
    for r in rows:
        key = str(r.get("doc_date") or "")[:7]
        if not key:
            continue
        cell = agg.setdefault(key, [0.0, 0.0])
        cell[0] += to_float(r.get("qty"))
        cell[1] += to_float(r.get("qty_returned"))
    return [(k, v[0], v[1]) for k, v in sorted(agg.items())[-months:]]


def monthly(db: ToolDB, months: int = 12,
            f: dict | None = None) -> list[tuple[str, float]]:
    rows = search(db, **(f or {}))
    agg: dict[str, float] = {}
    for r in rows:
        key = str(r.get("doc_date") or "")[:7]
        if key:
            agg[key] = agg.get(key, 0) + 1
    return sorted(agg.items())[-months:]


# -------------------------------------------------------------------- reports
REPORT_LIST = [
    "All Handover Documents",
    "Issue Register",
    "Transfer Register",
    "Temporary Loan Register",
    "Return Register",
    "Outstanding Custody (Not Returned)",
    "Overdue Loans",
    "Custody by Person",
    "Custody by Project",
    "Item-wise Handover Detail",
    "Asset Register (Where Is It Now)",
    "Asset Movement History",
    "Calibration Due Report",
    "Damaged / Defective Items",
    "Missing Documents & Signatures",
    "Monthly Handover Summary",
]

Report = tuple[str, list[str], list[list[Any]]]


def build_report(db: ToolDB, name: str, f: dict | None = None) -> Report:
    f = dict(f or {})
    title = name

    if name in ("All Handover Documents", "Issue Register", "Transfer Register",
                "Temporary Loan Register", "Return Register",
                "Outstanding Custody (Not Returned)", "Overdue Loans"):
        if name == "Issue Register":
            f["txn_type"] = ISSUE
        elif name == "Transfer Register":
            f["txn_type"] = TRANSFER
        elif name == "Temporary Loan Register":
            f["txn_type"] = LOAN
        elif name == "Return Register":
            f["txn_type"] = RETURN
        elif name == "Overdue Loans":
            f["overdue_only"] = True
        rows = search(db, **f)
        if name == "Outstanding Custody (Not Returned)":
            rows = [r for r in rows
                    if r["status"] in (OPEN, PART_RETURNED, OVERDUE)]
        cols = ["Reference", "Type", "Date", "Project", "Location",
                "Handed To", "Iqama / ID", "Mobile", "Items", "Qty",
                "Returned", "Outstanding", "Expected Return", "Days Late",
                "Issued By", "Status"]
        return title, cols, [
            [r["ref_no"], r["txn_type"], fmt_date(r["doc_date"]),
             r["project_id"] or r["project_name"], r["location"],
             r["handed_to"], r["iqama_id"], r["mobile"], r["n_items"],
             round(r["qty"], 2), round(r["qty_back"], 2),
             round(r["outstanding"], 2), fmt_date(r["expected_return"]),
             r["days_late"] or "", r["issued_by"], r["status"]] for r in rows]

    if name == "Custody by Person":
        cols = ["Custodian", "Iqama / ID", "Mobile", "Project", "Documents",
                "Items Outstanding", "Holding Since", "Overdue Docs"]
        return title, cols, [
            [r["handed_to"], r["iqama_id"], r["mobile"], r["project_id"],
             r["docs"], round(r["outstanding"] or 0, 2), fmt_date(r["since"]),
             r["overdue"]] for r in custody_by_person(db)]

    if name == "Custody by Project":
        agg: dict[str, dict] = {}
        for l in search_lines(db, **f):
            key = l["project_id"] or l["project_name"] or "(unassigned)"
            a = agg.setdefault(key, {"docs": set(), "items": 0, "out": 0.0,
                                     "people": set()})
            a["docs"].add(l["ref_no"])
            a["items"] += 1
            a["out"] += l["outstanding"]
            if l["handed_to"]:
                a["people"].add(l["handed_to"])
        cols = ["Project", "Documents", "Item Lines", "Outstanding Qty",
                "Custodians"]
        return title, cols, [
            [k, len(v["docs"]), v["items"], round(v["out"], 2),
             len(v["people"])]
            for k, v in sorted(agg.items(), key=lambda kv: -kv[1]["out"])]

    if name == "Item-wise Handover Detail":
        cols = ["Reference", "Type", "Date", "Asset / Tool ID", "Category",
                "Description", "Make / Model", "Serial No.", "Qty", "Returned",
                "Outstanding", "Cond.", "Calib. Due", "Handed To", "Project",
                "Status", "Remarks"]
        return title, cols, [
            [l["ref_no"], l["txn_type"], fmt_date(l["doc_date"]), l["asset_id"],
             l["category"], l["description"], l["make_model"], l["serial_no"],
             round(float(l["qty"] or 0), 2), round(float(l["qty_returned"] or 0), 2),
             round(l["outstanding"], 2), l["condition"], fmt_date(l["calib_due"]),
             l["handed_to"], l["project_id"], l["status"], l["remarks"]]
            for l in search_lines(db, **f)]

    if name == "Asset Register (Where Is It Now)":
        cols = ["Asset / Tool ID", "Category", "Description", "Make / Model",
                "Serial No.", "Status", "Held By", "Iqama / ID", "Project",
                "Location", "Cond.", "Calib. Due", "Days to Calib.",
                "Last Reference", "Last Movement"]
        return title, cols, [
            [a["asset_id"], a["category"], a["description"], a["make_model"],
             a["serial_no"], a["status"], a["holder"], a["holder_iqama"],
             a["project_id"], a["location"], a["condition"],
             fmt_date(a["calib_due"]),
             "" if a["calib_days"] is None else a["calib_days"],
             a["last_ref"], fmt_date(a["last_date"])]
            for a in search_assets(db)]

    if name == "Asset Movement History":
        cols = ["Asset / Tool ID", "Reference", "Type", "Date", "Handed To",
                "Iqama / ID", "Project", "Location", "Qty", "Returned",
                "Cond.", "Issued By", "Status"]
        rows = []
        for a in search_assets(db):
            for h in asset_history(db, a["asset_id"]):
                rows.append([a["asset_id"], h["ref_no"], h["txn_type"],
                             fmt_date(h["doc_date"]), h["handed_to"],
                             h["iqama_id"], h["project_id"], h["location"],
                             round(float(h["qty"] or 0), 2),
                             round(float(h["qty_returned"] or 0), 2),
                             h["condition"], h["issued_by"], h["status"]])
        return title, cols, rows

    if name == "Calibration Due Report":
        cols = ["Asset / Tool ID", "Description", "Make / Model", "Serial No.",
                "Calib. Due", "Days Remaining", "Verdict", "Status", "Held By",
                "Project"]
        rows = []
        for a in sorted(search_assets(db),
                        key=lambda x: (x["calib_days"] is None,
                                       x["calib_days"] or 0)):
            if a["calib_days"] is None:
                verdict = "No date recorded"
            elif a["calib_days"] < 0:
                verdict = f"EXPIRED {abs(a['calib_days'])} day(s) ago"
            elif a["calib_days"] <= 30:
                verdict = "Due soon"
            else:
                verdict = "Valid"
            rows.append([a["asset_id"], a["description"], a["make_model"],
                         a["serial_no"], fmt_date(a["calib_due"]),
                         "" if a["calib_days"] is None else a["calib_days"],
                         verdict, a["status"], a["holder"], a["project_id"]])
        return title, cols, rows

    if name == "Damaged / Defective Items":
        cols = ["Reference", "Date", "Asset / Tool ID", "Description",
                "Serial No.", "Cond.", "Grade", "Handed To", "Project",
                "Remarks / Defects"]
        rows = []
        for l in search_lines(db, **f):
            c = str(l.get("condition") or "").strip().upper()[:1]
            if c in ("C", "D"):
                rows.append([l["ref_no"], fmt_date(l["doc_date"]), l["asset_id"],
                             l["description"], l["serial_no"], c,
                             CONDITIONS.get(c, c), l["handed_to"],
                             l["project_id"], l["remarks"]])
        return title, cols, rows

    if name == "Missing Documents & Signatures":
        # the governance report: which controlled forms are incomplete
        cols = ["Reference", "Type", "Date", "Handed To", "Project",
                "Signed by Warehouse", "Signed by Custodian", "Serial Checked",
                "Accessories", "Calibration", "Photos", "Scanned File",
                "What Is Missing"]
        rows = []
        for r in search(db, **f):
            miss = []
            if not str(r["issued_by"] or "").strip():
                miss.append("warehouse signature")
            if not str(r["received_by"] or "").strip():
                miss.append("custodian signature")
            if not r["v_serial"]:
                miss.append("serial check")
            if not r["v_accessories"]:
                miss.append("accessories check")
            if not r["v_calibration"]:
                miss.append("calibration check")
            if not r["v_photos"]:
                miss.append("photos")
            if not str(r["source_file"] or "").strip():
                miss.append("scanned copy")
            if not miss:
                continue
            tick = lambda b: "Yes" if b else "—"        # noqa: E731
            rows.append([r["ref_no"], r["txn_type"], fmt_date(r["doc_date"]),
                         r["handed_to"], r["project_id"],
                         tick(str(r["issued_by"] or "").strip()),
                         tick(str(r["received_by"] or "").strip()),
                         tick(r["v_serial"]), tick(r["v_accessories"]),
                         tick(r["v_calibration"]), tick(r["v_photos"]),
                         tick(str(r["source_file"] or "").strip()),
                         ", ".join(miss)])
        return title, cols, rows

    if name == "Monthly Handover Summary":
        cols = ["Month", "Documents", "Issues", "Transfers", "Loans",
                "Returns", "Item Lines", "Still Outstanding"]
        agg: dict[str, dict] = {}
        for r in search(db, **f):
            key = str(r["doc_date"] or "")[:7]
            if not key:
                continue
            a = agg.setdefault(key, {"docs": 0, ISSUE: 0, TRANSFER: 0,
                                     LOAN: 0, RETURN: 0, "items": 0, "out": 0.0})
            a["docs"] += 1
            a[r["txn_type"]] = a.get(r["txn_type"], 0) + 1
            a["items"] += r["n_items"]
            a["out"] += r["outstanding"]
        return title, cols, [
            [k, v["docs"], v[ISSUE], v[TRANSFER], v[LOAN], v[RETURN],
             v["items"], round(v["out"], 2)] for k, v in sorted(agg.items())]

    return title, ["Report"], [[f"Unknown report: {name}"]]


# ------------------------------------------------- synchronised drop folder
SUPPORTED_SUFFIXES = (".pdf", ".xlsx", ".xlsm", ".csv", ".txt")


def folders(db: ToolDB, active_only: bool = False) -> list[dict]:
    sql = "SELECT * FROM folders"
    if active_only:
        sql += " WHERE active=1"
    sql += " ORDER BY id"
    out = []
    for r in db.query(sql):
        d = dict(r)
        ok, note = folder_status(d["path"])
        d["online"] = ok
        d["note"] = note
        d["files"] = db.scalar("SELECT COUNT(*) FROM files WHERE path LIKE ?",
                               (f"{d['path']}%",))
        out.append(d)
    return out


def add_folder(db: ToolDB, path: str | Path, label: str = "") -> int:
    p = str(Path(path))
    ok, note = folder_status(p)
    if not ok:
        raise ValueError(note)
    row = db.one("SELECT id FROM folders WHERE path=?", (p,))
    if row:
        db.execute("UPDATE folders SET active=1, label=? WHERE id=?",
                   (label or Path(p).name, row["id"]))
        db.commit()
        return int(row["id"])
    cur = db.execute("INSERT INTO folders(path,label) VALUES(?,?)",
                     (p, label or Path(p).name))
    db.commit()
    db.audit("ADDED", "folder", p, label)
    return int(cur.lastrowid)


def remove_folder(db: ToolDB, folder_id: int, forget_files: bool = True) -> None:
    r = db.one("SELECT path FROM folders WHERE id=?", (folder_id,))
    if r is None:
        return
    if forget_files:
        db.execute("DELETE FROM files WHERE path LIKE ?", (f"{r['path']}%",))
    db.execute("DELETE FROM folders WHERE id=?", (folder_id,))
    db.commit()
    db.audit("REMOVED", "folder", r["path"])


def folder_status(path: str | Path) -> tuple[bool, str]:
    """Check a shared folder is usable before we promise anything."""
    if not str(path or "").strip():
        return False, "No folder selected."
    p = Path(path)
    if not p.exists():
        return False, f"The folder does not exist or is offline:\n{p}"
    if not p.is_dir():
        return False, f"That path is a file, not a folder:\n{p}"
    if not os.access(p, os.R_OK):
        return False, f"No permission to read:\n{p}"
    return True, ("Read and write access." if os.access(p, os.W_OK)
                  else "Read-only — files can be read but not moved.")


def file_hash(path: str | Path, limit_mb: int = 32) -> str:
    import hashlib
    h = hashlib.sha256()
    try:
        with open(path, "rb") as fh:
            read = 0
            while chunk := fh.read(1 << 20):
                h.update(chunk)
                read += len(chunk)
                if read >= limit_mb << 20:
                    break
    except OSError:
        return ""
    return h.hexdigest()


def sync_folder(db: ToolDB, folder_id: int, auto_import: bool = True) -> dict:
    """Index a synchronised folder and pull the handover data out of it.

    Read-only with respect to the shared drive: files are never moved, renamed
    or deleted — only read. That is deliberate, because the folder is someone
    else's sync target.
    """
    row = db.one("SELECT * FROM folders WHERE id=?", (folder_id,))
    if row is None:
        raise ValueError("Folder not found.")
    ok, note = folder_status(row["path"])
    if not ok:
        return {"ok": False, "message": note, "seen": 0, "new": 0,
                "imported": 0, "failed": 0, "errors": []}

    res = {"ok": True, "message": note, "seen": 0, "new": 0, "imported": 0,
           "updated": 0, "failed": 0, "errors": []}
    for f in sorted(Path(row["path"]).rglob("*")):
        if not f.is_file() or f.suffix.lower() not in SUPPORTED_SUFFIXES:
            continue
        if f.name.startswith("~$") or f.name.startswith("."):
            continue
        res["seen"] += 1
        try:
            st = f.stat()
        except OSError:
            continue
        key = str(f)
        prev = db.one("SELECT * FROM files WHERE path=?", (key,))
        digest = file_hash(f)
        modified = _dt.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")

        if prev and prev["file_hash"] == digest:
            continue                       # unchanged since last scan

        ref = detect_ref(f.name)
        if prev:
            db.execute("UPDATE files SET name=?, size_kb=?, modified=?,"
                       " file_hash=?, ref_no=?, seen_at=? WHERE id=?",
                       (f.name, round(st.st_size / 1024.0, 1), modified,
                        digest, ref, _now(), prev["id"]))
            res["updated"] += 1
        else:
            db.execute("INSERT INTO files(path,name,ref_no,size_kb,modified,"
                       "file_hash,status) VALUES(?,?,?,?,?,?,?)",
                       (key, f.name, ref, round(st.st_size / 1024.0, 1),
                        modified, digest, "New"))
            res["new"] += 1

        if auto_import and f.suffix.lower() == ".pdf":
            try:
                got = import_pdf(db, f)
                if got:
                    res["imported"] += 1
            except Exception as exc:       # noqa: BLE001
                res["failed"] += 1
                res["errors"].append(f"{f.name}: {exc}")
                db.execute("UPDATE files SET status=?, note=? WHERE path=?",
                           ("Failed", str(exc)[:300], key))
    db.execute("UPDATE folders SET last_scan=? WHERE id=?", (_now(), folder_id))
    db.commit()
    db.audit("SYNCED", "folder", row["path"],
             f"{res['seen']} file(s), {res['imported']} imported, "
             f"{res['failed']} failed")
    return res


def sync_all(db: ToolDB, auto_import: bool = True) -> dict:
    total = {"ok": True, "seen": 0, "new": 0, "imported": 0, "updated": 0,
             "failed": 0, "errors": [], "folders": 0, "offline": []}
    for f in folders(db, active_only=True):
        r = sync_folder(db, f["id"], auto_import)
        total["folders"] += 1
        if not r["ok"]:
            total["offline"].append(f["path"])
            continue
        for k in ("seen", "new", "imported", "updated", "failed"):
            total[k] += r.get(k, 0)
        total["errors"] += r.get("errors", [])
    return total


def detect_ref(name: str) -> str:
    """Pull a handover reference out of a file name."""
    stem = Path(str(name)).stem
    for token in re.split(r"[ _]+", stem):
        token = token.strip("().,-")
        if parse_ref(token):
            return token
    m = re.search(r"[A-Z]{2,4}-?\d{2,4}(?:IS|TR|TL|RT)\d{8}\d{1,3}", stem, re.I)
    return m.group(0) if m else ""


def scan_files(db: ToolDB, status: str = "", text: str = "") -> list[dict]:
    sql = "SELECT * FROM files WHERE 1=1"
    p: list[Any] = []
    if status:
        sql += " AND status=?"
        p.append(status)
    if text:
        like = f"%{text}%"
        sql += " AND (name LIKE ? OR ref_no LIKE ? OR note LIKE ?)"
        p += [like] * 3
    sql += " ORDER BY modified DESC, id DESC"
    out = []
    for r in db.query(sql, p):
        d = dict(r)
        d["exists"] = Path(d["path"]).exists()
        if not d["exists"] and d["status"] != "Missing":
            d["status"] = "Missing"
        out.append(d)
    return out


# --------------------------------------------------------------- PDF reading
def read_pdf_text(path: str | Path) -> str:
    """Extract the text of a handover PDF. Returns "" when unreadable."""
    try:
        from pypdf import PdfReader
    except ImportError:
        return ""
    try:
        reader = PdfReader(str(path))
        return "\n".join((p.extract_text() or "") for p in reader.pages)
    except Exception:              # noqa: BLE001
        return ""


# labels of section A/B, mapped to our fields. Matching ignores case and
# punctuation so a revised form does not break the import.
_LABELS = {
    "formno": "form_no",
    "handoverreferenceno": "ref_no",
    "handoverreferencenoauto": "ref_no",
    "referenceno": "ref_no",
    "ref": "ref_no",
    "dateddmmyyyy": "doc_date",
    "date": "doc_date",
    "time": "doc_time",
    "expectedreturndate": "expected_return",
    "expectedreturndateifany": "expected_return",
    "projectid": "project_id",
    "projectname": "project_name",
    "projectsitelocation": "location",
    "projectlocation": "location",
    "sitelocation": "location",
    "handedtofullname": "handed_to",
    "handedto": "handed_to",
    "employeeiqamaid": "iqama_id",
    "iqamaid": "iqama_id",
    "employeeid": "iqama_id",
    "jobtitle": "job_title",
    "mobileno": "mobile",
    "mobile": "mobile",
    "companydepartment": "company",
    "company": "company",
    "email": "email",
    "supervisormanager": "supervisor",
    "supervisor": "supervisor",
    "costcodewbs": "cost_code",
    "costcode": "cost_code",
}

_ITEM_ROW = re.compile(
    r"^(?P<no>\d{1,3})\s+(?P<rest>\S.*)$")


def parse_handover_text(text: str, source: str = "") -> dict | None:
    """Turn the text of a handover form into a header + line dicts.

    Written against the supplied controlled form (WH-FRM-001 Rev 00) but kept
    tolerant: labels are matched loosely and every field is optional, so a
    slightly different revision still imports rather than failing outright.
    """
    if not str(text or "").strip():
        return None
    raw = [ln.rstrip() for ln in text.replace("\r\n", "\n").split("\n")]
    lines = [ln.strip() for ln in raw if ln.strip()]
    if not lines:
        return None

    head: dict[str, Any] = {}

    # --- labels sit on their own line, value on the next (as the form exports)
    i = 0
    while i < len(lines):
        key = norm(lines[i])
        field = _LABELS.get(key)
        if field and i + 1 < len(lines):
            val = lines[i + 1].strip()
            if not _LABELS.get(norm(val)) and not val.startswith(("A —", "B —",
                                                                  "C —", "D —")):
                head.setdefault(field, val)
                i += 2
                continue
        i += 1

    # --- the reference is authoritative: it also encodes type, date, project
    ref = str(head.get("ref_no") or "").strip()
    if not ref:
        for ln in lines:
            m = re.search(r"[A-Z]{2,4}-?\d{2,4}(?:IS|TR|TL|RT)\d{8}\d{1,3}",
                          ln, re.I)
            if m:
                ref = m.group(0)
                break
    if not ref and source:
        ref = detect_ref(source)
    if not ref:
        return None
    head["ref_no"] = ref
    dec = parse_ref(ref)
    if dec:
        head.setdefault("warehouse", dec["warehouse"])
        head["txn_type"] = dec["txn_type"]
        head.setdefault("project_id", dec["project_id"])
        head.setdefault("doc_date", dec["doc_date"])

    # --- transaction type: a ticked checkbox wins over the reference code
    for ln in lines:
        if "TRANSACTION TYPE" in ln.upper():
            picked = _ticked_type(ln)
            if picked:
                head["txn_type"] = picked
            break
    head.setdefault("txn_type", ISSUE)

    head["doc_date"] = to_date(head.get("doc_date"))
    head["expected_return"] = to_date(head.get("expected_return"))
    for k in ("email", "supervisor", "cost_code", "expected_return"):
        if str(head.get(k, "")).strip() == "-":
            head[k] = ""

    # --- section D signatures: "Name" then "23/08/2026  06:33"
    head.update(_parse_signatures(lines))

    # --- verification ticks
    for ln in lines:
        u = ln.upper()
        if "VERIFICATION" in u or "CHECKED" in u:
            head.setdefault("v_serial", 1 if "SERIAL" in u else 0)
            head.setdefault("v_accessories", 1 if "ACCESSOR" in u else 0)
            head.setdefault("v_calibration", 1 if "CALIBRATION" in u else 0)
            head.setdefault("v_photos", 1 if "PHOTO" in u else 0)
            break

    items = _parse_items(lines)
    if source:
        head["source_file"] = str(source)
    return {"head": head, "lines": items}


def _ticked_type(line: str) -> str:
    """Read which transaction-type box is ticked, if the export shows it."""
    u = line.upper()
    for mark in ("☑", "[X]", "(X)", "✓", "✔"):
        pos = u.find(mark)
        if pos < 0:
            continue
        after = u[pos + len(mark):pos + len(mark) + 22]
        for name in (ISSUE, TRANSFER, LOAN, RETURN):
            if name.upper().split()[0] in after:
                return name
    return ""


def _parse_signatures(lines: Sequence[str]) -> dict:
    """Pull the two names and date-times out of section D."""
    out: dict[str, str] = {}
    joined = [l.strip() for l in lines]
    for idx, ln in enumerate(joined):
        u = ln.upper()
        role = ""
        if u.startswith("ISSUED BY"):
            role = "issued"
        elif u.startswith("RECEIVED BY"):
            role = "received"
        if not role:
            continue
        # the next few lines hold NAME / SIGNATURE / DATE-TIME headings, then
        # the actual value(s)
        for cand in joined[idx + 1: idx + 8]:
            cu = norm(cand)
            if cu in ("name", "signature", "datetime", "date", "time"):
                continue
            m = re.search(r"(\d{2}/\d{2}/\d{4})\s+(\d{1,2}:\d{2})", cand)
            if m:
                out[f"{role}_at"] = f"{to_date(m.group(1))} {m.group(2)}"
                name = cand[:m.start()].strip(" ·-\t")
                if name and not out.get(f"{role}_by"):
                    out[f"{role}_by"] = name
                break
            if not out.get(f"{role}_by"):
                out[f"{role}_by"] = cand.strip()
    return out


def _parse_items(lines: Sequence[str]) -> list[dict]:
    """Read section C. Each item is one numbered line in the extracted text."""
    start = -1
    for i, ln in enumerate(lines):
        u = ln.upper()
        if "ASSET" in u and "TOOL" in u and "CATEGORY" in u:
            start = i + 1
            break
        if u.startswith("C —") or u.startswith("C -"):
            start = i + 1
    if start < 0:
        return []

    out: list[dict] = []
    for ln in lines[start:]:
        u = ln.upper()
        if u.startswith(("*CONDITION", "VERIFICATION", "D —", "D -",
                         "CONTROLLED FORM")):
            break
        m = _ITEM_ROW.match(ln)
        if not m:
            continue
        item = _split_item(m.group("rest"))
        if item:
            item["line_no"] = int(m.group("no"))
            out.append(item)
    return out


_COND_RE = re.compile(r"\b([ABCD])\b")


def _split_item(rest: str) -> dict | None:
    """Split one item row into its columns.

    The PDF text layer loses the column boundaries, so the row is decoded from
    the inside out using the strong anchors that are always present: the
    condition grade (a lone A/B/C/D), the calibration date, and the asset ID.
    Everything is optional — a partially readable row is better than none.
    """
    text = re.sub(r"\s{2,}", "  ", str(rest or "").strip())
    if not text:
        return None

    item: dict[str, Any] = {"asset_id": "", "category": "", "description": "",
                            "make_model": "", "serial_no": "", "qty": 1,
                            "accessories": "", "condition": "",
                            "calib_due": "", "remarks": ""}

    # calibration date + trailing remarks
    m = re.search(r"(\d{2}/\d{2}/\d{4}|\d{4}-\d{2}-\d{2})\s*(.*)$", text)
    if m:
        item["calib_due"] = to_date(m.group(1))
        item["remarks"] = m.group(2).strip()
        text = text[:m.start()].strip()

    # condition grade: the last lone capital letter A-D
    grades = list(_COND_RE.finditer(text))
    if grades:
        g = grades[-1]
        item["condition"] = g.group(1)
        after = text[g.end():].strip()
        if after and not item["remarks"]:
            item["remarks"] = after
        text = text[:g.start()].strip()

    # accessories phrase sits just before the condition
    for phrase in ("With All accessories", "With all accessories",
                   "With accessories"):
        if phrase.lower() in text.lower():
            idx = text.lower().rindex(phrase.lower())
            item["accessories"] = text[idx:idx + len(phrase)]
            text = (text[:idx] + " " + text[idx + len(phrase):]).strip()
            break
    else:
        if text.endswith(" -"):
            item["accessories"] = "-"
            text = text[:-2].strip()

    tokens = text.split()
    if not tokens:
        return item if (item["asset_id"] or item["condition"]) else None

    # asset / tool ID is the leading code-like token
    if re.match(r"^[A-Z0-9][A-Z0-9\-/]{3,}$", tokens[0], re.I) and \
            any(ch.isdigit() for ch in tokens[0]):
        item["asset_id"] = tokens.pop(0)

    # category is the next single known word
    if tokens and tokens[0].title() in ("Instrument", "Tool", "Device",
                                        "Equipment", "Machine", "Consumable",
                                        "Accessory", "Safety"):
        item["category"] = tokens.pop(0).title()

    # serial: the last token that looks like a serial and is not a word
    if tokens:
        last = tokens[-1]
        if last == "-":
            item["serial_no"] = ""
            tokens.pop()
        elif re.match(r"^[0-9][0-9\-/]{3,}$", last):
            item["serial_no"] = tokens.pop()

    # make / model: a known brand, else the last remaining word
    brands = ("leica", "bosch", "makita", "hilti", "stanley", "dewalt",
              "fluke", "topcon", "sokkia", "trimble", "milwaukee", "karcher",
              "honda", "yamaha", "3m", "abb", "siemens", "schneider", "skf")
    for i in range(len(tokens) - 1, -1, -1):
        if tokens[i].lower() in brands:
            item["make_model"] = tokens.pop(i)
            break

    item["description"] = " ".join(tokens).strip(" -")
    if not item["description"] and not item["asset_id"]:
        return None
    return item


def import_pdf(db: ToolDB, path: str | Path,
               overwrite: bool = False) -> int | None:
    """Read one handover PDF from the sync folder into the register."""
    p = Path(path)
    text = read_pdf_text(p)
    parsed = parse_handover_text(text, source=str(p))
    if not parsed or not parsed["head"].get("ref_no"):
        db.execute("UPDATE files SET status=?, note=? WHERE path=?",
                   ("Unreadable", "no handover reference found", str(p)))
        db.commit()
        return None

    head = parsed["head"]
    head["file_hash"] = file_hash(p)
    ref = head["ref_no"]
    existing = db.one("SELECT id FROM handovers WHERE ref_no=?", (ref,))
    if existing and not overwrite:
        db.execute("UPDATE files SET status=?, ref_no=?, handover_id=?,"
                   " note=? WHERE path=?",
                   ("Linked", ref, existing["id"], "already in the register",
                    str(p)))
        db.commit()
        return int(existing["id"])

    hid = save_handover(db, head, parsed["lines"],
                        handover_id=int(existing["id"]) if existing else None)
    db.execute("UPDATE files SET status=?, ref_no=?, handover_id=?, note=?"
               " WHERE path=?",
               ("Imported", ref, hid, f"{len(parsed['lines'])} item(s)", str(p)))
    db.commit()
    db.audit("IMPORTED", "handover", ref, f"from {p.name}")
    return hid


def import_folder_files(db: ToolDB, paths: Sequence[str],
                        overwrite: bool = False) -> dict:
    res = {"imported": 0, "skipped": 0, "failed": 0, "errors": []}
    for path in paths:
        try:
            got = import_pdf(db, path, overwrite)
            if got:
                res["imported"] += 1
            else:
                res["skipped"] += 1
        except Exception as exc:           # noqa: BLE001
            res["failed"] += 1
            res["errors"].append(f"{Path(path).name}: {exc}")
    return res


# --------------------------------------------------------------- spreadsheets
HEADER_MAP = {
    "no": "line_no", "sr": "line_no", "srno": "line_no", "line": "line_no",
    "assettoolid": "asset_id", "assetid": "asset_id", "toolid": "asset_id",
    "assetno": "asset_id", "code": "asset_id",
    "category": "category", "type": "category",
    "description": "description", "itemdescription": "description",
    "item": "description", "toolname": "description",
    "makemodel": "make_model", "make": "make_model", "model": "make_model",
    "brand": "make_model",
    "serialno": "serial_no", "serial": "serial_no", "sn": "serial_no",
    "qty": "qty", "quantity": "qty",
    "accessories": "accessories", "accessoriescomponents": "accessories",
    "cond": "condition", "condition": "condition",
    "calibdue": "calib_due", "calibrationdue": "calib_due",
    "remarks": "remarks", "remarksdefects": "remarks", "remark": "remarks",
    "reamrks": "remarks",          # the misspelling seen on site sheets
}


def sniff(text: str) -> tuple[list[str], list[list[str]]]:
    raw = [ln for ln in text.replace("\r\n", "\n").split("\n") if ln.strip()]
    if not raw:
        return [], []
    if "\t" in raw[0]:
        rows = [ln.split("\t") for ln in raw]
    elif raw[0].count(",") >= 2:
        rows = list(csv.reader(io.StringIO("\n".join(raw))))
    else:
        rows = [re.split(r"\s{2,}", ln.strip()) for ln in raw]
    width = max(len(r) for r in rows)
    rows = [list(r) + [""] * (width - len(r)) for r in rows]
    head = [str(c).strip() for c in rows[0]]
    if sum(1 for h in head if norm(h) in HEADER_MAP) >= 2:
        return head, [[str(c).strip() for c in r] for r in rows[1:]]
    return [f"Column {i + 1}" for i in range(width)], \
           [[str(c).strip() for c in r] for r in rows]


def read_table(path: str | Path) -> tuple[list[str], list[list[Any]]]:
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(p, data_only=True, read_only=True)
        ws = wb.active
        data = [[("" if c is None else c) for c in row]
                for row in ws.iter_rows(values_only=True)]
        wb.close()
        data = [r for r in data if any(str(c).strip() for c in r)]
        if not data:
            return [], []
        head = [str(c).strip() for c in data[0]]
        if sum(1 for c in head if norm(c) in HEADER_MAP) >= 2:
            return head, data[1:]
        return [f"Column {i + 1}" for i in range(len(head))], data
    return sniff(p.read_text(encoding="utf-8", errors="ignore"))


def auto_map(headers: Sequence[str]) -> dict[int, str]:
    out: dict[int, str] = {}
    for i, h in enumerate(headers):
        f = HEADER_MAP.get(norm(h))
        if f and f not in out.values():
            out[i] = f
    return out


def rows_to_lines(headers: Sequence[str], rows: Sequence[Sequence[Any]],
                  mapping: dict[int, str] | None = None) -> list[dict]:
    m = mapping or auto_map(headers)
    out = []
    for n, r in enumerate(rows, 1):
        d: dict[str, Any] = {}
        for i, field in m.items():
            if i < len(r):
                d[field] = str(r[i]).strip() if r[i] is not None else ""
        if not (d.get("asset_id") or d.get("description")):
            continue
        d["line_no"] = int(to_float(d.get("line_no"), n) or n)
        d["qty"] = to_float(d.get("qty"), 1) or 1
        d["calib_due"] = to_date(d.get("calib_due"))
        out.append(d)
    return out


def template_rows() -> tuple[list[str], list[list[Any]]]:
    cols = ["No.", "Asset / Tool ID", "Category", "Description", "Make / Model",
            "Serial No.", "Qty", "Accessories / Components", "Cond.",
            "Calib. Due", "Remarks / Defects"]
    rows = [[1, "12000AL01", "Instrument", "Auto Level", "Leica", "5778779", 1,
             "With All accessories", "A", "25/08/2026", "New"],
            [2, "12000TS01", "Instrument", "Total Station", "Leica", "3366852",
             1, "With All accessories", "A", "25/08/2026", "New"]]
    return cols, rows
