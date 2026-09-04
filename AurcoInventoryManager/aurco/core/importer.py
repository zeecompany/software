"""Excel/CSV import engine with column mapping (Item Master / Opening Stock / Transactions)."""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from .database import Database, today
from . import services as S

ITEM_TARGETS = {
    "code": "Item Code", "description": "Description", "short_desc": "Short Description",
    "category": "Category", "subcategory": "Subcategory", "uom": "UOM", "brand": "Brand",
    "model": "Model / Part No", "specification": "Specification", "barcode": "Barcode",
    "alt_code": "Alternate Code", "min_level": "Minimum Level", "max_level": "Maximum Level",
    "reorder_level": "Reorder Level", "critical_level": "Critical Level",
    "opening_balance": "Opening Balance", "unit_cost": "Unit Cost", "warehouse": "Warehouse",
    "location": "Location", "rack": "Rack / Bin", "remarks": "Remarks",
}
OPENING_TARGETS = {"code": "Item Code", "qty": "Quantity", "unit_cost": "Unit Cost",
                   "warehouse": "Warehouse", "location": "Location", "remarks": "Remarks"}
TXN_TARGETS = {"code": "Item Code", "txn_type": "Type (IN/OUT)", "qty": "Quantity",
               "date": "Date", "doc_no": "Document No", "party": "Supplier / Issued To",
               "unit_cost": "Unit Cost", "remarks": "Remarks"}

MODES = {"Item Master": ITEM_TARGETS, "Opening Stock": OPENING_TARGETS,
         "Stock Transactions": TXN_TARGETS}

NUMERIC = {"min_level", "max_level", "reorder_level", "critical_level", "opening_balance",
           "unit_cost", "qty"}


def read_table(path: str | Path, sheet: str | None = None,
               limit: int | None = None) -> tuple[list[str], list[list[Any]]]:
    """Return (headers, rows) from .xlsx/.xls/.csv."""
    p = Path(path)
    if p.suffix.lower() in (".csv", ".txt"):
        with open(p, newline="", encoding="utf-8-sig", errors="replace") as fh:
            rd = list(csv.reader(fh))
        if not rd:
            return [], []
        head = [str(h).strip() for h in rd[0]]
        rows = [r + [""] * (len(head) - len(r)) for r in rd[1:]]
        return head, rows[:limit] if limit else rows
    from openpyxl import load_workbook
    wb = load_workbook(p, data_only=True, read_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    it = ws.iter_rows(values_only=True)
    try:
        head_raw = next(it)
    except StopIteration:
        return [], []
    head = [str(h).strip() if h is not None else f"Column{i+1}" for i, h in enumerate(head_raw)]
    rows = []
    for i, r in enumerate(it):
        if limit and i >= limit:
            break
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        rows.append(["" if c is None else c for c in list(r) + [None] * (len(head) - len(r))])
    wb.close()
    return head, rows


def sheet_names(path: str | Path) -> list[str]:
    p = Path(path)
    if p.suffix.lower() in (".csv", ".txt"):
        return ["CSV"]
    from openpyxl import load_workbook
    wb = load_workbook(p, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def auto_map(headers: list[str], targets: dict[str, str]) -> dict[str, int]:
    """Best-effort guess: match on normalised target label / field name."""
    def norm(s: str) -> str:
        return "".join(ch for ch in str(s).lower() if ch.isalnum())

    hnorm = {norm(h): i for i, h in enumerate(headers)}
    mapping: dict[str, int] = {}
    for field, label in targets.items():
        for cand in (label, field, label.replace(" ", ""), field.replace("_", " ")):
            i = hnorm.get(norm(cand))
            if i is not None:
                mapping[field] = i
                break
        else:
            for hn, i in hnorm.items():
                if norm(field) and (norm(field) in hn or hn in norm(label)):
                    mapping[field] = i
                    break
    return mapping


def _num(v: Any) -> float:
    if v in (None, ""):
        return 0.0
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return 0.0


def run_import(db: Database, mode: str, headers: list[str], rows: list[list[Any]],
               mapping: dict[str, int], update_existing: bool = True) -> dict:
    """Execute the import. Returns {'created','updated','skipped','errors':[...]}."""
    res = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    def val(r, field):
        i = mapping.get(field)
        if i is None or i < 0 or i >= len(r):
            return ""
        v = r[i]
        return "" if v is None else v

    for n, r in enumerate(rows, 2):
        try:
            if mode == "Item Master":
                desc = str(val(r, "description")).strip()
                code = str(val(r, "code")).strip()
                if not desc and not code:
                    res["skipped"] += 1
                    continue
                data = {}
                for f in ITEM_TARGETS:
                    if f in mapping:
                        v = val(r, f)
                        data[f] = _num(v) if f in NUMERIC else str(v).strip()
                data.setdefault("uom", db.get_setting("default_uom", "PCS"))
                if not data.get("description"):
                    data["description"] = code
                existing = db.one("SELECT id FROM items WHERE code=?", (code,)) if code else None
                if existing:
                    if not update_existing:
                        res["skipped"] += 1
                        continue
                    data.pop("opening_balance", None)
                    S.save_item(db, data, existing["id"])
                    res["updated"] += 1
                else:
                    S.save_item(db, data)
                    res["created"] += 1
                _seed_lookup(db, data)

            elif mode == "Opening Stock":
                code = str(val(r, "code")).strip()
                qty = _num(val(r, "qty"))
                it = db.one("SELECT id, balance FROM items WHERE code=?", (code,))
                if it is None:
                    res["errors"].append(f"Row {n}: item '{code}' not found")
                    continue
                delta = qty - float(it["balance"] or 0)
                if abs(delta) < 1e-9:
                    res["skipped"] += 1
                    continue
                h = S.DocHeader(doc_type="ADJ", doc_date=today(),
                                reason="Opening balance import",
                                warehouse=str(val(r, "warehouse")),
                                remarks=str(val(r, "remarks")))
                S.post_adjustment(db, h, [S.Line(item_id=it["id"], qty=delta,
                                                 remarks=f"Imported opening {qty:g}")])
                res["updated"] += 1

            else:  # Stock Transactions
                code = str(val(r, "code")).strip()
                it = db.one("SELECT id FROM items WHERE code=?", (code,))
                if it is None:
                    res["errors"].append(f"Row {n}: item '{code}' not found")
                    continue
                qty = _num(val(r, "qty"))
                ttype = str(val(r, "txn_type")).strip().upper()
                date = str(val(r, "date") or today())[:10]
                line = S.Line(item_id=it["id"], qty=abs(qty), unit_cost=_num(val(r, "unit_cost")),
                              remarks=str(val(r, "remarks")))
                if ttype.startswith("IN") or ttype in ("RECEIPT", "GRN", "RECEIVE"):
                    h = S.DocHeader(doc_type="GRN", doc_date=date,
                                    supplier=str(val(r, "party")),
                                    reference=str(val(r, "doc_no")))
                    S.post_receipt(db, h, [line])
                else:
                    h = S.DocHeader(doc_type="DN", doc_date=date,
                                    issued_to=str(val(r, "party")),
                                    reference=str(val(r, "doc_no")))
                    S.post_issue(db, h, [line])
                res["created"] += 1
        except Exception as exc:  # noqa: BLE001 - collect and continue
            res["errors"].append(f"Row {n}: {exc}")
    db.commit()
    db.audit("IMPORTED", mode, "",
             f"created {res['created']}, updated {res['updated']}, errors {len(res['errors'])}")
    return res


def _seed_lookup(db: Database, data: dict) -> None:
    for key, table in (("category", "categories"), ("uom", "uoms"), ("warehouse", "warehouses")):
        v = (data.get(key) or "").strip()
        if v:
            db.execute(f"INSERT OR IGNORE INTO {table}(name) VALUES(?)", (v,))


def write_template(mode: str, out_path: str | Path) -> Path:
    """Create a ready-to-fill Excel template for the selected import mode."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    targets = MODES[mode]
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = mode[:31]
    ws.append(list(targets.values()))
    for i, c in enumerate(ws[1], 1):
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0B3D6B")
        ws.column_dimensions[get_column_letter(i)].width = 20
    if mode == "Item Master":
        ws.append(["ITM-00001", "Sample bolt M12 x 60mm", "Bolt M12", "Fasteners", "Bolts",
                   "PCS", "AURCO", "M12-60", "Grade 8.8", "8901234567890", "ALT-001",
                   40, 200, 60, 20, 150, 3.5, "Main Warehouse", "Rack A", "A-01-02", ""])
    wb.save(out)
    return out
