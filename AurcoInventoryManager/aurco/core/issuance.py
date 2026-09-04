"""COMPANY ISSUANCE REGISTER — material issued to other companies.

A third fully separate register (after the Admin Station), built around the
sheet the user keeps by hand:

    Date · Company Name · MR (if any) · Recipient · Iqama ID · Item issued ·
    Qty · Date of issuance · Date of Return · Evidence · Remarks

What this module adds over the spreadsheet:

  ·  **Photo proof is a first-class record.** Every issue can carry several
     images; files are COPIED into the register's own Evidence folder so the
     record survives even if the original photo is deleted or the phone is
     wiped. Missing proof is reported, never silently ignored.
  ·  **Temporary vs Permanent.** A temporary issue has an expected return date
     and becomes Overdue on its own; a permanent issue is never chased.
  ·  **Partial returns.** 3 of 5 back is a real state, with its own evidence.
  ·  **Custody by person.** Recipient + Iqama, so you can answer "what is still
     with Muhammad Shoaib?" instantly.

Separation is physical, exactly like the Admin Station:
    <storage>/Company Issuance/company_issuance.db   ← its own SQLite file
    <storage>/Company Issuance/Evidence/             ← its own photo store
No table, foreign key or import touches the inventory database, and nothing
here ever writes a stock ledger row.
"""
from __future__ import annotations

import csv
import datetime as _dt
import io
import os
import re
import shutil
import sqlite3
import stat
from pathlib import Path
from typing import Any, Iterable, Sequence

from . import config

FOLDER = "Company Issuance"
EVIDENCE_DIR = "Evidence"
DB_NAME = "company_issuance.db"
SCHEMA_VERSION = 1

IMAGE_SUFFIXES = (".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tif", ".tiff", ".webp")
DOC_SUFFIXES = (".pdf",)
EVIDENCE_SUFFIXES = IMAGE_SUFFIXES + DOC_SUFFIXES

DDL = """
PRAGMA journal_mode=WAL;
PRAGMA foreign_keys=ON;

CREATE TABLE IF NOT EXISTS settings (
    key   TEXT PRIMARY KEY,
    value TEXT
);

CREATE TABLE IF NOT EXISTS issues (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    issue_no       TEXT UNIQUE NOT NULL,
    record_date    TEXT DEFAULT '',      -- the "Date" column (entry date)
    company        TEXT DEFAULT '',
    mr_no          TEXT DEFAULT '',
    recipient      TEXT DEFAULT '',
    iqama          TEXT DEFAULT '',
    phone          TEXT DEFAULT '',
    item           TEXT DEFAULT '',
    item_code      TEXT DEFAULT '',
    uom            TEXT DEFAULT '',
    qty            REAL NOT NULL DEFAULT 0,
    issue_type     TEXT NOT NULL DEFAULT 'Temporary',   -- Temporary | Permanent
    issue_date     TEXT DEFAULT '',
    expected_return TEXT DEFAULT '',
    return_date    TEXT DEFAULT '',
    qty_returned   REAL NOT NULL DEFAULT 0,
    condition_out  TEXT DEFAULT '',
    condition_in   TEXT DEFAULT '',
    dn_no          TEXT DEFAULT '',      -- gate pass / DN reference
    project        TEXT DEFAULT '',
    location       TEXT DEFAULT '',
    unit_value     REAL NOT NULL DEFAULT 0,
    status         TEXT NOT NULL DEFAULT 'Issued',
    remarks        TEXT DEFAULT '',
    issued_by      TEXT DEFAULT '',
    received_back_by TEXT DEFAULT '',
    batch_id       INTEGER,
    created_by     TEXT DEFAULT '',
    created_at     TEXT DEFAULT (datetime('now','localtime')),
    updated_at     TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS ix_iss_company ON issues(company);
CREATE INDEX IF NOT EXISTS ix_iss_recip   ON issues(recipient);
CREATE INDEX IF NOT EXISTS ix_iss_status  ON issues(status);
CREATE INDEX IF NOT EXISTS ix_iss_date    ON issues(issue_date);
CREATE INDEX IF NOT EXISTS ix_iss_item    ON issues(item);

CREATE TABLE IF NOT EXISTS evidence (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    issue_id   INTEGER NOT NULL REFERENCES issues(id) ON DELETE CASCADE,
    kind       TEXT NOT NULL DEFAULT 'ISSUE',    -- ISSUE | RETURN
    file_path  TEXT NOT NULL,
    caption    TEXT DEFAULT '',
    added_by   TEXT DEFAULT '',
    added_at   TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS ix_ev_issue ON evidence(issue_id, kind);

CREATE TABLE IF NOT EXISTS batches (
    id       INTEGER PRIMARY KEY AUTOINCREMENT,
    ts       TEXT DEFAULT (datetime('now','localtime')),
    source   TEXT DEFAULT '',
    rows     INTEGER DEFAULT 0,
    skipped  INTEGER DEFAULT 0,
    username TEXT DEFAULT ''
);

CREATE TABLE IF NOT EXISTS audit (
    id        INTEGER PRIMARY KEY AUTOINCREMENT,
    ts        TEXT DEFAULT (datetime('now','localtime')),
    username  TEXT DEFAULT '',
    action    TEXT NOT NULL,
    entity    TEXT DEFAULT '',
    entity_id TEXT DEFAULT '',
    details   TEXT DEFAULT ''
);
CREATE INDEX IF NOT EXISTS ix_iss_audit ON audit(ts);
"""

# canonical field -> printed label, in the order of the user's sheet
FIELDS: list[tuple[str, str]] = [
    ("record_date", "Date"),
    ("company", "Company Name"),
    ("mr_no", "MR (If any)"),
    ("recipient", "Recipient"),
    ("iqama", "Iqama ID"),
    ("item", "Item Issued"),
    ("qty", "Qty"),
    ("issue_date", "Date of Issuance"),
    ("return_date", "Date of Return"),
    ("remarks", "Remarks"),
]
EXTRA_FIELDS: list[tuple[str, str]] = [
    ("issue_no", "Issue No"),
    ("item_code", "Item Code"),
    ("uom", "UOM"),
    ("issue_type", "Type"),
    ("expected_return", "Expected Return"),
    ("qty_returned", "Qty Returned"),
    ("condition_out", "Condition Out"),
    ("condition_in", "Condition In"),
    ("dn_no", "DN / Gate Pass"),
    ("project", "Project"),
    ("location", "Location"),
    ("phone", "Phone"),
    ("unit_value", "Unit Value"),
    ("issued_by", "Issued By"),
    ("received_back_by", "Received Back By"),
    ("status", "Status"),
]
ALL_FIELDS = FIELDS + EXTRA_FIELDS
LABELS = dict(ALL_FIELDS)
NUMERIC = {"qty", "qty_returned", "unit_value"}
DATE_FIELDS = {"record_date", "issue_date", "return_date", "expected_return"}

TEMPORARY = "Temporary"
PERMANENT = "Permanent"
ISSUE_TYPES = [TEMPORARY, PERMANENT]

ST_ISSUED = "Not Returned Yet"
ST_RETURNED = "Returned"
ST_PARTIAL = "Partially Returned"
ST_OVERDUE = "Overdue"
ST_PERMANENT = "Permanent (No Return)"
ST_LOST = "Lost / Written Off"
STATUSES = [ST_ISSUED, ST_PARTIAL, ST_RETURNED, ST_OVERDUE, ST_PERMANENT, ST_LOST]

STATUS_COLORS = {
    ST_ISSUED: "#e0a300",
    ST_PARTIAL: "#e8590c",
    ST_RETURNED: "#1a9c52",
    ST_OVERDUE: "#c92a2a",
    ST_PERMANENT: "#7048e8",
    ST_LOST: "#868e96",
}

CONDITIONS = ["", "New", "Good", "Used", "Needs Repair", "Damaged", "Scrap"]

# header aliases (case / punctuation insensitive) taken from the real sheet
HEADER_MAP = {
    "date": "record_date", "entrydate": "record_date", "recorddate": "record_date",
    "companyname": "company", "company": "company", "contractor": "company",
    "issuedto": "company", "party": "company", "vendor": "company",
    "mrifany": "mr_no", "mr": "mr_no", "mrno": "mr_no", "mrnumber": "mr_no",
    "materialrequest": "mr_no", "prno": "mr_no", "pr": "mr_no",
    "receipient": "recipient", "recipient": "recipient", "receiver": "recipient",
    "receivedby": "recipient", "person": "recipient", "name": "recipient",
    "iqamaid": "iqama", "iqama": "iqama", "id": "iqama", "idno": "iqama",
    "iqamano": "iqama", "iqamanumber": "iqama", "nationalid": "iqama",
    "itemissued": "item", "item": "item", "itemdescription": "item",
    "description": "item", "material": "item", "tool": "item", "equipment": "item",
    "itemcode": "item_code", "code": "item_code",
    "qty": "qty", "quantity": "qty", "nos": "qty",
    "uom": "uom", "unit": "uom",
    "dateofissuance": "issue_date", "issuedate": "issue_date",
    "dateofissue": "issue_date", "issuedon": "issue_date", "outdate": "issue_date",
    "dateofreturn": "return_date", "returndate": "return_date",
    "returnedon": "return_date", "indate": "return_date",
    "expectedreturn": "expected_return", "duedate": "expected_return",
    "expectedreturndate": "expected_return", "returndue": "expected_return",
    "qtyreturned": "qty_returned", "returnedqty": "qty_returned",
    "returnqty": "qty_returned", "qtyback": "qty_returned",
    "evidence": "_evidence", "proof": "_evidence", "photo": "_evidence",
    "picture": "_evidence", "attachment": "_evidence", "image": "_evidence",
    "remarks": "remarks", "remark": "remarks", "reamrks": "remarks",
    "status": "remarks", "note": "remarks", "notes": "remarks",
    "comment": "remarks", "comments": "remarks",
    "type": "issue_type", "issuetype": "issue_type", "basis": "issue_type",
    "dn": "dn_no", "dnno": "dn_no", "gatepass": "dn_no", "deliverynote": "dn_no",
    "project": "project", "site": "project",
    "location": "location", "place": "location",
    "phone": "phone", "mobile": "phone", "contact": "phone",
    "conditionout": "condition_out", "condition": "condition_out",
    "conditionin": "condition_in",
    "value": "unit_value", "unitvalue": "unit_value", "cost": "unit_value",
    "unitcost": "unit_value", "price": "unit_value",
    "issuedby": "issued_by", "storekeeper": "issued_by",
    "receivedbackby": "received_back_by",
}


def _now() -> str:
    return _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def norm(s: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def db_path() -> Path:
    return config.folder(FOLDER) / DB_NAME


def evidence_root() -> Path:
    p = config.folder(FOLDER) / EVIDENCE_DIR
    p.mkdir(parents=True, exist_ok=True)
    return p


# ------------------------------------------------------------ value parsing
def to_float(v: Any) -> float:
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    t = str(v).strip().lower()
    if t in ("-", "nil", "none", "n/a", "na"):
        return 0.0
    try:
        return float(re.sub(r"[^\d.\-]", "", t) or 0)
    except ValueError:
        return 0.0


_DATE_PATTERNS = ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y",
                  "%Y/%m/%d", "%d-%b-%Y", "%d %b %Y", "%b %d, %Y",
                  "%d-%b-%y", "%d-%m-%y", "%d/%m/%y", "%Y-%m-%d %H:%M:%S")


def to_date(v: Any) -> str:
    """Anything date-like -> ISO yyyy-mm-dd. Handles the sheet's 21-Dec-25."""
    if v in (None, ""):
        return ""
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    t = str(v).strip()
    if not t or t in ("-", "N/A", "n/a"):
        return ""
    for f in _DATE_PATTERNS:
        try:
            d = _dt.datetime.strptime(t, f).date()
            # two-digit years: 25 -> 2025, never 1925
            if d.year < 100:
                d = d.replace(year=d.year + 2000)
            return d.isoformat()
        except ValueError:
            continue
    try:
        n = float(t)
        if 20000 < n < 60000:            # Excel serial
            return (_dt.date(1899, 12, 30) + _dt.timedelta(days=int(n))).isoformat()
    except ValueError:
        pass
    return t


def _days_between(a: str, b: str) -> int | None:
    try:
        return (_dt.date.fromisoformat(b) - _dt.date.fromisoformat(a)).days
    except (ValueError, TypeError):
        return None


# ------------------------------------------------------------------ database
class IssuanceDB:
    """Standalone database for the Company Issuance Register."""

    def __init__(self, path: str | Path | None = None):
        self.path = Path(path or db_path())
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(str(self.path), check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA busy_timeout=15000")
        self.conn.execute("PRAGMA foreign_keys=ON")
        self.current_user = "admin"
        self._init()

    def _init(self) -> None:
        self.conn.executescript(DDL)
        self.conn.commit()
        self.set_setting("schema_version", str(SCHEMA_VERSION))

    # ----------------------------------------------------------- primitives
    def execute(self, sql: str, params: Sequence = ()) -> sqlite3.Cursor:
        return self.conn.execute(sql, params)

    def query(self, sql: str, params: Sequence = ()) -> list[sqlite3.Row]:
        return list(self.conn.execute(sql, params).fetchall())

    def one(self, sql: str, params: Sequence = ()):
        return self.conn.execute(sql, params).fetchone()

    def scalar(self, sql: str, params: Sequence = (), default=0):
        r = self.conn.execute(sql, params).fetchone()
        return default if r is None or r[0] is None else r[0]

    def commit(self) -> None:
        self.conn.commit()

    def close(self) -> None:
        try:
            self.conn.commit()
            self.conn.close()
        except Exception:
            pass

    # ------------------------------------------------------------- settings
    def get_setting(self, key: str, default: Any = None) -> Any:
        r = self.one("SELECT value FROM settings WHERE key=?", (key,))
        return r["value"] if r else default

    def get_bool(self, key: str, default: bool = False) -> bool:
        v = self.get_setting(key)
        return default if v is None else str(v) in ("1", "True", "true", "yes")

    def set_setting(self, key: str, value: Any) -> None:
        self.execute("INSERT INTO settings(key,value) VALUES(?,?) "
                     "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                     (key, str(value)))
        self.conn.commit()

    def audit(self, action: str, entity: str = "", entity_id: str = "",
              details: str = "") -> None:
        self.execute("INSERT INTO audit(ts,username,action,entity,entity_id,details)"
                     " VALUES(?,?,?,?,?,?)",
                     (_now(), self.current_user, action, entity, str(entity_id), details))
        self.conn.commit()

    # -------------------------------------------------------------- backups
    def backup(self, dest_folder: str | Path | None = None, note: str = "") -> Path:
        dest = Path(dest_folder or config.folder(FOLDER) / "Backups")
        dest.mkdir(parents=True, exist_ok=True)
        out = dest / f"company_issuance_{_dt.datetime.now():%Y%m%d_%H%M%S}.db"
        # two backups inside the same second must not collide -- the safety
        # copy taken by restore() would otherwise overwrite the very file it
        # is about to restore from, silently wiping it.
        n = 2
        while out.exists():
            out = dest / (f"company_issuance_{_dt.datetime.now():%Y%m%d_%H%M%S}"
                          f"_{n}.db")
            n += 1
        self.conn.commit()
        tgt = sqlite3.connect(str(out))
        with tgt:
            self.conn.backup(tgt)
        tgt.close()
        self.audit("BACKUP", "database", out.name, note)
        return out

    def restore(self, src: str | Path) -> None:
        src = Path(src)
        if not src.exists():
            raise FileNotFoundError(src)
        safety = self.backup(note=f"safety copy before restoring {src.name}")
        if safety.resolve() == src.resolve():
            raise ValueError("Refusing to restore a file over itself.")
        self.conn.close()
        shutil.copy2(src, self.path)
        self.conn = sqlite3.connect(str(self.path), check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._init()
        self.audit("RESTORE", "database", src.name)


_iss_db: IssuanceDB | None = None


def get_issuance_db() -> IssuanceDB:
    global _iss_db
    if _iss_db is None:
        _iss_db = IssuanceDB()
    return _iss_db


def set_issuance_db(db: IssuanceDB | None) -> None:
    global _iss_db
    _iss_db = db


def reset_issuance_db() -> None:
    global _iss_db
    if _iss_db is not None:
        _iss_db.close()
    _iss_db = None


# --------------------------------------------------------------- numbering
def next_issue_no(db: IssuanceDB) -> str:
    year = _dt.date.today().strftime("%Y")
    prefix = db.get_setting("issue_prefix", "ISS")
    pad = int(db.get_setting("issue_pad", 5) or 5)
    like = f"{prefix}-{year}-%"
    mx = 0
    for r in db.query("SELECT issue_no FROM issues WHERE issue_no LIKE ?", (like,)):
        tail = str(r["issue_no"]).rsplit("-", 1)[-1]
        if tail.isdigit():
            mx = max(mx, int(tail))
    return f"{prefix}-{year}-{str(mx + 1).zfill(pad)}"


# ------------------------------------------------------------ status engine
def compute_status(rec: dict, today: str | None = None) -> str:
    """Derive the status from the quantities, dates and issue type.

    A manually set Lost / Written Off is respected; everything else is always
    recalculated so the register can never drift out of step with its numbers.
    """
    if str(rec.get("status", "")) == ST_LOST:
        return ST_LOST
    today = today or _dt.date.today().isoformat()
    qty = to_float(rec.get("qty"))
    back = to_float(rec.get("qty_returned"))
    if str(rec.get("issue_type", TEMPORARY)) == PERMANENT:
        return ST_PERMANENT
    if qty > 0 and back >= qty - 1e-9:
        return ST_RETURNED
    if back > 0:
        return ST_PARTIAL
    due = str(rec.get("expected_return") or "")
    if due and len(due) == 10 and due < today:
        return ST_OVERDUE
    return ST_ISSUED


def outstanding_qty(rec: dict) -> float:
    if str(rec.get("issue_type", TEMPORARY)) == PERMANENT:
        return 0.0
    return max(0.0, to_float(rec.get("qty")) - to_float(rec.get("qty_returned")))


def days_out(rec: dict, today: str | None = None) -> int | None:
    """How long the material has been (or was) out."""
    start = str(rec.get("issue_date") or rec.get("record_date") or "")
    end = str(rec.get("return_date") or "") or (today or _dt.date.today().isoformat())
    if len(start) != 10:
        return None
    return _days_between(start, end)


def days_overdue(rec: dict, today: str | None = None) -> int:
    if compute_status(rec, today) != ST_OVERDUE:
        return 0
    today = today or _dt.date.today().isoformat()
    d = _days_between(str(rec.get("expected_return") or ""), today)
    return max(0, d or 0)


# ------------------------------------------------------------------ evidence
def store_evidence(db: IssuanceDB, issue_id: int, src: str | Path,
                   kind: str = "ISSUE", caption: str = "") -> Path:
    """Copy a proof file into the register's own Evidence folder.

    Copying rather than linking is deliberate: the original photo usually lives
    on a phone or a temp folder and would be gone within days.
    """
    src = Path(src)
    if not src.exists():
        raise FileNotFoundError(f"Evidence file not found: {src}")
    no = db.scalar("SELECT issue_no FROM issues WHERE id=?", (issue_id,), default="") \
        or f"ID{issue_id}"
    safe = re.sub(r"[^A-Za-z0-9._-]", "_", str(no))
    folder = evidence_root() / safe
    folder.mkdir(parents=True, exist_ok=True)
    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = folder / f"{kind.lower()}_{stamp}_{re.sub(r'[^A-Za-z0-9._-]', '_', src.name)}"
    n = 1
    while dest.exists():
        dest = dest.with_name(f"{dest.stem}_{n}{dest.suffix}")
        n += 1
    shutil.copy2(src, dest)
    db.execute("INSERT INTO evidence(issue_id,kind,file_path,caption,added_by,added_at)"
               " VALUES(?,?,?,?,?,?)",
               (issue_id, kind.upper(), str(dest), caption, db.current_user, _now()))
    db.commit()
    db.audit("ATTACHED", "evidence", no, f"{kind}: {dest.name}")
    return dest


def evidence_for(db: IssuanceDB, issue_id: int, kind: str = "") -> list[dict]:
    sql = "SELECT * FROM evidence WHERE issue_id=?"
    p: list[Any] = [issue_id]
    if kind:
        sql += " AND kind=?"
        p.append(kind.upper())
    sql += " ORDER BY id"
    return [dict(r) for r in db.query(sql, p)]



def _safe_remove(path: str | Path, reason: str = "") -> bool:
    """Never hard-delete a proof file.

    Evidence is the whole point of this register, so a 'delete' archives the
    picture instead. Returns True only when the file really went.

    FAIL-SAFE: if the protection layer cannot be reached for any reason (no
    inventory database open, a stale handle, an import problem) we archive the
    file ourselves rather than falling back to unlink. Losing a photograph
    because of an unrelated error is never the right outcome.
    """
    p = Path(path)
    if not p.exists():
        return False
    try:
        from .database import get_db
        from . import protection as P
        return P.guarded_unlink(get_db(), p, reason)
    except Exception:
        pass
    # last resort: move it aside ourselves, still without deleting
    try:
        archive = p.parent / "_Archive"
        archive.mkdir(parents=True, exist_ok=True)
        stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = archive / f"{p.stem}_{stamp}{p.suffix}"
        n = 1
        while dest.exists():
            dest = archive / f"{p.stem}_{stamp}_{n}{p.suffix}"
            n += 1
        try:
            os.chmod(p, stat.S_IWRITE | stat.S_IREAD)
        except OSError:
            pass
        shutil.move(str(p), str(dest))
    except OSError:
        pass
    return False


def delete_evidence(db: IssuanceDB, ev_id: int, remove_file: bool = False) -> None:
    row = db.one("SELECT * FROM evidence WHERE id=?", (ev_id,))
    if row is None:
        return
    if remove_file:
        _safe_remove(row["file_path"], "evidence removed from record")
    db.execute("DELETE FROM evidence WHERE id=?", (ev_id,))
    db.commit()
    db.audit("DELETED", "evidence", row["issue_id"], Path(row["file_path"]).name)


def evidence_counts(db: IssuanceDB, issue_id: int) -> tuple[int, int]:
    """(issue proofs, return proofs) that still exist on disk."""
    out = ret = 0
    for e in evidence_for(db, issue_id):
        if not Path(e["file_path"]).exists():
            continue
        if e["kind"] == "RETURN":
            ret += 1
        else:
            out += 1
    return out, ret


def missing_evidence(db: IssuanceDB) -> list[dict]:
    """Issues with no usable proof — the whole point of the register."""
    rows = search(db)
    bad = []
    for r in rows:
        n_out, _ = evidence_counts(db, r["id"])
        if n_out == 0 and not str(r.get("dn_no") or "").strip():
            bad.append(r)
    return bad


# ---------------------------------------------------------------------- CRUD
def save_issue(db: IssuanceDB, data: dict, issue_id: int | None = None,
               evidence_files: Sequence[str] | None = None,
               require_evidence: bool | None = None) -> int:
    """Create or update one issuance line.

    `require_evidence` defaults to the register's setting. When on, a NEW issue
    must arrive with at least one photo or a DN / gate-pass reference.
    """
    fields = [f for f, _ in ALL_FIELDS]
    rec = {f: data.get(f, "") for f in fields}
    for f in NUMERIC:
        rec[f] = to_float(rec.get(f))
    for f in DATE_FIELDS:
        rec[f] = to_date(rec.get(f))
    rec["issue_type"] = rec.get("issue_type") or TEMPORARY
    if rec["issue_type"] not in ISSUE_TYPES:
        rec["issue_type"] = TEMPORARY
    if not rec["issue_date"]:
        rec["issue_date"] = rec["record_date"] or _dt.date.today().isoformat()
    if not rec["record_date"]:
        rec["record_date"] = rec["issue_date"]
    # a return date implies the goods came back
    if rec["return_date"] and rec["qty_returned"] <= 0:
        rec["qty_returned"] = rec["qty"]
    if rec["qty_returned"] > rec["qty"] > 0:
        rec["qty_returned"] = rec["qty"]
    rec["status"] = compute_status(dict(rec, status=data.get("status", "")))

    if require_evidence is None:
        require_evidence = db.get_bool("require_evidence", True)
    if require_evidence and issue_id is None:
        has_ref = bool(str(rec.get("dn_no") or "").strip())
        if not evidence_files and not has_ref:
            raise ValueError(
                "Photo proof is required for every issuance.\n\n"
                "Attach at least one picture, or record a DN / gate-pass number "
                "as the reference.")

    if issue_id:
        sets = ", ".join(f"{f}=?" for f in fields)
        db.execute(f"UPDATE issues SET {sets}, updated_at=? WHERE id=?",
                   [rec[f] for f in fields] + [_now(), issue_id])
        db.commit()
        db.audit("EDITED", "issue", rec.get("issue_no") or issue_id, rec.get("item", ""))
    else:
        rec["issue_no"] = rec.get("issue_no") or next_issue_no(db)
        cols = ", ".join(fields + ["created_by", "created_at", "updated_at", "batch_id"])
        marks = ", ".join("?" * (len(fields) + 4))
        cur = db.execute(f"INSERT INTO issues({cols}) VALUES({marks})",
                         [rec[f] for f in fields] +
                         [db.current_user, _now(), _now(), data.get("batch_id")])
        issue_id = int(cur.lastrowid)
        db.commit()
        db.audit("CREATED", "issue", rec["issue_no"],
                 f"{rec['item']} x{rec['qty']:g} to {rec['company']}")

    for f in (evidence_files or []):
        try:
            store_evidence(db, issue_id, f, "ISSUE")
        except (OSError, FileNotFoundError):
            continue
    return int(issue_id)


def record_return(db: IssuanceDB, issue_id: int, qty: float,
                  return_date: str = "", condition: str = "",
                  received_by: str = "", remarks: str = "",
                  evidence_files: Sequence[str] | None = None) -> dict:
    """Book a (possibly partial) return against an issue."""
    row = get_issue(db, issue_id)
    if row is None:
        raise ValueError("That issuance record no longer exists.")
    qty = to_float(qty)
    if qty <= 0:
        raise ValueError("Returned quantity must be greater than zero.")
    already = to_float(row["qty_returned"])
    total = to_float(row["qty"])
    if already + qty > total + 1e-9:
        raise ValueError(
            f"Only {total - already:g} of {total:g} is still outstanding — "
            f"cannot return {qty:g}.")
    new_back = already + qty
    rdate = to_date(return_date) or _dt.date.today().isoformat()
    note = (str(row["remarks"] or "").strip() + " " + (remarks or "").strip()).strip()
    db.execute(
        "UPDATE issues SET qty_returned=?, return_date=?, condition_in=?,"
        " received_back_by=?, remarks=?, updated_at=? WHERE id=?",
        (new_back, rdate, condition or row["condition_in"],
         received_by or row["received_back_by"], note, _now(), issue_id))
    db.commit()
    fresh = get_issue(db, issue_id) or {}
    db.execute("UPDATE issues SET status=? WHERE id=?",
               (compute_status(fresh), issue_id))
    db.commit()
    for f in (evidence_files or []):
        try:
            store_evidence(db, issue_id, f, "RETURN")
        except (OSError, FileNotFoundError):
            continue
    db.audit("RETURNED", "issue", row["issue_no"],
             f"{qty:g} of {total:g} returned on {rdate}")
    return get_issue(db, issue_id) or {}


def mark_lost(db: IssuanceDB, issue_id: int, reason: str = "") -> None:
    row = get_issue(db, issue_id)
    if row is None:
        return
    note = (str(row["remarks"] or "") + f" [written off: {reason}]").strip()
    db.execute("UPDATE issues SET status=?, remarks=?, updated_at=? WHERE id=?",
               (ST_LOST, note, _now(), issue_id))
    db.commit()
    db.audit("EDITED", "issue", row["issue_no"], f"written off: {reason}")


def reopen(db: IssuanceDB, issue_id: int) -> None:
    """Undo a write-off / recompute the status from the numbers."""
    row = get_issue(db, issue_id)
    if row is None:
        return
    fresh = dict(row)
    fresh["status"] = ""
    db.execute("UPDATE issues SET status=?, updated_at=? WHERE id=?",
               (compute_status(fresh), _now(), issue_id))
    db.commit()
    db.audit("EDITED", "issue", row["issue_no"], "reopened")


def refresh_statuses(db: IssuanceDB) -> int:
    """Recompute every status — an issue goes Overdue simply because time passed."""
    n = 0
    for r in db.query("SELECT * FROM issues"):
        rec = dict(r)
        new = compute_status(rec)
        if new != rec["status"]:
            db.execute("UPDATE issues SET status=? WHERE id=?", (new, rec["id"]))
            n += 1
    if n:
        db.commit()
    return n


def get_issue(db: IssuanceDB, issue_id: int) -> dict | None:
    r = db.one("SELECT * FROM issues WHERE id=?", (issue_id,))
    return dict(r) if r else None


def delete_issues(db: IssuanceDB, ids: Iterable[int],
                  remove_files: bool = False) -> int:
    ids = list(ids)
    if not ids:
        return 0
    if remove_files:
        for i in ids:
            for e in evidence_for(db, i):
                _safe_remove(e["file_path"], "issuance record deleted")
    marks = ",".join("?" * len(ids))
    db.execute(f"DELETE FROM evidence WHERE issue_id IN ({marks})", ids)
    db.execute(f"DELETE FROM issues WHERE id IN ({marks})", ids)
    db.commit()
    db.audit("DELETED", "issue", ",".join(str(i) for i in ids), f"{len(ids)} record(s)")
    return len(ids)


def search(db: IssuanceDB, text: str = "", company: str = "", recipient: str = "",
           status: str = "", issue_type: str = "", item: str = "",
           date_from: str = "", date_to: str = "", only_open: bool = False,
           only_overdue: bool = False, missing_proof: bool = False) -> list[dict]:
    sql = "SELECT * FROM issues WHERE 1=1"
    p: list[Any] = []
    if text.strip():
        like = f"%{text.strip()}%"
        sql += (" AND (item LIKE ? OR company LIKE ? OR recipient LIKE ?"
                " OR iqama LIKE ? OR mr_no LIKE ? OR remarks LIKE ?"
                " OR issue_no LIKE ? OR dn_no LIKE ? OR item_code LIKE ?"
                " OR project LIKE ?)")
        p += [like] * 10
    for col, val in (("company", company), ("recipient", recipient),
                     ("status", status), ("issue_type", issue_type)):
        if val:
            sql += f" AND {col}=?"
            p.append(val)
    if item:
        sql += " AND item=?"
        p.append(item)
    if date_from:
        sql += " AND issue_date>=?"
        p.append(date_from)
    if date_to:
        sql += " AND issue_date<=?"
        p.append(date_to)
    if only_open:
        sql += (" AND issue_type=? AND qty_returned < qty")
        p.append(TEMPORARY)
    if only_overdue:
        sql += " AND status=?"
        p.append(ST_OVERDUE)
    sql += " ORDER BY date(issue_date) DESC, id DESC"
    rows = [dict(r) for r in db.query(sql, p)]
    if missing_proof:
        keep = []
        for r in rows:
            n_out, _ = evidence_counts(db, r["id"])
            if n_out == 0 and not str(r.get("dn_no") or "").strip():
                keep.append(r)
        rows = keep
    return rows


def distinct(db: IssuanceDB, column: str) -> list[str]:
    if column not in {f for f, _ in ALL_FIELDS}:
        return []
    return [r[0] for r in db.query(
        f"SELECT DISTINCT {column} FROM issues WHERE COALESCE({column},'')<>''"
        f" ORDER BY {column}")]


# ------------------------------------------------------------------ analytics
def _where(f: dict | None) -> tuple[str, list]:
    f = f or {}
    sql, p = "", []
    if f.get("text"):
        like = f"%{str(f['text']).strip()}%"
        sql += (" AND (item LIKE ? OR company LIKE ? OR recipient LIKE ?"
                " OR iqama LIKE ? OR mr_no LIKE ? OR remarks LIKE ?"
                " OR issue_no LIKE ? OR dn_no LIKE ?)")
        p += [like] * 8
    for col in ("company", "recipient", "status", "issue_type"):
        if f.get(col):
            sql += f" AND {col}=?"
            p.append(f[col])
    if f.get("date_from"):
        sql += " AND issue_date>=?"
        p.append(f["date_from"])
    if f.get("date_to"):
        sql += " AND issue_date<=?"
        p.append(f["date_to"])
    if f.get("only_open"):
        sql += " AND issue_type=? AND qty_returned < qty"
        p.append(TEMPORARY)
    if f.get("only_overdue"):
        sql += " AND status=?"
        p.append(ST_OVERDUE)
    return sql, p


def dashboard(db: IssuanceDB, f: dict | None = None) -> dict:
    w, p = _where(f)

    def sc(expr: str, extra: str = "", params: list | None = None):
        return db.scalar(f"SELECT {expr} FROM issues WHERE 1=1{w}{extra}",
                         p + (params or []), default=0)

    total = int(sc("COUNT(*)"))
    qty = float(sc("COALESCE(SUM(qty),0)"))
    back = float(sc("COALESCE(SUM(qty_returned),0)"))
    out_qty = float(sc("COALESCE(SUM(CASE WHEN issue_type='Temporary'"
                       " THEN qty - qty_returned ELSE 0 END),0)"))
    month = _dt.date.today().strftime("%Y-%m")
    rows = [dict(r) for r in db.query(
        f"SELECT * FROM issues WHERE 1=1{w}", p)]
    no_proof = sum(1 for r in rows
                   if evidence_counts(db, r["id"])[0] == 0
                   and not str(r.get("dn_no") or "").strip())
    overdue_rows = [r for r in rows if compute_status(r) == ST_OVERDUE]
    ages = [d for d in (days_out(r) for r in rows
                        if compute_status(r) in (ST_ISSUED, ST_PARTIAL, ST_OVERDUE))
            if d is not None]
    return {
        "records": total,
        "companies": int(sc("COUNT(DISTINCT company)", " AND company<>''")),
        "recipients": int(sc("COUNT(DISTINCT recipient)", " AND recipient<>''")),
        "items": int(sc("COUNT(DISTINCT item)", " AND item<>''")),
        "qty_issued": qty,
        "qty_returned": back,
        "qty_out": out_qty,
        "value_out": float(sc("COALESCE(SUM(CASE WHEN issue_type='Temporary'"
                              " THEN (qty - qty_returned) * unit_value ELSE 0 END),0)")),
        "open_lines": int(sc("COUNT(*)",
                             " AND issue_type='Temporary' AND qty_returned < qty")),
        "returned_lines": int(sc("COUNT(*)", " AND qty>0 AND qty_returned >= qty")),
        "partial_lines": int(sc("COUNT(*)",
                                " AND qty_returned > 0 AND qty_returned < qty")),
        "permanent_lines": int(sc("COUNT(*)", " AND issue_type='Permanent'")),
        "overdue": len(overdue_rows),
        "overdue_qty": sum(outstanding_qty(r) for r in overdue_rows),
        "worst_overdue": max([days_overdue(r) for r in overdue_rows] or [0]),
        "missing_proof": no_proof,
        "proof_pct": (100.0 * (total - no_proof) / total) if total else 100.0,
        "this_month": int(sc("COUNT(*)", " AND substr(issue_date,1,7)=?", [month])),
        "return_rate": (back / qty * 100.0) if qty else 0.0,
        "avg_days_out": (sum(ages) / len(ages)) if ages else 0.0,
        "evidence_files": int(db.scalar("SELECT COUNT(*) FROM evidence", default=0)),
    }


def by_column(db: IssuanceDB, column: str, limit: int = 10, measure: str = "qty",
              f: dict | None = None) -> list[tuple[str, float]]:
    if column not in {x for x, _ in ALL_FIELDS}:
        return []
    agg = {
        "qty": "SUM(qty)",
        "count": "COUNT(*)",
        "outstanding": "SUM(CASE WHEN issue_type='Temporary'"
                       " THEN qty - qty_returned ELSE 0 END)",
        "value": "SUM(qty * unit_value)",
    }.get(measure, "SUM(qty)")
    w, p = _where(f)
    rows = db.query(
        f"SELECT COALESCE(NULLIF({column},''),'(blank)') k, {agg} v FROM issues"
        f" WHERE 1=1{w} GROUP BY k ORDER BY v DESC LIMIT ?", p + [limit])
    return [(r["k"], float(r["v"] or 0)) for r in rows]


def monthly_issue_return(db: IssuanceDB, months: int = 12,
                         f: dict | None = None) -> list[tuple[str, float, float]]:
    w, p = _where(f)
    rows = db.query(
        f"SELECT substr(issue_date,1,7) m, COALESCE(SUM(qty),0) q,"
        f" COALESCE(SUM(qty_returned),0) r FROM issues"
        f" WHERE length(issue_date)>=7{w} GROUP BY m ORDER BY m DESC LIMIT ?",
        p + [months])
    return [(r["m"][-2:], float(r["q"]), float(r["r"])) for r in reversed(rows)]


def status_split(db: IssuanceDB, f: dict | None = None) -> list[tuple[str, float]]:
    w, p = _where(f)
    rows = db.query(
        f"SELECT status k, COUNT(*) v FROM issues WHERE 1=1{w}"
        f" GROUP BY k ORDER BY v DESC", p)
    return [(r["k"] or "(none)", float(r["v"])) for r in rows]


def ageing(db: IssuanceDB, f: dict | None = None) -> list[tuple[str, float]]:
    """Outstanding quantity bucketed by how long it has been out."""
    w, p = _where(f)
    rows = [dict(r) for r in db.query(
        f"SELECT * FROM issues WHERE issue_type='Temporary' AND qty_returned < qty{w}",
        p)]
    buckets = {"0-7 d": 0.0, "8-30 d": 0.0, "31-60 d": 0.0,
               "61-90 d": 0.0, "90+ d": 0.0}
    for r in rows:
        d = days_out(r)
        if d is None:
            continue
        o = outstanding_qty(r)
        if d <= 7:
            buckets["0-7 d"] += o
        elif d <= 30:
            buckets["8-30 d"] += o
        elif d <= 60:
            buckets["31-60 d"] += o
        elif d <= 90:
            buckets["61-90 d"] += o
        else:
            buckets["90+ d"] += o
    return list(buckets.items())


def company_matrix(db: IssuanceDB, f: dict | None = None
                   ) -> tuple[list[str], list[list[Any]]]:
    """Company x status cross-tab — who is holding what."""
    rows = search(db, **{k: v for k, v in (f or {}).items()
                         if k in ("text", "company", "recipient", "status",
                                  "issue_type", "date_from", "date_to")})
    companies = sorted({r["company"] or "(blank)" for r in rows})
    cols = ["Company"] + STATUSES + ["Total Qty", "Still Out"]
    out: list[list[Any]] = []
    for c in companies:
        mine = [r for r in rows if (r["company"] or "(blank)") == c]
        line: list[Any] = [c]
        for st in STATUSES:
            line.append(sum(1 for r in mine if compute_status(r) == st))
        line.append(round(sum(to_float(r["qty"]) for r in mine), 2))
        line.append(round(sum(outstanding_qty(r) for r in mine), 2))
        out.append(line)
    if out:
        tot: list[Any] = ["TOTAL"]
        for i in range(1, len(cols)):
            tot.append(round(sum(float(r[i] or 0) for r in out), 2))
        out.append(tot)
    return cols, out


# -------------------------------------------------------------------- reports
REPORT_LIST = [
    "Full Issuance Register",
    "Currently Outstanding",
    "Overdue Returns",
    "Returned Items",
    "Permanent Issues",
    "By Company",
    "By Recipient (Custody)",
    "By Item",
    "Missing Photo Proof",
    "Ageing of Outstanding",
    "Monthly Issue vs Return",
    "Company × Status Matrix",
    "Evidence Index",
    "Written Off / Lost",
]


def build_report(db: IssuanceDB, name: str, f: dict | None = None
                 ) -> tuple[str, list[str], list[list[Any]]]:
    f = f or {}
    rows = search(db, **{k: v for k, v in f.items()
                         if k in ("text", "company", "recipient", "status",
                                  "issue_type", "item", "date_from", "date_to",
                                  "only_open", "only_overdue")})

    def base_cols() -> list[str]:
        return ["Issue No", "Date", "Company", "MR", "Recipient", "Iqama ID",
                "Item Issued", "Qty", "Date of Issuance", "Date of Return",
                "Returned", "Still Out", "Type", "Proof", "Status", "Remarks"]

    def base_rows(src) -> list[list[Any]]:
        out = []
        for r in src:
            n_out, n_in = evidence_counts(db, r["id"])
            proof = (f"{n_out} photo" + ("s" if n_out != 1 else "")) if n_out else (
                r["dn_no"] or "— none —")
            out.append([r["issue_no"], r["record_date"], r["company"], r["mr_no"],
                        r["recipient"], r["iqama"], r["item"],
                        round(to_float(r["qty"]), 2), r["issue_date"],
                        r["return_date"], round(to_float(r["qty_returned"]), 2),
                        round(outstanding_qty(r), 2), r["issue_type"], proof,
                        compute_status(r), r["remarks"]])
        return out

    if name == "Full Issuance Register":
        return name, base_cols(), base_rows(rows)

    if name == "Currently Outstanding":
        sel = [r for r in rows if outstanding_qty(r) > 0
               and compute_status(r) not in (ST_LOST,)]
        cols = ["Issue No", "Company", "Recipient", "Iqama ID", "Item Issued",
                "Qty", "Returned", "Still Out", "Issued On", "Due", "Days Out",
                "Status"]
        data = [[r["issue_no"], r["company"], r["recipient"], r["iqama"], r["item"],
                 round(to_float(r["qty"]), 2), round(to_float(r["qty_returned"]), 2),
                 round(outstanding_qty(r), 2), r["issue_date"],
                 r["expected_return"] or "-", days_out(r) if days_out(r) is not None
                 else "-", compute_status(r)] for r in sel]
        if data:
            data.append(["", "TOTAL", f"{len(data)} line(s)", "", "",
                         round(sum(x[5] for x in data), 2),
                         round(sum(x[6] for x in data), 2),
                         round(sum(x[7] for x in data), 2), "", "", "", ""])
        return name, cols, data

    if name == "Overdue Returns":
        sel = [r for r in rows if compute_status(r) == ST_OVERDUE]
        sel.sort(key=lambda r: -days_overdue(r))
        cols = ["Issue No", "Company", "Recipient", "Iqama ID", "Item Issued",
                "Still Out", "Issued On", "Was Due", "Days Overdue", "Phone",
                "Remarks"]
        return name, cols, [
            [r["issue_no"], r["company"], r["recipient"], r["iqama"], r["item"],
             round(outstanding_qty(r), 2), r["issue_date"], r["expected_return"],
             days_overdue(r), r["phone"], r["remarks"]] for r in sel]

    if name == "Returned Items":
        sel = [r for r in rows if compute_status(r) == ST_RETURNED]
        cols = ["Issue No", "Company", "Recipient", "Item Issued", "Qty",
                "Issued On", "Returned On", "Days Out", "Condition In",
                "Received Back By", "Proof"]
        return name, cols, [
            [r["issue_no"], r["company"], r["recipient"], r["item"],
             round(to_float(r["qty"]), 2), r["issue_date"], r["return_date"],
             days_out(r) if days_out(r) is not None else "-", r["condition_in"],
             r["received_back_by"],
             f"{evidence_counts(db, r['id'])[1]} photo(s)"] for r in sel]

    if name == "Permanent Issues":
        sel = [r for r in rows if r["issue_type"] == PERMANENT]
        cols = ["Issue No", "Date", "Company", "Recipient", "Iqama ID",
                "Item Issued", "Qty", "Value", "DN / Gate Pass", "Proof", "Remarks"]
        return name, cols, [
            [r["issue_no"], r["issue_date"], r["company"], r["recipient"], r["iqama"],
             r["item"], round(to_float(r["qty"]), 2),
             round(to_float(r["qty"]) * to_float(r["unit_value"]), 2), r["dn_no"],
             f"{evidence_counts(db, r['id'])[0]} photo(s)", r["remarks"]] for r in sel]

    if name == "Written Off / Lost":
        sel = [r for r in rows if compute_status(r) == ST_LOST]
        return name, base_cols(), base_rows(sel)

    if name in ("By Company", "By Recipient (Custody)", "By Item"):
        key = {"By Company": "company", "By Recipient (Custody)": "recipient",
               "By Item": "item"}[name]
        label = {"company": "Company", "recipient": "Recipient",
                 "item": "Item Issued"}[key]
        agg: dict[str, dict] = {}
        for r in rows:
            k = r.get(key) or "(blank)"
            a = agg.setdefault(k, {"lines": 0, "qty": 0.0, "back": 0.0, "out": 0.0,
                                   "overdue": 0, "value": 0.0, "noproof": 0})
            a["lines"] += 1
            a["qty"] += to_float(r["qty"])
            a["back"] += to_float(r["qty_returned"])
            a["out"] += outstanding_qty(r)
            a["value"] += outstanding_qty(r) * to_float(r["unit_value"])
            if compute_status(r) == ST_OVERDUE:
                a["overdue"] += 1
            if evidence_counts(db, r["id"])[0] == 0 and not str(r["dn_no"] or "").strip():
                a["noproof"] += 1
        cols = [label, "Lines", "Qty Issued", "Qty Returned", "Still Out",
                "Overdue", "Value Out", "No Proof"]
        data = [[k, a["lines"], round(a["qty"], 2), round(a["back"], 2),
                 round(a["out"], 2), a["overdue"], round(a["value"], 2), a["noproof"]]
                for k, a in sorted(agg.items(), key=lambda kv: -kv[1]["out"])]
        if data:
            data.append(["TOTAL", sum(x[1] for x in data),
                         round(sum(x[2] for x in data), 2),
                         round(sum(x[3] for x in data), 2),
                         round(sum(x[4] for x in data), 2),
                         sum(x[5] for x in data),
                         round(sum(x[6] for x in data), 2),
                         sum(x[7] for x in data)])
        return name, cols, data

    if name == "Missing Photo Proof":
        sel = [r for r in rows
               if evidence_counts(db, r["id"])[0] == 0
               and not str(r["dn_no"] or "").strip()]
        cols = ["Issue No", "Date", "Company", "Recipient", "Item Issued", "Qty",
                "Status", "Remarks"]
        return name, cols, [
            [r["issue_no"], r["issue_date"], r["company"], r["recipient"], r["item"],
             round(to_float(r["qty"]), 2), compute_status(r), r["remarks"]]
            for r in sel]

    if name == "Ageing of Outstanding":
        cols = ["Age Bucket", "Outstanding Qty"]
        data = [[k, round(v, 2)] for k, v in ageing(db, f)]
        data.append(["TOTAL", round(sum(x[1] for x in data), 2)])
        return name, cols, data

    if name == "Monthly Issue vs Return":
        agg: dict[str, dict] = {}
        for r in rows:
            k = (r.get("issue_date") or "")[:7] or "(no date)"
            a = agg.setdefault(k, {"lines": 0, "qty": 0.0, "back": 0.0})
            a["lines"] += 1
            a["qty"] += to_float(r["qty"])
            a["back"] += to_float(r["qty_returned"])
        cols = ["Month", "Lines", "Qty Issued", "Qty Returned", "Net Out"]
        data = [[k, a["lines"], round(a["qty"], 2), round(a["back"], 2),
                 round(a["qty"] - a["back"], 2)] for k, a in sorted(agg.items())]
        if data:
            data.append(["TOTAL", sum(x[1] for x in data),
                         round(sum(x[2] for x in data), 2),
                         round(sum(x[3] for x in data), 2),
                         round(sum(x[4] for x in data), 2)])
        return name, cols, data

    if name == "Company × Status Matrix":
        cols, data = company_matrix(db, f)
        return name, cols, data

    if name == "Evidence Index":
        cols = ["Issue No", "Date", "Company", "Item Issued", "Kind", "File",
                "Added On", "Exists"]
        data = []
        for r in rows:
            for e in evidence_for(db, r["id"]):
                data.append([r["issue_no"], r["issue_date"], r["company"], r["item"],
                             e["kind"], Path(e["file_path"]).name, e["added_at"],
                             "Yes" if Path(e["file_path"]).exists() else "MISSING"])
        return name, cols, data

    return name, base_cols(), base_rows(rows)


# --------------------------------------------------------------------- import
def sniff(text: str) -> tuple[list[str], list[list[str]]]:
    raw = [ln for ln in text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
           if ln.strip()]
    if not raw:
        return [], []
    if "\t" in raw[0]:
        rows = [ln.split("\t") for ln in raw]
    elif raw[0].count(",") >= 2:
        rows = list(csv.reader(io.StringIO("\n".join(raw))))
    elif raw[0].count("|") >= 2:
        rows = [[c.strip() for c in ln.strip().strip("|").split("|")] for ln in raw]
    else:
        rows = [re.split(r"\s{2,}", ln.strip()) for ln in raw]
    rows = [[str(c).strip() for c in r] for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return [], []
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    head = rows[0]
    if sum(1 for c in head if norm(c) in HEADER_MAP) >= max(2, width // 3):
        return head, rows[1:]
    return [f"Column {i + 1}" for i in range(width)], rows


def read_file(path: str | Path) -> tuple[list[str], list[list[Any]]]:
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(p, data_only=True, read_only=True)
        ws = wb.active
        data = [[("" if c is None else c) for c in row]
                for row in ws.iter_rows(values_only=True)]
        wb.close()
        data = [r for r in data if any(str(c).strip() for c in r)]
        if not data:
            return [], []
        head = [str(c).strip() for c in data[0]]
        if sum(1 for c in head if norm(c) in HEADER_MAP) >= 2:
            return head, data[1:]
        return [f"Column {i + 1}" for i in range(len(head))], data
    return sniff(p.read_text(encoding="utf-8", errors="ignore"))


def auto_map(headers: Sequence[str]) -> dict[int, str]:
    out: dict[int, str] = {}
    used: set[str] = set()
    for i, h in enumerate(headers):
        f = HEADER_MAP.get(norm(h))
        if f and f not in used:
            out[i] = f
            used.add(f)
    return out


def preview(headers: Sequence[str], rows: Sequence[Sequence[Any]],
            mapping: dict[int, str], defaults: dict | None = None) -> list[dict]:
    """Turn raw rows into canonical records (nothing saved yet).

    The sheet's Remarks column doubles as a status ("Returned" / "Not Return
    yet"), so it is read for meaning as well as kept verbatim.
    """
    defaults = defaults or {}
    out = []
    for r in rows:
        rec = {f: "" for f, _ in ALL_FIELDS}
        rec.update({k: v for k, v in defaults.items() if v})
        ev_hint = ""
        for i, field in mapping.items():
            if i >= len(r):
                continue
            val = r[i]
            if field == "_evidence":
                ev_hint = str(val or "").strip()
                continue
            rec[field] = val
        for f in NUMERIC:
            rec[f] = to_float(rec.get(f))
        for f in DATE_FIELDS:
            rec[f] = to_date(rec.get(f))
        if not rec["issue_date"]:
            rec["issue_date"] = rec["record_date"]
        if not rec["record_date"]:
            rec["record_date"] = rec["issue_date"]
        remark = str(rec.get("remarks") or "").strip()
        low = remark.lower()
        if not rec["issue_type"]:
            rec["issue_type"] = PERMANENT if "permanent" in low else TEMPORARY
        if rec["return_date"] and rec["qty_returned"] <= 0:
            rec["qty_returned"] = rec["qty"]
        if "returned" in low and "not" not in low and rec["qty_returned"] <= 0:
            rec["qty_returned"] = rec["qty"]
        # an evidence cell holding a DN number is a reference, a path is a file
        if ev_hint and ev_hint not in ("-", "N/A"):
            if Path(ev_hint).suffix.lower() in EVIDENCE_SUFFIXES:
                rec["_evidence_file"] = ev_hint
            else:
                rec["dn_no"] = rec.get("dn_no") or ev_hint
        rec["status"] = compute_status(rec)
        if not any(str(rec.get(k, "")).strip() for k in ("company", "item", "recipient")) \
                and not rec["qty"]:
            continue
        out.append(rec)
    return out


def import_records(db: IssuanceDB, records: Sequence[dict], source: str = "",
                   skip_duplicates: bool = True) -> tuple[int, int]:
    cur = db.execute(
        "INSERT INTO batches(ts,source,rows,skipped,username) VALUES(?,?,?,?,?)",
        (_now(), str(source), 0, 0, db.current_user))
    batch_id = int(cur.lastrowid)

    existing: set[tuple] = set()
    if skip_duplicates:
        for r in db.query("SELECT company,item,issue_date,qty,recipient FROM issues"):
            existing.add((norm(r["company"]), norm(r["item"]), r["issue_date"] or "",
                          round(to_float(r["qty"]), 3), norm(r["recipient"])))

    ins = skipped = 0
    for rec in records:
        key = (norm(rec.get("company")), norm(rec.get("item")),
               rec.get("issue_date") or "", round(to_float(rec.get("qty")), 3),
               norm(rec.get("recipient")))
        if skip_duplicates and key in existing:
            skipped += 1
            continue
        existing.add(key)
        data = dict(rec)
        data["batch_id"] = batch_id
        ev = data.pop("_evidence_file", "")
        new_id = save_issue(db, data, require_evidence=False)
        if ev and Path(ev).exists():
            try:
                store_evidence(db, new_id, ev, "ISSUE", "imported")
            except OSError:
                pass
        ins += 1
    db.execute("UPDATE batches SET rows=?, skipped=? WHERE id=?",
               (ins, skipped, batch_id))
    db.commit()
    db.audit("IMPORTED", "issues", batch_id,
             f"{ins} inserted, {skipped} duplicate(s) skipped from {source}")
    return ins, skipped


def undo_batch(db: IssuanceDB, batch_id: int) -> int:
    ids = [r["id"] for r in db.query("SELECT id FROM issues WHERE batch_id=?",
                                     (batch_id,))]
    n = delete_issues(db, ids)
    db.audit("DELETED", "batch", batch_id, f"{n} record(s) removed (undo import)")
    return n


def batches(db: IssuanceDB) -> list[dict]:
    return [dict(r) for r in db.query(
        "SELECT b.*, (SELECT COUNT(*) FROM issues i WHERE i.batch_id=b.id) live"
        " FROM batches b ORDER BY b.id DESC LIMIT 100")]


def template_rows() -> tuple[list[str], list[list[Any]]]:
    cols = [lbl for _, lbl in FIELDS] + ["Evidence"]
    today = _dt.date.today()
    return cols, [
        [today.isoformat(), "Alnoor", "1728", "Muhammad Shoaib", "2563232723",
         "Cable Puller", 24, today.isoformat(), "", "Not Return yet", "photo1.jpg"],
        [today.isoformat(), "Broad", "-", "Muhammad Adnan", "2563232723",
         "Heat Gun", 1, today.isoformat(), today.isoformat(), "Returned", "-"],
    ]
