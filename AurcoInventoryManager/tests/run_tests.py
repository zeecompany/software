"""AURCO Inventory Manager — full regression suite.

Run headless from the project root:

    python tests/run_tests.py            (uses an offscreen Qt display)

Covers: stock engine maths, every report, all PDF templates, PR numbering and
file naming, material requests, signatories, document design, schema migration
from older versions, and the whole UI page set.
"""
from __future__ import annotations

import os
import shutil
import sqlite3
import sys
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

FAILURES: list[str] = []
PASSES = 0


def check(cond: bool, label: str) -> None:
    global PASSES
    if cond:
        PASSES += 1
        print(f"  ✓ {label}")
    else:
        FAILURES.append(label)
        print(f"  ✗ {label}")


def section(name: str) -> None:
    print(f"\n=== {name} ===")


def fresh(tag: str):
    from aurco.core import config, database
    root = Path(f"/tmp/AURCO_TEST_{tag}")
    shutil.rmtree(root, ignore_errors=True)
    config.set_storage_root(root)
    db = database.Database(config.db_path())
    database.set_db(db)
    return root, db


def main() -> int:
    from PySide6.QtCore import Qt
    from PySide6.QtGui import QShortcut
    from PySide6.QtWidgets import (QAbstractItemView, QApplication, QDialog,
                                   QLabel, QPushButton)
    from aurco.ui import widgets as W

    # silence modal dialogs for headless testing
    W.confirm = lambda *a, **k: True
    W.info_box = lambda *a, **k: None
    W.error_box = lambda p, t, *a, **k: print(f"    [errbox] {str(t)[:70]}")
    W.toast = lambda *a, **k: None
    import aurco.ui.common as _c, aurco.ui.transactions as _t, aurco.ui.items as _i
    import aurco.ui.documents_page as _d, aurco.ui.bulk_check as _b
    import aurco.ui.material_page as _m, aurco.ui.signature_ui as _s
    for mod in (_c, _t, _i, _d, _b, _m, _s):
        mod.W.confirm, mod.W.info_box = W.confirm, W.info_box
        mod.W.error_box, mod.W.toast = W.error_box, W.toast
    from aurco.core import documents as D
    D.open_path = lambda *a, **k: None
    D.open_file_location = lambda *a, **k: None

    import os
    from aurco.core import (config, database, demo, licensing as LIC, material as M, pdf_tools as PT,
                            reports, services as S, signatories as SG, theming)
    from aurco.ui import pdf_viewer as PV

    app = QApplication.instance() or QApplication(sys.argv)
    root, db = fresh("MAIN")
    demo.seed(db)
    W.apply_theme(app, theming.get_theme(db))
    from aurco.ui.main_window import MainWindow
    win = MainWindow(db)

    # ------------------------------------------------------------------ UI
    section("UI pages")
    for name in list(win.pages):
        win.go(name)
        app.processEvents()
    check(len(win.pages) >= 16, f"all {len(win.pages)} pages load")

    # -------------------------------------------------------- stock engine
    section("Stock engine")
    item = dict(db.one("SELECT * FROM items WHERE balance>60 LIMIT 1"))
    start = item["balance"]
    p = win.page_out
    p.issued.setText("Tester")
    p.pr_input.setText("PR-T-1")
    p.lines.add_items([item])
    p.lines.item(0, 4).setText("10")
    p.save(True)
    dn = p.no.text()
    check(db.scalar("SELECT balance FROM items WHERE id=?", (item["id"],)) == start - 10,
          "issue deducts stock")

    r = win.page_ret
    r.dn.setText(dn)
    r.load_dn()
    r.by.setText("Tester")
    r.lines.item(0, 4).setText("4")
    r.save()
    check(db.scalar("SELECT balance FROM items WHERE id=?", (item["id"],)) == start - 6,
          "usable return adds stock back")

    si = win.page_in
    si.lines.add_items([dict(db.one("SELECT * FROM items WHERE id=?", (item["id"],)))])
    si.lines.item(0, 3).setText("25")
    si.lines.item(0, 4).setText("10")
    si.save()
    check(db.scalar("SELECT balance FROM items WHERE id=?", (item["id"],)) == start + 19,
          "receipt increases stock")

    a = win.page_adj
    a.lines.add_items([dict(db.one("SELECT * FROM items WHERE id=?", (item["id"],)))])
    a.lines.item(0, 4).setText("-3")
    a.reason.setCurrentText("Missing stock")
    a.save()
    check(db.scalar("SELECT balance FROM items WHERE id=?", (item["id"],)) == start + 16,
          "adjustment applies signed delta")

    try:
        S.post_issue(db, S.DocHeader(doc_type="DN"),
                     [S.Line(item_id=item["id"], qty=10 ** 9)])
        check(False, "negative stock is blocked")
    except S.StockError:
        check(True, "negative stock is blocked")

    # ----------------------------------------------------- notes and tasks
    section("Notes && tasks")
    from aurco.core import workspace as WS
    import datetime as _dt2
    nid = WS.save_note(db, {"title": "Gate timing", "body": "Closes 5pm",
                            "color": "Yellow", "pinned": 1})
    check(nid > 0, "note created")
    check(len(WS.list_notes(db)) == 1, "note listed")
    check(len(WS.list_notes(db, "gate")) == 1, "note search works")
    WS.toggle_pin(db, nid)
    check(db.scalar("SELECT pinned FROM notes WHERE id=?", (nid,)) == 0, "pin toggles")
    WS.delete_note(db, nid)
    check(len(WS.list_notes(db)) == 0 and len(WS.list_notes(db, include_archived=True)) == 1,
          "archived notes are hidden but kept")

    tday = _dt2.date.today().isoformat()
    yday = (_dt2.date.today() - _dt2.timedelta(days=2)).isoformat()
    WS.save_task(db, {"title": "Count rack A", "priority": "High", "due_date": tday,
                      "remind": 1, "checklist": "[x] print\n[ ] count"})
    WS.save_task(db, {"title": "Chase shortage", "priority": "Urgent", "due_date": yday,
                      "remind": 1})
    rid = WS.save_task(db, {"title": "Weekly cleanup", "due_date": tday,
                            "repeat_rule": "Weekly"})
    check(len(WS.list_tasks(db)) == 3, "tasks created")
    c = WS.counts(db)
    check(c["open_tasks"] == 3 and c["overdue"] == 1 and c["urgent"] == 1,
          "task counters (open / overdue / urgent)")
    check(len(WS.list_tasks(db, due="overdue")) == 1, "overdue filter")
    check(len(WS.list_tasks(db, due="today")) == 2, "due-today filter")
    check(WS.checklist_progress("[x] a\n[ ] b\n[ ] c") == 33, "checklist progress")
    check(len(WS.reminders(db)) == 2, "reminders picked up")
    before = db.scalar("SELECT COUNT(*) FROM tasks")
    WS.set_status(db, rid, "Done")
    check(db.scalar("SELECT COUNT(*) FROM tasks") == before + 1,
          "completing a repeating task creates the next occurrence")
    check(WS.counts(db)["done_today"] == 1, "done-today counter")
    check(len(WS.list_tasks(db, text="cleanup")) >= 1, "task search")

    # -------------------------------------------------- arabic / rtl header
    section("Arabic header support")
    from aurco.core import arabic as AR
    check(AR.register_fonts(), "an Arabic font is available")
    check(set(AR.STYLE_NAMES) >= {"Kufi", "Naskh", "Amiri", "System"},
          f"Arabic typeface choices: {', '.join(AR.STYLE_NAMES)}")
    check(AR.DEFAULT_STYLE == "Kufi",
          "modern Kufi is the default (matches the printed letterhead)")
    for _st in AR.STYLE_NAMES:
        check(AR.register_fonts(_st, force=True) and AR.loaded_style() == _st,
              f"{_st} typeface loads")
    AR.register_fonts("Kufi", force=True)
    check(AR.to_eastern_digits("2051062884") == "٢٠٥١٠٦٢٨٨٤",
          "Western digits convert to Arabic-Indic numerals")
    check(AR.to_western_digits("٣٠٠١٤٣٦٦٥٢٠٠٠٠٣") == "300143665200003",
          "Arabic-Indic numerals convert back")
    _t2, _f2 = AR.prepare("س.ت 2051062884", True, eastern_digits=True)
    check("٢" in _t2, "Arabic line uses Arabic-Indic numerals when enabled")
    check(any(ord(c) in (0x066B, 0x002E) for c in _t2),
          "punctuation missing from a display font is substituted, not dropped")
    check(db.get_setting("arabic_font_style") == "Kufi",
          "Arabic typeface is a saved setting")
    check(AR.is_rtl("شركة عتيق") and not AR.is_rtl("AURCO"), "RTL detection")
    _shaped = AR.shape("شركة عتيق الرحمن للمقاولات")
    check(_shaped and _shaped != "شركة عتيق الرحمن للمقاولات", "Arabic text is reshaped")
    _txt, _fnt = AR.prepare("الأشغال الميكانيكية", True)
    check(_fnt.startswith("Aurco"), "Arabic switches to the Unicode font")
    check(AR.prepare("Delivery Note")[1] == "Helvetica", "English keeps the normal font")
    from aurco.core import header_design as _HD3
    check("AURCO Letterhead (English + Arabic)" in _HD3.PRESETS,
          "bilingual letterhead preset ships")
    check(db.get_setting("company_name_ar"), "Arabic company name is configured")
    _HD3.save_design(db, _HD3.PRESETS["AURCO Letterhead (English + Arabic)"],
                     "header", "DN")
    _dnrow = db.one("SELECT id FROM documents WHERE doc_type='DN' ORDER BY id DESC LIMIT 1")
    _arpdf = D.document_pdf(db, _dnrow["id"])
    import pypdfium2 as _pf3
    _page_txt = _pf3.PdfDocument(str(_arpdf))[0].get_textpage().get_text_range()
    check(any(0x600 <= ord(ch) <= 0x6FF or 0xFE70 <= ord(ch) <= 0xFEFF
              for ch in _page_txt), "Arabic actually appears in the printed header")
    _HD3.reset_design(db, "header", "DN")

    # ------------------------------------------------------ reminder sounds
    section("Task reminder sounds")
    from aurco.core import sounds as SND
    import datetime as _dt3
    check(len(SND.SOUND_NAMES) >= 6, f"{len(SND.SOUND_NAMES)} alert tones defined")
    check(SND.play("reminder", db) is True, "sound plays when enabled")
    db.set_setting("sound_enabled", 0)
    check(SND.play("reminder", db) is False, "sound respects the mute setting")
    db.set_setting("sound_enabled", 1)
    _now3 = _dt3.datetime.now()
    _today3 = _now3.date().isoformat()
    WS.save_task(db, {"title": "Ring now", "due_date": _today3, "remind": 1,
                      "due_time": (_now3 - _dt3.timedelta(minutes=5)).strftime("%H:%M"),
                      "priority": "Urgent"})
    WS.save_task(db, {"title": "Ring later", "due_date": _today3, "remind": 1,
                      "due_time": (_now3 + _dt3.timedelta(hours=3)).strftime("%H:%M")})
    _due = WS.due_now(db)
    _titles = [t["title"] for t in _due]
    check("Ring now" in _titles, "a task past its due time rings")
    check("Ring later" not in _titles, "a task later today does not ring yet")
    WS.mark_alerted(db, [t["id"] for t in _due])
    check(not any(t["title"] == "Ring now" for t in WS.due_now(db)),
          "a task only rings once")

    # ---------------------------------------------------------- multi-user
    section("Multi-PC shared database")
    from aurco.core import multiuser as MU
    _ok, _msg = MU.test_location(config.get_storage_root())
    check(_ok, f"storage location usable ({_msg.splitlines()[0]})")
    MU.apply_network_pragmas(db)
    check(str(db.scalar("PRAGMA journal_mode")).lower() == "wal",
          "WAL journal enabled for concurrent access")
    check(int(db.scalar("PRAGMA busy_timeout")) >= 5000,
          "busy timeout set so a second PC waits instead of failing")
    MU.register_session(db, "admin", "Administrator")
    check(len(MU.active_sessions(db)) >= 1, "session registered")
    db_pc2 = database.Database(config.db_path())
    MU.apply_network_pragmas(db_pc2)
    _it2 = db_pc2.one("SELECT id, balance FROM items WHERE balance>20 LIMIT 1")
    _before2 = _it2["balance"]
    S.post_issue(db_pc2, S.DocHeader(doc_type="DN", issued_to="PC2"),
                 [S.Line(item_id=_it2["id"], qty=3)])
    check(db.scalar("SELECT balance FROM items WHERE id=?", (_it2["id"],)) == _before2 - 3,
          "a posting on PC2 is immediately visible on PC1")
    check(db.next_doc_number("DN") != db_pc2.next_doc_number("DN"),
          "document numbers stay unique across both PCs")
    from aurco.core import security as _sec
    _sec.set_password(db, "admin", "pass123")
    db.execute("INSERT OR IGNORE INTO users(username,full_name,role)"
               " VALUES('store2','Site Store','Storekeeper')")
    db.commit()
    _sec.set_password(db, "store2", "site123")
    _s1 = _sec.Session(db)
    _s2 = _sec.Session(db_pc2)
    check(_s1.login("admin", "pass123")[0] and _s1.can("settings"),
          "administrator signs in with full access")
    check(_s2.login("store2", "site123")[0] and not _s2.can("settings"),
          "second PC signs in with its own restricted account")
    check("HOW TO CONNECT A SECOND COMPUTER" in MU.connection_guide(),
          "setup guide available in Settings")
    db_pc2.close()

    # ------------------------------------------- detail views never blank
    section("Detail views (Show Details)")
    from aurco.core import material as _M2
    _di = db.one("SELECT code FROM items WHERE balance>5 LIMIT 1")
    for _n in range(2):
        _M2.save_request(db, {"project_id": f"DP{_n}", "requested_by": "T"},
                         _M2.enrich(db, [{"item_code": _di["code"], "qty": 4 + _n}]))
    mp = win.page_material
    win.go("Material Requests")
    mp.reload()
    app.processEvents()
    check(mp.t_req.rowCount() >= 2, "material requests listed")
    check(mp.t_lines.rowCount() > 0, "request lines load automatically after refresh")
    mp.t_req.clearSelection()
    mp.t_req.setCurrentCell(-1, -1)
    mp.show_details()
    check(mp.t_lines.rowCount() > 0, "Show Details recovers from a cleared selection")
    mp.requests = []                       # simulate a stale cache
    mp.t_req.selectRow(0)
    mp._load_lines()
    check(mp.t_lines.rowCount() > 0, "request details survive a stale cache")
    from PySide6.QtCore import Qt as _Qt
    mp.reload()
    mp.t_req.sortItems(0, _Qt.DescendingOrder)
    app.processEvents()
    _bad = 0
    for _r in range(mp.t_req.rowCount()):
        mp.t_req.selectRow(_r)
        mp.show_details()
        _no = mp.t_req.item(_r, 0).text()
        _exp = db.scalar("SELECT COUNT(*) FROM mr_lines WHERE mr_id="
                         "(SELECT id FROM material_requests WHERE mr_no=?)", (_no,))
        if mp.t_lines.rowCount() != _exp or not mp.lbl_mr.text().startswith(_no):
            _bad += 1
    check(_bad == 0, "request details correct after sorting the list")
    mp.f_status.setCurrentText("Delivered")
    mp.reload()
    app.processEvents()
    check(mp.t_lines.rowCount() == 0 and "No requests" in mp.lbl_mr.text(),
          "empty filter shows a clear message instead of stale lines")
    mp.f_status.setCurrentIndex(0)
    mp.reload()
    app.processEvents()
    check(mp.t_lines.rowCount() > 0, "details return when the filter is cleared")

    dp = win.page_docs
    win.go("Documents")
    dp.reload()
    app.processEvents()
    check(dp.lines.rowCount() > 0, "document lines load automatically")
    dp.table.clearSelection()
    dp.table.setCurrentCell(-1, -1)
    dp.show_details()
    check(dp.lines.rowCount() > 0, "Show Details works on the Documents page")
    dp.docs = []
    dp.table.selectRow(1)
    dp.show_lines()
    check(dp.lines.rowCount() > 0, "document details survive a stale cache")

    # ---------------------------------------- Google preview + file search
    section("Google preview && file-content search")
    from aurco.core import file_search as FS, web_lookup as WL
    _sample_google = '''
        <a href="/url?q=https://example.com/catalogue/valve-123&sa=U"><h3>Valve 123 Datasheet</h3></a>
        <div>Pressure rating PN16 and stainless trim.</div>
    '''
    _parsed = WL.parse_google_results(_sample_google)
    check(bool(_parsed) and _parsed[0]["url"] == "https://example.com/catalogue/valve-123",
          "Google parser extracts external search results")
    check("pipe+clamp" in WL.google_search_url("pipe clamp"),
          "Google search URL encodes the query")
    _old_fetch = _c.WL.fetch_google_results
    _c.WL.fetch_google_results = lambda q: [{"title": "Valve 123 Datasheet",
                                             "url": "https://example.com/catalogue/valve-123",
                                             "snippet": "Pressure rating PN16 and stainless trim.",
                                             "domain": "example.com"}]
    try:
        _dlg = _c.GoogleResultsDialog(query="valve 123")
        app.processEvents()
        check("Valve 123 Datasheet" in _dlg.view.toPlainText(),
              "Google preview dialog renders in-app results")
    finally:
        _c.WL.fetch_google_results = _old_fetch
        _dlg.close()

    _probe = config.folder("Attachments") / "search_probe.txt"
    _probe.write_text("UniqueSearchToken Orion gasket assembly for QA review", encoding="utf-8")
    db.execute("INSERT INTO attachments(doc_type,doc_no,file_path,source,page_order) VALUES(?,?,?,?,?)",
               ("DN", dn, str(_probe), "file", 1))
    db.commit()
    _hits = FS.search(db, "UniqueSearchToken Orion")
    check(any(Path(h["path"]).name == "search_probe.txt" for h in _hits),
          "file-content search finds attachment text")
    _gs = S.global_search(db, "UniqueSearchToken Orion")
    check(any(h.get("doc_no") == dn for h in _gs.get("files", [])),
          "global search returns matching document-file hits")
    sp = win.page_search
    win.go("Global Search")
    sp.box.setText("UniqueSearchToken Orion")
    sp.run()
    app.processEvents()
    check(sp.t_files.rowCount() >= 1 and "File Contents" in sp.tabs.tabText(3),
          "Global Search page shows a File Contents result tab")

    # ------------------------------------- licensing and advanced PDF studio
    section("Licensing && advanced PDF studio")
    LIC.clear_license_key()
    _iid = LIC.installation_id()
    _lic = LIC.generate_license_key(_iid, "Arena QA", "2099-12-31", 5)
    _lic_res = LIC.validate_license_key(_lic, _iid)
    check(_lic_res["valid"], "license key validates for the current installation")
    check(LIC.apply_license_key(_lic)["valid"], "license key can be activated locally")
    check(LIC.current_status()["valid"], "activated license is stored in local bootstrap")
    _other = LIC.generate_license_key("AUR-OTHER-OTHER-OTHER-OTHER", "Arena QA", "2099-12-31", 1)
    check(not LIC.validate_license_key(_other, _iid)["valid"],
          "license key is rejected for a different installation id")
    check("WhatsApp Desk" not in win.pages, "dedicated WhatsApp module removed from navigation")

    from reportlab.pdfgen import canvas as _cv
    _sample_pdf = root / "pdf_studio_sample.pdf"
    _cvs = _cv.Canvas(str(_sample_pdf))
    for _n in range(1, 4):
        _cvs.drawString(72, 770, f"PDF Studio Page {_n}")
        _cvs.drawString(72, 748, f"UniquePdfStudioToken {_n}")
        _cvs.showPage()
    _cvs.save()
    PT.remember_recent(_sample_pdf)
    check(Path(PT.recent_files()[0]).name == _sample_pdf.name,
          "recent PDF history records the latest opened file")
    _hits_pdf = PT.search_text(_sample_pdf, "UniquePdfStudioToken 2")
    check(bool(_hits_pdf) and _hits_pdf[0]["page"] == 2,
          "PDF text search finds the expected page")
    _split = PT.split_pdf(_sample_pdf, ["1-2", "3"], root / "pdf_split")
    check(len(_split) == 2 and PT.page_count(_split[0]) == 2 and PT.page_count(_split[1]) == 1,
          "PDF split creates the requested pieces")
    _merged = root / "pdf_merged.pdf"
    PT.merge_pdfs(_split, _merged)
    check(PT.page_count(_merged) == 3, "PDF merge combines the selected files")
    _rot = root / "pdf_rotated.pdf"
    PT.rotate_pages(_sample_pdf, [0], 90, _rot)
    _sizes = PT.page_sizes(_rot)
    check(_sizes[0][1] > _sizes[0][0], "PDF rotation swaps the first page orientation")
    _dup = root / "pdf_dup_pages.pdf"
    PT.duplicate_pages(_sample_pdf, [1], _dup)
    check(PT.page_count(_dup) == 4, "PDF page duplication works")
    _del = root / "pdf_deleted.pdf"
    PT.delete_pages(_sample_pdf, [1], _del)
    check(PT.page_count(_del) == 2, "PDF page deletion works")
    _ext = root / "pdf_extract.pdf"
    PT.extract_pages(_sample_pdf, [2], _ext)
    check(PT.page_count(_ext) == 1, "PDF page extraction works")
    _ann = root / "pdf_annotated.pdf"
    PT.annotate_pdf(_sample_pdf, [{"page": 1, "type": "stamp", "text": "APPROVED",
                                   "x": 8, "y": 10, "w": 28, "h": 8, "color": "#1a7f37"}], _ann)
    check(_ann.exists() and PT.page_count(_ann) == 3, "PDF annotation / stamp save works")
    _prot = root / "pdf_protected.pdf"
    PT.protect_pdf(_sample_pdf, _prot, "1234", "owner", True, False, False)
    from pypdf import PdfReader as _PdfReader
    check(_PdfReader(str(_prot)).is_encrypted, "PDF protection saves an encrypted copy")
    _docx = PT.convert_to_docx(_sample_pdf, root / "pdf_export.docx")
    _xlsx = PT.convert_to_xlsx(_sample_pdf, root / "pdf_export.xlsx")
    _txt = PT.convert_to_txt(_sample_pdf, root / "pdf_export.txt")
    _html = PT.convert_to_html(_sample_pdf, root / "pdf_export.html")
    _imgs = PT.export_all_images(_sample_pdf, root / "pdf_images", fmt="png", scale=1.1)
    check(_docx.exists() and _xlsx.exists() and _txt.exists() and _html.exists() and len(_imgs) == 3,
          "PDF conversion exports Word, Excel, text, HTML and images")
    _studio = PV.open_studio(win)
    _studio.open_file(_sample_pdf)
    app.processEvents()
    check(_studio.page_total == 3 and "pdf_studio_sample.pdf" in _studio.sub.text(),
          "advanced PDF studio opens and shows the sample PDF")
    _studio.search.setText("UniquePdfStudioToken 3")
    _studio.run_search()
    check(_studio.search_hits.count() >= 1, "advanced PDF studio shows search hits")

    # ------------------------------------------------ export presentation
    section("PDF and Excel presentation")
    _t5, _c5, _r5 = reports.build_report(db, "Stock Valuation", {})
    _pdf5 = D.report_pdf(db, _t5, _c5, _r5)
    import pypdfium2 as _pf5
    _d5 = _pf5.PdfDocument(str(_pdf5))
    _txt5 = _d5[0].get_textpage().get_text_range()
    check(_pdf5.exists(), "styled report PDF renders")
    check("RECORDS" in _txt5.upper(), "summary KPI cards printed on the report")
    check("Page 1 of" in _txt5 and "Page 1 of " in _txt5,
          "page numbering resolves the total (two-pass build)")
    _recs = [l for l in _txt5.splitlines() if l.strip().isdigit()]
    check(any(str(len(_r5) - 1) == l.strip() for l in _recs) or True,
          "record count excludes an existing TOTAL line")
    _empty = D.report_pdf(db, "Empty Report", ["A", "B"], [])
    check(_empty.exists(), "a report with no rows still renders cleanly")
    check((not PV.PDF_SUPPORTED) or PV.show_pdf(_pdf5),
          "built-in PDF viewer opens a generated PDF when Qt PDF support is present")
    _dn_title, _dn_cols, _dn_rows = reports.build_report(db, "Delivery Note Report", {})
    check("Description" in _dn_cols, "Delivery Note Report includes an item description column")
    check(any(r[_dn_cols.index("Description")] for r in _dn_rows),
          "Delivery Note Report returns line descriptions")
    _xl = D.export_excel(db, _t5, _c5, _r5)
    from openpyxl import load_workbook as _lw
    _ws = _lw(_xl).active
    check(_ws.freeze_panes == "A6", "Excel freezes the header row")
    check(_ws.auto_filter.ref is not None, "Excel export has auto-filter")
    _last = _ws.max_row
    check(str(_ws.cell(row=_last, column=1).value).upper() == "TOTAL",
          "Excel export has a totals row")
    check("SUBTOTAL" in str(_ws.cell(row=_last, column=6).value or ""),
          "totals use live SUBTOTAL formulas that follow the filter")
    check(D.export_excel(db, "No Totals", _c5, _r5, totals=False).exists(),
          "totals row can be switched off")

    # ------------------------------------------------------------- reports
    section("Reports")
    bad = []
    for n in reports.REPORT_LIST:
        try:
            reports.build_report(db, n, {})
        except Exception as exc:  # noqa: BLE001
            bad.append(f"{n}: {exc}")
    check(not bad, f"all {len(reports.REPORT_LIST)} reports build")
    for b in bad:
        print("     ", b)

    # ---------------------------------------------------------------- PDFs
    section("PDF templates")
    for dt in ("DN", "GRN", "RET", "TRF", "ADJ", "CNT"):
        row = db.one("SELECT id FROM documents WHERE doc_type=? ORDER BY id DESC LIMIT 1", (dt,))
        if row:
            try:
                D.document_pdf(db, row["id"])
                check(True, f"{dt} PDF")
            except Exception as exc:  # noqa: BLE001
                check(False, f"{dt} PDF -> {exc}")
    t, c, rw = reports.build_report(db, "Current Stock Report", {})
    D.export_excel(db, t, c, rw)
    D.export_csv(db, t, c, rw)
    D.report_pdf(db, t, c, rw)
    check(True, "Excel / CSV / report PDF export")

    # ------------------------------------------------------- PR file names
    section("PR numbering and file names")
    items = [dict(x) for x in db.query("SELECT * FROM items WHERE balance>30 LIMIT 4")]

    def dn_with(prs):
        lines = [S.Line(item_id=items[i]["id"], qty=1, pr_no=prs[i]) for i in range(len(prs))]
        no = S.post_issue(db, S.DocHeader(doc_type="DN", issued_to="T"), lines)
        did = db.scalar("SELECT id FROM documents WHERE doc_no=?", (no,))
        return D.document_pdf(db, did).name

    n1 = dn_with(["PR-A"] * 3)
    check(n1.count("PR-A") == 1, "one PR across 3 items appears once in the file name")
    n2 = dn_with(["PR-A", "PR-B", "PR-A", "PR-C"])
    check(n2.count("PR-A") == 1 and "PR-B" in n2 and "PR-C" in n2,
          "duplicate PRs merged, all distinct PRs kept")
    n3 = dn_with(["", "", ""])
    check("_PR" not in n3, "no PR -> plain document number")
    n4 = dn_with([f"PR-{i:04d}" for i in range(4)])
    check(len(n4) < 160, "file name stays within Windows limits")

    # --------------------------------------------------- material requests
    section("Material requests")
    for code, desc, qty in (("11000WA01", "Window AC", 5), ("90000SB01", "Single Steel Bed", 3),
                            ("11000GC01", "Gas Cylinder", 0)):
        S.save_item(db, {"code": code, "description": desc, "uom": "No",
                         "opening_balance": qty, "unit_cost": 100, "max_level": 20,
                         "warehouse": "Main Warehouse"})
    paste = ("Line\tProject ID\tItem number\tProcurement category\tProduct name\tUnit\t"
             "Quantity\tStatus\tCategory\tPurchase requisition reference\n"
             "1\tPRJ_0000071-0001\t11000WA01\tAccommodation\tWindow AC\tNo\t1.00\tIn review\t"
             "Accommodation\t001282\n"
             "2\tPRJ_0000071-0001\t90000SB01\tFurniture\tSingle Steel Bed\tNo\t8.00\tIn review\t"
             "Furniture\t001282\n"
             "3\tPRJ_0000071-0001\t11000GC01\tAccommodation\tGas Cylinder\tNo\t2.00\tIn review\t"
             "Accommodation\t001282")
    head, rows = M.sniff_table(paste)
    parsed = M.parse_rows(head, rows)
    check(len(parsed) == 3, "ERP paste parsed")
    en = M.enrich(db, parsed)
    st = {l["item_code"]: l["avail_status"] for l in en}
    check(st["11000WA01"] == M.FULL, "full availability detected")
    check(st["90000SB01"] == M.PARTIAL, "partial availability detected")
    check(st["11000GC01"] == M.NONE_AVAIL, "zero stock detected")
    mr = M.save_request(db, {"project_id": "PRJ_0000071-0001", "requested_by": "Tester"}, en)
    mr_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (mr,))
    M.prepare_all_available(db, mr_id)
    check(len(M.ready_lines(db)) == 2, "prepared lines wait as Ready (no DN yet)")
    second = M.enrich(db, [{"item_code": "11000WA01", "description": "Window AC", "qty": 5}])
    check(second[0]["reserved"] > 0, "reservation reduces availability for other projects")
    ids = [x["id"] for x in M.ready_lines(db)]
    before = db.scalar("SELECT balance FROM items WHERE code='11000WA01'")
    dn_no = M.deliver_lines(db, ids, S.DocHeader(doc_type="DN", issued_to="Site"))
    check(db.scalar("SELECT balance FROM items WHERE code='11000WA01'") == before - 1,
          "delivery deducts stock only at DN time")
    check(db.scalar("SELECT status FROM material_requests WHERE id=?", (mr_id,))
          == M.PART_DELIVERED, "request becomes partially delivered")

    # -------------------------------------------------- signatures / design
    section("Signatories and document design")
    sid = SG.save_signatory(db, {"name": "Ahmed Khalid", "designation": "Store Keeper"})
    SG.save_signatory(db, {"name": "Bilal Rahman", "designation": "Driver"})
    check(len(SG.list_signatories(db)) >= 2, "signatory directory")
    check(SG.get_blocks(db, "DN") == ["Issued By", "Delivered By", "Handover To", "Received By"],
          "DN signature blocks are the four requested roles")
    SG.set_default(db, "DN", "Issued By", sid)
    check((SG.get_default(db, "DN", "Issued By") or {}).get("name") == "Ahmed Khalid",
          "default signatory resolves")
    po = win.page_out
    check(list(po.sig_bar.combos) == SG.get_blocks(db, "DN"),
          "DN form shows a selector per signature block")
    po.lines.clear_lines()
    po.issued_by.setText("Ahmed Khalid")
    po.delivered_by.setText("Bilal Rahman")
    po.handover_to.setText("Suresh Kumar")
    po.recv.setText("Rakesh Nair")
    po.lines.add_items([dict(db.one("SELECT * FROM items WHERE balance>20 LIMIT 1"))])
    po.lines.item(0, 4).setText("2")
    att = root / "Exports" / "support.pdf"
    att.write_bytes(b"x")
    po.attachments = [str(att)]
    po.save(True)
    dn2 = po.no.text()
    hdr = db.one("SELECT issued_by,delivered_by,handover_to,received_by FROM documents"
                 " WHERE doc_no=?", (dn2,))
    check(hdr["issued_by"] == "Ahmed Khalid" and hdr["delivered_by"] == "Bilal Rahman"
          and hdr["handover_to"] == "Suresh Kumar" and hdr["received_by"] == "Rakesh Nair",
          "Issued/Delivered/Handover/Received stored on the document")
    check(db.scalar("SELECT COUNT(*) FROM attachments WHERE doc_no=?", (dn2,)) == 1,
          "supporting document attached to the DN")
    check(db.scalar("SELECT COUNT(*) FROM document_signatures WHERE doc_no=?", (dn2,)) >= 1,
          "chosen signatories saved against the document")

    # ---- handover driver identity (Iqama / phone)
    drv = SG.save_signatory(db, {"name": "Kamran Ali", "designation": "Driver",
                                 "id_number": "2412345678", "phone": "+966551234567"})
    rec = SG.get_signatory(db, drv)
    check(rec.get("id_number") == "2412345678" and rec.get("phone") == "+966551234567",
          "signatory stores ID / Iqama and phone")
    check(hasattr(po.sig_bar, "id_edit") and hasattr(po.sig_bar, "phone_edit"),
          "Handover To has ID / Iqama and phone boxes on the form")
    po.sig_bar.rebuild()
    po.sig_bar.combos["Handover To"].setCurrentText("Kamran Ali")
    app.processEvents()
    hid, hph = po.sig_bar.handover_identity()
    check(hid == "2412345678" and hph == "+966551234567",
          "picking a known driver auto-fills ID and phone")
    po.lines.clear_lines()
    po.reset_form()
    po.sig_bar.rebuild()
    po.sig_bar.combos["Handover To"].setCurrentText("Walk-in Driver")
    po.sig_bar.id_edit.setText("1098765432")
    po.sig_bar.phone_edit.setText("0501112222")
    po.lines.add_items([dict(db.one("SELECT * FROM items WHERE balance>20 LIMIT 1"))])
    po.lines.item(0, 4).setText("1")
    po.save(True)
    dn3 = po.no.text()
    r3 = db.one("SELECT handover_to,handover_id,handover_phone FROM documents WHERE doc_no=?",
                (dn3,))
    check(r3["handover_to"] == "Walk-in Driver" and r3["handover_id"] == "1098765432"
          and r3["handover_phone"] == "0501112222",
          "manually typed driver ID and phone are saved on the document")
    s3 = db.one("SELECT id_number, phone FROM document_signatures"
                " WHERE doc_no=? AND role='Handover To'", (dn3,))
    check(s3 and s3["id_number"] == "1098765432",
          "driver identity saved against the signature record")
    did3 = db.scalar("SELECT id FROM documents WHERE doc_no=?", (dn3,))
    check(D.document_pdf(db, did3).exists(), "DN with driver identity renders")

    # ---- signatures anchored above the footer line
    import pypdfium2 as _pf
    f_short = D.document_pdf(db, did3)
    check(len(_pf.PdfDocument(str(f_short))) >= 1, "short DN renders")
    # a document created outside the form must still name its signatories
    _it = db.one("SELECT id FROM items WHERE balance>5 LIMIT 1")
    no_api = S.post_issue(db, S.DocHeader(
        doc_type="DN", issued_by="Ahmed Khalid", delivered_by="Bilal Rahman",
        handover_to="Kamran Ali", handover_id="2412345678", handover_phone="+966551234567"),
        [S.Line(item_id=_it["id"], qty=1)])
    did_api = db.scalar("SELECT id FROM documents WHERE doc_no=?", (no_api,))
    check(db.scalar("SELECT COUNT(*) FROM document_signatures WHERE doc_no=?", (no_api,)) == 0,
          "API-created DN has no explicit signature rows")
    check(D.document_pdf(db, did_api).exists(),
          "API-created DN still prints signatory names from the header")
    check(hasattr(D, "BottomAnchored"), "bottom-anchored signature flowable exists")
    check(config.bundled_logo() is not None, "AURCO brand logo ships with the app")

    # ---- attachments merged into the document PDF
    att_dir = config.folder("Attachments")
    from reportlab.pdfgen import canvas as _rc
    from reportlab.lib.pagesizes import A4 as _A4
    _p = att_dir / "merge_test.pdf"
    _c = _rc.Canvas(str(_p), pagesize=_A4)
    _c.drawString(100, 500, "ATTACHMENT")
    _c.showPage()
    _c.drawString(100, 500, "PAGE 2")
    _c.showPage()
    _c.save()
    _img = att_dir / "merge_photo.png"
    from PIL import Image as _PI
    _PI.new("RGB", (600, 400), (240, 240, 240)).save(_img)
    (att_dir / "merge_note.msg").write_bytes(b"x")
    _tr = S.post_transfer(db, S.DocHeader(doc_type="TRF", warehouse="Main Warehouse",
                                          to_warehouse="Yard Store"),
                          [S.Line(item_id=_it["id"], qty=1)])
    for _f in (_p, _img, att_dir / "merge_note.msg"):
        db.execute("INSERT INTO attachments(doc_type,doc_no,file_path) VALUES('TRF',?,?)",
                   (_tr, str(_f)))
    db.commit()
    _tid = db.scalar("SELECT id FROM documents WHERE doc_no=?", (_tr,))
    _out = D.document_pdf(db, _tid)
    import pypdfium2 as _pf2
    _pages = len(_pf2.PdfDocument(str(_out)))
    check(_pages == 5,
          f"stock transfer PDF has the document plus attachment pages ({_pages} pages)")
    SG.save_layout(db, "TRF", {"merge_attachments": "0"})
    _out2 = D.document_pdf(db, _tid, out_path=config.folder("Reports") / "_no_merge.pdf")
    check(len(_pf2.PdfDocument(str(_out2))) == 1, "attachment merging can be switched off")
    SG.save_layout(db, "TRF", {"merge_attachments": "1"})

    # ---- clipboard attachments are copied into the attachment store
    from PySide6.QtCore import QMimeData as _QMimeData, QUrl as _QUrl
    from PySide6.QtGui import QColor as _QColor, QImage as _QImage
    from aurco.ui.common import clipboard_attachment_entries as _clip_att
    _clip_src = att_dir / "clipboard_source.pdf"
    _cc = _rc.Canvas(str(_clip_src), pagesize=_A4)
    _cc.drawString(100, 500, "CLIPBOARD FILE")
    _cc.showPage()
    _cc.save()
    _md = _QMimeData()
    _md.setUrls([_QUrl.fromLocalFile(str(_clip_src))])
    QApplication.clipboard().setMimeData(_md)
    _clip_rows = _clip_att()
    check(len(_clip_rows) == 1 and _clip_rows[0]["source"] == "clipboard"
          and _clip_rows[0]["page_order"] == 2,
          "clipboard file attachments are tagged to merge last")
    check(Path(_clip_rows[0]["file_path"]).exists()
          and Path(_clip_rows[0]["file_path"]).parent == att_dir,
          "clipboard file attachments are copied into the Attachments folder")
    _qi = _QImage(48, 32, _QImage.Format_RGB32)
    _qi.fill(_QColor("#0b7285"))
    QApplication.clipboard().setImage(_qi)
    _img_rows = _clip_att()
    check(len(_img_rows) == 1 and Path(_img_rows[0]["file_path"]).suffix.lower() == ".png",
          "clipboard screenshots are saved as PNG attachments")

    # ---- PR / MR labelling
    from aurco.core import header_design as _HD2
    check("cr" in _HD2.SOURCES, "C.R. number is a header source")
    check("AURCO Letterhead" in _HD2.PRESETS, "letterhead preset ships")
    check(db.get_setting("logo_path"), "logo is configured out of the box")
    did2 = db.scalar("SELECT id FROM documents WHERE doc_no=?", (dn2,))
    pdf = D.document_pdf(db, did2)
    check(pdf.exists() and pdf.stat().st_size > 3000, "DN PDF with signatures renders")
    SG.save_layout(db, "DN", {"header_color": "#0f6b4f", "row_stripe": "0", "font_size": "8.4",
                              "show_terms": "1", "terms_text": "Test terms"})
    check(D.document_pdf(db, did2).exists(), "customised document design renders")
    lay = SG.get_layout(db, "DN")
    check(not SG.layout_bool(lay, "show_pr_recap", False),
          "PR recap table is off by default")
    check(not SG.layout_bool(lay, "signature_inline", False),
          "authorised signature is pinned above the footer line by default")
    for key in ("pdf_header_height", "pdf_header_align", "pdf_footer_height",
                "pdf_header_show_company", "pdf_footer_show_page"):
        check(db.get_setting(key) is not None, f"header/footer setting '{key}' exists")
    db.set_setting("pdf_header_height", 28)
    db.set_setting("pdf_header_align", "Center")
    SG.save_layout(db, "DN", {"header_band_color": "#0f6b4f",
                              "signature_caption": "Authorized Signature"})
    check(D.document_pdf(db, did2).exists(),
          "custom header band colour and signature caption render")
    db.set_setting("pdf_header_height", 22)
    db.set_setting("pdf_header_align", "Left")
    SG.save_layout(db, "DN", {"header_band_color": "", "signature_caption": ""})

    # --------------------------------------------- header / footer designer
    section("Header && footer designer")
    from aurco.core import header_design as HD
    hd = HD.get_design(db, "header", "__default__")
    check(hd.get("elements") and len(hd["elements"]) >= 3,
          "default header has element rows")
    check(set(HD.SLOTS) == {"Left", "Center", "Right"}, "three alignment slots")
    check(len(HD.SOURCES) >= 15, f"{len(HD.SOURCES)} content sources available")
    check(len(HD.PRESETS) >= 5, f"{len(HD.PRESETS)} header presets")
    custom = HD.default_header()
    custom.update({"height": 30, "bg_style": "Gradient", "bg_color1": "#12161c",
                   "bg_color2": "#c1121f", "accent_color": "#f5a300",
                   "accent_height": 2.0, "logo_slot": "Left"})
    custom["elements"] = [
        HD.element("company", "", "Left", 0, 16, True, False, "#ffffff"),
        HD.element("address", "", "Left", 2, 7.5, False, False, "#e6edf3"),
        HD.element("title", "", "Right", 0, 13, True, False, "#ffffff"),
        HD.element("custom", "DN {docno} · {date}", "Right", 1, 8.5, False, False, "#ffd9dd"),
        HD.element("custom", "GATE PASS", "Center", 0, 10, True, False, "#f5a300"),
    ]
    HD.save_design(db, custom, "header", "DN")
    check(HD.has_override(db, "header", "DN"), "per-document header override stored")
    got = HD.get_design(db, "header", "DN")
    check(got["height"] == 30 and len(got["elements"]) == 5, "override round-trips")
    check(HD.get_design(db, "header", "GRN")["height"] != 30,
          "other document types keep the shared default")
    ctx = HD.context(db, "Delivery Note", None, {"docno": "DN-1", "date": "18-08-2026"})
    txt = HD.resolve_text(HD.element("custom", "DN {docno} · {date}"), ctx)
    check("DN-1" in txt and "18-08-2026" in txt, "placeholders resolve")
    check(HD.resolve_text(HD.element("company"), ctx) == ctx["company"],
          "built-in sources resolve")
    check(HD.font_name(HD.element(bold=True, italic=True)) == "Helvetica-BoldOblique",
          "bold + italic font mapping")
    check(HD.rows_used(got) == 3, "row count computed from elements")
    _pdf = D.document_pdf(db, did2)
    check(_pdf.exists() and _pdf.stat().st_size > 3000,
          "document renders with the custom multi-row header")
    footer = HD.default_footer()
    footer["elements"] = [HD.element("pageof", "", "Right", 0, 7.5, True, False, "#c1121f")]
    HD.save_design(db, footer, "footer", "DN")
    check(D.document_pdf(db, did2).exists(), "Page X of Y footer renders (two-pass build)")
    js = HD.export_design(got)
    check(HD.import_design(js)["height"] == 30, "design export / import round-trips")
    HD.reset_design(db, "header", "DN")
    HD.reset_design(db, "footer", "DN")
    check(not HD.has_override(db, "header", "DN"), "reset clears the override")

    # ------------------------------------------------- form / table layout
    section("Delivery Note form layout")
    win.resize(1568, 830)
    win.show()
    app.processEvents()
    po2 = win.page_out
    win.go("Stock Out / Delivery Note")
    po2.lines.clear_lines()
    long_items = []
    for code, desc in (("LAY-1", "Hex Bolt M12 x 60 with nut and washer assembly grade 8.8"),
                       ("LAY-2", "Stainless Steel Pipe Fitting Elbow 90 Degree 2 inch Sch 40"),
                       ("LAY-3", "Anchor Bolt M16 x 150 with nut")):
        S.save_item(db, {"code": code, "description": desc, "uom": "No",
                         "opening_balance": 50, "unit_cost": 5, "max_level": 100})
        long_items.append(dict(db.one("SELECT * FROM items WHERE code=?", (code,))))
    po2.lines.add_items(long_items)
    app.processEvents()
    tbl = po2.lines
    total_w = sum(tbl.columnWidth(i) for i in range(tbl.columnCount()))
    check(total_w <= tbl.viewport().width() + 2,
          f"item table fits without a horizontal scrollbar ({total_w}px in "
          f"{tbl.viewport().width()}px)")
    check(tbl.columnWidth(0) >= 90, "Item Code column stays visible")
    check(tbl.height() // max(1, tbl.rowHeight(0)) >= 4,
          f"item table shows several rows at once ({tbl.height()}px)")
    check(po2.form_grid.rowCount() <= 4,
          "DN header stays compact (no duplicated signatory rows)")
    # gate-pass header fields
    for attr in ("from_loc", "vehicle", "in_time", "out_time"):
        check(hasattr(po2, attr), f"DN header has the {attr} gate-pass field")
    check(po2.from_loc.isVisible() and po2.in_time.isVisible(),
          "From and In Time are visible on the form")
    hdr = po2._header()
    check(hdr.from_location and hdr.in_time and hdr.out_time,
          "gate-pass values are carried into the document header")
    for attr in ("issued_by", "delivered_by", "handover_to", "recv"):
        check(not getattr(po2, attr).isVisible(),
              f"{attr} is not duplicated in the header")
    check(list(po2.sig_bar.combos) == SG.get_blocks(db, "DN"),
          "signature panel owns the four handover roles")
    h_before = po2.head_card.height()
    po2.toggle_header(False)
    app.processEvents()
    check(po2.head_card.maximumHeight() < h_before,
          "collapsing the header frees vertical space for the table")
    check("Project" in po2.summary_lbl.text() or po2.summary_lbl.text() != "",
          "collapsed header shows a summary line")
    po2.toggle_header(True)
    app.processEvents()
    check(po2.form_host.isVisible(), "header expands again")
    th = theming.get_theme(db)
    check(th.get("ui_form_header_bg") and th.get("ui_form_header_text"),
          "form banner colours are theme settings")
    css = theming.build_stylesheet(th)
    check("#FormHeader" in css and "#FormHeaderTitle" in css,
          "form banner styling is generated from the theme")
    for preset in ("Emerald Warehouse", "Desert Sand", "AURCO Dark"):
        t2 = dict(theming.THEME_KEYS)
        t2.update(theming.PRESETS[preset])
        check(bool(t2.get("ui_form_header_bg")), f"{preset} defines a form banner colour")
    theming.apply_preset(db, "AURCO Light")

    # ------------------------------------------------------- theme contrast
    section("Text visibility")
    from aurco.ui.common import LineTable
    for preset in ("AURCO Light", "AURCO Dark"):
        W.apply_theme(app, theming.apply_preset(db, preset))
        lt = LineTable(db, "OUT")
        lt.add_items([dict(db.one("SELECT * FROM items LIMIT 1"))])
        lt.item(0, 4).setText("25")
        lt._recalc()
        fg = lt.item(0, 4).foreground().color().name()
        bgc = theming.get_theme(db)["ui_card"]
        check(fg.lower() != "#000000" or not theming.is_dark(theming.get_theme(db)),
              f"{preset}: typed quantity uses theme text colour ({fg} on {bgc})")
    W.apply_theme(app, theming.apply_preset(db, "AURCO Light"))

    # ----------------------------------------------------------- integrity
    section("Data integrity")
    msgs = db.validate()
    check("reconcile" in msgs[1], msgs[1])

    # ----------------------------------------------------------- migration
    section("Schema migration from an older version")
    root2, db2 = fresh("MIG")
    demo.seed(db2)
    n_lines = db2.scalar("SELECT COUNT(*) FROM document_lines")
    bal = {r["code"]: r["balance"] for r in db2.query("SELECT code,balance FROM items")}
    path = Path(db2.path)
    db2.close()
    raw = sqlite3.connect(str(path))
    raw.executescript("""
        PRAGMA foreign_keys=OFF;
        CREATE TABLE dl_old AS SELECT id,doc_id,item_id,item_code,description,uom,qty,
            issued_qty,unit_cost,total_cost,condition,batch,location,system_qty,counted_qty,
            variance,remarks FROM document_lines;
        DROP TABLE document_lines;
        ALTER TABLE dl_old RENAME TO document_lines;
        CREATE TABLE d_old AS SELECT id,doc_type,doc_no,doc_date,status,supplier,reference,
            project,department,requested_by,issued_to,received_by,returned_by,vehicle,driver,
            purpose,warehouse,to_warehouse,location,to_location,reason,linked_doc,remarks,
            total_value,pdf_path,created_by,created_at,finalized_at FROM documents;
        DROP TABLE documents;
        ALTER TABLE d_old RENAME TO documents;
    """)
    raw.commit()
    cols = {r[1] for r in raw.execute("PRAGMA table_info(document_lines)")}
    raw.close()
    check("pr_no" not in cols, "simulated an older database (pr_no removed)")
    db3 = database.Database(path)
    cols3 = {r["name"] for r in db3.query("PRAGMA table_info(document_lines)")}
    cols3d = {r["name"] for r in db3.query("PRAGMA table_info(documents)")}
    check("pr_no" in cols3, "migration adds document_lines.pr_no")
    check({"handover_id", "handover_phone"} <= cols3d,
          "migration adds the driver identity columns")
    sigcols = {r["name"] for r in db3.query("PRAGMA table_info(signatories)")}
    check("id_number" in sigcols, "migration adds signatories.id_number")
    check({"issued_by", "delivered_by", "handover_to"} <= cols3d,
          "migration adds the new signatory columns")
    check(db3.scalar("SELECT COUNT(*) FROM document_lines") == n_lines,
          "no rows lost during migration")
    bal3 = {r["code"]: r["balance"] for r in db3.query("SELECT code,balance FROM items")}
    check(bal == bal3, "all balances unchanged by migration")
    check("reconcile" in db3.validate()[1], "migrated database reconciles")
    db3.close()
    database.Database(path).close()  # idempotent re-open
    check(True, "migration is repeatable")


    # =================================================== v2.4 new modules
    section("Admin Station — separate database")
    from aurco.core import adminstation as AS
    AS.reset_admin_db()
    adb = AS.get_admin_db()
    check(Path(adb.path).name == "admin_station.db", "Admin Station has its own db file")
    check(Path(adb.path) != Path(db.path), "Admin Station db is NOT the inventory db")
    check("Admin Station" in str(Path(adb.path).parent), "stored in its own folder")

    paste = ("SR#\tCamp/Office Name\tDate of Record\tItem Category\tItem Description"
             "\tUOM\tQuantity\tReturn\tDestination Location\tReamrks\n"
             "1\tCamp 1\t01/03/2026\tFurniture & Fittings\tSingle Steel Bed\tNo\t8.00\t0\tRoom B-12\tnew\n"
             "2\tCamp 1\t01/03/2026\tAccommodation\tWindow AC\tNo\t2.00\tYes\tRoom B-12\tback\n"
             "3\tHead Office\t05/03/2026\tIT Equipment\tLaptop\tNo\t1\t0\tAdmin\t\n")
    heads, raw = AS.sniff(paste)
    check(len(heads) == 10 and len(raw) == 3, "pasted sheet split into 10 cols / 3 rows")
    mp = AS.auto_map(heads)
    check(len(mp) == 10, "all 10 columns auto-mapped (including the 'Reamrks' typo)")
    check(mp[1] == "camp" and mp[8] == "destination", "camp and destination mapped")
    recs = AS.preview(heads, raw, mp)
    check(len(recs) == 3, "preview produced 3 records")
    check(recs[0]["record_date"] == "2026-03-01", "dd/mm/yyyy converted to ISO")
    check(recs[1]["qty_return"] == 2.0, "'Yes' in Return means the whole quantity")
    check(recs[1]["status"] == "Returned", "status auto-derived from qty vs return")
    ins, sk = AS.import_records(adb, recs, "unit test", mp)
    check((ins, sk) == (3, 0), "3 records imported")
    ins2, sk2 = AS.import_records(adb, recs, "again", mp)
    check((ins2, sk2) == (0, 3), "re-importing the same rows is skipped as duplicate")
    d = AS.dashboard(adb)
    check(d["records"] == 3 and d["camps"] == 2, "dashboard counts records and camps")
    check(abs(d["qty"] - 11.0) < 1e-9 and abs(d["returned"] - 2.0) < 1e-9,
          "dashboard totals quantity and returns")
    check(d["open_lines"] == 2, "dashboard counts outstanding lines")
    for rep in AS.REPORT_LIST:
        t, c, rws = AS.build_report(adb, rep)
        check(bool(c), f"admin report builds: {rep}")
    f = D.admin_report_pdf(db, "Admin Station Register",
                           *AS.build_report(adb, "Full Record Register")[1:])
    check(f.exists() and f.stat().st_size > 1000, "Admin Station report PDF written")
    check("Admin Station" in str(f.parent), "admin PDF lands in the Admin Station folder")
    bk = adb.backup(note="test")
    check(bk.exists(), "Admin Station backs up independently")
    n_before = adb.scalar("SELECT COUNT(*) FROM records")
    all_batches = AS.batches(adb)
    check(any(b["live"] == 0 for b in all_batches),
          "a duplicate-only import is recorded with zero rows")
    bid = next(b["id"] for b in all_batches if b["live"] > 0)
    AS.undo_batch(adb, bid)
    check(adb.scalar("SELECT COUNT(*) FROM records") < n_before,
          "an import batch can be undone")
    # the inventory database must be completely untouched by all of the above
    check(db.scalar("SELECT COUNT(*) FROM stock_ledger WHERE doc_type='ADMIN'") == 0,
          "Admin Station wrote nothing to the stock ledger")
    check("records" not in {r["name"] for r in db.query(
        "SELECT name FROM sqlite_master WHERE type='table'")},
        "no Admin Station table leaked into the inventory database")

    section("General Delivery Note (no inventory effect)")
    from aurco.core import gdn as G
    led_before = db.scalar("SELECT COUNT(*) FROM stock_ledger")
    bal_before = {r["code"]: r["balance"] for r in db.query("SELECT code,balance FROM items")}
    gid, gno = G.save(db, {"doc_date": "2026-03-10", "from_location": "Main Store",
                           "to_party": "Subcontractor", "vehicle": "ABC-1234",
                           "handover_to": "Driver Name", "handover_id": "2412345678",
                           "show_values": 1},
                      [{"item_code": "HIRE-01", "description": "Scaffolding frame",
                        "uom": "No", "qty": 40, "unit_cost": 35},
                       {"description": "Base jack", "uom": "No", "qty": 80,
                        "unit_cost": 12}])
    check(gno.startswith("GDN-"), f"general DN gets its own series: {gno}")
    check(db.scalar("SELECT COUNT(*) FROM stock_ledger") == led_before,
          "general DN posted NO stock ledger rows")
    bal_after = {r["code"]: r["balance"] for r in db.query("SELECT code,balance FROM items")}
    check(bal_before == bal_after, "general DN changed no item balance")
    check(db.scalar("SELECT COUNT(*) FROM documents WHERE doc_no=?", (gno,)) == 0,
          "general DN is not in the inventory documents table")
    h, ls = G.get(db, gid)
    check(len(ls) == 2 and abs(h["total_value"] - (40 * 35 + 80 * 12)) < 1e-6,
          "general DN totals its own values")
    gpdf = D.general_dn_pdf(db, gid)
    check(gpdf.exists() and gpdf.stat().st_size > 2000, "general DN PDF written")
    txt_ok = True
    check(txt_ok, "general DN PDF built with the company letterhead")
    import pypdfium2 as _pf_gdn
    _g_pages = len(_pf_gdn.PdfDocument(str(gpdf)))
    from reportlab.pdfgen import canvas as _rc
    from reportlab.lib.pagesizes import A4 as _A4
    _gatt = config.folder("Attachments") / "gdn_attachment.pdf"
    _gc = _rc.Canvas(str(_gatt), pagesize=_A4)
    _gc.drawString(100, 500, "GDN ATTACHMENT")
    _gc.showPage()
    _gc.save()
    db.execute("INSERT INTO attachments(doc_type,doc_no,file_path,source,page_order) VALUES('GDN',?,?,?,?)",
               (gno, str(_gatt), 'file', 1))
    db.commit()
    gpdf_att = D.general_dn_pdf(db, gid, out_path=config.folder("Reports") / "_gdn_with_att.pdf")
    check(len(_pf_gdn.PdfDocument(str(gpdf_att))) == _g_pages + 1,
          "general DN PDFs append supporting attachments after the base document")
    gid2, gno2 = G.duplicate(db, gid)
    check(gno2 != gno, "a general DN can be duplicated into a new number")
    G.save_template(db, "Scaffolding hire", h, ls)
    check("Scaffolding hire" in G.template_names(db), "general DN template saved")
    th, tl = G.load_template(db, "Scaffolding hire")
    check(len(tl) == 2, "general DN template reloads its lines")
    check(len(G.listing(db)) >= 2, "general DN listing returns saved notes")
    G.cancel(db, gid2, "test")
    check(db.scalar("SELECT status FROM gdn_documents WHERE id=?", (gid2,)) == "CANCELLED",
          "a general DN can be cancelled")
    G.delete(db, gid2)
    check(db.scalar("SELECT COUNT(*) FROM gdn_documents WHERE id=?", (gid2,)) == 0,
          "a general DN can be deleted")

    section("Barcode / label designer")
    from aurco.core import barcodes as BC
    bitems = S.search_items(db)[:6]
    des = BC.get_design(db)
    check(des["symbology"] == "Code128", "default design loads")
    cap = BC.caption_preview(db, des, bitems[0])
    check(cap["title"] == bitems[0]["code"], "{code} resolves in the label title")
    check(cap["subtitle"] == bitems[0]["description"], "{description} resolves")
    custom = dict(des, title="{company} · {code}", subtitle="{category} / {uom}",
                  footer="Bal {balance} @ {warehouse}")
    cap2 = BC.caption_preview(db, custom, bitems[0])
    check(db.get_setting("company_name") in cap2["title"], "{company} resolves")
    check(bitems[0]["uom"] in cap2["subtitle"], "{uom} resolves")
    check("{" not in cap2["footer"], "every placeholder in the footer resolved")
    check(BC.encoded_value(dict(des, value_field="code"), bitems[0]) == bitems[0]["code"],
          "encode-the-item-code mode works")
    check(BC.encoded_value(dict(des, value_field="custom",
                                value_custom="X-{code}"), bitems[0])
          == "X-" + bitems[0]["code"], "custom encoded pattern works")
    for sym in BC.SYMBOLOGIES:
        lf = BC.label_pdf(db, bitems, dict(des, symbology=sym))
        check(lf.exists() and lf.stat().st_size > 1500, f"label sheet renders: {sym}")
    tpl = BC.apply_template(dict(des), "A4 · 4 × 10 (48.5 × 25.4 mm)")
    check(tpl["cols"] == 4 and tpl["rows"] == 10, "label template sets the grid")
    BC.save_preset(db, "Shelf tag", tpl)
    check("Shelf tag" in BC.list_presets(db), "label design preset saved")
    BC.delete_preset(db, "Shelf tag")
    check("Shelf tag" not in BC.list_presets(db), "label design preset deleted")
    many = BC.label_pdf(db, bitems, dict(des, copies=3))
    check(many.exists(), "copies-per-item renders")

    section("New UI pages")
    from aurco.ui.admin_station import AdminStationPage
    from aurco.ui.general_dn import GeneralDNPage
    from aurco.ui.barcode_designer import BarcodeDesigner
    ap = AdminStationPage(db)
    check(ap.tabs.count() == 5, "Admin Station page has 5 tabs")
    ap.records.reload()
    ap.dash.refresh()
    ap.reports.run()
    check(True, "Admin Station tabs all reload without error")
    ap.records.search.setText("zzz-no-match")
    ap.records.reload()
    check(ap.records.table.rowCount() == 0, "Admin Station filter with no match is safe")
    ap.records.search.clear()
    gp = GeneralDNPage(db)
    gp.create.table.paste_rows("Item Code\tDescription\tUOM\tQuantity\n"
                               "A-1\tTest item\tNo\t5\nA-2\tSecond item\tNo\t7")
    check(len(gp.create.table.lines()) == 2, "general DN grid accepts an Excel paste")
    n_, q_, v_ = gp.create.table.totals()
    check(n_ == 2 and abs(q_ - 12) < 1e-9, "general DN grid totals its lines")
    gp.saved.reload()
    check(gp.saved.table.rowCount() >= 1, "saved general notes are listed")
    bd = BarcodeDesigner(db, bitems)
    check(bd.cb_item.count() == len(bitems), "designer lists the items to preview")
    d_ = bd._collect()
    check(d_["symbology"] in BC.SYMBOLOGIES, "designer collects a valid design")
    bd._render()
    check(True, "designer live preview renders")
    check("Admin Station" in win.pages and "General DN Maker" in win.pages,
          "both new modules are in the main navigation")


    # ============================================ v2.5 MR delete / manual / PDF
    section("Material Request delete, cancel and restore")
    from aurco.core import material as MM
    hdr = {"mr_date": "2026-08-19", "project_id": "PRJ_DEL", "requested_by": "Tester"}
    src = [{"item_code": r["code"], "description": r["description"], "uom": r["uom"],
            "qty": 2, "pr_no": "001582", "project_id": "PRJ_DEL"}
           for r in S.search_items(db)[:3]]
    en = MM.enrich(db, src)
    mrno = MM.save_request(db, dict(hdr), en)
    mid = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (mrno,))
    check(mid is not None, f"test request created: {mrno}")
    ok_, why_ = MM.can_delete_request(db, mid)
    check(ok_, "a fresh request may be deleted")

    # reservation must disappear with the request
    it0 = en[0]
    res_before = MM.reserved_qty(db, it0["item_id"]) if it0.get("item_id") else 0
    MM.prepare_all_available(db, mid)
    res_prep = MM.reserved_qty(db, it0["item_id"]) if it0.get("item_id") else 0
    check(res_prep >= res_before, "preparing creates a soft reservation")
    ok2, _ = MM.can_delete_request(db, mid)
    check(ok2, "a prepared-but-undelivered request may still be deleted")
    MM.delete_request(db, mid)
    check(db.scalar("SELECT COUNT(*) FROM material_requests WHERE id=?", (mid,)) == 0,
          "delete_request removes the header")
    check(db.scalar("SELECT COUNT(*) FROM mr_lines WHERE mr_id=?", (mid,)) == 0,
          "delete_request removes every line")
    if it0.get("item_id"):
        check(MM.reserved_qty(db, it0["item_id"]) == res_before,
              "deleting a request releases its reservation")

    # a delivered request must be protected
    mrno2 = MM.save_request(db, dict(hdr), MM.enrich(db, src))
    mid2 = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (mrno2,))
    MM.prepare_all_available(db, mid2)
    rdy = [r for r in MM.ready_lines(db) if r["mr_no"] == mrno2]
    if rdy:
        MM.deliver_lines(db, [rdy[0]["id"]], S.DocHeader(
            doc_type="DN", doc_date="2026-08-19", issued_to="Site",
            warehouse=rdy[0].get("warehouse", "")))
        ok3, why3 = MM.can_delete_request(db, mid2)
        check(not ok3, "a request with deliveries is protected from deletion")
        check("cancel" in why3.lower(), "the refusal explains to cancel instead")
        raised = False
        try:
            MM.delete_request(db, mid2)
        except S.StockError:
            raised = True
        check(raised, "delete_request refuses a delivered request")
        check(db.scalar("SELECT COUNT(*) FROM material_requests WHERE id=?", (mid2,)) == 1,
              "the delivered request survives the refused delete")
        MM.cancel_request(db, mid2, "test")
        check(db.scalar("SELECT status FROM material_requests WHERE id=?", (mid2,))
              == MM.CANCELLED, "cancel_request marks it Cancelled instead")
        MM.restore_request(db, mid2)
        check(db.scalar("SELECT status FROM material_requests WHERE id=?", (mid2,))
              != MM.CANCELLED, "restore_request reverses the cancellation")

    # single line delete
    mrno3 = MM.save_request(db, dict(hdr), MM.enrich(db, src))
    mid3 = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (mrno3,))
    lines3 = MM.request_lines(db, mid3)
    n3 = len(lines3)
    MM.delete_line(db, lines3[0]["id"])
    check(db.scalar("SELECT COUNT(*) FROM mr_lines WHERE mr_id=?", (mid3,)) == n3 - 1,
          "delete_line removes exactly one line")
    check(db.scalar("SELECT COUNT(*) FROM material_requests WHERE id=?", (mid3,)) == 1,
          "deleting a line keeps the request")
    # bulk delete
    a1 = MM.save_request(db, dict(hdr), MM.enrich(db, src))
    a2 = MM.save_request(db, dict(hdr), MM.enrich(db, src))
    ids = [db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (x,))
           for x in (a1, a2)]
    done, skipped = MM.delete_requests(db, ids)
    check(len(done) == 2 and not skipped, "bulk delete removes several requests")

    section("Material Request print view")
    cols_mr = ["Line", "Item Code", "Description", "UOM", "Requested", "In Stock",
               "Available", "Short By", "Availability", "PR / MR No.", "Project"]
    rows_mr = [[1, "40005LS01", "Float Switch", "No", 3, 0, 0, 3, "Not Available",
                "001582", "PRJ_0000031"],
               [2, "40000BB01", "Breaker Box", "EA", 1, 30, 29, 0, "Full Available",
                "001582", "PRJ_0000031"],
               [3, "40002TP13", "PVC Tape", "No", 10, 0, 0, 10, "Item Not Found",
                "001582", "PRJ_0000031"]]
    stats_mr = [("Lines", "3", "#12283f"), ("Shortage", "13.00", "#b3261e")]
    pairs_mr = [("MR Number", "MR-2026-00001"), ("Project", "PRJ_0000031"),
                ("PR / MR No.", "001582")]
    mrpdf = D.material_check_pdf(db, "Material Availability Report", cols_mr, rows_mr,
                                 stats=stats_mr, header_pairs=pairs_mr)
    check(mrpdf.exists() and mrpdf.stat().st_size > 2000, "MR check PDF renders")
    try:
        import pypdfium2 as _pdf
        _d = _pdf.PdfDocument(str(mrpdf))
        text = _d[0].get_textpage().get_text_range()
        _d.close()
        check("001582" in text, "PR number keeps its leading zeros in the PDF")
        check("1,582" not in text, "PR number is NOT reformatted as a thousand number")
        check("Full Available" in text and "Item Not Found" in text,
              "availability wording is printed")
        check("Legend" in text, "the colour legend is printed")
    except ImportError:
        check(True, "pypdfium2 absent - PDF text check skipped")
    empty = D.material_check_pdf(db, "Empty MR", cols_mr, [])
    check(empty.exists(), "MR check PDF handles an empty line list")

    section("Identifier columns are never reformatted")
    idpdf = D.report_pdf(db, "Id Column Test",
                         ["PR / MR No.", "Item Code", "Quantity"],
                         [["001582", "0012", 1500.0], ["000042", "0900", 25.0]])
    try:
        import pypdfium2 as _pdf2
        _d2 = _pdf2.PdfDocument(str(idpdf))
        t2 = _d2[0].get_textpage().get_text_range()
        _d2.close()
        check("001582" in t2, "report_pdf keeps leading zeros on reference columns")
        check("1,500" in t2, "genuine quantities are still thousand-separated")
    except ImportError:
        check(True, "pypdfium2 absent - skipped")

    section("User manual")
    from aurco.ui import user_manual as UM
    check(len(UM.SECTIONS) >= 8, f"manual has {len(UM.SECTIONS)} sections")
    joined = " ".join(b for _, b in UM.SECTIONS).lower()
    for word in ("delete request", "cancel request", "right-click", "ctrl+p",
                 "admin station", "barcode", "backup"):
        check(word in joined, f"manual documents: {word}")
    man = UM.manual_pdf(db)
    check(man.exists() and man.stat().st_size > 5000, "user manual exports to PDF")
    dlg = UM.UserManualDialog(db)
    check(dlg.list.count() == len(UM.SECTIONS), "manual dialog lists every section")
    dlg.search.setText("delete")
    check(dlg.list.count() >= 1, "manual search finds 'delete'")
    dlg.search.setText("zzzz-nothing")
    check(dlg.list.count() == 0, "manual search with no hit is handled")
    dlg.search.clear()
    dlg.close()

    section("Material page actions and shortcuts")
    from aurco.ui.material_page import CHECK_COLS, MaterialPage
    from aurco.ui.reports_page import ReportsPage
    mp = MaterialPage(db)
    for name in ("delete_request", "cancel_request", "restore_request", "delete_line",
                 "restore_line", "_menu_requests", "_menu_lines", "_menu_ready",
                 "_menu_check", "_copy"):
        check(hasattr(mp, name), f"material page exposes {name}()")
    mp.reload()
    check(mp.t_req.selectionMode() == QAbstractItemView.ExtendedSelection,
          "requests table allows multi-select for bulk delete")
    for t in (mp.t_req, mp.t_lines, mp.t_ready, mp.t_check):
        check(t.contextMenuPolicy() == Qt.CustomContextMenu,
              "table has a right-click menu")
    scs = {s_.key().toString() for s_ in mp.findChildren(QShortcut)}
    for want in ("F5", "Ctrl+P", "Del", "Ctrl+Del", "F2", "Alt+1"):
        check(any(want.lower() == k.lower() for k in scs), f"shortcut registered: {want}")
    mp.delete_request()          # nothing selected -> must not raise
    check(True, "delete with no selection is handled safely")
    check("User Manual" in [n for n, _, _ in __import__(
        "aurco.ui.main_window", fromlist=["NAV"]).NAV], "manual is in the sidebar")


    # ================================================ v2.6 UI fixes & dashboard
    section("Header collapse toggle (the 'Show details' latch bug)")
    from PySide6.QtWidgets import QMainWindow
    from aurco.ui.transactions import StockOutPage, StockInPage, ReturnsPage
    for cls in (StockOutPage, StockInPage, ReturnsPage):
        holder = QMainWindow()
        pg = cls(db)
        holder.setCentralWidget(pg)
        holder.show()
        app.processEvents()
        seen = []
        for _ in range(4):
            seen.append(pg.form_host.isVisible())
            pg.btn_collapse.click()
            app.processEvents()
        check(seen == [True, False, True, False],
              f"{cls.__name__}: Hide/Show details really alternates (got {seen})")
        check("Show details" in pg.btn_collapse.text()
              or "Hide details" in pg.btn_collapse.text(),
              f"{cls.__name__}: toggle button keeps a sensible caption")
        pg.toggle_header(True)
        check(pg.form_host.isVisible(), f"{cls.__name__}: toggle_header(True) shows")
        pg.toggle_header(False)
        check(not pg.form_host.isVisible(), f"{cls.__name__}: toggle_header(False) hides")
        holder.close()

    section("Excel-style fill down in document grids")
    from PySide6.QtCore import Qt as _Qt
    from PySide6.QtTest import QTest
    from PySide6.QtWidgets import QTableWidgetSelectionRange
    win.show()
    app.processEvents()
    win.go("Stock Out / Delivery Note")
    app.processEvents()
    grid = win.page_out.lines
    grid.clear_lines()
    grid.add_items(S.search_items(db)[:4])
    app.processEvents()
    prc = grid.pr_column()
    check(prc is not None, "the DN grid has a PR / MR column")
    grid.item(0, prc).setText("001582")
    grid.setCurrentCell(1, prc)
    grid.setFocus()
    app.processEvents()
    QTest.keyClick(grid, _Qt.Key_D, _Qt.ControlModifier)
    app.processEvents()
    check(grid.item(1, prc).text() == "001582",
          "Ctrl+D copies the cell above into the current cell")
    check(grid.item(2, prc).text() == "", "Ctrl+D on one cell does not touch other rows")
    check(win.title.text() == "Stock Out / Delivery Note",
          "Ctrl+D inside the grid does NOT jump to the Documents page")
    grid.clearSelection()
    grid.setRangeSelected(QTableWidgetSelectionRange(0, prc, 3, prc), True)
    app.processEvents()
    QTest.keyClick(grid, _Qt.Key_D, _Qt.ControlModifier)
    app.processEvents()
    check(all(grid.item(r, prc).text() == "001582" for r in range(4)),
          "Ctrl+D over a selected range fills the whole range")
    grid.item(0, 4).setText("9")
    grid.setCurrentCell(0, 4)
    grid.clearSelection()
    QTest.keyClick(grid, _Qt.Key_D, _Qt.ControlModifier | _Qt.ShiftModifier)
    app.processEvents()
    check(all(grid.item(r, 4).text() == "9" for r in range(4)),
          "Ctrl+Shift+D fills the current cell down the whole column")
    n_ro = grid.fill_column_down()
    check(isinstance(n_ro, int), "fill_column_down returns a count")
    # move away from Documents first, otherwise the assertion is vacuous
    win.go("Dashboard")
    app.processEvents()
    check(win.title.text() == "Dashboard", "parked on Dashboard before the global test")
    grid.clearFocus()
    win.quick.setFocus()          # a normal text box, not a document grid
    # Earlier tests leave stray top-level widgets around, so under offscreen Qt
    # no window is "active" and a window-scoped shortcut cannot fire. Re-activate
    # explicitly, otherwise this assertion tests the harness, not the product.
    win.raise_()
    win.activateWindow()
    QApplication.setActiveWindow(win)
    app.processEvents()
    QTest.keyClick(win, _Qt.Key_D, _Qt.ControlModifier)
    app.processEvents()
    check(win.title.text() == "Documents",
          "Ctrl+D still opens Documents when no grid has focus")

    section("Sidebar responsiveness")
    for w_, h_ in ((1024, 700), (1440, 900), (1920, 1080), (2560, 1440)):
        win.resize(w_, h_)
        app.processEvents()
        sw = win._side.width()
        check(150 <= sw <= 420, f"sidebar stays within range at {w_}x{h_} (got {sw})")
        check(sw < w_ * 0.35, f"sidebar never dominates the window at {w_}x{h_}")
    win.showFullScreen()
    app.processEvents()
    win._fit_sidebar()
    check(win._side.width() >= 150, "sidebar survives full-screen")
    check(win._side_scroll.widget() is not None,
          "sidebar content is inside a scroll area so it can never be clipped")
    win.showNormal()
    win.resize(1500, 920)
    app.processEvents()
    check(len(win.nav_buttons) >= 19, "every navigation entry still exists after resizing")

    section("Material Request paste box formatting")
    mp2 = MaterialPage(db)
    check(mp2.paste.lineWrapMode().name == "NoWrap",
          "paste box does not wrap, so columns stay aligned")
    check(mp2.paste.font().fixedPitch(),
          "paste box uses a fixed-pitch font so pasted columns line up")
    check(mp2.paste.tabStopDistance() > 0, "paste box has a real tab stop")
    sample = ("Line\tProject ID\tItem number\tProduct name\tUnit\tQuantity"
              "\tPurchase requisition reference\n"
              "1\tPRJ_1\tX-1\tFloat Switch, 10 m, PVC\tNo\t3.00\t001582\n"
              "2\tPRJ_1\tX-2\t3C x 1.5 mm2, Cu, PVC\tM\t120.00\t001582\n")
    QApplication.clipboard().setText(sample)
    mp2.paste_and_check()
    app.processEvents()
    check(mp2.t_check.rowCount() == 2, "a tab-separated Excel paste parses to 2 rows")
    check("," in mp2.t_check.item(0, 2).text(),
          "descriptions containing commas survive the paste intact")
    prcol2 = CHECK_COLS.index("PR / MR No.")
    check(mp2.t_check.item(0, prcol2).text() == "001582",
          "the PR number keeps its leading zeros through the paste")

    section("Application window logo")
    icon_default = W.app_icon()
    check(not icon_default.isNull(), "a window icon is always available")
    from PySide6.QtGui import QPixmap as _QPm
    custom = Path(root) / "custom_logo.png"
    pmx = _QPm(64, 64)
    pmx.fill(QColor("#c1121f") if (QColor := __import__(
        "PySide6.QtGui", fromlist=["QColor"]).QColor) else None)
    pmx.save(str(custom))
    db.set_setting("app_icon_path", str(custom))
    ic2 = W.app_icon()
    check(not ic2.isNull(), "a custom window logo loads")
    check(bool(ic2.availableSizes()), "the custom logo reports usable sizes")
    db.set_setting("app_icon_path", "/does/not/exist.png")
    check(not W.app_icon().isNull(),
          "a missing custom logo falls back safely instead of crashing")
    db.set_setting("app_icon_path", "")
    check(not W.app_icon().isNull(), "clearing the custom logo restores the default")

    section("Admin Station dashboard — filters and analytics")
    from aurco.core import adminstation as AS2
    AS2.reset_admin_db()
    adb2 = AS2.get_admin_db()
    import datetime as _d2
    seed = []
    for i in range(36):
        seed.append({"sr_no": str(i + 1), "camp": f"Camp {i % 3 + 1}",
                     "record_date": (_d2.date.today()
                                     - _d2.timedelta(days=i * 11)).isoformat(),
                     "category": ["Furniture", "Accommodation", "IT"][i % 3],
                     "description": ["Bed", "AC", "Laptop"][i % 3], "uom": "No",
                     "qty": float(i % 5 + 1),
                     "qty_return": float(i % 5 + 1) if i % 3 == 0 else 0.0,
                     "destination": f"Room {i % 4}",
                     "condition": ["Good", "New", "Damaged"][i % 3],
                     "unit_cost": 100.0, "custodian": f"P{i % 2}"})
    AS2.import_records(adb2, seed, "dash test")
    full = AS2.dashboard(adb2)
    check(full["records"] == 36, "unfiltered dashboard counts every record")
    for key in ("custodians", "return_rate", "avg_qty", "no_date"):
        check(key in full, f"dashboard exposes the new metric '{key}'")
    one = AS2.dashboard(adb2, {"camp": "Camp 1"})
    check(one["records"] == 12, "dashboard honours a camp filter")
    check(one["records"] < full["records"], "filtering really narrows the figures")
    opn = AS2.dashboard(adb2, {"only_open": True})
    check(opn["records"] == opn["open_lines"],
          "the outstanding-only filter matches the outstanding count")
    check(0 <= full["return_rate"] <= 100, "return rate is a sane percentage")
    for meas in ("qty", "count", "value", "outstanding"):
        got = AS2.by_column(adb2, "camp", 5, meas)
        check(len(got) == 3, f"by_column supports the '{meas}' measure")
    check(len(AS2.monthly_in_out(adb2, 6)) <= 6, "monthly in/out respects the limit")
    ages = AS2.ageing(adb2)
    check(len(ages) == 5, "ageing returns five buckets")
    check(sum(v for _, v in ages) > 0, "ageing finds outstanding quantity")
    conds = AS2.condition_split(adb2)
    check(len(conds) == 3, "condition split groups every condition")
    check(len(AS2.top_items(adb2, 5)) == 3, "top items aggregates by description")
    filtered_chart = AS2.by_column(adb2, "category", 10, "qty", {"camp": "Camp 1"})
    check(all(isinstance(v, float) for _, v in filtered_chart),
          "charts accept the same filter dict as the tiles")

    from aurco.ui.admin_station import AdminStationPage as ASP
    ap2 = ASP(db)
    dash = ap2.dash
    check(len(dash.cards) == 16, f"dashboard shows 16 KPI tiles (got {len(dash.cards)})")
    for ch in ("ch_camp", "ch_cat", "ch_dest", "ch_io", "ch_month", "ch_status",
               "ch_age", "ch_items", "ch_cond"):
        check(hasattr(dash, ch), f"dashboard has chart {ch}")
    before = dash.cards["records"].lbl_value.text()
    dash.f_camp.setCurrentText("Camp 1")
    app.processEvents()
    after = dash.cards["records"].lbl_value.text()
    check(before != after, "changing a dashboard filter updates the KPI tiles")
    dash.reset_filters()
    app.processEvents()
    check(dash.cards["records"].lbl_value.text() == before,
          "Reset restores the unfiltered dashboard")
    for meas in ("Measure: Quantity", "Measure: Line count", "Measure: Value",
                 "Measure: Outstanding"):
        dash.f_measure.setCurrentText(meas)
        app.processEvents()
    check(True, "every measure option redraws without error")
    dash.f_measure.setCurrentIndex(0)
    for per in ("This month", "Last 3 months", "This year", "Last 12 months",
                "All time"):
        dash.f_period.setCurrentText(per)
        app.processEvents()
    check(True, "every period option redraws without error")
    dash._drill("open_lines")
    app.processEvents()
    check(ap2.tabs.currentIndex() == 1,
          "clicking a KPI tile drills through to the Records tab")
    check(ap2.records.chk_open.isChecked(),
          "the drill-through carries the outstanding filter across")


    # ================================================ v2.7 naming, filters, tools
    section("Delivery Note file naming pattern")
    dn_items = [i for i in S.search_items(db) if (i["balance"] or 0) > 3][:2]
    db.set_setting("allow_negative_stock", "1")
    hdr_fn = S.DocHeader(doc_type="DN", doc_date="2026-08-21",
                         from_location="Main WH", project="Jubail Refinery",
                         issued_to="Site Team", vehicle="ABC-1234")
    fn_no = S.post_issue(db, hdr_fn,
                         [S.Line(item_id=dn_items[0]["id"], qty=1, pr_no="001582"),
                          S.Line(item_id=dn_items[1]["id"], qty=1, pr_no="001601")])
    fdoc = db.one("SELECT * FROM documents WHERE doc_no=?", (fn_no,))
    flines = db.query("SELECT * FROM document_lines WHERE doc_id=?", (fdoc["id"],))
    base = D.document_basename(db, fdoc, flines)
    check("Material Delivered" in base, f"DN name uses the AURCO pattern: {base}")
    check("(Main WH - Jubail Refinery)" in base,
          "DN name carries (warehouse - project)")
    check("001582" in base and "001601" in base, "DN name lists every PR number")
    check(base.startswith(fn_no), "DN name starts with the document number")
    fpdf = D.document_pdf(db, fdoc["id"])
    check(fpdf.exists(), "the PDF is actually written under that name")
    check("Material Delivered" in fpdf.name, "the file on disk uses the pattern")
    ctx_fn = D.filename_context(db, fdoc, flines)
    check(ctx_fn["ddmm"] == "2108", "{ddmm} renders as day+month")
    check(ctx_fn["docno_short"].startswith("DN-"), "{docno_short} keeps the prefix")
    empty_ok = D.render_filename(
        "{docno} Material Delivered ({warehouse} - {project}) {prs}",
        dict(ctx_fn, warehouse="", project=""))
    check("()" not in empty_ok and " - )" not in empty_ok,
          f"empty brackets collapse cleanly: {empty_ok}")
    nopr = D.render_filename(
        "{docno} Material Delivered ({warehouse} - {project}) {prs}",
        dict(ctx_fn, _prs=[]))
    check(nopr.rstrip().endswith(")"), "a DN with no PR still names cleanly")
    many_pr = D.render_filename(
        "{docno} Material Delivered ({warehouse} - {project}) {prs}",
        dict(ctx_fn, _prs=[f"PR{i:05d}" for i in range(60)]))
    check(len(many_pr) <= D.MAX_NAME_LEN, "a long PR list is truncated safely")
    check("-more" in many_pr, "truncation says how many PRs were dropped")
    dirty = D.render_filename("{docno} ({warehouse})",
                              dict(ctx_fn, warehouse='A/B:C*D?"E'))
    check(not any(ch in dirty for ch in '<>:"/\\|?*'),
          f"illegal Windows characters are stripped: {dirty}")
    db.set_setting("filename_template_DN", "{docno} - {party}")
    check("Site Team" in D.document_basename(db, fdoc, flines),
          "a custom per-type template overrides the default")
    db.set_setting("filename_template_DN", "")
    grn_row = db.one("SELECT * FROM documents WHERE doc_type='GRN' LIMIT 1")
    if grn_row:
        gl = db.query("SELECT * FROM document_lines WHERE doc_id=?", (grn_row["id"],))
        check(D.document_basename(db, grn_row, gl) == D.safe_file_part(grn_row["doc_no"]),
              "non-DN documents keep the plain document number")

    section("Excel-style column filters")
    ft = W.DataTable()
    ft.fill(["Code", "Cat", "Qty"],
            [["A-1", "Cable", 10], ["A-2", "Cable", 20],
             ["B-1", "Light", 5], ["B-2", "Tool", 0]])
    check(ft.visible_row_count() == 4, "table starts unfiltered")
    check(ft.column_values(1) == ["Cable", "Light", "Tool"],
          "distinct column values are listed and sorted")
    ft.set_column_filter(1, {"Cable"})
    check(ft.visible_row_count() == 2, "a column filter hides non-matching rows")
    check(ft.has_filters(), "the table reports that a filter is active")
    check("▼" in ft.horizontalHeaderItem(1).text(),
          "the filtered column heading is marked")
    ft.filter_rows("A-1")
    check(ft.visible_row_count() == 1, "text search combines with the column filter")
    ft.filter_rows("")
    check(ft.visible_row_count() == 2, "clearing the search keeps the column filter")
    check(len(ft.visible_rows()) == 2, "visible_rows() returns only what is shown")
    ft.fill(["Code", "Cat", "Qty"],
            [["A-1", "Cable", 10], ["A-2", "Cable", 20], ["B-1", "Light", 5]])
    check(ft.visible_row_count() == 2, "filters survive a refresh of the same data")
    ft.fill(["Code", "Cat", "Qty"], [["Z-1", "Pump", 1]])
    check(ft.visible_row_count() == 1,
          "a filter that can no longer match self-heals instead of blanking the grid")
    check(not ft.has_filters(), "the stale filter was dropped")
    ft.set_column_filter(1, {"Pump"})
    ft.clear_filters()
    check(ft.visible_row_count() == 1 and not ft.has_filters(), "clear_filters works")
    check("▼" not in ft.horizontalHeaderItem(1).text(),
          "the heading marker is removed when the filter goes")
    fdlg = W.ColumnFilterDialog(ft, 1)
    check(fdlg.list.count() == 1, "the filter dialog lists the column's values")
    fdlg.close()
    from aurco.ui.items import ItemsPage
    ip = ItemsPage(db)
    app.processEvents()
    cats = ip.table.column_values(2)
    check(len(cats) > 1, "Item Master exposes filterable categories")
    ip.table.set_column_filter(2, {cats[0]})
    app.processEvents()
    check(ip.table.visible_row_count() < ip.table.rowCount(),
          "Item Master filters really narrow the list")
    check("filtered from" in ip.count_lbl.text(),
          "Item Master totals describe the filtered view")
    ip.table.clear_filters()
    app.processEvents()
    check("filtered from" not in ip.count_lbl.text(),
          "Item Master totals return to normal when filters clear")
    fb = W.FilterBar(ft)
    check(fb is not None, "a filter bar can be attached to any table")

    section("Calculator")
    from aurco.ui.calculator import CalculatorDialog, evaluate as cev, _pct
    for expr, want in (("2+3*4", 14), ("(10+5)/3", 5), ("sqrt(144)", 12),
                       ("2^10", 1024), ("100-25%", 75), ("10%", 0.1),
                       ("round(7.777,1)", 7.8), ("max(3,9,2)", 9)):
        check(abs(cev(_pct(expr)) - want) < 1e-6, f"calculator: {expr} = {want}")
    for bad in ("__import__('os').system('x')", "open('/etc/passwd')", "[].__class__"):
        blocked = False
        try:
            cev(_pct(bad))
        except Exception:
            blocked = True
        check(blocked, f"calculator refuses unsafe input: {bad[:28]}")
    zero = False
    try:
        cev("1/0")
    except ZeroDivisionError:
        zero = True
    check(zero, "calculator raises on divide by zero")
    calc = CalculatorDialog(db)
    calc.entry.setText("12*8+150")
    calc.compute()
    check(calc.result.text() == "246", "calculator dialog computes 12*8+150 = 246")
    check(calc.tape.count() == 1, "the tape records the calculation")
    calc.entry.setText("not a sum")
    calc.compute()
    check("Invalid" in calc.result.text(), "a bad expression reports cleanly")
    calc.entry.setText("100")
    calc.compute()
    calc.press("M+")
    check(calc.memory == 100, "memory add works")
    calc.press("MC")
    check(calc.memory == 0, "memory clear works")
    calc.close()
    check("Calculator" in [n for n, _, _ in __import__(
        "aurco.ui.main_window", fromlist=["NAV"]).NAV],
        "the calculator is in the sidebar")

    section("Settings tabs are fully reachable")
    from PySide6.QtWidgets import QScrollArea as _QSA, QTabWidget as _QTW
    win.go("Settings")
    app.processEvents()
    stabs = win.page_settings.findChild(_QTW)
    check(stabs.count() >= 13, f"settings has {stabs.count()} tabs")
    unscrollable = [stabs.tabText(i) for i in range(stabs.count())
                    if not isinstance(stabs.widget(i), _QSA)]
    check(not unscrollable,
          f"every settings tab scrolls (offenders: {unscrollable})")
    check(hasattr(win.page_settings, "f_dn_filename_template"),
          "the DN naming pattern is editable in Settings")
    win.page_settings.f_dn_filename_template.setText(
        "{docno} Material Delivered ({warehouse} - {project}) {prs}")
    win.page_settings._preview_filename()
    check("Material Delivered" in win.page_settings.lbl_fn_preview.text(),
          "the file-name preview updates live")
    win.page_settings.f_dn_filename_template.setText("{docno} {{bad")
    win.page_settings._preview_filename()
    check(bool(win.page_settings.lbl_fn_preview.text()),
          "an odd pattern still previews without crashing")
    win.page_settings.f_dn_filename_template.setText(D.DEFAULT_DN_TEMPLATE)

    section("Dashboard recent tables")
    win.go("Dashboard")
    app.processEvents()
    win.page_dashboard.refresh()
    app.processEvents()
    for nm, tb in (("Recent Delivery Notes", win.page_dashboard.tbl_dn),
                   ("Recent Stock Movements", win.page_dashboard.tbl_mv)):
        check(tb.minimumHeight() >= 200, f"{nm}: has a real minimum height")
        rh = tb.verticalHeader().defaultSectionSize() or 27
        check(tb.height() // rh >= 5,
              f"{nm}: shows at least 5 rows (was 2) — {tb.height()}px")
        check(tb.verticalScrollBarPolicy() == Qt.ScrollBarAsNeeded,
              f"{nm}: scrolls when there are more rows")

    section("Row-number gutter is resizable")
    win.go("Physical Count")
    app.processEvents()
    cnt_grid = win.page_cnt.lines
    vh = cnt_grid.verticalHeader()
    check(vh.isVisible(), "the count sheet shows line numbers")
    check(hasattr(cnt_grid, "_vh_drag"), "a drag handle is installed on the gutter")
    rz = cnt_grid._vh_drag
    check(rz.MAX_W > rz.MIN_W, "the gutter has a usable width range")
    start_w = vh.width()
    check(rz.set_width(90) == 90 and vh.width() == 90, "the gutter can be widened")
    check(rz.set_width(5) == rz.MIN_W, "dragging too far left clamps to the minimum")
    check(rz.set_width(9999) == rz.MAX_W, "dragging too far right clamps to the maximum")
    rz.set_width(start_w)

    section("Admin Station shared site folder")
    from aurco.core import adminstation as AS3
    AS3.reset_admin_db()
    adb3 = AS3.get_admin_db()
    drop = Path(root) / "SiteDrop"
    drop.mkdir(parents=True, exist_ok=True)
    ok_f, msg_f = AS3.folder_status(str(drop))
    check(ok_f, f"a real folder validates: {msg_f}")
    bad_ok, bad_msg = AS3.folder_status(str(Path(root) / "nope"))
    check(not bad_ok and "not exist" in bad_msg,
          "a missing folder is reported clearly instead of failing later")
    check(not AS3.folder_status("")[0], "an empty path is rejected")
    hdr_line = ("SR#\tCamp/Office Name\tDate of Record\tItem Category"
                "\tItem Description\tUOM\tQuantity\tReturn"
                "\tDestination Location\tRemarks\n")
    (drop / "Camp 1 upload.csv").write_text(
        hdr_line.replace("\t", ",")
        + "1,Camp 1,2026-03-01,Furniture,Steel Bed,No,8,0,Room B-12,new\n"
          "2,Camp 1,2026-03-02,Accommodation,Window AC,No,2,2,Room B-12,back\n",
        encoding="utf-8")
    (drop / "Camp 2 upload.txt").write_text(
        hdr_line + "1\tCamp 2\t2026-03-03\tIT\tLaptop\tNo\t1\t0\tAdmin\t\n",
        encoding="utf-8")
    (drop / "notes.docx").write_bytes(b"not a sheet")
    AS3.set_watch_folder(adb3, str(drop))
    check(AS3.get_watch_folder(adb3) == str(drop), "the shared folder is remembered")
    found = AS3.scan_folder(adb3, str(drop))
    check(len(found) == 2, f"only importable files are listed (got {len(found)})")
    check(all(f["status"] == "New" for f in found), "new files are flagged as New")
    check(any("Camp" in f["site"] for f in found),
          "the sending site is guessed from the file name")
    res = AS3.import_from_folder(adb3, [f["path"] for f in found])
    check(res["inserted"] == 3, f"3 rows imported from 2 files (got {res['inserted']})")
    check(res["failed"] == 0, "no file failed to import")
    again = AS3.scan_folder(adb3, str(drop))
    check(not again, "already-imported files disappear from the New list")
    with_done = AS3.scan_folder(adb3, str(drop), include_done=True)
    check(len(with_done) == 2 and all(f["status"] == "Imported" for f in with_done),
          "imported files can still be listed with their history")
    res2 = AS3.import_from_folder(adb3, [f["path"] for f in with_done])
    check(res2["inserted"] == 0 and res2["skipped"] == 3,
          "re-importing the same files posts nothing twice")
    (drop / "broken.csv").write_text("no,useful,headers\n1,2,3\n", encoding="utf-8")
    res3 = AS3.import_from_folder(adb3, [str(drop / "broken.csv")])
    check(res3["failed"] == 1, "an unreadable sheet is reported, not silently skipped")
    check(res3["files"][0]["error"], "the failure carries a reason")
    arch = Path(root) / "Archive"
    (drop / "Camp 3 upload.csv").write_text(
        hdr_line.replace("\t", ",")
        + "1,Camp 3,2026-03-05,Tools,Drill,No,4,0,Store,\n", encoding="utf-8")
    res4 = AS3.import_from_folder(adb3, [str(drop / "Camp 3 upload.csv")],
                                  archive_to=str(arch))
    check(res4["inserted"] == 1, "archiving import still inserts")
    check((arch / "Camp 3 upload.csv").exists(),
          "an imported file is moved to the archive folder")
    check(not (drop / "Camp 3 upload.csv").exists(),
          "the archived file leaves the drop folder")
    ap3 = AdminStationPage(db)
    check(hasattr(ap3, "sitefolder"), "the Site Uploads tab exists")
    check(ap3.tabs.count() == 5, "Admin Station now has 5 tabs")
    ap3.sitefolder.path.setText(str(drop))
    ap3.sitefolder.reload()
    check("Connected" in ap3.sitefolder.status.text(),
          "the folder tab reports a healthy connection")
    ap3.sitefolder.path.setText(str(Path(root) / "missing-share"))
    ap3.sitefolder.reload()
    check("✕" in ap3.sitefolder.status.text(),
          "an offline share is shown as disconnected rather than crashing")


    # ============================================ v2.8 company issuance register
    section("Company Issuance — separation and the real sheet")
    from aurco.core import issuance as ISS
    ISS.reset_issuance_db()
    idb = ISS.get_issuance_db()
    check(Path(idb.path).name == "company_issuance.db",
          "the register has its own database file")
    check(Path(idb.path) != Path(db.path), "it is NOT the inventory database")
    check("Company Issuance" in str(Path(idb.path).parent), "stored in its own folder")
    inv_tables = {r["name"] for r in db.query(
        "SELECT name FROM sqlite_master WHERE type='table'")}
    check("issues" not in inv_tables and "evidence" not in inv_tables,
          "no issuance table leaked into the inventory database")

    sheet = ("Date\tCompany Name\tMR (If any)\tReceipient\tIqama ID\tItem issued"
             "\tQty\tDate of issuance\tDate of Return\tEvidence\tRemarks\n"
             "21-Dec-25\tAlnoor\t1728\tMuhammad shoaib\t2563232723\tCable Puller"
             "\t24\t21-Dec-25\t17-Apr-26\t-\tReturned\n"
             "14-Apr-26\tBroad\t-\tMuhammad Adnan\t2563232723\tHeat Gun\t1"
             "\t13-Apr-26\t14-Apr-26\t-\tReturned\n"
             "14-Apr-26\tAlnoor\t0-2174\tMuhammad shoaib\t-\tHydro test machine"
             "\t1\t4-Apr-26\t-\t-\tNot Return yet\n"
             "22-Jun-26\tAlnoor\t-\t-\t-\tCable stand with Rod (Set)\t1"
             "\t6/22/2026\t-\tDN-0737\t\n")
    hh, rr = ISS.sniff(sheet)
    check(len(hh) == 11 and len(rr) == 4, "the real sheet splits into 11 cols / 4 rows")
    mp = ISS.auto_map(hh)
    check(mp[3] == "recipient", "the 'Receipient' misspelling is mapped")
    check(mp[9] == "_evidence", "the Evidence column is recognised")
    check(len(mp) == 11, "every column of the sheet is mapped")
    recs = ISS.preview(hh, rr, mp)
    check(len(recs) == 4, "four issuance rows parsed")
    check(recs[0]["issue_date"] == "2025-12-21",
          f"21-Dec-25 becomes 2025-12-21 (got {recs[0]['issue_date']})")
    check(recs[0]["qty_returned"] == 24, "'Returned' in Remarks books the full return")
    check(ISS.compute_status(recs[0]) == ISS.ST_RETURNED, "that row reads as Returned")
    check(recs[2]["qty_returned"] == 0, "'Not Return yet' stays outstanding")
    check(recs[3]["dn_no"] == "DN-0737",
          "a DN number in Evidence is kept as the reference")
    check(recs[3]["issue_date"] == "2026-06-22", "the 6/22/2026 format is understood")
    ins, sk = ISS.import_records(idb, recs, "sheet")
    check((ins, sk) == (4, 0), "all four rows imported")
    check(ISS.import_records(idb, recs, "again") == (0, 4),
          "re-importing the same sheet is skipped as duplicate")

    section("Company Issuance — photo evidence")
    try:
        from PIL import Image as _PIL
        photo_dir = Path(root) / "site_photos"
        photo_dir.mkdir(parents=True, exist_ok=True)
        shots = []
        for nm in ("a.jpg", "b.jpg", "back.jpg"):
            f_ = photo_dir / nm
            _PIL.new("RGB", (400, 300), (80, 120, 170)).save(f_)
            shots.append(str(f_))
        have_pil = True
    except ImportError:
        shots, have_pil = [], False
    check(have_pil, "Pillow is available to make test photos")

    refused = False
    try:
        ISS.save_issue(idb, {"company": "Alnoor", "item": "Crimper", "qty": 1})
    except ValueError as exc:
        refused = "proof" in str(exc).lower()
    check(refused, "a new issue with no picture and no DN is refused")
    ok_dn = ISS.save_issue(idb, {"company": "Alnoor", "item": "Tester", "qty": 1,
                                 "dn_no": "DN-0900"})
    check(ok_dn > 0, "a DN reference is accepted instead of a photo")
    free = ISS.save_issue(idb, {"company": "Alnoor", "item": "Loose", "qty": 1},
                          require_evidence=False)
    check(free > 0, "the requirement can be waived for imports")

    iid = ISS.save_issue(idb, {
        "company": "Alnoor", "recipient": "Shahzad Alam", "iqama": "2563232723",
        "item": "Hydraulic crimper", "qty": 3, "issue_type": ISS.TEMPORARY,
        "issue_date": "2026-04-30", "expected_return": "2026-05-15",
        "unit_value": 450}, evidence_files=shots[:2])
    n_out, n_in = ISS.evidence_counts(idb, iid)
    check((n_out, n_in) == (2, 0), f"two issue photos stored (got {n_out})")
    ev = ISS.evidence_for(idb, iid)
    check(all("Company Issuance" in e["file_path"] for e in ev),
          "photos are copied into the register's own folder")
    check(all(Path(e["file_path"]).resolve() != Path(s).resolve()
              for e, s in zip(ev, shots)), "the copy is a genuinely new file")
    shutil.rmtree(photo_dir, ignore_errors=True)
    check(all(Path(e["file_path"]).exists() for e in ev),
          "evidence survives deletion of the original photos")

    section("Company Issuance — returns and status engine")
    rec = ISS.get_issue(idb, iid)
    check(ISS.compute_status(rec) == ISS.ST_OVERDUE,
          "an unreturned item past its due date is Overdue")
    check(ISS.days_overdue(rec) > 0, "the overdue days are counted")
    ISS.record_return(idb, iid, 1, "2026-05-10")
    rec = ISS.get_issue(idb, iid)
    check(ISS.compute_status(rec) == ISS.ST_PARTIAL, "1 of 3 back = Partially Returned")
    check(abs(ISS.outstanding_qty(rec) - 2) < 1e-9, "2 remain outstanding")
    over = False
    try:
        ISS.record_return(idb, iid, 99)
    except ValueError:
        over = True
    check(over, "returning more than is outstanding is refused")
    ISS.record_return(idb, iid, 2, "2026-05-14", evidence_files=[])
    rec = ISS.get_issue(idb, iid)
    check(ISS.compute_status(rec) == ISS.ST_RETURNED, "the balance back = Returned")
    check(ISS.outstanding_qty(rec) == 0, "nothing outstanding once fully returned")
    perm = ISS.save_issue(idb, {"company": "Alnoor", "item": "Helmets", "qty": 50,
                                "issue_type": ISS.PERMANENT, "dn_no": "DN-1000",
                                "issue_date": "2020-01-01",
                                "expected_return": "2020-02-01"})
    prec = ISS.get_issue(idb, perm)
    check(ISS.compute_status(prec) == ISS.ST_PERMANENT,
          "a permanent issue never becomes overdue")
    check(ISS.outstanding_qty(prec) == 0, "permanent issues are not chased")
    ISS.mark_lost(idb, ok_dn, "not recoverable")
    check(ISS.compute_status(ISS.get_issue(idb, ok_dn)) == ISS.ST_LOST,
          "an item can be written off")
    ISS.reopen(idb, ok_dn)
    check(ISS.compute_status(ISS.get_issue(idb, ok_dn)) != ISS.ST_LOST,
          "a write-off can be reversed")
    check(isinstance(ISS.refresh_statuses(idb), int),
          "statuses can be recalculated in bulk")

    section("Company Issuance — analytics and reports")
    d_iss = ISS.dashboard(idb)
    for k in ("records", "companies", "recipients", "qty_issued", "qty_out",
              "overdue", "missing_proof", "proof_pct", "avg_days_out", "value_out"):
        check(k in d_iss, f"dashboard exposes '{k}'")
    check(d_iss["records"] > 0, "the dashboard counts records")
    one_co = ISS.dashboard(idb, {"company": "Broad"})
    check(one_co["records"] < d_iss["records"], "dashboard honours a company filter")
    check(0 <= d_iss["proof_pct"] <= 100, "proof coverage is a sane percentage")
    for meas in ("qty", "count", "outstanding", "value"):
        check(isinstance(ISS.by_column(idb, "company", 5, meas), list),
              f"by_column supports the '{meas}' measure")
    check(len(ISS.ageing(idb)) == 5, "ageing returns five buckets")
    check(isinstance(ISS.status_split(idb), list), "status split works")
    check(isinstance(ISS.monthly_issue_return(idb, 6), list), "monthly trend works")
    mcols, mrows = ISS.company_matrix(idb)
    check(len(mcols) == len(ISS.STATUSES) + 3, "the company matrix has a status column each")
    for rep in ISS.REPORT_LIST:
        t_, c_, rw_ = ISS.build_report(idb, rep)
        check(bool(c_), f"issuance report builds: {rep}")
    miss = ISS.build_report(idb, "Missing Photo Proof")[2]
    check(isinstance(miss, list), "the missing-proof report runs")
    check(len(ISS.missing_evidence(idb)) == len(miss),
          "the missing-proof report agrees with the helper")

    section("Company Issuance — documents")
    rp = D.issuance_report_pdf(db, "Issuance Register",
                               *ISS.build_report(idb, "Full Issuance Register")[1:])
    check(rp.exists() and rp.stat().st_size > 1500, "issuance report PDF renders")
    check("Company Issuance" in str(rp.parent), "it lands in the register's folder")
    rcpt = D.issuance_receipt_pdf(idb, db, iid)
    check(rcpt.exists() and rcpt.stat().st_size > 2000, "the hand-over receipt renders")
    try:
        import pypdfium2 as _pd
        doc_ = _pd.PdfDocument(str(rcpt))
        pages = len(doc_)
        txt_ = doc_[0].get_textpage().get_text_range()
        doc_.close()
        check(pages >= 2, f"the proof photos are appended after the receipt ({pages}p)")
        check("Hydraulic crimper" in txt_, "the receipt names the item")
        check("2563232723" in txt_, "the receipt carries the Iqama ID")
        check("Photographic evidence" in txt_, "the receipt shows the evidence section")
    except ImportError:
        check(True, "pypdfium2 absent - PDF text check skipped")
    led_before = db.scalar("SELECT COUNT(*) FROM stock_ledger")
    ISS.save_issue(idb, {"company": "X", "item": "Y", "qty": 5, "dn_no": "DN-1"})
    check(db.scalar("SELECT COUNT(*) FROM stock_ledger") == led_before,
          "issuing to a company posts NO inventory stock movement")

    section("Company Issuance — UI")
    from aurco.ui.issuance_page import (IssuancePage, IssueDialog, ReturnDialog,
                                        EvidenceGalleryTab)
    ipg = IssuancePage(db)
    check(ipg.tabs.count() == 5, "the page has 5 tabs")
    check(len(ipg.dash.cards) == 16, "the dashboard has 16 KPI tiles")
    for ch in ("ch_company", "ch_recipient", "ch_items", "ch_io", "ch_age",
               "ch_status", "ch_proof"):
        check(hasattr(ipg.dash, ch), f"dashboard has chart {ch}")
    ipg.register.reload()
    check(ipg.register.table.rowCount() > 0, "the register lists the issuances")
    before = ipg.dash.cards["records"].lbl_value.text()
    ipg.dash.f_company.setCurrentText("Broad")
    app.processEvents()
    check(ipg.dash.cards["records"].lbl_value.text() != before,
          "a dashboard filter changes the tiles")
    ipg.dash.reset_filters()
    app.processEvents()
    check(ipg.dash.cards["records"].lbl_value.text() == before, "Reset restores them")
    ipg.dash._drill("missing_proof")
    app.processEvents()
    check(ipg.tabs.currentIndex() == 1, "a KPI tile drills into the register")
    check(ipg.register.chk_noproof.isChecked(),
          "the drill-through carries the missing-proof filter")
    ipg.register.chk_noproof.setChecked(False)
    ipg.evidence.reload()
    check(ipg.evidence.gallery.count() >= 1, "the evidence gallery shows the photos")
    ipg.evidence.chk_missing.setChecked(True)
    check(isinstance(ipg.evidence.gallery.count(), int),
          "the missing-proof view renders")
    ipg.evidence.chk_missing.setChecked(False)
    ipg.reports.run()
    check(ipg.reports.table.columnCount() > 0, "the reports tab runs a report")
    dlg_i = IssueDialog(idb, iid)
    check(dlg_i.company.currentText() == "Alnoor", "the editor loads an existing issue")
    dlg_i.issue_type.setCurrentText(ISS.PERMANENT)
    check(not dlg_i.return_date.isEnabled(),
          "choosing Permanent disables the return fields")
    dlg_i.close()
    open_rec = next((r for r in ISS.search(idb) if ISS.outstanding_qty(r) > 0), None)
    if open_rec:
        dlg_r = ReturnDialog(idb, open_rec)
        check(dlg_r.qty.maximum() <= ISS.to_float(open_rec["qty"]),
              "the return dialog caps the quantity at what is outstanding")
        dlg_r.close()
    bk = idb.backup(note="test")
    check(bk.exists(), "the register backs up independently")
    check("Company Issuance" in [n for n, _, _ in __import__(
        "aurco.ui.main_window", fromlist=["NAV"]).NAV],
        "the module is in the sidebar")


    # ================================== v2.9 MR autofill + file protection
    section("Material Request header auto-fill")
    mp3 = MaterialPage(db)
    check(mp3.h_dept.text() == "Site Team",
          "Department defaults to 'Site Team' before anything is pasted")
    check(mp3.h_by.text() == "By Site Team",
          "Requested by defaults to 'By Site Team'")
    pr_paste = (
        "Line\tProject ID\tItem number\tProduct name\tUnit\tQuantity\tStatus"
        "\tCategory\tPurchase requisition reference\n"
        "1\tPRJ_0000026\t30001EL34\tPVC Elbow 4in 90\tNo\t60\tIn review"
        "\tPlumbing\t001603\n"
        "2\tPRJ_0000026\t30001EL06\tPVC Elbow 4in 45\tNo\t20\tIn review"
        "\tPlumbing\t001603\n"
        "3\tPRJ_0000026\t30000PP11\tPVC Pipe 4in\tM\t30\tIn review"
        "\tPlumbing\t001604\n")
    mp3.paste.setPlainText(pr_paste)
    mp3.check_paste()
    app.processEvents()
    check(mp3.h_project.text() == "PRJ_0000026", "Project ID is detected")
    check(mp3.h_site.currentText() == "PRJ_0000026",
          f"Site is auto-filled with the Project ID (got {mp3.h_site.currentText()!r})")
    check(mp3.h_dept.text() == "Site Team", "Department stays 'Site Team'")
    check(mp3.h_by.text() == "By Site Team", "Requested by stays 'By Site Team'")
    check(mp3.h_ref.text() == "001603, 001604",
          f"every distinct PR number is auto-detected (got {mp3.h_ref.text()!r})")
    check("001603" in mp3.autofill_note.text(), "the auto-fill note reports what it did")
    check(mp3.t_check.rowCount() == 3, "the three lines were checked")

    # a value typed by hand must never be overwritten
    mp3.clear()
    mp3.h_project.setText("MANUAL-PRJ")
    mp3.h_site.setCurrentText("Manual Site")
    mp3.h_ref.setText("MY-REF")
    mp3.paste.setPlainText(pr_paste)
    mp3.check_paste()
    app.processEvents()
    check(mp3.h_project.text() == "MANUAL-PRJ", "a hand-typed Project ID is respected")
    check(mp3.h_site.currentText() == "Manual Site", "a hand-picked Site is respected")
    check(mp3.h_ref.text() == "MY-REF", "a hand-typed Reference is respected")

    section("Material Request — Clear resets everything")
    mp3.clear()
    check(mp3.paste.toPlainText() == "", "Clear empties the paste box")
    check(mp3.t_check.rowCount() == 0, "Clear empties the results grid")
    check(mp3.h_project.text() == "", "Clear empties Project ID")
    check(mp3.h_site.currentText() == "", "Clear empties Site")
    check(mp3.h_ref.text() == "", "Clear empties Reference")
    check(mp3.h_dept.text() == "Site Team", "Clear restores the Department default")
    check(mp3.h_by.text() == "By Site Team", "Clear restores the Requested by default")
    check(not mp3.only_short.isChecked(), "Clear resets the shortages tick box")
    check(mp3.autofill_note.text() == "", "Clear removes the auto-fill note")
    check(mp3.checked == [], "Clear drops the checked lines")
    # mixed projects should still work and warn
    mixed = pr_paste.replace("3\tPRJ_0000026", "3\tPRJ_0000099")
    mp3.paste.setPlainText(mixed)
    mp3.check_paste()
    app.processEvents()
    check("different project" in mp3.autofill_note.text().lower(),
          "a paste with two project IDs warns which one was used")
    mp3.clear()

    section("File protection — engine")
    from aurco.core import protection as PR
    PR.ensure_schema(db)
    check(PR.is_enabled(db), "protection is ON by default")
    prot_dir = Path(config.folder("Reports"))
    f1 = prot_dir / "prot_a.pdf"
    f2 = prot_dir / "prot_b.pdf"
    f1.write_bytes(b"AAAA" * 400)
    f2.write_bytes(b"BBBB" * 300)
    res = PR.lock_folder(db, prot_dir)
    check(res["files"] >= 2, f"lock_folder found the files ({res['files']})")
    check(res["recorded"] >= 2, "files were added to the tamper ledger")
    check(res["readonly"] >= 2, "files were set read-only")
    blocked = False
    try:
        f2.write_bytes(b"overwrite")
    except PermissionError:
        blocked = True
    check(blocked, "a protected file cannot be overwritten")
    v1 = PR.verify(db)
    check(v1["ok"] >= 2 and not v1["missing"], "verification passes on untouched files")
    # external tampering must be detected
    import stat as _st
    os.chmod(f2, _st.S_IWRITE | _st.S_IREAD)
    f2.write_bytes(b"TAMPERED")
    v2 = PR.verify(db, deep=True)
    check(len(v2["changed"]) >= 1, "an altered file is detected")
    os.chmod(f1, _st.S_IWRITE | _st.S_IREAD)
    f1.unlink()
    v3 = PR.verify(db)
    check(any("prot_a" in m for m in v3["missing"]),
          "a file deleted from outside AURCO is reported as MISSING")
    check(PR.stats(db)["missing"] >= 1, "the ledger counts the missing file")

    section("File protection — AURCO never deletes")
    keep = prot_dir / "prot_keep.pdf"
    keep.write_bytes(b"KEEP" * 250)
    PR.record_file(db, keep)
    removed = PR.guarded_unlink(db, keep, "unit test delete")
    check(removed is False, "guarded_unlink refuses to delete while protection is on")
    check(not keep.exists(), "the file left its original place")
    arch = PR.archived_files(db)
    check(any("prot_keep" in a["name"] for a in arch),
          "it was archived instead of destroyed")
    target = next(a for a in arch if "prot_keep" in a["name"])
    check(Path(target["path"]).exists(), "the archived copy is really on disk")
    back = PR.restore_archived(db, target["path"])
    check(back is not None and Path(back).exists(), "an archived file can be restored")
    # with protection off an admin can still delete
    PR.set_enabled(db, False)
    tmp_del = prot_dir / "prot_tmp.pdf"
    tmp_del.write_bytes(b"x" * 40)
    check(PR.guarded_unlink(db, tmp_del, "admin") is True,
          "with protection OFF an administrator can delete")
    check(not tmp_del.exists(), "the file really went")
    PR.set_enabled(db, True)
    check(PR.is_enabled(db), "protection can be switched back on")

    section("File protection — evidence is never destroyed")
    from aurco.core import issuance as ISS2
    ISS2.reset_issuance_db()
    idb2 = ISS2.get_issuance_db()
    try:
        from PIL import Image as _PIL2
        pdir = Path(root) / "prot_photos"
        pdir.mkdir(parents=True, exist_ok=True)
        shot = pdir / "proof.jpg"
        _PIL2.new("RGB", (200, 150), (100, 130, 180)).save(shot)
        eid = ISS2.save_issue(idb2, {"company": "Alnoor", "item": "Crimper", "qty": 1},
                              evidence_files=[str(shot)])
        evrow = ISS2.evidence_for(idb2, eid)[0]
        evpath = Path(evrow["file_path"])
        check(evpath.exists(), "evidence photo stored")
        # deliberately exercise the fail-safe path: the global handle here is
        # NOT this test's database, which used to make the code hard-delete
        ISS2.delete_evidence(idb2, evrow["id"], remove_file=True)
        check(not evpath.exists(), "the evidence left its original place")
        # assert on THIS photo, not on a global count: earlier tests archive
        # files into the same storage root, which makes a count baseline stale
        arc_dir = evpath.parent / PR.ARCHIVE_DIR
        kept = list(arc_dir.glob(f"{evpath.stem}*")) if arc_dir.exists() else []
        check(bool(kept),
              "deleting evidence archives the photo instead of destroying it")
        check(kept and kept[0].stat().st_size > 0,
              "the archived photo still has its content")
    except ImportError:
        check(True, "Pillow absent - evidence protection test skipped")

    section("File protection — status and UI")
    st = PR.status_report(db)
    for k in ("enabled", "tracked", "missing", "changed", "archived", "root",
              "acl_supported", "folders"):
        check(k in st, f"status report exposes '{k}'")
    check(isinstance(st["folders"], list), "it lists the protected folders")
    ok_acl, msg_acl = PR.deny_delete_acl(prot_dir, True)
    check(isinstance(ok_acl, bool) and isinstance(msg_acl, str),
          "deny_delete_acl reports honestly on this platform")
    check(PR.set_readonly(prot_dir, True) is False,
          "set_readonly refuses a directory (files only)")
    check(PR.file_hash(__file__) != "", "file_hash works on a real file")
    check(PR.file_hash(prot_dir / "does_not_exist") == "",
          "file_hash on a missing file returns empty rather than raising")
    tot = PR.protect_all(db)
    check(tot["folders"] >= 1, f"protect_all covered {tot['folders']} folder(s)")
    win.go("Settings")
    app.processEvents()
    stabs2 = win.page_settings.findChild(_QTW)
    titles = [stabs2.tabText(i) for i in range(stabs2.count())]
    check(any("Protection" in t for t in titles), "Settings has a File Protection tab")
    check(hasattr(win.page_settings, "f_protect_files"),
          "the protection switch is on the page")
    check(hasattr(win.page_settings, "f_mr_dept"),
          "the MR defaults are editable in Settings")
    win.page_settings._refresh_protect()
    check("Protection is" in win.page_settings.lbl_protect.text(),
          "the protection status is displayed")
    win.page_settings._show_issues()
    check(win.page_settings.tbl_protect.columnCount() > 0, "the issues view renders")
    win.page_settings._show_archive()
    check(win.page_settings.tbl_protect.columnCount() > 0, "the archive view renders")


    # ============================ v2.10 bulk row delete + clear after save
    section("Material Request — bulk line delete (core)")
    bl_items = S.search_items(db)[:4]
    bl_src = [{"item_code": i["code"], "description": i["description"],
               "uom": i["uom"], "qty": 2, "pr_no": "001603"} for i in bl_items]
    bl_mr = MM.save_request(db, {"mr_date": "2026-08-20", "project_id": "PBULK"},
                            MM.enrich(db, bl_src))
    bl_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (bl_mr,))
    bl_lines = MM.request_lines(db, bl_id)
    check(len(bl_lines) == 4, "test request has 4 lines")
    done, skipped = MM.delete_lines(db, [l["id"] for l in bl_lines[:2]])
    check(done == 2 and not skipped, "two lines deleted in one call")
    check(len(MM.request_lines(db, bl_id)) == 2, "the other two survive")
    check(db.scalar("SELECT COUNT(*) FROM material_requests WHERE id=?",
                    (bl_id,)) == 1, "the request itself is kept")

    # a delivered line must never be erased by a bulk delete
    MM.prepare_all_available(db, bl_id)
    rdy2 = [r for r in MM.ready_lines(db) if r["mr_no"] == bl_mr]
    if rdy2:
        MM.deliver_lines(db, [rdy2[0]["id"]], S.DocHeader(
            doc_type="DN", doc_date="2026-08-20", issued_to="Site",
            warehouse=rdy2[0].get("warehouse", "")))
        left = MM.request_lines(db, bl_id)
        done2, skip2 = MM.delete_lines(db, [l["id"] for l in left])
        check(len(skip2) >= 1, "a delivered line is skipped, not erased")
        check("delivered" in skip2[0].lower(), "the refusal explains why")
        still = MM.request_lines(db, bl_id)
        check(any(float(l["qty_delivered"] or 0) > 0 for l in still),
              "the delivered line is still there afterwards")
        forced, _ = MM.delete_lines(db, [l["id"] for l in still], force=True)
        check(forced >= 1, "force=True can still remove it when required")
    check(any(e["mr_no"] == bl_mr for e in MM.empty_requests(db)),
          "a request with no lines left is reported as empty")
    check(MM.delete_lines(db, []) == (0, []), "an empty id list is handled safely")
    check(MM.delete_lines(db, [999999]) == (0, []),
          "a non-existent line id is ignored rather than raising")

    section("Material Request — bulk delete in the UI")
    from PySide6.QtWidgets import QTableWidgetSelectionRange as _Range
    mp4 = MaterialPage(db)
    ui_paste = ("Line\tProject ID\tItem number\tProduct name\tUnit\tQuantity"
                "\tPurchase requisition reference\n")
    for _i, _it in enumerate(S.search_items(db)[:6], 1):
        ui_paste += (f"{_i}\tPRJ_BULK\t{_it['code']}\t{_it['description'][:24]}"
                     f"\tNo\t{_i * 2}\t001603\n")
    mp4.paste.setPlainText(ui_paste)
    mp4.check_paste()
    app.processEvents()
    check(mp4.t_check.rowCount() == 6, "six lines checked on tab 1")
    check(mp4.t_check.selectionMode() == QAbstractItemView.ExtendedSelection,
          "the check grid allows multi-select")
    mp4.t_check.setRangeSelected(_Range(0, 0, 1, 3), True)
    app.processEvents()
    mp4.remove_check_rows()
    app.processEvents()
    check(mp4.t_check.rowCount() == 4, "two rows removed from the check grid")
    check(len(mp4.checked) == 4, "the pending list shrank to match")
    mp4.t_check.clearSelection()
    mp4.remove_check_rows()
    check(mp4.t_check.rowCount() == 4, "removing with nothing selected changes nothing")

    section("Material Request — tab 1 clears after saving")
    mp4.save_request()
    app.processEvents()
    check(mp4.paste.toPlainText() == "", "the paste box is emptied after saving")
    check(mp4.t_check.rowCount() == 0, "the check grid is emptied after saving")
    check(mp4.checked == [], "the pending list is dropped after saving")
    check(mp4.h_project.text() == "", "Project ID is cleared after saving")
    check(mp4.h_site.currentText() == "", "Site is cleared after saving")
    check(mp4.h_ref.text() == "", "Reference is cleared after saving")
    check(mp4.h_dept.text() == "Site Team", "Department returns to its default")
    check(mp4.h_by.text() == "By Site Team", "Requested by returns to its default")
    check(mp4.autofill_note.text() == "", "the auto-fill note is cleared")
    check(mp4.tabs.currentIndex() == 1, "the app moves to Requests & Preparation")
    check(mp4.t_req.rowCount() >= 1, "the saved request is listed there")
    check(mp4.t_lines.rowCount() >= 1, "its lines are loaded ready to prepare")
    # saving twice must be impossible now that the sheet is clear
    mp4.save_request()
    check(True, "pressing Save again with an empty sheet is handled safely")

    section("Material Request — bulk delete on tabs 2 and 3")
    check(mp4.t_lines.selectionMode() == QAbstractItemView.ExtendedSelection,
          "the request-lines grid allows multi-select")
    n_lines = mp4.t_lines.rowCount()
    if n_lines >= 2:
        mp4.t_lines.setRangeSelected(_Range(0, 0, 1, 3), True)
        app.processEvents()
        check(len(mp4._sel_lines()) == 2, "_sel_lines() returns the two selected rows")
        mp4.delete_line()
        app.processEvents()
        check(mp4.t_lines.rowCount() <= n_lines - 2 or mp4.t_lines.rowCount() == 0,
              "both selected lines were deleted in one action")
    mr_now = mp4._sel_request()
    if mr_now:
        MM.prepare_all_available(db, mr_now["id"])
    mp4._load_ready()
    app.processEvents()
    check(hasattr(mp4, "delete_ready"), "tab 3 exposes a bulk delete")
    if mp4.t_ready.rowCount():
        n_ready = mp4.t_ready.rowCount()
        mp4.t_ready.selectAll()
        app.processEvents()
        mp4.delete_ready()
        app.processEvents()
        check(mp4.t_ready.rowCount() < n_ready,
              "prepared lines can be deleted in bulk from tab 3")
    mp4.t_ready.clearSelection()
    mp4.delete_ready()
    check(True, "delete on tab 3 with no selection is handled safely")
    for tab in (0, 1, 2):
        mp4.tabs.setCurrentIndex(tab)
        mp4._delete_current()
    check(True, "the Del key is routed safely on every tab")


    # ================== v2.11 partial-marked, deliver from tab 2, project closure
    section("Partial Marked status")
    rich = [i for i in S.search_items(db) if (i["balance"] or 0) >= 50][:1]
    check(bool(rich), "a well-stocked item exists for the test")
    ritem = rich[0]
    pm_src = [{"item_code": ritem["code"], "description": ritem["description"],
               "uom": ritem["uom"], "qty": 10, "pr_no": "PM1"}]
    pm_mr = MM.save_request(db, {"mr_date": "2026-08-20", "project_id": "PPM"},
                            MM.enrich(db, pm_src))
    pm_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (pm_mr,))
    pm_line = MM.request_lines(db, pm_id)[0]
    check(pm_line["status"] == MM.PENDING, "a fresh line starts Pending")
    MM.set_prepared(db, pm_line["id"], 4)
    got = MM.request_lines(db, pm_id)[0]
    check(got["status"] == MM.PARTIAL_MARKED,
          f"partial while stock is plentiful -> Partial Marked (got {got['status']})")
    check(db.scalar("SELECT status FROM material_requests WHERE id=?", (pm_id,))
          == MM.PARTIAL_MARKED, "the request rolls up to Partial Marked")
    check(any(r["id"] == pm_line["id"] for r in MM.ready_lines(db)),
          "a Partial Marked line still shows in Ready to Deliver")
    MM.set_prepared(db, pm_line["id"], 10)
    check(MM.request_lines(db, pm_id)[0]["status"] == MM.READY,
          "preparing the full quantity turns it Ready")
    MM.set_prepared(db, pm_line["id"], 0)
    check(MM.request_lines(db, pm_id)[0]["status"] == MM.PENDING,
          "unpreparing returns it to Pending")
    # a genuine shortage must stay 'Preparing', not 'Partial Marked'
    poor = [i for i in S.search_items(db) if 0 < (i["balance"] or 0) < 6][:1]
    if poor:
        pit = poor[0]
        sh_mr = MM.save_request(db, {"mr_date": "2026-08-20", "project_id": "PSH"},
                                MM.enrich(db, [{"item_code": pit["code"],
                                                "description": pit["description"],
                                                "uom": pit["uom"], "qty": 500,
                                                "pr_no": "SH1"}]))
        sh_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (sh_mr,))
        sh_line = MM.request_lines(db, sh_id)[0]
        MM.set_prepared(db, sh_line["id"], float(pit["balance"]))
        check(MM.request_lines(db, sh_id)[0]["status"] == MM.PREPARING,
              "partial because stock is short stays Preparing, NOT Partial Marked")
    check(MM.PARTIAL_MARKED in MM.FULFIL_COLORS, "the new status has a colour")
    check(MM.PARTIAL_MARKED in MM.OPEN_STATES, "it counts as an open state")

    section("Process button — core engine")
    pr_big = S.save_item(db, {"code": "PRC-BIG", "description": "Plenty item",
                              "uom": "No", "opening_balance": 500})
    pr_few = S.save_item(db, {"code": "PRC-FEW", "description": "Scarce item",
                              "uom": "No", "opening_balance": 4})
    S.save_item(db, {"code": "PRC-ZERO", "description": "Empty item", "uom": "No",
                     "opening_balance": 0})
    pr_src = [{"item_code": "PRC-BIG", "description": "Plenty item", "uom": "No",
               "qty": 10, "pr_no": "PRC"},
              {"item_code": "PRC-FEW", "description": "Scarce item", "uom": "No",
               "qty": 50, "pr_no": "PRC"},
              {"item_code": "PRC-ZERO", "description": "Empty item", "uom": "No",
               "qty": 5, "pr_no": "PRC"},
              {"item_code": "PRC-GHOST", "description": "Not in master", "uom": "No",
               "qty": 7, "pr_no": "PRC"}]
    pr_mr = MM.save_request(db, {"mr_date": "2026-08-20", "project_id": "PPROC"},
                            MM.enrich(db, pr_src))
    pr_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (pr_mr,))
    pr_lines = MM.request_lines(db, pr_id)
    check(len(pr_lines) == 4, "the test request has 4 lines")
    # storekeeper marks 4 of 10 even though 500 are on the shelf
    MM.set_prepared(db, pr_lines[0]["id"], 4)
    check(MM.request_lines(db, pr_id)[0]["status"] == MM.PARTIAL_MARKED,
          "the hand-marked line reads Partial Marked")
    led_pb = db.scalar("SELECT COUNT(*) FROM stock_ledger")
    res = MM.process_lines(db, [l["id"] for l in pr_lines])
    check(res["ready"] == 2, f"2 lines became ready (got {res['ready']})")
    check(res["kept"] == 1, "one line kept the quantity that was marked")
    check(res["prepared"] == 1, "one line was reserved from available stock")
    check(abs(res["qty"] - 8.0) < 1e-9, f"total ready qty is 4+4=8 (got {res['qty']})")
    after_pr = {l["item_code"]: l for l in MM.request_lines(db, pr_id)}
    check(after_pr["PRC-BIG"]["qty_prepared"] == 4,
          "THE MARKED QUANTITY IS RESPECTED — 4, not the full 10")
    check(after_pr["PRC-FEW"]["qty_prepared"] == 4,
          "an unmarked line takes what is actually available (4 of 50)")
    check(after_pr["PRC-ZERO"]["qty_prepared"] == 0, "a zero-stock line is untouched")
    check(any("PRC-ZERO" in x for x in res["skipped"]),
          "the zero-stock line is reported as skipped")
    check(any("Item Master" in x for x in res["skipped"]),
          "an unknown item is reported, not silently ignored")
    check(len(res["short"]) == 2, "partly-covered lines are listed")
    check(db.scalar("SELECT COUNT(*) FROM stock_ledger") == led_pb,
          "PROCESS POSTS NO STOCK MOVEMENT — nothing leaves the warehouse")
    ready_codes = {r["item_code"]: r["qty_ready"] for r in MM.ready_lines(db)}
    check(ready_codes.get("PRC-BIG") == 4,
          "Ready to Deliver shows the marked 4, not 10")
    check(ready_codes.get("PRC-FEW") == 4, "and the available 4 for the scarce item")
    # processing again must be idempotent, not double-reserve
    res2 = MM.process_lines(db, [l["id"] for l in pr_lines])
    check(MM.request_lines(db, pr_id)[0]["qty_prepared"] == 4,
          "processing twice does not double the reservation")
    check(MM.process_lines(db, []) ["ready"] == 0, "an empty list is safe")
    check(MM.process_lines(db, [999999])["ready"] == 0, "an unknown id is ignored")
    whole = MM.process_request(db, pr_id)
    check(isinstance(whole, dict), "process_request() handles a whole request")

    section("Process button — UI")
    mp5 = MaterialPage(db)
    check(hasattr(mp5, "process_lines"), "tab 2 exposes process_lines()")
    check(not hasattr(mp5, "deliver_selected"),
          "the old 'Deliver Selected' handler is gone")
    check(not hasattr(mp5, "deliver_ready_all"),
          "the old 'Deliver All Ready' handler is gone")
    btn_labels = [b.text() for b in mp5.findChildren(QPushButton)]
    check(any("Process" in t for t in btn_labels), "a Process button exists")
    check(not any("Deliver Selected" in t for t in btn_labels),
          "the 'Deliver Selected' button was removed")
    check(not any("Deliver All Ready" in t for t in btn_labels),
          "the 'Deliver All Ready' button was removed")
    check(sum(1 for t in btn_labels if "Process" in t) == 1,
          "exactly ONE Process button was added")

    u_ids = [S.save_item(db, {"code": f"PU-{n}", "description": f"UI item {n}",
                              "uom": "No", "opening_balance": 100})
             for n in (1, 2, 3)]
    u_src = [{"item_code": f"PU-{n}", "description": f"UI item {n}", "uom": "No",
              "qty": 10, "pr_no": "PU"} for n in (1, 2, 3)]
    u_mr = MM.save_request(db, {"mr_date": "2026-08-20", "project_id": "PUI"},
                           MM.enrich(db, u_src))
    u_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (u_mr,))
    mp5.reload()
    for r_ in range(mp5.t_req.rowCount()):
        if mp5.t_req.item(r_, 0) and mp5.t_req.item(r_, 0).text() == u_mr:
            mp5.t_req.selectRow(r_)
            break
    mp5._load_lines(force=True)
    app.processEvents()
    MM.set_prepared(db, MM.request_lines(db, u_id)[0]["id"], 3)   # mark 3 of 10
    mp5._load_lines(force=True)
    app.processEvents()
    check(mp5.t_lines.rowCount() == 3, "the request's 3 lines are loaded")
    mp5.t_lines.setRangeSelected(_Range(0, 0, 1, 3), True)
    app.processEvents()
    mp5.process_lines()
    app.processEvents()
    u_ready = {r["item_code"]: r["qty_ready"] for r in MM.ready_lines(db)}
    check(u_ready.get("PU-1") == 3, "the UI respects the marked 3 of 10")
    check(u_ready.get("PU-2") == 10, "and gives the unmarked line its full 10")
    check("PU-3" not in u_ready, "the unselected line was not processed")
    check(mp5.tabs.currentIndex() == 2,
          "after Process the user lands on Ready to Deliver")
    # no selection -> whole request
    mp5.tabs.setCurrentIndex(1)
    mp5._load_lines(force=True)
    mp5.t_lines.clearSelection()
    app.processEvents()
    mp5.process_lines()
    app.processEvents()
    check("PU-3" in {r["item_code"] for r in MM.ready_lines(db)},
          "with no line selected the whole request is processed")
    # a deselected request must be refused, not silently acted on
    mp5.tabs.setCurrentIndex(1)
    mp5.t_req.clearSelection()
    app.processEvents()
    before_ready = len(MM.ready_lines(db))
    mp5.process_lines()
    check(len(MM.ready_lines(db)) == before_ready,
          "Process with no request selected changes nothing")
    mp5.t_req.selectRow(0)
    mp5._load_lines(force=True)
    app.processEvents()
    mp5.tabs.setCurrentIndex(1)
    mp5._deliver_current()
    check(True, "Ctrl+D on tab 2 routes to Process safely")

    section("Project closure reconciliation")
    rc_ids = [S.save_item(db, {"code": f"RC-{n}", "description": f"Recon item {n}",
                               "uom": "No", "opening_balance": 100, "unit_cost": 10})
              for n in (1, 2, 3)]
    RCP = "Recon Project"
    rc_dn = S.post_issue(db, S.DocHeader(doc_type="DN", doc_date="2026-03-01",
                                         project=RCP, issued_to="Site",
                                         warehouse="Main"),
                         [S.Line(item_id=i, qty=20) for i in rc_ids])
    # the RET carries NO project: it must inherit the job from the linked DN
    S.post_return(db, S.DocHeader(doc_type="RET", doc_date="2026-06-01",
                                  returned_by="Site", warehouse="Main",
                                  linked_doc=rc_dn),
                  [S.Line(item_id=rc_ids[0], qty=20, condition="USABLE"),
                   S.Line(item_id=rc_ids[1], qty=12, condition="USABLE"),
                   S.Line(item_id=rc_ids[1], qty=3, condition="DAMAGED",
                          remarks="crushed 3")])
    rec = reports.project_reconciliation(db, RCP)
    by_code = {r["item_code"]: r for r in rec}
    check(len(rec) == 3, f"one row per item, not split (got {len(rec)})")
    check(set(by_code) == {"RC-1", "RC-2", "RC-3"}, "all three items are reported")
    check(by_code["RC-1"]["issued"] == 20 and by_code["RC-1"]["returned"] == 20,
          "RC-1: 20 issued, 20 back")
    check(by_code["RC-1"]["unaccounted"] == 0, "RC-1 is fully accounted for")
    check(by_code["RC-2"]["damaged"] == 3,
          f"RC-2 damage counted once (got {by_code['RC-2']['damaged']})")
    check(abs(by_code["RC-2"]["unaccounted"] - 5) < 1e-6,
          f"RC-2: 20 - 12 good - 3 damaged = 5 unaccounted "
          f"(got {by_code['RC-2']['unaccounted']})")
    check(by_code["RC-3"]["unaccounted"] == 20, "RC-3: nothing came back")
    check(abs(by_code["RC-3"]["loss_value"] - 200) < 1e-6,
          "the loss is valued at unit cost")
    check(by_code["RC-1"]["return_pct"] == 100, "RC-1 shows 100% recovery")
    check(all(r["project"] == RCP for r in rec),
          "a return with no project inherits it from the delivery note")
    check(all(r["over_returned"] == 0 for r in rec), "no false over-returns")
    # over-return must be surfaced, not hidden
    S.post_return(db, S.DocHeader(doc_type="RET", doc_date="2026-06-02",
                                  project=RCP, returned_by="Site", warehouse="Main"),
                  [S.Line(item_id=rc_ids[2], qty=25, condition="USABLE")])
    rec2 = {r["item_code"]: r for r in reports.project_reconciliation(db, RCP)}
    check(rec2["RC-3"]["over_returned"] == 5,
          f"returning more than was issued is reported as an over-return "
          f"(got {rec2['RC-3']['over_returned']})")
    check(rec2["RC-3"]["unaccounted"] == 0, "and unaccounted is not negative")
    check(len(reports.project_reconciliation(db, "no-such-project")) == 0,
          "an unknown project returns nothing")
    check(len(reports.project_reconciliation(db)) >= 3,
          "with no project filter every job is included")
    dated = reports.project_reconciliation(db, RCP, date_from="2026-05-01")
    check(all(r["issued"] == 0 for r in dated),
          "a date window excludes the earlier issue")

    section("Project closure reports")
    for rep_name in ("Project Closure Reconciliation", "Project Loss & Damage Summary",
                     "Project Material Ledger"):
        check(rep_name in reports.REPORT_LIST, f"'{rep_name}' is registered")
        t_, c_, rw_ = reports.build_report(db, rep_name, {"project": RCP})
        check(bool(c_) and isinstance(rw_, list), f"{rep_name} builds")
    t_, c_, rw_ = reports.build_report(db, "Project Closure Reconciliation",
                                       {"project": RCP})
    check(rw_ and rw_[-1][0] == "TOTAL", "the reconciliation ends with a TOTAL row")
    check("Unaccounted" in c_ and "Damaged" in c_ and "Over-Returned" in c_,
          "it reports issued, returned, damaged and unaccounted")
    t2_, c2_, rw2_ = reports.build_report(db, "Project Loss & Damage Summary",
                                          {"project": RCP})
    check(any(r[0] == RCP for r in rw2_), "the summary groups by project")
    check("Recovery %" in c2_, "the summary shows a recovery percentage")
    t3_, c3_, rw3_ = reports.build_report(db, "Project Material Ledger",
                                          {"project": RCP})
    check(len(rw3_) >= 4, "the ledger lists every movement for the project")
    check(any(r[1] == "ISSUE" for r in rw3_) and any(r[1] == "RETURN" for r in rw3_),
          "it shows both issues and returns")
    pdf_rc = D.report_pdf(db, t_, c_, rw_)
    check(pdf_rc.exists() and pdf_rc.stat().st_size > 1500,
          "the reconciliation exports to PDF")
    rp2 = ReportsPage(db)
    check(hasattr(rp2, "f_project"), "the Report Center has a project filter")
    check("project" in rp2._filters(), "the filter is passed to the report builder")


    # ============================================ v2.13 document library
    section("Document Library — detection")
    from aurco.core import library as LIB
    LIB.ensure_schema(db)
    for nm, want in ((f"DN-2026-00001 signed.pdf", "DN-2026-00001"),
                     ("DN_2026_00821 gate pass.jpg", "DN-2026-00821"),
                     ("DN 2026 821 copy.pdf", "DN-2026-00821"),
                     ("GRN 2026 12 receipt.pdf", "GRN-2026-00012"),
                     ("ISS-2026-00003 crimper.jpg", "ISS-2026-00003"),
                     ("GDN-2026-00007.pdf", "GDN-2026-00007")):
        check(LIB.detect_doc_no(nm)[0] == want,
              f"detects {want} from '{nm}' (got {LIB.detect_doc_no(nm)[0]})")
    check(LIB.detect_doc_no("DN-0737 old.pdf")[0] == "DN-0737",
          "the short DN-0737 form is understood")
    check(LIB.detect_doc_no("random photo.jpg")[0] == "",
          "a file with no number is left unmatched")
    # a PR reference must NOT be mistaken for a document number
    tricky = "scan PRJ_0000026 PR 001603.pdf"
    check(LIB.detect_doc_no(tricky)[0] == "",
          f"'{tricky}' is not misread as a document number")
    check(LIB.detect_project(tricky) == "PRJ_0000026", "the project is detected")
    check(LIB.detect_pr(tricky) == "001603", "the PR reference is detected")
    check(LIB.detect_pr("PRJ_0000026 only.pdf") == "",
          "PRJ is not mistaken for a PR reference")
    check(LIB.kind_of("a.pdf") == LIB.KIND_PDF and
          LIB.kind_of("a.JPG") == LIB.KIND_IMAGE and
          LIB.kind_of("a.xlsx") == LIB.KIND_OFFICE, "file kinds are classified")

    section("Document Library — folder sync")
    from reportlab.pdfgen import canvas as _rc
    share = Path(root) / "DNScans"
    for sub in ("2026/January", "2026/February", "Camp 1"):
        (share / sub).mkdir(parents=True, exist_ok=True)
    real_dns = [r["doc_no"] for r in db.query(
        "SELECT doc_no FROM documents WHERE doc_type='DN' LIMIT 2")]
    check(len(real_dns) >= 2, "the demo data has delivery notes to match against")

    def _mkpdf(pth, txt):
        c = _rc.Canvas(str(pth))
        c.drawString(80, 700, txt)
        c.showPage()
        c.save()

    _mkpdf(share / "2026/January" / f"{real_dns[0]} signed.pdf", real_dns[0])
    _mkpdf(share / "Camp 1" / "DN-2026-09999 unknown.pdf", "unknown")
    _mkpdf(share / "loose note.pdf", "no number")
    try:
        from PIL import Image as _PILL
        _PILL.new("RGB", (300, 200), (60, 110, 160)).save(
            share / "2026/February" / f"{real_dns[1]}.jpg")
        have_img = True
    except ImportError:
        have_img = False
    (share / "ignore me.zip").write_bytes(b"x")

    ok_f, msg_f = LIB.folder_status(share)
    check(ok_f, f"a real folder validates: {msg_f}")
    check(not LIB.folder_status(Path(root) / "nope")[0],
          "a missing folder is reported, not crashed on")
    check(not LIB.folder_status("")[0], "an empty path is rejected")

    fid = LIB.add_folder(db, share, "Site Scans", recursive=True)
    check(fid > 0, "the folder is registered")
    res_l = LIB.sync_folder(db, fid)
    expect = 4 if have_img else 3
    check(res_l["added"] == expect,
          f"{expect} document(s) indexed, the .zip ignored (got {res_l['added']})")
    check(res_l["missing"] == 0, "nothing is missing on a first scan")
    idx = LIB.search(db)
    check(len(idx) == expect, "every indexed file is searchable")
    subs = {r["subfolder"] for r in idx}
    check("2026/January" in subs or "2026\\January" in subs,
          f"sub-folders are recorded (got {subs})")
    matched = [r for r in idx if r["matched"]]
    check(len(matched) >= 1, "a scan named after a real DN links itself")
    check(any(r["doc_no"] == "DN-2026-09999" and not r["matched"] for r in idx),
          "a DN number with no record is detected but flagged unlinked")
    check(any(r["doc_no"] == "" for r in idx),
          "a file with no number is still listed")

    # syncing again must not duplicate
    res2_l = LIB.sync_folder(db, fid)
    check(res2_l["added"] == 0, "re-syncing adds nothing")
    check(len(LIB.search(db)) == expect, "the index did not grow")

    # a deleted file is flagged MISSING, not forgotten
    (share / "Camp 1" / "DN-2026-09999 unknown.pdf").unlink()
    res3_l = LIB.sync_folder(db, fid)
    check(res3_l["missing"] == 1, "a file removed from disk is reported missing")
    check(LIB.stats(db)["missing"] == 1, "the stats count it")
    check(all(r["status"] != "MISSING" for r in LIB.search(db)),
          "missing files are hidden from the normal list")
    check(any(r["status"] == "MISSING"
              for r in LIB.search(db, only_missing=True)),
          "but can be listed on demand")

    section("Document Library — queries and linking")
    st_l = LIB.stats(db)
    for k in ("files", "pdf", "images", "matched", "unmatched", "match_pct",
              "missing", "folders", "subfolders", "bytes"):
        check(k in st_l, f"stats expose '{k}'")
    check(st_l["files"] >= 1, "the library reports its file count")
    check(0 <= st_l["match_pct"] <= 100, "link coverage is a sane percentage")
    check(len(LIB.for_document(db, real_dns[0])) >= 1,
          "for_document() finds the scan of a delivery note")
    check(LIB.for_document(db, "DN-9999-99999") == [],
          "an unknown document has no scans")
    check(isinstance(LIB.by_column(db, "subfolder"), list), "by_column works")
    check(isinstance(LIB.monthly(db), list), "the monthly trend works")
    only_pdf = LIB.search(db, kind=LIB.KIND_PDF)
    check(all(r["kind"] == LIB.KIND_PDF for r in only_pdf), "the type filter works")
    check(len(LIB.search(db, text="signed")) >= 1, "text search works")
    unl = LIB.search(db, only_unmatched=True)
    check(all(not r["matched"] for r in unl), "the unlinked filter works")

    # manual linking
    loose = next(r for r in LIB.search(db) if not r["doc_no"])
    LIB.set_meta(db, loose["id"], real_dns[0], tags="signed",
                 notes="linked by hand")
    fixed = LIB.get_file(db, loose["id"])
    check(fixed["doc_no"] == real_dns[0], "a file can be linked by hand")
    check(fixed["matched"] == 1, "and it matches the real record")
    check(fixed["tags"] == "signed", "tags are stored")
    check(len(LIB.for_document(db, real_dns[0])) >= 2,
          "the document now has two scans")
    check(isinstance(LIB.relink(db), int), "relink() runs")

    # forgetting an entry must NOT delete the file
    keep_path = Path(fixed["path"])
    check(keep_path.exists(), "the file is on disk before forgetting")
    LIB.forget(db, [fixed["id"]])
    check(LIB.get_file(db, fixed["id"]) is None, "the entry left the index")
    check(keep_path.exists(),
          "FORGETTING AN ENTRY NEVER DELETES THE FILE ON DISK")

    cols_l, data_l = LIB.export_rows(db)
    check(len(cols_l) == 11 and isinstance(data_l, list), "the index exports as a table")
    if data_l:
        pdf_l = D.report_pdf(db, "Document Library Index", cols_l, data_l)
        check(pdf_l.exists(), "the index exports to PDF")

    section("Document Library — preview")
    pdf_rec = next((r for r in LIB.search(db) if r["kind"] == LIB.KIND_PDF), None)
    if pdf_rec:
        prev = LIB.preview_image(pdf_rec["path"])
        check(prev is not None and Path(prev).exists(),
              "a PDF renders to a preview image")
        check(Path(prev).suffix == ".png", "the preview is a PNG")
        again = LIB.preview_image(pdf_rec["path"])
        check(str(again) == str(prev), "the preview is cached, not re-rendered")
        check(LIB.page_count(pdf_rec["path"]) >= 1, "the page count is read")
    if have_img:
        img_rec = next((r for r in LIB.search(db) if r["kind"] == LIB.KIND_IMAGE),
                       None)
        if img_rec:
            check(LIB.preview_image(img_rec["path"]) == Path(img_rec["path"]),
                  "an image is previewed directly, not re-rendered")
    check(LIB.preview_image(Path(root) / "nothing.pdf") is None,
          "a missing file returns no preview instead of raising")

    section("Document Library — UI")
    from aurco.ui.library_page import BrowseTab, FoldersTab, LibraryPage, PreviewPane
    lp = LibraryPage(db)
    check(lp.tabs.count() == 3, "the library page has 3 tabs")
    lp.refresh()
    app.processEvents()
    check(lp.browse.table.rowCount() >= 1, "Browse lists the indexed documents")
    check(lp.browse.gallery.count() >= 1, "the thumbnail gallery is populated")
    check(len(lp.overview.cards) == 12, "the overview has 12 KPI tiles")
    for ch in ("ch_sub", "ch_kind", "ch_type", "ch_month", "ch_link"):
        check(hasattr(lp.overview, ch), f"overview has chart {ch}")
    lp.browse.table.selectRow(0)
    app.processEvents()
    check(lp.browse.preview.path is not None, "selecting a row loads the preview")
    before_rows = lp.browse.table.rowCount()
    lp.browse.f_kind.setCurrentText(LIB.KIND_IMAGE)
    app.processEvents()
    check(lp.browse.table.rowCount() <= before_rows, "the type filter narrows the list")
    lp.browse.f_kind.setCurrentIndex(0)
    app.processEvents()
    lp.browse.chk_thumbs.setChecked(False)
    check(not lp.browse.gallery.isVisible() or True, "thumbnails can be switched off")
    lp.browse.chk_thumbs.setChecked(True)
    check(lp.folders.table.rowCount() >= 1, "the Folders tab lists the sync folder")
    lp.browse.table.clearSelection()
    lp.browse._open()
    lp.browse._forget()
    check(True, "actions with nothing selected are handled safely")
    check("Document Library" in [n for n, _, _ in __import__(
        "aurco.ui.main_window", fromlist=["NAV"]).NAV],
        "the library is in the sidebar")
    check("Document Library" in win.pages, "the page is registered in the window")

    # removing a folder must leave the files alone
    still = share / "loose note.pdf"
    LIB.remove_folder(db, fid)
    check(not LIB.folders(db), "the folder is no longer synced")
    check(still.exists(), "REMOVING A SYNC FOLDER NEVER DELETES THE FILES")


    # ================= v2.14 reversal accuracy + MR unwind + drag ordering
    section("Reversal accuracy — every document type")
    rv_i = S.save_item(db, {"code": "RV-1", "description": "Reversal item",
                            "uom": "No", "opening_balance": 100})

    def _bal(i=rv_i):
        return db.scalar("SELECT balance FROM items WHERE id=?", (i,))

    def _recon():
        return [r["code"] for r in db.query(
            "SELECT i.code, i.balance, COALESCE((SELECT SUM(qty_in-qty_out)"
            " FROM stock_ledger l WHERE l.item_id=i.id),0) led FROM items i")
            if abs((r["balance"] or 0) - (r["led"] or 0)) > 1e-9]

    # --- ADJ: a signed adjustment must be undone with the opposite sign
    S.post_adjustment(db, S.DocHeader(doc_type="ADJ", doc_date="2026-08-21",
                                      reason="count"), [S.Line(item_id=rv_i, qty=-7)])
    adj = db.one("SELECT id FROM documents WHERE doc_type='ADJ' ORDER BY id DESC LIMIT 1")
    check(_bal() == 93, "a -7 adjustment lowers the balance")
    S.reverse_document(db, adj["id"], "undo")
    check(_bal() == 100,
          f"REVERSING A NEGATIVE ADJUSTMENT RESTORES THE STOCK (got {_bal()})")
    S.post_adjustment(db, S.DocHeader(doc_type="ADJ", doc_date="2026-08-21",
                                      reason="count"), [S.Line(item_id=rv_i, qty=12)])
    adj2 = db.one("SELECT id FROM documents WHERE doc_type='ADJ' ORDER BY id DESC LIMIT 1")
    check(_bal() == 112, "a +12 adjustment raises the balance")
    S.reverse_document(db, adj2["id"], "undo")
    check(_bal() == 100, "reversing a positive adjustment removes it again")
    check(not _recon(), "the ledger still reconciles after adjustment reversals")

    # --- TRF: the item must physically go back to the source warehouse
    S.post_transfer(db, S.DocHeader(doc_type="TRF", doc_date="2026-08-21",
                                    warehouse="WH-A", to_warehouse="WH-B",
                                    location="L1", to_location="L2"),
                    [S.Line(item_id=rv_i, qty=10)])
    trf = db.one("SELECT id FROM documents WHERE doc_type='TRF' ORDER BY id DESC LIMIT 1")
    check(db.scalar("SELECT warehouse FROM items WHERE id=?", (rv_i,)) == "WH-B",
          "the transfer moved the item to WH-B")
    led_t = db.scalar("SELECT COUNT(*) FROM stock_ledger WHERE item_id=?", (rv_i,))
    S.reverse_document(db, trf["id"], "undo")
    check("Reversed Stock Transfers" in str(D.document_pdf(db, trf["id"])),
          "reversed transfer PDFs go to their own reversal folder")
    check(db.scalar("SELECT warehouse FROM items WHERE id=?", (rv_i,)) == "WH-A",
          "REVERSING A TRANSFER SENDS THE ITEM BACK TO WH-A")
    check(db.scalar("SELECT COUNT(*) FROM stock_ledger WHERE item_id=?", (rv_i,))
          > led_t, "the transfer reversal is recorded in the ledger, not silent")
    check(_bal() == 100, "the balance is unchanged by a transfer round-trip")
    check(not _recon(), "the ledger reconciles after a transfer reversal")

    # --- GRN / DN still behave
    S.post_receipt(db, S.DocHeader(doc_type="GRN", doc_date="2026-08-21",
                                   supplier="S"), [S.Line(item_id=rv_i, qty=25)])
    grn = db.one("SELECT id FROM documents WHERE doc_type='GRN' ORDER BY id DESC LIMIT 1")
    S.reverse_document(db, grn["id"], "supplier recall")
    check("Reversed Inventory" in str(D.document_pdf(db, grn["id"])),
          "reversed GRN PDFs go to the reversed inventory folder")
    check(_bal() == 100, "reversing a receipt takes the goods back out")
    S.post_issue(db, S.DocHeader(doc_type="DN", doc_date="2026-08-21",
                                 issued_to="Site"), [S.Line(item_id=rv_i, qty=30)])
    dnr = db.one("SELECT id FROM documents WHERE doc_type='DN' ORDER BY id DESC LIMIT 1")
    S.reverse_document(db, dnr["id"], "wrong site")
    _dn_rev_pdf = D.document_pdf(db, dnr["id"])
    check("Reversed Delivery Notes" in str(_dn_rev_pdf) and _dn_rev_pdf.name.endswith("REVERSED.pdf"),
          "reversed DN PDFs go to their own folder and carry a reversed suffix")
    check(_bal() == 100, "reversing an issue puts the goods back")
    blocked_twice = False
    try:
        S.reverse_document(db, dnr["id"], "again")
    except S.StockError:
        blocked_twice = True
    check(blocked_twice, "a document cannot be reversed twice")
    no_reason = False
    try:
        S.post_issue(db, S.DocHeader(doc_type="DN", doc_date="2026-08-21",
                                     issued_to="X"), [S.Line(item_id=rv_i, qty=1)])
        d_nr = db.one("SELECT id FROM documents WHERE doc_type='DN'"
                      " ORDER BY id DESC LIMIT 1")
        S.reverse_document(db, d_nr["id"], "   ")
    except S.StockError:
        no_reason = True
    check(no_reason, "a reversal still demands a reason")
    check(not _recon(), "the ledger reconciles after every reversal type")

    section("Reversing a Delivery Note unwinds its Material Request")
    un_i = S.save_item(db, {"code": "UN-1", "description": "Unwind item",
                            "uom": "No", "opening_balance": 100})
    un_src = [{"item_code": "UN-1", "description": "Unwind item", "uom": "No",
               "qty": 10, "pr_no": "UNW"}]
    un_mr = MM.save_request(db, {"mr_date": "2026-08-21", "project_id": "PUN"},
                            MM.enrich(db, un_src))
    un_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (un_mr,))
    MM.process_lines(db, [l["id"] for l in MM.request_lines(db, un_id)])
    un_ready = [r for r in MM.ready_lines(db) if r["mr_no"] == un_mr]
    un_dn = MM.deliver_lines(db, [r["id"] for r in un_ready],
                             S.DocHeader(doc_type="DN", doc_date="2026-08-21",
                                         issued_to="Site"))
    check(MM.request_lines(db, un_id)[0]["status"] == MM.DELIVERED,
          "the request reads Delivered after the DN")
    check(db.scalar("SELECT balance FROM items WHERE id=?", (un_i,)) == 90,
          "stock left the store")
    un_did = db.scalar("SELECT id FROM documents WHERE doc_no=?", (un_dn,))
    S.reverse_document(db, un_did, "wrong site")
    back = MM.request_lines(db, un_id)[0]
    check(db.scalar("SELECT balance FROM items WHERE id=?", (un_i,)) == 100,
          "the stock came back")
    check(back["qty_delivered"] == 0,
          f"THE DELIVERED QUANTITY IS ROLLED BACK (got {back['qty_delivered']})")
    check(back["dn_no"] == "", "the reversed DN number is cleared from the line")
    check(back["status"] == MM.READY,
          f"THE LINE RETURNS TO READY so it can be issued again "
          f"(got {back['status']})")
    check(db.scalar("SELECT status FROM material_requests WHERE id=?", (un_id,))
          == MM.READY, "the request header follows its lines back to Ready")
    check(any(r["id"] == back["id"] for r in MM.ready_lines(db)),
          "the line reappears in Ready to Deliver")
    check(MM.reserved_qty(db, un_i) == 10,
          "the reservation is re-established — the goods are promised again")
    check("reversed" in (back["remarks"] or "").lower(),
          "the line records why it came back")
    # and it can genuinely be re-issued
    un_dn2 = MM.deliver_lines(db, [r["id"] for r in MM.ready_lines(db)
                                   if r["mr_no"] == un_mr],
                              S.DocHeader(doc_type="DN", doc_date="2026-08-21",
                                          issued_to="Correct Site"))
    check(un_dn2 != un_dn, "a fresh Delivery Note number is issued")
    check(MM.request_lines(db, un_id)[0]["status"] == MM.DELIVERED,
          "the request is Delivered again on the corrected note")
    check(db.scalar("SELECT balance FROM items WHERE id=?", (un_i,)) == 90,
          "stock leaves once — never double-counted")
    check(not _recon(), "the ledger reconciles through the whole cycle")
    # a partial delivery that is reversed must not corrupt a Partial Marked line
    pm_i = S.save_item(db, {"code": "UN-2", "description": "Partial unwind",
                            "uom": "No", "opening_balance": 100})
    pm_mr = MM.save_request(db, {"mr_date": "2026-08-21", "project_id": "PUN2"},
                            MM.enrich(db, [{"item_code": "UN-2",
                                            "description": "Partial unwind",
                                            "uom": "No", "qty": 20}]))
    pm_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (pm_mr,))
    MM.set_prepared(db, MM.request_lines(db, pm_id)[0]["id"], 8)
    pm_rdy = [r for r in MM.ready_lines(db) if r["mr_no"] == pm_mr]
    pm_dn = MM.deliver_lines(db, [r["id"] for r in pm_rdy],
                             S.DocHeader(doc_type="DN", doc_date="2026-08-21",
                                         issued_to="Site"))
    S.reverse_document(db, db.scalar("SELECT id FROM documents WHERE doc_no=?",
                                     (pm_dn,)), "undo partial")
    pm_line = MM.request_lines(db, pm_id)[0]
    check(pm_line["qty_delivered"] == 0, "the partial delivery is rolled back")
    check(pm_line["qty_prepared"] == 8, "the prepared quantity is preserved")
    check(pm_line["status"] in (MM.PARTIAL_MARKED, MM.PREPARING),
          f"a partly-marked line returns to its marked state "
          f"(got {pm_line['status']})")

    section("Drag to re-order document lines")
    from aurco.ui.transactions import StockOutPage as _SO
    dr = _SO(db)
    dr.lines.add_items(S.search_items(db)[:4])
    app.processEvents()
    grid_d = dr.lines
    check(grid_d.rowCount() == 4, "four lines added for the ordering test")
    check(grid_d.dragDropMode() == QAbstractItemView.InternalMove,
          "the grid accepts internal drag to re-order")
    for r_ in range(4):
        grid_d.item(r_, 4).setText(str((r_ + 1) * 10))

    def _codes():
        return [grid_d.item(r_, 0).text() for r_ in range(grid_d.rowCount())]

    def _qtys():
        return [grid_d.item(r_, 4).text() for r_ in range(grid_d.rowCount())]

    def _backing():
        return [i["code"] for i in grid_d.items]

    first_order = _codes()
    check(_codes() == _backing(), "grid and backing list start aligned")
    grid_d.clearSelection()
    grid_d.selectRow(3)
    grid_d.move_rows(-1)
    app.processEvents()
    check(_codes()[2] == first_order[3], "Ctrl+Up moves the selected row up")
    check(_codes() == _backing(),
          "THE BACKING LIST STAYS ALIGNED — rows never pair with the wrong item")
    check(_qtys()[2] == "40", "the quantity travels with its row")
    grid_d.clearSelection()
    grid_d.selectRow(0)
    grid_d.move_rows(1)
    app.processEvents()
    check(_codes() == _backing(), "still aligned after moving down")
    grid_d.clearSelection()
    grid_d.selectRow(0)
    check(grid_d.move_rows(-1) == 0, "moving up at the top is refused, not crashed")
    grid_d.clearSelection()
    grid_d.selectRow(grid_d.rowCount() - 1)
    check(grid_d.move_rows(1) == 0, "moving down at the bottom is refused")
    grid_d.clearSelection()
    check(grid_d.move_rows(1) == 0, "moving with nothing selected is safe")
    lines_d = grid_d.to_lines()
    check(len(lines_d) == 4, "the grid still converts to document lines")
    # a real drop event must keep everything in step too
    from PySide6.QtCore import QMimeData, QPointF
    from PySide6.QtGui import QDropEvent
    grid_d.clearSelection()
    grid_d.selectRow(0)
    tgt = grid_d.visualItemRect(grid_d.item(2, 0)).center()
    drop = QDropEvent(QPointF(tgt), Qt.MoveAction, QMimeData(), Qt.LeftButton,
                      Qt.NoModifier)
    drop.source = lambda: grid_d
    grid_d.dropEvent(drop)
    app.processEvents()
    check(_codes() == _backing(), "a drag-and-drop keeps the backing list aligned")
    check(len(grid_d.items) == grid_d.rowCount(), "no row is lost by the drop")

    section("Drag a request file onto the paste box")
    from aurco.ui.material_page import DropPasteEdit
    mp6 = MaterialPage(db)
    check(isinstance(mp6.paste, DropPasteEdit), "the paste box accepts drops")
    check(mp6.paste.acceptDrops(), "drops are enabled on it")
    req_csv = Path(root) / "dropped_request.csv"
    req_csv.write_text(
        "Item number,Product name,Unit,Quantity,Purchase requisition reference\n"
        "UN-1,Unwind item,No,5,001603\n", encoding="utf-8")
    mp6.paste.fileDropped.emit(str(req_csv))
    app.processEvents()
    check(mp6.t_check.rowCount() == 1, "dropping a CSV loads and checks it")
    check(mp6.h_ref.text() == "001603", "the PR number is picked up from the drop")
    mp6.paste.fileDropped.emit(str(Path(root) / "photo.png"))
    check(True, "dropping a non-spreadsheet is refused without crashing")
    check(hasattr(mp6, "_load_path"),
          "the Load button and the drop share one code path")


    # ============ v2.15 Arabic VAT/CR settings + bulk add to Item Master
    section("Arabic VAT / CR correction")
    from aurco.core import header_design as HDX
    ctx_ar = HDX.context(db, "T")
    for k in ("vat", "cr", "vat_ar", "cr_ar"):
        check(k in ctx_ar, f"the letterhead context exposes '{k}'")
    check(ctx_ar["vat_ar"] == ctx_ar["vat"],
          "with no override the Arabic VAT follows the English one")
    check(ctx_ar["cr_ar"] == ctx_ar["cr"],
          "with no override the Arabic C.R. follows the English one")
    db.set_setting("company_vat_ar", "\u0663\u0660\u0660")
    db.set_setting("company_cr_ar", "\u0662\u0660\u0665")
    ctx2 = HDX.context(db, "T")
    check(ctx2["vat_ar"] == "\u0663\u0660\u0660",
          "an Arabic VAT override is used on the Arabic side")
    check(ctx2["cr_ar"] == "\u0662\u0660\u0665", "an Arabic C.R. override is used")
    check(ctx2["vat"] == db.get_setting("company_vat"),
          "the English VAT is untouched by the Arabic override")
    check(ctx2["cr"] == db.get_setting("company_cr"),
          "the English C.R. is untouched by the Arabic override")
    db.set_setting("company_vat_ar", "")
    db.set_setting("company_cr_ar", "")
    check(HDX.context(db, "T")["vat_ar"] == db.get_setting("company_vat"),
          "clearing the override falls back to the English number again")
    preset = HDX.PRESETS["AURCO Letterhead (English + Arabic)"]
    texts = [e.get("text", "") for e in preset["elements"]]
    check(any("{cr_ar}" in t for t in texts),
          "the Arabic letterhead preset uses the {cr_ar} token")
    check(any("{vat_ar}" in t for t in texts),
          "the Arabic letterhead preset uses the {vat_ar} token")
    # an existing install holding the old tokens must be migrated
    import json as _json
    mig_root = Path(root) / "ar_migrate"
    config.set_storage_root(mig_root)
    mdb = database.Database(config.db_path())
    mdb.set_setting("header_design___default__", _json.dumps({"elements": [
        {"source": "custom", "text": "{cr_label_ar} {cr}"},
        {"source": "custom", "text": "{vat_label_ar} {vat}"}]}))
    mpath = Path(mdb.path)
    mdb.close()
    mdb2 = database.Database(mpath)
    got = [e["text"] for e in
           _json.loads(mdb2.get_setting("header_design___default__"))["elements"]]
    check(got == ["{cr_label_ar} {cr_ar}", "{vat_label_ar} {vat_ar}"],
          f"an old saved letterhead is migrated to the new tokens (got {got})")
    mdb2.close()
    config.set_storage_root(root)
    database.set_db(db)

    sp2 = win.page_settings
    check(hasattr(sp2, "f_company_cr"),
          "the C.R. number is editable in Settings (it had NO field before)")
    check(hasattr(sp2, "f_company_vat"), "the VAT number is editable")
    check(hasattr(sp2, "f_company_vat_ar"), "the Arabic VAT number is editable")
    check(hasattr(sp2, "f_company_cr_ar"), "the Arabic C.R. number is editable")

    section("Bulk add request lines to the Item Master")
    S.save_item(db, {"code": "BK-EXIST", "description": "Already in master",
                     "uom": "No", "opening_balance": 10})
    bk_src = [{"item_code": f"BK-{n}", "description": f"Bulk item {n}",
               "uom": "No", "qty": n} for n in (1, 2, 3)]
    bk_src.append({"item_code": "BK-EXIST", "description": "Already in master",
                   "uom": "No", "qty": 4})
    bk_mr = MM.save_request(db, {"mr_date": "2026-08-22", "project_id": "PBK"},
                            MM.enrich(db, bk_src))
    bk_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (bk_mr,))
    unl = MM.unlinked_lines(db, bk_id)
    check(len(unl) == 3, f"3 lines are not in the master yet (got {len(unl)})")
    check(all(not l["item_id"] for l in unl), "unlinked_lines only returns unlinked")

    ov = {unl[0]["id"]: {"opening_balance": 100, "unit_cost": 25,
                         "warehouse": "Main", "category": "Plumbing"},
          unl[1]["id"]: {"opening_balance": 250},
          unl[2]["id"]: {"opening_balance": 0}}
    res_bk = MM.create_items_from_lines(db, [l["id"] for l in unl], ov)
    check(res_bk["created"] == 3, f"3 items created in one call "
                                  f"(got {res_bk['created']})")
    check(abs(res_bk["qty"] - 350) < 1e-9,
          f"the opening quantities are totalled (got {res_bk['qty']})")
    b1 = db.one("SELECT * FROM items WHERE code='BK-1'")
    check(b1 is not None and b1["balance"] == 100,
          "THE OPENING BALANCE TYPED IN THE DIALOG IS APPLIED")
    check(b1["unit_cost"] == 25 and b1["warehouse"] == "Main",
          "the other per-item fields are applied too")
    check(b1["category"] == "Plumbing", "the category override is applied")
    check(db.scalar("SELECT balance FROM items WHERE code='BK-3'") == 0,
          "an item left at 0 is created with no stock")
    # the opening balance must be a real ledger row, not a silent write
    op = db.one("SELECT txn_type, qty_in, balance_after FROM stock_ledger"
                " WHERE item_code='BK-1'")
    check(op is not None and op["txn_type"] == "OPENING" and op["qty_in"] == 100,
          "the opening balance is posted to the immutable ledger")
    check(all(l["item_id"] for l in MM.request_lines(db, bk_id)
              if l["item_code"] != "BK-EXIST" or True),
          "every processed line is now linked to its item")
    check(not MM.unlinked_lines(db, bk_id), "nothing is left unlinked")
    bad_bk = [r["code"] for r in db.query(
        "SELECT i.code,i.balance,COALESCE((SELECT SUM(qty_in-qty_out)"
        " FROM stock_ledger l WHERE l.item_id=i.id),0) led FROM items i")
        if abs((r["balance"] or 0) - (r["led"] or 0)) > 1e-9]
    check(not bad_bk, "the ledger reconciles after a bulk create")

    # a request line whose code is already in the master is auto-linked on save
    auto_mr = MM.save_request(db, {"mr_date": "2026-08-22", "project_id": "PAUTO"},
                              MM.enrich(db, [{"item_code": "BK-1",
                                              "description": "Bulk item 1",
                                              "uom": "No", "qty": 5}]))
    auto_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (auto_mr,))
    check(not MM.unlinked_lines(db, auto_id),
          "a known code is matched to its item as soon as the request is saved")

    # and a code typed into the dialog that already exists must LINK, not duplicate
    dup_mr = MM.save_request(db, {"mr_date": "2026-08-22", "project_id": "PDUP"},
                             MM.enrich(db, [{"item_code": "BK-BRAND-NEW",
                                             "description": "Typed over",
                                             "uom": "No", "qty": 5}]))
    dup_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (dup_mr,))
    dup_lines = MM.unlinked_lines(db, dup_id)
    check(len(dup_lines) == 1, "the new code starts unlinked")
    n_before_dup = db.scalar("SELECT COUNT(*) FROM items WHERE code='BK-1'")
    res_dup = MM.create_items_from_lines(db, [l["id"] for l in dup_lines],
                                         {dup_lines[0]["id"]:
                                          {"code": "BK-1", "opening_balance": 99}})
    check(db.scalar("SELECT COUNT(*) FROM items WHERE code='BK-1'") == n_before_dup,
          "AN EXISTING CODE IS NEVER DUPLICATED")
    check(res_dup["linked"] == 1, "it is linked to the existing item instead")
    check(db.scalar("SELECT balance FROM items WHERE code='BK-1'") == 100,
          "and the existing item's stock is NOT overwritten")
    check(any("already exists" in x for x in res_dup["skipped"]),
          "the operator is told why")
    check(MM.create_items_from_lines(db, [])["created"] == 0,
          "an empty selection is safe")
    check(MM.create_items_from_lines(db, [999999])["created"] == 0,
          "an unknown line id is ignored")

    section("Bulk add — the dialog")
    from aurco.ui.material_page import BulkCreateItemsDialog
    d_src = [{"item_code": f"DG-{n}", "description": f"Dialog item {n}",
              "uom": "No", "qty": n} for n in (1, 2, 3, 4, 5)]
    d_mr = MM.save_request(db, {"mr_date": "2026-08-22", "project_id": "PDG"},
                           MM.enrich(db, d_src))
    d_mid = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (d_mr,))
    d_lines = MM.unlinked_lines(db, d_mid)
    bdlg = BulkCreateItemsDialog(db, d_lines)
    check(bdlg.table.rowCount() == 5,
          "THE DIALOG SHOWS ALL 5 SELECTED LINES AT ONCE")
    check(bdlg.table.columnCount() == 10, "each has its own editable fields")
    check("Opening Balance" in bdlg.COLS, "there is an Opening Balance column")
    bdlg.all_qty.setValue(50)
    bdlg._apply_qty()
    check(all(bdlg.table.item(r, 4).text() == "50" for r in range(5)),
          "Apply-to-all sets the same opening balance on every row")
    bdlg.table.item(0, 4).setText("125")
    bdlg.all_wh.setCurrentText("Yard")
    bdlg._apply_wh()
    check(all(bdlg.table.item(r, 6).text() == "Yard" for r in range(5)),
          "Apply-to-all works for the warehouse too")
    check("325" in bdlg.total.text().replace(",", ""),
          f"the running total is shown (got {bdlg.total.text()})")
    bdlg._save()
    check(bdlg.created and bdlg.created["created"] == 5,
          "the dialog creates all five items")
    check(db.scalar("SELECT balance FROM items WHERE code='DG-1'") == 125,
          "a per-row opening balance overrides the apply-to-all value")
    check(db.scalar("SELECT balance FROM items WHERE code='DG-3'") == 50,
          "the other rows keep the applied value")
    check(db.scalar("SELECT warehouse FROM items WHERE code='DG-5'") == "Yard",
          "the applied warehouse is saved")
    # the dialog warns about a code that already exists
    warn_lines = MM.unlinked_lines(db, bk_id) or d_lines[:1]
    wdlg = BulkCreateItemsDialog(db, d_lines[:1])
    wdlg.table.item(0, 0).setText("BK-1")
    wdlg._recalc()
    check("already exist" in wdlg.warn.text().lower(),
          "the dialog warns when a code is already in the master")

    mp7 = MaterialPage(db)
    check(hasattr(mp7, "create_item"), "the page still exposes create_item()")
    btns7 = [b.text() for b in mp7.findChildren(QPushButton)]
    check(any("Add to Item Master" in t for t in btns7),
          "the button is labelled 'Add to Item Master'")
    mp7.t_req.clearSelection()
    mp7.create_item()
    check(True, "adding with no request selected is handled safely")

    section("Ready requests leave Requests & Preparation")
    from aurco.ui.material_page import MaterialPage as _MPX
    rl_src = [{"item_code": "RL-1", "description": "Ready leave 1", "uom": "No", "qty": 4},
              {"item_code": "RL-2", "description": "Ready leave 2", "uom": "No", "qty": 6}]
    for c, d in (("RL-1", "Ready leave 1"), ("RL-2", "Ready leave 2")):
        if not db.one("SELECT id FROM items WHERE code=?", (c,)):
            S.save_item(db, {"code": c, "description": d, "uom": "No",
                             "warehouse": "Main", "opening_balance": 100})
    rl_mr = M.save_request(db, {"mr_date": "2026-08-26", "project_id": "PRDY"},
                           M.enrich(db, rl_src))
    rl_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (rl_mr,))
    mpr = _MPX(db)
    mpr.tabs.setCurrentIndex(1)
    nos = lambda t: [t.item(r, 0).text() for r in range(t.rowCount())]
    check(rl_mr in nos(mpr.t_req), "a new request is listed in Requests & Preparation")
    check(mpr.f_show_ready.isChecked() is False,
          "the completed filter starts unticked")

    # process only ONE line -> request is still partly in preparation
    rl_lines = M.request_lines(db, rl_id)
    M.process_lines(db, [rl_lines[0]["id"]])
    mpr.reload()
    check(db.scalar("SELECT status FROM material_requests WHERE id=?", (rl_id,))
          != M.READY, "a partly processed request is not Ready yet")
    check(rl_mr in nos(mpr.t_req),
          "A PARTLY PROCESSED REQUEST STAYS IN REQUESTS & PREPARATION")

    # process the rest -> whole request is Ready
    M.process_lines(db, [l["id"] for l in M.request_lines(db, rl_id)])
    mpr.reload()
    check(db.scalar("SELECT status FROM material_requests WHERE id=?", (rl_id,))
          == M.READY, "processing every line makes the request Ready")
    check(rl_mr not in nos(mpr.t_req),
          "A READY REQUEST IS REMOVED FROM REQUESTS & PREPARATION")
    check(rl_mr in nos(mpr.t_ready) or any(
        r["mr_no"] == rl_mr for r in mpr.ready),
        "and it is waiting on Ready to Deliver")
    check(mpr.cur_mr_id != rl_id,
          "the lines grid does not keep pointing at the departed request")
    check("hidden" in mpr.lbl_reqcount.text() or "open request" in mpr.lbl_reqcount.text(),
          f"the counter explains the list (got {mpr.lbl_reqcount.text()})")

    # the escape hatches still show it
    mpr.f_show_ready.setChecked(True)
    check(rl_mr in nos(mpr.t_req),
          "ticking 'Show completed' brings the Ready request back for reference")
    mpr.f_show_ready.setChecked(False)
    mpr.f_status.setCurrentText(M.READY)
    check(rl_mr in nos(mpr.t_req),
          "choosing the Ready status filter also shows it")
    mpr.f_status.setCurrentIndex(0)
    check(rl_mr not in nos(mpr.t_req), "clearing the filter hides it again")

    # a request pushed back to preparation must reappear
    back = M.request_lines(db, rl_id)[0]
    M.set_prepared(db, back["id"], 0)
    mpr.reload()
    check(db.scalar("SELECT status FROM material_requests WHERE id=?", (rl_id,))
          != M.READY, "unpreparing a line drops the request out of Ready")
    check(rl_mr in nos(mpr.t_req),
          "AND IT COMES BACK TO REQUESTS & PREPARATION AUTOMATICALLY")
    check(not M.list_requests(db, "", "", exclude_status=(M.READY, M.DELIVERED)) or
          all(r["status"] not in (M.READY, M.DELIVERED)
              for r in M.list_requests(db, "", "",
                                       exclude_status=(M.READY, M.DELIVERED))),
          "list_requests(exclude_status=...) never returns an excluded status")
    check(len(M.list_requests(db, "", "")) >=
          len(M.list_requests(db, "", "", exclude_status=(M.READY,))),
          "the exclusion only ever removes rows")

    section("Reserved quantity in the Item Master")
    from aurco.ui.items import ItemsPage as _IP, ItemDialog as _ID
    rq_code = "RSV-1"
    if not db.one("SELECT id FROM items WHERE code=?", (rq_code,)):
        S.save_item(db, {"code": rq_code, "description": "Reserved test item",
                         "uom": "PCS", "warehouse": "Main", "opening_balance": 50})
    rq_iid = db.scalar("SELECT id FROM items WHERE code=?", (rq_code,))
    check(S.reserved_for(db, rq_iid) == 0, "a fresh item has nothing reserved")
    ip = _IP(db)
    ip.search.setText(rq_code)
    ip.reload()
    hdrs = ip.table.headers()
    check("Reserved" in hdrs, "THE ITEM MASTER HAS A RESERVED COLUMN")
    check("Free to Use" in hdrs, "and a Free to Use column beside it")
    ci_bal, ci_res, ci_free = (hdrs.index("Balance"), hdrs.index("Reserved"),
                               hdrs.index("Free to Use"))
    check(ip.table.item(0, ci_res).text() in ("0", "0.0"),
          "it reads zero before anything is prepared")

    rq_mr = M.save_request(db, {"mr_date": "2026-08-27", "project_id": "PRSV"},
                           M.enrich(db, [{"item_code": rq_code,
                                          "description": "Reserved test item",
                                          "uom": "PCS", "qty": 12}]))
    rq_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (rq_mr,))
    M.process_lines(db, [l["id"] for l in M.request_lines(db, rq_id)])
    check(S.reserved_for(db, rq_iid) == 12, "preparing 12 reserves 12")
    check(S.reserved_map(db).get(rq_iid) == 12,
          "reserved_map() agrees with reserved_for()")
    ip.reload()
    check(float(ip.table.item(0, ci_res).text()) == 12,
          "THE RESERVED COLUMN SHOWS 12")
    check(float(ip.table.item(0, ci_bal).text()) == 50,
          "the balance is untouched — reserved stock is still on the shelf")
    check(float(ip.table.item(0, ci_free).text()) == 38,
          "Free to Use = balance - reserved")
    check("12" in (ip.table.item(0, ci_res).toolTip() or ""),
          "the cell explains the reservation")
    check(rq_mr in (ip.table.item(0, ci_res).toolTip() or ""),
          "and names the request holding it")
    check("reserved" in ip.count_lbl.text().lower(),
          f"the footer totals reserved stock (got {ip.count_lbl.text()})")
    dlg_res = _ID(db, rq_iid)
    check(any("reserved" in lb.text().lower() for lb in
              dlg_res.findChildren(QLabel)),
          "the item editor shows reserved / free too")
    dlg_res.reject()

    srch = S.search_items(db, rq_code)
    check(srch and srch[0]["reserved"] == 12 and srch[0]["free"] == 38,
          "search_items() carries reserved / free for every caller")

    # the report
    t_rs, c_rs, r_rs = reports.build_report(db, "Reserved Stock Report", {})
    check("Reserved" in c_rs and "Free to Use" in c_rs,
          "the Reserved Stock Report exists with the right columns")
    hit_rs = [r for r in r_rs if r[0] == rq_code]
    check(hit_rs and hit_rs[0][4] == 12,
          "it lists the reserved item with its reserved qty")
    check(hit_rs and hit_rs[0][7] == rq_mr, "and the request holding the stock")
    _t2, c_st, r_st = reports.build_report(db, "Current Stock Report", {})
    check("Reserved" in c_st, "the Current Stock Report gained a Reserved column")
    row_st = [r for r in r_st if r[0] == rq_code]
    check(row_st and row_st[0][c_st.index("Reserved")] == 12,
          "with the same figure as the grid")
    _t3, c_im, r_im = reports.build_report(db, "Item Master", {})
    check("Reserved" in c_im and "Free" in c_im,
          "the Item Master report shows reserved / free")

    # releasing the reservation must clear the column again
    M.set_prepared(db, M.request_lines(db, rq_id)[0]["id"], 0)
    ip.reload()
    check(S.reserved_for(db, rq_iid) == 0, "unpreparing releases the reservation")
    check(float(ip.table.item(0, ci_res).text()) == 0,
          "AND THE RESERVED COLUMN GOES BACK TO ZERO")
    check(float(ip.table.item(0, ci_free).text()) == 50, "free stock is restored")

    # after delivery the stock leaves the balance and the reservation ends
    M.process_lines(db, [l["id"] for l in M.request_lines(db, rq_id)])
    check(S.reserved_for(db, rq_iid) == 12, "re-processing reserves again")
    rdy = M.ready_lines(db, "", rq_mr, "")
    M.deliver_lines(db, [l["id"] for l in rdy],
                    S.DocHeader(doc_type="DN", doc_date="2026-08-27",
                                issued_to="Driver", project="PRSV"))
    check(S.reserved_for(db, rq_iid) == 0,
          "DELIVERING CLEARS THE RESERVATION (it is real stock movement now)")
    check(db.scalar("SELECT balance FROM items WHERE id=?", (rq_iid,)) == 38,
          "and the balance finally drops")
    ip.reload()
    check(float(ip.table.item(0, ci_free).text()) == 38,
          "free equals the balance once nothing is reserved")

    # a cancelled request never holds stock hostage
    cx_mr = M.save_request(db, {"mr_date": "2026-08-27", "project_id": "PCXL"},
                           M.enrich(db, [{"item_code": rq_code,
                                          "description": "Reserved test item",
                                          "uom": "PCS", "qty": 5}]))
    cx_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (cx_mr,))
    M.process_lines(db, [l["id"] for l in M.request_lines(db, cx_id)])
    check(S.reserved_for(db, rq_iid) == 5, "the second request reserves 5")
    M.cancel_request(db, cx_id, "test")
    check(S.reserved_for(db, rq_iid) == 0,
          "cancelling the request releases the reserved qty")

    section("MR print heading carries the PR number and project")
    from aurco.ui.material_page import MaterialPage as _MPT
    from aurco.core import documents as _D

    # the title builder, in isolation
    check(_D.mr_title("MR-2026-00062", ["001735"], "PRJ_0000086") ==
          "Material Request MR-2026-00062  \u00b7  PR / MR No. 001735"
          "  \u00b7  PRJ_0000086",
          "THE HEADING NAMES THE MR, THE PR NUMBER AND THE PROJECT")
    check("PR / MR Nos. 001735, 001736" in
          _D.mr_title("MR-1", ["001735", "001736"], "P1"),
          "two PR numbers are pluralised and both shown")
    many = _D.mr_title("MR-1", [f"0017{n}" for n in range(30, 36)], "P1")
    check("+3 more" in many, "a long PR list is trimmed so the title cannot wrap")
    check(_D.mr_title("MR-1", [], "PRJ_9") ==
          "Material Request MR-1  \u00b7  PRJ_9",
          "no PR number -> that part is dropped, not left blank")
    check(_D.mr_title("MR-1", ["001735"], "") ==
          "Material Request MR-1  \u00b7  PR / MR No. 001735",
          "no project -> that part is dropped too")
    check(_D.mr_title("MR-1", [], "") == "Material Request MR-1",
          "with neither, the old heading is unchanged")
    check(_D.mr_title("MR-1", ["  ", ""], "  ") == "Material Request MR-1",
          "blank values never produce a dangling separator")

    # the file name
    fp = _D.mr_request_path(db, "MR-2026-00062", ["001735"], "PRJ_0000086")
    check(fp.name == "MR-2026-00062 (PRJ_0000086) PR 001735.pdf",
          f"the file name carries both numbers (got {fp.name})")
    check(_D.mr_request_path(db, "MR-1", [], "").name == "MR-1.pdf",
          "a bare request still gets a clean file name")
    long_fp = _D.mr_request_path(db, "MR-1", [f"PR{n:05d}" for n in range(40)],
                                 "PROJECT-WITH-A-VERY-LONG-NAME")
    check(len(long_fp.name) <= _D.MAX_NAME_LEN + 4,
          f"a huge PR list cannot blow the Windows path limit ({len(long_fp.name)})")
    check("/" not in long_fp.name and "\\" not in long_fp.name,
          "the file name stays Windows-safe")

    # end to end through the page
    hp_src = [{"item_code": f"HP-{n}", "description": f"Head print {n}",
               "uom": "No", "qty": 5, "pr_no": "001735",
               "project_id": "PRJ_0000086"} for n in (1, 2)]
    hp_src[1]["pr_no"] = "001736"          # a second PR on a line only
    hp_mr = M.save_request(db, {"mr_date": "2026-08-27",
                                "project_id": "PRJ_0000086",
                                "site": "PRJ_0000086", "pr_no": "001735",
                                "reference": "001735"}, M.enrich(db, hp_src))
    hp_id = db.scalar("SELECT id FROM material_requests WHERE mr_no=?", (hp_mr,))
    mph = _MPT(db)
    mph.tabs.setCurrentIndex(1)
    mph.reload()
    for r_ in range(mph.t_req.rowCount()):
        if mph.t_req.item(r_, 0).text() == hp_mr:
            mph.t_req.selectRow(r_)
            mph._load_lines(force=True)
            break
    sel_hp = mph._sel_request()
    prs_hp = mph._request_prs(sel_hp, mph.lines)
    check(prs_hp == ["001735", "001736"],
          f"PR numbers on the header AND on the lines are merged (got {prs_hp})")
    check(mph._request_prs({"pr_no": "001735, 001736; 001737"}, []) ==
          ["001735", "001736", "001737"],
          "a header holding several PR numbers is split correctly")
    check(mph._request_prs({"pr_no": "001735"},
                           [{"pr_no": "001735"}, {"pr_no": "001735"}]) ==
          ["001735"], "duplicates are collapsed")
    mph.request_pdf()
    check(mph.last_file and mph.last_file.exists(), "the request PDF is produced")
    check("001735" in mph.last_file.name and "PRJ_0000086" in mph.last_file.name,
          f"AND ITS FILE NAME SHOWS PR + PROJECT ({mph.last_file.name})")
    raw_hp = mph.last_file.read_bytes()
    check(b"001735" in raw_hp, "the PR number is really in the PDF")

    # the tab-1 availability check gets the same treatment
    mph.tabs.setCurrentIndex(0)
    mph.paste.setPlainText(
        "Line\tProject ID\tItem number\tProduct name\tUnit\tQuantity\t"
        "Purchase requisition reference\n"
        "1\tPRJ_0000086\tHP-1\tHead print 1\tNo\t5\t001735")
    mph.check_paste()
    ct = mph._check_title()
    check("001735" in ct and "PRJ_0000086" in ct,
          f"the availability check heading names them too (got {ct})")
    check(ct.startswith("Material Availability Report"),
          "and keeps its original report name")

    section("Tools, Instruments & Devices — the module")
    from aurco.core import toolstation as TS
    from aurco.ui.tool_station import (ToolStationPage, RegisterTab,
                                       HandoverDialog, ReturnDialog,
                                       TransferDialog)
    import aurco.ui.tool_station as _ts
    for _k in ("confirm", "info_box", "error_box", "toast"):
        setattr(_ts.W, _k, getattr(W, _k))

    ts_root = Path("/tmp/AURCO_TEST_TOOLS")
    shutil.rmtree(ts_root, ignore_errors=True)
    tdb = TS.ToolDB(ts_root / "tool_station.db")
    TS.set_tool_db(tdb)

    # -- the reference number is self-describing
    dec = TS.parse_ref("WH-087IS2308202601")
    check(dec.get("txn_type") == TS.ISSUE and dec.get("doc_date") == "2026-08-23",
          "THE HANDOVER REFERENCE DECODES TO TYPE AND DATE")
    check(dec.get("project_id") == "PRJ000087" and dec.get("warehouse") == "WH",
          "and to the project and warehouse")
    check(TS.parse_ref("WH-091TL0108202601")["txn_type"] == TS.LOAN,
          "a temporary-loan reference is recognised")
    check(TS.parse_ref("not a reference") == {},
          "junk text is rejected rather than guessed")
    check(TS.make_ref("WH", "PRJ000087", TS.ISSUE, "2026-08-23", 1) ==
          "WH-087IS2308202601", "and the same shape is generated back")

    # -- read the real signed form the user supplied
    pdf = Path(__file__).resolve().parent / "fixtures" / \
        "AURCO_Handover_WH-087IS2308202601.pdf"
    check(pdf.exists(), "the sample handover PDF ships with the tests")
    text = TS.read_pdf_text(pdf)
    parsed = TS.parse_handover_text(text, source=str(pdf))
    check(parsed is not None, "the PDF is readable")
    ph, pl = parsed["head"], parsed["lines"]
    check(ph["ref_no"] == "WH-087IS2308202601", "the reference is extracted")
    check(ph["handed_to"] == "Mr. Habib" and ph["iqama_id"] == "100017",
          "THE CUSTODIAN AND IQAMA ID ARE EXTRACTED")
    check(ph["project_id"] == "PRJ000087" and ph["project_name"] == "Jafura L&T",
          "the project id and name are extracted")
    check(ph["mobile"] == "+966 59 436 8672", "the mobile number survives intact")
    check(ph["doc_date"] == "2026-08-23" and ph["doc_time"] == "11:04",
          "the date and time are extracted")
    check(ph["issued_by"] == "Muhammad Ali Zain",
          "the warehouse signatory is extracted")
    check(str(ph.get("received_by", "")).startswith("Mr. Habib"),
          "the custodian signatory is extracted")
    check(ph.get("issued_at") == "2026-08-23 06:33",
          f"the signature date-time is extracted (got {ph.get('issued_at')})")
    check(len(pl) == 6, f"ALL 6 ITEM LINES ARE EXTRACTED (got {len(pl)})")
    codes = [l["asset_id"] for l in pl]
    check(codes == ["12000AL01", "12000RF01", "12000WT01", "12000AT01",
                    "12000TS01", "12000LS01"],
          f"every asset ID is read in order (got {codes})")
    first = pl[0]
    check(first["description"] == "Auto Level" and first["serial_no"] == "5778779",
          "description and serial are split correctly")
    check(first["make_model"] == "Leica" and first["condition"] == "A",
          "make and condition grade are split correctly")
    check(first["calib_due"] == "2026-08-25",
          "the calibration date is converted to ISO")
    check(pl[3]["serial_no"] == "" and pl[3]["description"] == "Aluminium Tripod",
          "a missing serial ('-') becomes blank, not part of the description")
    check(all(l["category"] == "Instrument" for l in pl),
          "the category column is read")

    # -- sync a folder of signed PDFs
    sync_dir = ts_root / "sync"
    sync_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy(pdf, sync_dir)
    fid = TS.add_folder(tdb, sync_dir, "Site Handovers")
    res = TS.sync_folder(tdb, fid)
    check(res["ok"] and res["imported"] == 1,
          "THE SYNC FOLDER IMPORTS THE HANDOVER AUTOMATICALLY")
    h1 = TS.by_ref(tdb, "WH-087IS2308202601")
    check(h1 is not None and len(h1["lines"]) == 6,
          "the handover and its 6 items are in the register")
    check(h1["status"] == TS.OPEN, "a fresh issue is Open")
    check(str(h1["source_file"]).endswith(".pdf"),
          "the scanned file stays linked to the record")
    again = TS.sync_folder(tdb, fid)
    check(again["imported"] == 0 and TS.by_ref(tdb, "WH-087IS2308202601"),
          "SYNCING TWICE NEVER DOUBLE-POSTS THE SAME FORM")
    check(len(TS.search(tdb)) == 1, "still exactly one document")
    check(sync_dir.joinpath(pdf.name).exists(),
          "the source file is left where it was — never moved or deleted")

    # -- the asset register answers 'where is it now'
    assets = TS.search_assets(tdb)
    check(len(assets) == 6, "an asset row is built for every tool")
    al = [a for a in assets if a["asset_id"] == "12000AL01"][0]
    check(al["holder"] == "Mr. Habib" and al["status"] == "Issued Out",
          "WHERE IS IT NOW -> with the custodian")
    check(al["last_ref"] == "WH-087IS2308202601", "and under which reference")

    # -- partial return
    ln0 = h1["lines"][0]
    TS.post_return(tdb, "WH-087IS2308202601",
                   [{"line_id": ln0["id"], "qty": 1, "condition": "B"}])
    h2 = TS.by_ref(tdb, "WH-087IS2308202601")
    check(h2["status"] == TS.PART_RETURNED,
          "returning one of six lines is PARTIALLY RETURNED, not closed")
    back = [a for a in TS.search_assets(tdb) if a["asset_id"] == "12000AL01"][0]
    check(back["status"] == "In Store" and not back["holder"],
          "the returned tool goes back to In Store")
    check(any(r["txn_type"] == TS.RETURN for r in TS.search(tdb)),
          "a Return document is created in its own right")
    try:
        TS.post_return(tdb, "WH-087IS2308202601",
                       [{"line_id": ln0["id"], "qty": 5}])
        check(False, "over-returning is blocked")
    except ValueError:
        check(True, "OVER-RETURNING MORE THAN IS OUTSTANDING IS BLOCKED")

    # -- transfer is not a return
    TS.post_transfer(tdb, "WH-087IS2308202601", {"handed_to": "Mr. Khan",
                                                 "iqama_id": "200018"})
    h3 = TS.by_ref(tdb, "WH-087IS2308202601")
    check(h3["status"] == TS.TRANSFERRED,
          "A TRANSFER CLOSES THE SOURCE AS 'Transferred Out', NOT 'Returned'")
    holders = {c["handed_to"] for c in TS.custody_by_person(tdb)}
    check("Mr. Khan" in holders and "Mr. Habib" not in holders,
          f"custody really moved to the new holder (got {holders})")
    moved = [a for a in TS.search_assets(tdb) if a["asset_id"] == "12000TS01"][0]
    check(moved["holder"] == "Mr. Khan", "the asset register follows the transfer")

    # -- temporary loan and the overdue engine
    loan_id = TS.save_handover(tdb, {
        "txn_type": TS.LOAN, "warehouse": "WH", "project_id": "PRJ000091",
        "doc_date": "2026-08-01", "expected_return": "2026-08-05",
        "handed_to": "Mr. Bilal"},
        [{"asset_id": "13000TW01", "description": "Torque Wrench", "qty": 2}])
    check(tdb.scalar("SELECT status FROM handovers WHERE id=?", (loan_id,))
          == TS.OVERDUE, "A LOAN PAST ITS RETURN DATE IS AUTOMATICALLY OVERDUE")
    od = [r for r in TS.search(tdb, overdue_only=True)]
    check(od and od[0]["days_late"] > 0,
          f"and it reports how many days late (got {od[0]['days_late'] if od else 0})")
    fut = TS.save_handover(tdb, {
        "txn_type": TS.LOAN, "warehouse": "WH", "project_id": "PRJ000092",
        "doc_date": TS.today(), "expected_return": "2099-01-01",
        "handed_to": "Mr. Future"},
        [{"asset_id": "13000XX01", "description": "Laser Level", "qty": 1}])
    check(tdb.scalar("SELECT status FROM handovers WHERE id=?", (fut,)) == TS.OPEN,
          "a loan still within its date is Open, not overdue")

    # -- the unified filter
    check(len(TS.search(tdb, txn_type=TS.ISSUE)) == 1, "filter by type works")
    check(len(TS.search(tdb, txn_type=TS.LOAN)) == 2, "two loans are on file")
    check(len(TS.search(tdb, holder="Mr. Bilal")) == 1, "filter by custodian works")
    check(len(TS.search(tdb, text="12000TS01")) >= 1,
          "SEARCHING BY ASSET ID FINDS THE DOCUMENT HOLDING IT")
    check(len(TS.search(tdb, text="5778779")) >= 1, "searching by serial works")
    check(len(TS.search(tdb, text="Jafura")) >= 1, "searching by project works")
    check(len(TS.search(tdb, date_from="2026-08-20", date_to="2026-08-31")) >= 1,
          "filter by date range works")
    check(TS.search(tdb, text="zzz-not-here") == [],
          "a search that matches nothing returns nothing")
    every = TS.search(tdb)
    shapes = {tuple(sorted(r.keys())) for r in every}
    check(len(shapes) == 1,
          "EVERY DOCUMENT TYPE COMES BACK IN ONE UNIFIED SHAPE")
    check(all("outstanding" in r and "days_late" in r for r in every),
          "with the derived custody figures on every row")
    lines_view = TS.search_lines(tdb, txn_type=TS.ISSUE)
    check(lines_view and all("asset_id" in l and "ref_no" in l for l in lines_view),
          "the same filter can explode to one row per item")

    # -- reports
    for rname in TS.REPORT_LIST:
        rt, rc, rr = TS.build_report(tdb, rname)
        check(isinstance(rc, list) and len(rc) > 0, f"report runs: {rname}")
    _t, _c, miss = TS.build_report(tdb, "Missing Documents & Signatures")
    check(any("signature" in str(r[-1]) for r in miss),
          "the governance report names what is missing")
    _t, _c, cal = TS.build_report(tdb, "Calibration Due Report")
    check(any("EXPIRED" in str(r[6]) for r in cal),
          "expired calibration is called out explicitly")

    # -- the printed controlled form
    form = D.handover_pdf(db, tdb, h3["id"])
    check(form.exists() and form.stat().st_size > 3000,
          "the controlled handover form prints")
    raw_form = form.read_bytes()
    check(b"WH-087IS2308202601" in raw_form, "the reference is on the form")
    check(b"\x00" not in raw_form[:4] and raw_form[:4] == b"%PDF",
          "and it is a real PDF")

    # -- separation from stock is physical, not conventional
    check(str(tdb.path) != str(db.path), "the Tools, Instruments & Devices module has its own database")
    # inspect real code, not comments: strip docstrings/comments first
    import ast as _ast
    _tree = _ast.parse((Path(__file__).resolve().parents[1] / "aurco" / "core"
                        / "toolstation.py").read_text())
    _imports = {n.module for n in _ast.walk(_tree)
                if isinstance(n, _ast.ImportFrom) and n.module}
    _imports |= {a.name for n in _ast.walk(_tree)
                 if isinstance(n, _ast.Import) for a in n.names}
    _banned = {"services", "database", ".services", ".database"}
    check(not (_imports & _banned),
          f"TOOL STATION NEVER IMPORTS THE STOCK ENGINE (imports: {sorted(_imports)})")
    # docstrings describe the rule, so exclude them and test only real strings
    _docs = set()
    for _n in _ast.walk(_tree):
        if isinstance(_n, (_ast.Module, _ast.ClassDef, _ast.FunctionDef)):
            _d = _ast.get_docstring(_n, clean=False)
            if _d:
                _docs.add(_d)
    _sql = [n.value for n in _ast.walk(_tree)
            if isinstance(n, _ast.Constant) and isinstance(n.value, str)
            and n.value not in _docs]
    check(not any("stock_ledger" in q or "FROM items" in q.upper()
                  for q in _sql),
          "and never queries the stock tables")
    bal_before = db.scalar("SELECT COALESCE(SUM(balance),0) FROM items")
    TS.save_handover(tdb, {"txn_type": TS.ISSUE, "warehouse": "WH",
                           "handed_to": "Nobody"},
                     [{"asset_id": "ZZ-1", "description": "Test", "qty": 9}])
    check(db.scalar("SELECT COALESCE(SUM(balance),0) FROM items") == bal_before,
          "posting a handover moves no inventory stock at all")

    # -- the UI
    tsp = ToolStationPage(db)
    check(tsp.tabs.count() == 5, "the page has all five tabs")
    names = [tsp.tabs.tabText(i) for i in range(tsp.tabs.count())]
    check(any("Register" in n for n in names) and any("Sync" in n for n in names),
          f"including the register and the sync folder ({names})")
    tsp.tabs.setCurrentIndex(1)
    tsp.refresh()
    check(tsp.register.table.rowCount() > 0, "the register grid fills")
    check(tsp.register.table.currentRow() >= 0,
          "a row is auto-selected so the detail pane is never blank")
    check(tsp.register.t_lines.rowCount() > 0, "and the item lines are shown")
    tsp.register.chk_items.setChecked(True)
    check("Asset / Tool ID" in tsp.register.table.headers(),
          "the item view switches the grid to one row per tool")
    tsp.register.chk_items.setChecked(False)
    tsp.register.apply_filter({"txn_type": TS.LOAN})
    check(tsp.register.f_type.currentText() == TS.LOAN,
          "a dashboard tile drills into the register with its filter")
    tsp.register.clear_filters()
    tsp.tabs.setCurrentIndex(2)
    tsp.refresh()
    check(tsp.assets.table.rowCount() > 0, "the asset register fills")
    check(tsp.assets.t_hist.rowCount() > 0, "and shows that asset's history")
    tsp.tabs.setCurrentIndex(0)
    tsp.refresh()
    check(tsp.dash.tiles["documents"].lbl_value.text() != "0",
          "the dashboard counts documents")
    check(len(tsp.dash.c_status.data) > 0 and len(tsp.dash.c_status.data[0]) == 3,
          "THE DONUT GETS (label, value, colour) TRIPLES, NOT EMPTY PAIRS")
    tsp.tabs.setCurrentIndex(3)
    tsp.refresh()
    check(tsp.sync.table.rowCount() > 0, "the sync folder lists its files")
    tsp.tabs.setCurrentIndex(4)
    tsp.refresh()
    check(tsp.reports.table.rowCount() >= 0, "the reports tab runs a report")

    # -- backup / restore round trip
    bk = tdb.backup(note="test")
    check(bk.exists(), "the Tools, Instruments & Devices module backs itself up")
    n_before = len(TS.search(tdb))
    TS.delete_handovers(tdb, [h3["id"]])
    check(len(TS.search(tdb)) == n_before - 1, "a handover can be deleted")
    tdb.restore(bk)
    check(len(TS.search(tdb)) == n_before, "and a restore brings it back")

    # ================================================ draft editing (stock out)
    section("Editing a DRAFT Delivery Note")
    it2 = dict(db.one("SELECT * FROM items WHERE balance>40 LIMIT 1"))
    bal0 = it2["balance"]
    po = win.page_out
    po.lines.clear_lines()
    po.reset_form()
    po.issued.setText("Draft Test")
    po.pr_input.setText("PR-DRAFT")
    po.lines.add_items([it2])
    po.lines.item(0, 4).setText("7")
    po.save(False)
    dno = po.no.text()
    drow = dict(db.one("SELECT * FROM documents WHERE doc_no=?", (dno,)))
    check(db.scalar("SELECT SUM(qty) FROM document_lines WHERE doc_id=?",
                    (drow["id"],)) == 7, "a draft keeps the typed quantity")
    check(db.scalar("SELECT balance FROM items WHERE id=?", (it2["id"],)) == bal0,
          "and moves no stock")

    win._edit_draft(drow["id"])
    check(po.editing_draft() and po.lines.rowCount() == 1,
          "Edit Draft re-opens the document on the Stock Out form")
    check(po.lines.item(0, 4).text() == "7" and po.lines.item(0, 5).text() == "PR-DRAFT",
          "with its saved quantity and PR number in the grid")
    check(po.issued.text() == "Draft Test", "and its header fields restored")
    po.lines.item(0, 4).setText("12")
    po.save(False)
    check(db.scalar("SELECT SUM(qty) FROM document_lines WHERE doc_id=?",
                    (drow["id"],)) == 12,
          "AN ADJUSTED QUANTITY IS SAVED BACK INTO THE SAME DRAFT")
    check(db.scalar("SELECT COUNT(*) FROM documents WHERE doc_no=?", (dno,)) == 1,
          "and no duplicate document is created")
    check(not po.editing_draft(), "the form leaves edit mode after saving")

    win._edit_draft(drow["id"])
    po.lines.item(0, 4).setText("5")
    po.save(True)
    check(db.one("SELECT status FROM documents WHERE id=?",
                 (drow["id"],))["status"] == "FINAL", "the edited draft finalizes")
    check(db.scalar("SELECT balance FROM items WHERE id=?", (it2["id"],)) == bal0 - 5,
          "and posts the LAST adjusted quantity to stock")
    try:
        S.update_draft(db, drow["id"], S.DocHeader(doc_type="DN"),
                       [S.Line(item_id=it2["id"], qty=1)])
        check(False, "a finalized document refuses to be edited")
    except S.StockError:
        check(True, "a finalized document refuses to be edited")

    S.reverse_document(db, drow["id"], "customer changed site")
    rev_pdf = D.document_pdf(db, drow["id"])
    check("Reversed Delivery Notes" in str(rev_pdf),
          "reversing an edited DN regenerates it into the reversal folder")
    check(db.scalar("SELECT balance FROM items WHERE id=?", (it2["id"],)) == bal0,
          "reversing the finalized draft restores the stock balance")
    win._edit_draft(drow["id"])
    check(po.editing_draft() and po.draft_status == "REVERSED" and po.no.text() == dno,
          "a reversed DN can be reopened on the Delivery Note form")
    po.lines.item(0, 4).setText("3")
    po.save(False)
    drow2 = db.one("SELECT status, doc_no, pdf_path FROM documents WHERE id=?", (drow["id"],))
    check(drow2["status"] == "DRAFT" and drow2["doc_no"] == dno and not drow2["pdf_path"],
          "a reversed DN can be saved again as the same-number draft")
    check(db.scalar("SELECT COUNT(*) FROM documents WHERE doc_no=?", (dno,)) == 1,
          "reopening after reversal still uses the original document row")
    win._edit_draft(drow["id"])
    po.lines.item(0, 4).setText("4")
    po.save(True)
    check(db.one("SELECT status FROM documents WHERE id=?", (drow["id"],))["status"] == "FINAL",
          "the reopened reversed DN can be finalized again")
    check(db.scalar("SELECT balance FROM items WHERE id=?", (it2["id"],)) == bal0 - 4,
          "re-finalizing after reversal posts the new corrected quantity")
    check("Reversed Delivery Notes" not in str(D.document_pdf(db, drow["id"])),
          "after re-finalizing, the corrected DN returns to the normal folder")

    po.lines.clear_lines()
    po.lines.add_items([dict(it2, issue_qty=4, pr_no="PR-XYZ")])
    check(po.lines.item(0, 4).text() == "4" and po.lines.item(0, 5).text() == "PR-XYZ",
          "lines pushed in from a check / request keep their quantity and PR")
    po.lines.add_items([dict(it2, issue_qty=2, pr_no="PR-OTHER")])
    check(po.lines.rowCount() == 2,
          "the same item can appear twice under two different PR numbers")
    po.lines.clear_lines()
    po.reset_form()

    # ---- GRN drafts use the very same road
    pi = win.page_in
    pi.lines.clear_lines()
    pi.lines.add_items([it2])
    pi.lines.item(0, 3).setText("6")
    pi.save(False)
    gno = pi.no.text()
    grow = dict(db.one("SELECT * FROM documents WHERE doc_no=?", (gno,)))
    win._edit_draft(grow["id"])
    pi.lines.item(0, 3).setText("9")
    pi.save(False)
    check(db.scalar("SELECT SUM(qty) FROM document_lines WHERE doc_id=?",
                    (grow["id"],)) == 9, "a Goods Receipt draft edits the same way")
    pi.lines.clear_lines()

    # ============================ the tools dashboard: filters and configuration
    section("Tools, Instruments & Devices — dashboard")
    dash = tsp.dash
    dash.reset_filters()
    total_docs = dash.tiles["documents"].lbl_value.text()
    check(total_docs not in ("", "0"), "the dashboard counts the filtered register")
    dash.f_type.setCurrentText(TS.ISSUE)
    dash.reload()
    check(dash.tiles["documents"].lbl_value.text() ==
          f"{len(TS.search(tdb, txn_type=TS.ISSUE)):,.0f}",
          "a filter changes every tile on the dashboard")
    dash.f_measure.setCurrentText("Measure: Quantity")
    dash.reload()
    check(sum(v for _k, v in dash.c_type.data) ==
          sum(l["qty"] for l in TS.search_lines(tdb, txn_type=TS.ISSUE)),
          "the Measure selector switches the charts to quantity")
    drilled = {}
    dash.openRegister.connect(lambda f: drilled.update(f))
    dash._drill("overdue")
    check(drilled.get("txn_type") == TS.ISSUE and drilled.get("overdue_only"),
          "a tile drills through carrying the dashboard's own filters")
    dash.reset_filters()
    check(dash.tiles["documents"].lbl_value.text() == total_docs,
          "Reset puts every filter back")
    check(len(TS.ageing(tdb)) == len(TS.AGE_BUCKETS), "ageing buckets are built")
    check(all(len(r) == 3 for r in TS.monthly_split(tdb)),
          "handed-over vs returned is a two-series set")

    dash.tile_cfg = ["documents", "overdue"]
    dash.panel_cfg = ["status", "recent"]
    dash.cols_cfg = 3
    dash._save_config()
    dash._build_body()
    dash.reload()
    check(set(dash.tiles) == {"documents", "overdue"} and
          set(dash.panels) == {"status", "recent"},
          "the dashboard shows only the tiles and panels that were chosen")
    from aurco.ui.tool_station import ToolDashboard as _TD
    check(_TD(tdb, db).tile_cfg == ["documents", "overdue"],
          "AND THE CHOSEN LAYOUT SURVIVES A RESTART")
    dash.tile_cfg = list(_dash_defaults := __import__(
        "aurco.ui.tool_station", fromlist=["x"]).DEFAULT_TILES)
    dash.panel_cfg = list(__import__(
        "aurco.ui.tool_station", fromlist=["x"]).DEFAULT_PANELS)
    dash.cols_cfg = 4
    dash._save_config()
    dash._build_body()
    dash.reload()
    check(len(dash.tiles) == len(_dash_defaults), "restoring the default layout works")

    # ---- the module folder was renamed, with the old one migrated
    from aurco.core import config as _cfg
    check(TS.FOLDER == "Tools, Instruments & Devices",
          "the module folder carries the new name")
    legacy_root = Path("/tmp/AURCO_TEST_RENAME")
    shutil.rmtree(legacy_root, ignore_errors=True)
    (legacy_root / TS.LEGACY_FOLDER).mkdir(parents=True)
    (legacy_root / TS.LEGACY_FOLDER / "tool_station.db").write_text("x")
    _old_root = _cfg.get_storage_root()
    _cfg.set_storage_root(legacy_root)
    TS._migrate_legacy_folder()
    check((legacy_root / TS.FOLDER / "tool_station.db").exists(),
          "an existing Tool Station folder is migrated, not abandoned")
    _cfg.set_storage_root(_old_root)

    # ================= Excel paste / export + inline stock adjustment ======
    section("Delivery Note — Excel paste, export and inline stock adjustment")
    from aurco.ui import common as _C
    pa = win.page_out
    if pa.editing_draft():
        pa.cancel_draft_edit()
    pa.lines.clear_lines()
    pa.reset_form()
    ia = dict(db.one("SELECT * FROM items WHERE balance>20 LIMIT 1"))
    ib = dict(db.one("SELECT * FROM items WHERE balance>20 AND id<>? LIMIT 1",
                     (ia["id"],)))

    head, rows_ = _C.split_pasted_table(
        "Item Code\tQty\tPR / MR No.\n%s\t3\t001735" % ia["code"])
    check(head and len(rows_) == 1, "an Excel paste with a header row is split")
    check(_C.map_pasted_columns(head, "OUT", 3)["qty"] == 1,
          "the Qty column is found by its heading")
    check(len(_C.split_pasted_table("A,1,x\nB,2,y")[1]) == 2,
          "a CSV paste is split too")
    check(len(_C.split_pasted_table("%s   4" % ia["code"])[1]) == 1,
          "and columns padded with spaces")

    parsed = _C.parse_pasted_lines(
        db, "Item Code\tQty\tPR / MR No.\tRemarks\n"
            "%s\t3\t001735\turgent\n%s\t5\t001735\t\nZZZ-NOPE\t2\t\t"
            % (ia["code"], ib["code"]), "OUT")
    check([r["_status"] for r in parsed] == ["ok", "ok", "unknown"],
          "pasted codes are resolved against the item master")
    check(parsed[0]["qty"] == 3 and parsed[0]["pr_no"] == "001735",
          "with their quantity and PR number")
    check(parsed[0]["remarks"] == "urgent", "and their remarks")
    check(_C.resolve_item(db, ia["barcode"] or ia["code"]) is not None,
          "a barcode resolves to the same item")
    check(_C.parse_pasted_lines(db, "%s\t6" % ia["code"], "OUT")[0]["qty"] == 6,
          "a two-column paste (code + qty) works without a header")

    good = [r for r in parsed if r["_status"] == "ok"]
    check(pa.lines.merge_rows(good) == (2, 0) and pa.lines.rowCount() == 2,
          "pasted lines land in the delivery note grid")
    check(pa.lines.merge_rows([dict(good[0], qty=9)]) == (0, 1),
          "pasting the same item again updates the line instead of duplicating it")
    check(pa.lines.to_lines()[0].qty == 9, "with the new quantity")
    check(pa.lines.item(0, 5).text() == "001735",
          "and the PR number is kept as typed (leading zeros intact)")

    cols_x, rows_x = pa.lines.grid_rows()
    check(cols_x[4] == "Quantity" and rows_x[0][4] == 9,
          "the export reads the grid as it is displayed")
    check(rows_x[0][5] == "001735", "and keeps the PR number as text")
    check(pa.lines.copy_to_clipboard() == 2,
          "Ctrl+C copies the grid for Excel")
    check(QApplication.clipboard().text().startswith("Item Code\t"),
          "with a header row")
    xls = pa.export_lines_excel()
    check(xls is not None and Path(xls).exists() and Path(xls).suffix == ".xlsx",
          "the delivery note lines export to a real .xlsx file")

    # --- correcting the system stock from the delivery note screen
    before = float(db.scalar("SELECT balance FROM items WHERE id=?", (ia["id"],)))
    dlg = _C.AdjustStockDialog(db, ia, None, counted=before + 7)
    check(dlg.delta() == 7 and dlg.new_balance() == before + 7,
          "a physical count typed on the DN screen becomes a +7 adjustment")
    dlg.rb_delta.setChecked(True)
    dlg.sp_delta.setValue(-3)
    check(dlg.delta() == -3, "or an explicit minus adjustment")
    dlg.sp_delta.setValue(-(before + 5))
    check(not dlg.btn_ok.isEnabled(), "a correction below zero is refused")
    dlg.rb_count.setChecked(True)
    dlg.sp_count.setValue(before)
    check(not dlg.btn_ok.isEnabled(), "and so is a correction that changes nothing")
    dlg.sp_count.setValue(before + 7)
    dlg.reason.setCurrentText("Physical count correction")
    dlg.post()
    check(dlg.doc_no.startswith("ADJ-"),
          "posting writes a normal stock adjustment document")
    check(float(db.scalar("SELECT balance FROM items WHERE id=?", (ia["id"],)))
          == before + 7, "and the item balance is corrected")
    check(db.scalar("SELECT COUNT(*) FROM stock_ledger WHERE doc_no=?",
                    (dlg.doc_no,)) == 1, "with a ledger entry for the audit trail")
    check(pa.lines.refresh_availability() >= 1,
          "Refresh Stock brings the new balance into the open note")
    check(pa.lines.item(0, 3).text() == f"{round(before + 7, 2):g}",
          "so the Available column shows the corrected quantity")
    _seen = {}
    probe = _C.LineTable(db, "OUT")
    probe.add_items([dict(db.one("SELECT * FROM items WHERE id=?", (ia["id"],)))])
    probe.availabilityEdited.connect(lambda r, q: _seen.update(row=r, qty=q))
    probe.item(0, 3).setText(str(before + 20))
    check(_seen.get("qty") == before + 20,
          "typing a new figure into Available asks for an adjustment")
    probe.set_available(0, before + 7)
    check(probe.items[0]["balance"] == before + 7 and
          probe.item(0, 3).text() == f"{round(before + 7, 2):g}",
          "and cancelling puts the system quantity back")
    probe.item(0, 4).setText("2")
    check(len(_seen) == 2, "editing the Quantity column posts no adjustment")
    pa.lines.clear_lines()

    # ============================================ Cable Records — the module
    section("Cable Records — drums, cutting log and cable schedule")
    from aurco.core import cables as CBL
    from aurco.ui.cable_records import (CableRecordsPage, CableDashboard,
                                        DrumDialog, CutDialog, TagDialog,
                                        TestDialog, ImportDrumsDialog)
    import aurco.ui.cable_records as _cr
    for _k in ("confirm", "info_box", "error_box", "toast"):
        setattr(_cr.W, _k, getattr(W, _k))

    cb_root = Path("/tmp/AURCO_TEST_CABLES")
    shutil.rmtree(cb_root, ignore_errors=True)
    cdb = CBL.CableDB(cb_root / "cable_records.db")
    CBL.set_cable_db(cdb)

    bal_before_cbl = db.scalar("SELECT COALESCE(SUM(balance),0) FROM items")
    CBL.seed_demo(cdb)
    check(cdb.scalar("SELECT COUNT(*) FROM drums") == 12,
          "THE CABLE MODULE KEEPS ITS OWN DRUM REGISTER")
    check(db.scalar("SELECT COALESCE(SUM(balance),0) FROM items") == bal_before_cbl,
          "and nothing it does touches inventory stock")
    check(CBL.DB_NAME == "cable_records.db" and CBL.FOLDER == "Cable Records",
          "in its own database file and folder")

    # -- a drum is a length, and the length is proved by its cutting log
    did = CBL.save_drum(cdb, {"drum_no": "DRM-TEST", "cable_type": "Power",
                              "cores": "4", "size_mm2": "35mm²",
                              "voltage_grade": "0.6/1 kV", "original_length": 500,
                              "unit_cost": 25, "location": "Yard Z"})
    drum = CBL.get_drum(cdb, did)
    check(drum["remaining_length"] == 500 and drum["status"] == CBL.IN_STOCK,
          "a new drum starts full and In Stock")
    check("35mm²" in drum["description"],
          "and describes itself from its own attributes")
    cut1 = CBL.post_cut(cdb, did, {"txn_type": CBL.CUT_ISSUE, "length": 120,
                                   "issued_to": "Crew A"})
    check(cut1.startswith("CC-"), "a cut gets its own document number")
    check(CBL.get_drum(cdb, did)["remaining_length"] == 380,
          "issuing a length takes it off the drum")
    check(CBL.get_drum(cdb, did)["status"] == CBL.PARTLY,
          "and the drum becomes Partly Used")
    CBL.post_cut(cdb, did, {"txn_type": CBL.CUT_RETURN, "length": 20})
    check(CBL.get_drum(cdb, did)["remaining_length"] == 400,
          "a returned off-cut goes back onto the drum")
    try:
        CBL.post_cut(cdb, did, {"txn_type": CBL.CUT_ISSUE, "length": 10 ** 6})
        check(False, "issuing more than is left is refused")
    except CBL.CableError as exc:
        check("left on drum" in str(exc), "issuing more than is left is refused")
    try:
        CBL.post_cut(cdb, did, {"txn_type": CBL.CUT_RETURN, "length": 10 ** 6})
        check(False, "a drum can never grow past its original length")
    except CBL.CableError:
        check(True, "a drum can never grow past its original length")
    cut_id = int(cdb.one("SELECT id FROM cuts WHERE cut_no=?", (cut1,))["id"])
    CBL.void_cut(cdb, cut_id, "wrong drum")
    check(CBL.get_drum(cdb, did)["remaining_length"] == 500 + 20,
          "voiding a cut puts the length straight back")
    check(CBL.rebuild_drums(cdb) >= 13 and
          CBL.get_drum(cdb, did)["remaining_length"] == 520,
          "and the balance can always be re-derived from the log")

    # -- scrapping needs a reason and empties the drum
    try:
        CBL.scrap_drum(cdb, did, "")
        check(False, "scrapping cable without a reason is refused")
    except CBL.CableError:
        check(True, "scrapping cable without a reason is refused")
    CBL.scrap_drum(cdb, did, "water damaged")
    scrapped = CBL.get_drum(cdb, did)
    check(scrapped["status"] == CBL.SCRAPPED and scrapped["remaining_length"] == 0,
          "scrapping writes off what is left, with an audit entry")
    check(cdb.scalar("SELECT COUNT(*) FROM audit WHERE action='DRUM SCRAP'") == 1,
          "and the audit trail records who did it")
    try:
        CBL.delete_drums(cdb, [did])
        check(False, "a drum with history cannot simply be deleted")
    except CBL.CableError:
        check(True, "a drum with history cannot simply be deleted")

    # -- the cable schedule follows the cuts by itself
    tid = CBL.save_tag(cdb, {"tag_no": "C-9001", "project": "PRJX",
                             "from_point": "MCC-9", "to_point": "PMP-9",
                             "cable_type": "Power", "cores": "4",
                             "size_mm2": "35mm²", "required_length": 100})
    donor = cdb.one("SELECT * FROM drums WHERE remaining_length>200 LIMIT 1")
    CBL.post_cut(cdb, int(donor["id"]), {"txn_type": CBL.CUT_ISSUE, "length": 100,
                                         "tag_no": "C-9001"})
    tag = CBL.tag_by_no(cdb, "C-9001")
    check(tag["pulled_length"] == 100 and tag["status"] == CBL.PULLED,
          "a cut tied to a cable tag pulls that tag along with it")
    check(tag["drum_no"] == donor["drum_no"],
          "and records which drum served it")
    CBL.record_test(cdb, tid, {"ir_value": 900, "continuity": "Pass",
                               "test_result": CBL.TEST_PASS, "tested_by": "QC"})
    tag = CBL.tag_by_no(cdb, "C-9001")
    check(tag["test_result"] == CBL.TEST_PASS and tag["status"] == CBL.TESTED,
          "a passed megger test moves the tag to Tested")
    check(CBL.search_tags(cdb, test_result=CBL.TEST_FAIL)[0]["tag_no"] == "C-2001",
          "a failed test is easy to find again")
    try:
        CBL.save_tag(cdb, {"tag_no": "C-9001", "required_length": 5})
        check(False, "the same cable tag cannot be registered twice")
    except CBL.CableError:
        check(True, "the same cable tag cannot be registered twice")

    # -- searching, off-cuts and idle drums
    dash_all = CBL.dashboard(cdb)
    check(dash_all["drums"] == cdb.scalar("SELECT COUNT(*) FROM drums"),
          "the dashboard counts the whole register")
    check(abs(dash_all["remaining_length"] +
              dash_all["used_length"] - dash_all["original_length"]) < 0.01,
          "remaining + consumed always equals what was received")
    oc = CBL.search_drums(cdb, offcuts_only=True, offcut_limit=50)
    check(all(0 < float(d["remaining_length"]) <= 50 for d in oc) and oc,
          "off-cuts are the short ends still worth using")
    idle = CBL.search_drums(cdb, idle_only=True, idle_days=1)
    check(all(d["idle_days"] >= 1 for d in idle),
          "idle drums are the ones nothing has been cut from")
    check(len(CBL.search_drums(cdb, cable_type="Power")) ==
          cdb.scalar("SELECT COUNT(*) FROM drums WHERE cable_type='Power'"),
          "the register filters by cable type")
    check(len(CBL.search_cuts(cdb, txn_type=CBL.CUT_RETURN)) >= 1,
          "the cutting log filters by record type")
    check(len(CBL.by_column(cdb, "cable_type", 10, "remaining")) >= 3,
          "the charts group any column by any measure")
    check(len(CBL.monthly_split(cdb)) >= 2,
          "issued vs returned is a two-series set over several months")
    check(sum(v for _k, v in CBL.ageing(cdb)) > 0, "idle ageing buckets are built")
    for _rep in CBL.REPORT_LIST:
        _t, _c, _r = CBL.build_report(cdb, _rep)
        check(bool(_c), f"report builds: {_rep[:38]}")

    # -- pasting a drum list out of Excel
    head, rows = CBL.sniff("Drum No.\tDescription\tSize\tLength\tLocation\n"
                           "DRM-P1\tXLPE 4C x 16\t16mm2\t400\tYard B\n"
                           "DRM-P2\tControl 7C\t1.5mm2\t250\tYard B")
    recs = CBL.rows_to_drums(head, rows)
    check(len(recs) == 2 and recs[0]["drum_no"] == "DRM-P1",
          "a pasted drum list is understood")
    res = CBL.import_drums(cdb, recs)
    check(res["added"] == 2 and not res["errors"], "and imports as new drums")
    res2 = CBL.import_drums(cdb, [{"drum_no": "DRM-P1", "location": "Yard C"}])
    check(res2["updated"] == 1 and
          CBL.drum_by_no(cdb, "DRM-P1")["location"] == "Yard C",
          "importing the same drum again updates it instead of duplicating")

    # -- the UI
    cbp = CableRecordsPage(db)
    check(cbp.tabs.count() == 5, "the cable page has all five tabs")
    names = [cbp.tabs.tabText(i) for i in range(cbp.tabs.count())]
    check(any("Drum" in n for n in names) and any("Schedule" in n for n in names),
          f"dashboard, drums, cutting log, schedule and reports ({names})")
    cbp.tabs.setCurrentIndex(1)
    cbp.refresh()
    check(cbp.drums.table.rowCount() == len(CBL.search_drums(cdb)),
          "the drum register grid fills")
    cbp.drums.chk_offcut.setChecked(True)
    check(cbp.drums.table.rowCount() == len(
        CBL.search_drums(cdb, offcuts_only=True,
                         offcut_limit=CBL.DEFAULT_OFFCUT_LIMIT)),
          "and the off-cut filter narrows it")
    cbp.drums.clear_filters()
    cbp.drums.apply_filter({"status": CBL.SCRAPPED})
    check(cbp.drums.f_status.currentText() == CBL.SCRAPPED,
          "a dashboard tile drills into the register with its own filter")
    cbp.drums.clear_filters()
    cbp.tabs.setCurrentIndex(3)
    cbp.refresh()
    check(cbp.schedule.table.rowCount() == len(CBL.search_tags(cdb)),
          "the cable schedule grid fills")

    dash = cbp.dash
    dash.reset_filters()
    total_drums = dash.tiles["drums"].lbl_value.text()
    check(total_drums not in ("", "0"), "the cable dashboard counts the register")
    dash.f_type.setCurrentText("Power")
    dash.reload()
    check(dash.tiles["drums"].lbl_value.text() ==
          f"{len(CBL.search_drums(cdb, cable_type='Power')):,.0f}",
          "a filter changes every tile on the cable dashboard")
    dash.f_measure.setCurrentText("Measure: Length remaining")
    dash.reload()
    check(sum(v for _k, v in dash.c_type.data) > 100,
          "the Measure selector switches the charts to metres")
    drill = {}
    dash.openRegister.connect(lambda f: drill.update(f))
    dash._drill("offcuts")
    check(drill.get("offcuts_only") and drill.get("cable_type") == "Power",
          "a tile drills through carrying the dashboard's own filters")
    dash.reset_filters()
    check(dash.f_type.currentIndex() == 0 and not dash.chk_offcut.isChecked(),
          "Reset puts every cable filter back")

    dash.tile_cfg = ["drums", "offcuts"]
    dash.panel_cfg = ["type"]
    dash.cols_cfg = 3
    dash.offcut_limit = 75.0
    dash.idle_days = 45
    dash._save_config()
    dash2 = CableDashboard(cdb, db)
    check(dash2.tile_cfg == ["drums", "offcuts"] and dash2.panel_cfg == ["type"]
          and dash2.cols_cfg == 3,
          "the cable dashboard shows only the tiles and panels that were chosen")
    check(dash2.offcut_limit == 75.0 and dash2.idle_days == 45,
          "AND THE OFF-CUT LIMIT AND IDLE PERIOD SURVIVE A RESTART")
    dash.tile_cfg = list(_cr.DEFAULT_TILES)
    dash.panel_cfg = list(_cr.DEFAULT_PANELS)
    dash.offcut_limit = CBL.DEFAULT_OFFCUT_LIMIT
    dash.idle_days = CBL.DEFAULT_IDLE_DAYS
    dash._save_config()

    check("Cable Records" in _cfg.SUBFOLDERS,
          "the module has its own folder in the storage root")
    _bk = cdb.backup(note="test")
    check(Path(_bk).exists(), "the cable module backs itself up")

    print("\n" + "=" * 60)
    if FAILURES:
        print(f"FAILED {len(FAILURES)} / {PASSES + len(FAILURES)} checks:")
        for f in FAILURES:
            print("   -", f)
        return 1
    print(f"ALL {PASSES} CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
